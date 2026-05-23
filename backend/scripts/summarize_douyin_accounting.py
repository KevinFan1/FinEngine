"""Summarize Excel rows by month, shop name, subject, and reclassification.

Usage:
    cd backend
    python -m scripts.summarize_douyin_accounting /path/to/input.xlsx
    python -m scripts.summarize_douyin_accounting /path/to/input.xlsx --month 202604
    python -m scripts.summarize_douyin_accounting /path/to/input.xlsx --output /path/to/output.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import xlrd


REQUIRED_HEADERS = [
    "账户",
    "月份",
    "抖店名称",
    "科目",
    "动账金额",
    "备注",
    "重分类",
    "主体",
]

GROUP_HEADERS = ["月份", "抖店名称", "科目", "重分类", "动账金额合计"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="筛选指定月份，并按抖店名称、科目、重分类汇总动账金额。"
    )
    parser.add_argument("input_file", type=Path, help="输入 Excel 文件路径")
    parser.add_argument(
        "--month",
        default="202604",
        help="要筛选的月份，默认 202604",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="要读取的工作表名称，默认读取第一个工作表",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 Excel 文件路径，默认生成到输入文件同目录",
    )
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def parse_amount(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")

    # Handle accounting-style negatives like (123.45).
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"无法解析动账金额: {value!r}") from exc


def find_header_indexes(header_row: Iterable[object]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for idx, cell_value in enumerate(header_row):
        header = normalize_text(cell_value)
        if header:
            indexes[header] = idx

    missing = [header for header in REQUIRED_HEADERS if header not in indexes]
    if missing:
        raise ValueError(f"缺少必要表头: {', '.join(missing)}")
    return indexes


def build_default_output_path(input_file: Path, month: str) -> Path:
    return input_file.with_name(f"{input_file.stem}_{month}_汇总.xlsx")


def load_sheet_rows(input_file: Path, sheet_name: str | None) -> list[tuple[object, ...]]:
    suffix = input_file.suffix.lower()

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(filename=input_file, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        return list(worksheet.iter_rows(values_only=True))

    if suffix == ".xls":
        workbook = xlrd.open_workbook(input_file)
        worksheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)
        return [tuple(worksheet.row_values(row_idx)) for row_idx in range(worksheet.nrows)]

    raise ValueError(f"暂不支持的文件类型: {input_file.suffix}")


def summarize_rows(input_file: Path, sheet_name: str | None, month: str) -> tuple[list[list[object]], int, int]:
    rows = load_sheet_rows(input_file, sheet_name)
    if not rows:
        raise ValueError("Excel 文件为空，未读取到任何内容。")

    header_indexes = find_header_indexes(rows[0])
    month_text = str(month).strip()
    grouped: defaultdict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    total_data_rows = 0
    filtered_rows = 0

    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        total_data_rows += 1
        row_month = normalize_text(row[header_indexes["月份"]])
        if row_month != month_text:
            continue

        filtered_rows += 1
        shop_name = normalize_text(row[header_indexes["抖店名称"]])
        subject = normalize_text(row[header_indexes["科目"]])
        reclassify = normalize_text(row[header_indexes["重分类"]])
        amount = parse_amount(row[header_indexes["动账金额"]])
        grouped[(shop_name, subject, reclassify)] += amount

    summary_rows: list[list[object]] = []
    for shop_name, subject, reclassify in sorted(grouped):
        summary_rows.append(
            [
                month_text,
                shop_name,
                subject,
                reclassify,
                float(grouped[(shop_name, subject, reclassify)]),
            ]
        )

    return summary_rows, total_data_rows, filtered_rows


def write_output(output_file: Path, summary_rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总结果"
    worksheet.append(GROUP_HEADERS)

    for row in summary_rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=5).number_format = "#,##0.00"

    column_widths = {
        "A": 12,
        "B": 28,
        "C": 24,
        "D": 20,
        "E": 18,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def main() -> None:
    args = parse_args()
    input_file = args.input_file.expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_file}")

    output_file = (
        args.output.expanduser().resolve()
        if args.output is not None
        else build_default_output_path(input_file, str(args.month).strip())
    )

    summary_rows, total_data_rows, filtered_rows = summarize_rows(
        input_file=input_file,
        sheet_name=args.sheet,
        month=str(args.month).strip(),
    )
    write_output(output_file, summary_rows)

    print(f"读取完成: {input_file}")
    print(f"数据行数: {total_data_rows}")
    print(f"月份={args.month} 的行数: {filtered_rows}")
    print(f"汇总分组数: {len(summary_rows)}")
    print(f"输出文件: {output_file}")


if __name__ == "__main__":
    main()
