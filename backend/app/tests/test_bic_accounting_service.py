from decimal import Decimal
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
import pytest

from app.api.v1.bic_accounting import _normalize_bic_summary
from app.models.bic_accounting import BicDetail, BicSourceRow, BicTask, BicUploadFile
from app.services.oss_service import OSSObjectUnavailableError, SOURCE_FILE_UNAVAILABLE_MESSAGE
from app.services.bic_accounting_service import (
    BIC_RECONCILIATION_SOURCE_EXPORT_HEADERS,
    BicAccountingService,
    _dedupe_bic_rows,
    _service_provider_filter,
    _upsert_source_rows_by_flow,
)
from app.services.transaction_accounting_service import TransactionAccountingService
from app.tasks.processors.douyin import DOUYIN_BIC_HEADERS


def _write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _row(headers: list[str], **overrides: object) -> list[object]:
    values = {header: None for header in headers}
    values.update(overrides)
    return [values[header] for header in headers]


def _source_model(
    *,
    task_id: int,
    file_id: int,
    detail_id: int,
    amount: str,
    flow_no: str,
    settlement_no: str = "SETTLE-1",
) -> BicSourceRow:
    return BicSourceRow(
        task_id=task_id,
        file_id=file_id,
        detail_id=detail_id,
        org_id=3,
        shop_id=5,
        platform_code="douyin",
        shop_name="示例店铺",
        accounting_year=2026,
        accounting_month=5,
        accounting_period=202605,
        service_provider="服务商A",
        qic_warehouse="华东仓",
        source_row_number=2,
        settlement_no=settlement_no,
        order_code="ORDER-1",
        related_order_no="REL-1",
        related_waybill_no="",
        fee_item="质检费(通过)",
        settlement_amount=Decimal(amount),
        billing_params="金额",
        billing_completed_time=datetime(2026, 5, 1, 10, 0, 0),
        business_node="质检通过",
        business_occurred_time=datetime(2026, 5, 1, 9, 59, 59),
        settled_at=datetime(2026, 5, 1, 10, 0, 1),
        status="结算成功",
        transaction_account="货款",
        transaction_flow_no=flow_no,
        remark="",
        is_mudaibao="否",
        is_child_order="否",
    )


class _BicExecuteSession:
    def __init__(self, *, task: BicTask, upload_file: BicUploadFile) -> None:
        self.task = task
        self.upload_file = upload_file
        self.rollback_count = 0

    async def get(self, model: type, _item_id: int):
        if model is BicTask:
            return self.task
        if model is BicUploadFile:
            return self.upload_file
        raise AssertionError(f"unexpected model: {model!r}")

    async def flush(self) -> None:
        return None

    async def rollback(self) -> None:
        self.rollback_count += 1

    async def refresh(self, _instance: object) -> None:
        return None


class _BicPersistSession:
    def __init__(self) -> None:
        self.added_rows: list[object] = []
        self.statements: list[object] = []
        self.execute_count = 0
        self.flush_count = 0
        self._next_id = 1

    async def execute(self, statement: object):
        self.statements.append(statement)
        self.execute_count += 1

        class _Result:
            def scalar(self_inner):
                return 0

            def scalars(self_inner):
                return self_inner

            def all(self_inner):
                return []

            def mappings(self_inner):
                return self_inner

        return _Result()

    def add_all(self, rows: list[object]) -> None:
        self.added_rows.extend(rows)

    async def flush(self) -> None:
        self.flush_count += 1
        for row in self.added_rows:
            if isinstance(row, (BicDetail, BicSourceRow)) and getattr(row, "id", None) is None:
                row.id = self._next_id
                self._next_id += 1


class _BicSourceRowsQuerySession:
    def __init__(self) -> None:
        self.statements: list[object] = []
        self.active_rows: list[BicSourceRow] = []

    async def execute(self, statement: object):
        self.statements.append(statement)

        class _Result:
            def __init__(self_inner, rows: list[BicSourceRow]):
                self_inner._rows = rows

            def scalar(self_inner):
                return 0

            def scalars(self_inner):
                return self_inner

            def all(self_inner):
                return self_inner._rows

            def mappings(self_inner):
                return self_inner

        return _Result(self.active_rows)

    def add_all(self, rows: list[object]) -> None:
        self.active_rows.extend([row for row in rows if isinstance(row, BicSourceRow)])


def test_bic_accounting_parse_file_keeps_only_quality_inspection_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "douyin_bic_qic.xlsx"
    _write_workbook(
        file_path,
        DOUYIN_BIC_HEADERS,
        [
            _row(
                DOUYIN_BIC_HEADERS,
                费用项="质检费(通过)",
                服务商="服务商A",
                QIC仓="华东仓",
                结算金额="12.30",
                计费完成时间="2026-04-25T19:05:14",
                业务发生时间="2026-04-25T19:04:59",
                结算时间="2026-04-25T19:05:15",
            ),
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(拒绝)", 服务商="服务商A", QIC仓="华东仓", 结算金额="99.00"),
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(通过)", 服务商="服务商B", QIC仓="华南仓", 结算金额="7.70"),
        ],
    )

    result = BicAccountingService.parse_file(str(file_path))

    assert result["total_rows"] == 3
    assert result["success_rows"] == 2
    assert result["failed_rows"] == 0
    assert [(row["service_provider"], row["qic_warehouse"], row["amount"]) for row in result["bic_rows"]] == [
        ("服务商A", "华东仓", Decimal("12.30")),
        ("服务商B", "华南仓", Decimal("7.70")),
    ]
    assert result["bic_rows"][0]["billing_completed_time"] == datetime(2026, 4, 25, 19, 5, 14)
    assert result["bic_rows"][0]["business_occurred_time"] == datetime(2026, 4, 25, 19, 4, 59)
    assert result["bic_rows"][0]["settled_at"] == datetime(2026, 4, 25, 19, 5, 15)


def test_bic_accounting_parse_file_marks_missing_header_as_file_error(tmp_path: Path) -> None:
    file_path = tmp_path / "invalid_bic.xlsx"
    _write_workbook(file_path, ["错误表头"], [["x"]])

    result = BicAccountingService.parse_file(str(file_path))

    assert result["total_rows"] == 0
    assert result["success_rows"] == 0
    assert result["failed_rows"] == 1
    assert result["fatal_error"] is True
    assert result["errors"][0].startswith("缺少BIC必要表头:")


def test_bic_accounting_parse_file_explains_no_matching_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "douyin_bic_no_qic.xlsx"
    _write_workbook(
        file_path,
        DOUYIN_BIC_HEADERS,
        [
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(拒绝)", 服务商="服务商A", QIC仓="华东仓", 结算金额="99.00"),
        ],
    )

    result = BicAccountingService.parse_file(str(file_path))

    assert result["total_rows"] == 1
    assert result["success_rows"] == 0
    assert result["failed_rows"] == 0
    assert result["warnings"] == ["未找到费用项为“质检费(通过)”的记录"]


def test_bic_task_summary_hides_removed_report_fields() -> None:
    normalized = _normalize_bic_summary(
        {
            "type": "bic",
            "total_rows": 8,
            "groups": 2,
            "report_groups": 3,
            "report_ids": [1, 2],
            "report_id": 1,
            "报表分组数": 3,
            "报表记录ID列表": [1, 2],
            "首个报表记录ID": 1,
        }
    )

    assert normalized == {
        "文件类型": "BIC",
        "总行数": 8,
        "明细分组数": 2,
    }


@pytest.mark.asyncio
async def test_bic_accounting_result_summary_uses_chinese_keys(tmp_path: Path) -> None:
    file_path = tmp_path / "douyin_bic_qic.xlsx"
    _write_workbook(
        file_path,
        DOUYIN_BIC_HEADERS,
        [
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(通过)", 服务商="服务商A", QIC仓="华东仓", 结算金额="12.30"),
        ],
    )
    task = BicTask(id=1, file_id=2, org_id=3, user_id=4, status="processing", progress=5)
    upload_file = BicUploadFile(
        id=2,
        org_id=3,
        user_id=4,
        shop_id=5,
        original_name="26年02月_bic_抖音旗舰店.xlsx",
        oss_key="oss-key",
        platform_code="douyin",
        shop_name="抖音旗舰店",
        accounting_year=2026,
        accounting_month=2,
        status="uploaded",
    )

    summary = await BicAccountingService.persist_task_result(
        _BicPersistSession(),  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        file_path=str(file_path),
        platform_code="douyin",
        shop_id=5,
        shop_name="抖音旗舰店",
    )

    assert summary == {
        "文件类型": "BIC",
        "总行数": 1,
        "符合条件行数": 1,
        "失败行数": 0,
        "明细分组数": 1,
        "源数据行数": 1,
    }
    assert not any(key in summary for key in ["total_rows", "success_rows", "failed_rows", "detail_ids"])


@pytest.mark.asyncio
async def test_bic_persist_deletes_stale_source_rows_without_expanding_synced_ids(tmp_path: Path) -> None:
    file_path = tmp_path / "douyin_bic_qic.xlsx"
    _write_workbook(
        file_path,
        DOUYIN_BIC_HEADERS,
        [
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(通过)", 服务商="服务商A", QIC仓="华东仓", 结算金额="12.30", 动账流水号="FLOW-1"),
            _row(DOUYIN_BIC_HEADERS, 费用项="质检费(通过)", 服务商="服务商A", QIC仓="华东仓", 结算金额="7.70", 动账流水号="FLOW-2"),
        ],
    )
    task = BicTask(id=1, file_id=2, org_id=3, user_id=4, status="processing", progress=5)
    upload_file = BicUploadFile(
        id=2,
        org_id=3,
        user_id=4,
        shop_id=5,
        original_name="26年02月_bic_抖音旗舰店.xlsx",
        oss_key="oss-key",
        platform_code="douyin",
        shop_name="抖音旗舰店",
        accounting_year=2026,
        accounting_month=2,
        status="uploaded",
    )
    session = _BicPersistSession()

    await BicAccountingService.persist_task_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        file_path=str(file_path),
        platform_code="douyin",
        shop_id=5,
        shop_name="抖音旗舰店",
    )

    statement_sql = "\n".join(str(statement.compile(compile_kwargs={"literal_binds": True})) for statement in session.statements)
    assert "DELETE FROM fin_bic_source_rows WHERE fin_bic_source_rows.task_id = 1" in statement_sql
    assert "NOT IN" not in statement_sql


@pytest.mark.asyncio
async def test_bic_summary_export_rejects_over_row_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_list_details(_db, **_kwargs):
        return [{"id": 1}], 2

    monkeypatch.setattr(BicAccountingService, "list_details", staticmethod(fake_list_details))
    monkeypatch.setattr("app.services.bic_accounting_service.BIC_EXCEL_EXPORT_ROW_LIMIT", 1)

    with pytest.raises(ValueError, match="BIC汇总导出数据量 2 行"):
        await BicAccountingService.export_details(
            None,  # type: ignore[arg-type]
            scope="all",
            user=SimpleNamespace(role="member", org_id=3),
        )


@pytest.mark.asyncio
async def test_bic_summary_current_page_export_rejects_over_row_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_list_details(_db, **_kwargs):
        return [{"id": 1}, {"id": 2}], 2

    monkeypatch.setattr(BicAccountingService, "list_details", staticmethod(fake_list_details))
    monkeypatch.setattr("app.services.bic_accounting_service.BIC_EXCEL_EXPORT_ROW_LIMIT", 1)

    with pytest.raises(ValueError, match="BIC汇总导出数据量 2 行"):
        await BicAccountingService.export_details(
            None,  # type: ignore[arg-type]
            scope="current_page",
            user=SimpleNamespace(role="member", org_id=3),
        )


@pytest.mark.asyncio
async def test_bic_source_export_rejects_over_row_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_list_source_rows(_db, **_kwargs):
        return [{"id": 1}], 2

    monkeypatch.setattr(BicAccountingService, "list_source_rows", staticmethod(fake_list_source_rows))
    monkeypatch.setattr("app.services.bic_accounting_service.BIC_EXCEL_EXPORT_ROW_LIMIT", 1)

    with pytest.raises(ValueError, match="BIC明细导出数据量 2 行"):
        await BicAccountingService.export_source_rows(
            None,  # type: ignore[arg-type]
            scope="selected",
            user=SimpleNamespace(role="member", org_id=3),
            ids=[1, 2],
        )


@pytest.mark.asyncio
async def test_bic_source_current_page_export_rejects_over_row_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_list_source_rows(_db, **_kwargs):
        return [{"id": 1}, {"id": 2}], 2

    monkeypatch.setattr(BicAccountingService, "list_source_rows", staticmethod(fake_list_source_rows))
    monkeypatch.setattr("app.services.bic_accounting_service.BIC_EXCEL_EXPORT_ROW_LIMIT", 1)

    with pytest.raises(ValueError, match="BIC明细导出数据量 2 行"):
        await BicAccountingService.export_source_rows(
            None,  # type: ignore[arg-type]
            scope="current_page",
            user=SimpleNamespace(role="member", org_id=3),
        )


@pytest.mark.asyncio
async def test_bic_reconciliation_export_rejects_over_row_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_list_source_rows(_db, **_kwargs):
        return [{"id": 1}], 2

    monkeypatch.setattr(BicAccountingService, "list_source_rows", staticmethod(fake_list_source_rows))
    monkeypatch.setattr("app.services.bic_accounting_service.BIC_EXCEL_EXPORT_ROW_LIMIT", 1)

    with pytest.raises(ValueError, match="BIC对账表导出数据量 2 行"):
        await BicAccountingService.export_reconciliation(
            None,  # type: ignore[arg-type]
            user=SimpleNamespace(role="member", org_id=3),
            accounting_year=2026,
            accounting_month=5,
            service_provider="服务商A",
        )


@pytest.mark.asyncio
async def test_bic_reconciliation_export_uses_fuzzy_service_provider_filter(monkeypatch: pytest.MonkeyPatch) -> None:
    calls: list[dict[str, object]] = []

    async def fake_list_source_rows(_db, **kwargs):
        calls.append(kwargs)
        return [{"shop_name": "示例店铺", "qic_warehouse": "华东仓", "settlement_amount": Decimal("1.00")}], 1

    monkeypatch.setattr(BicAccountingService, "list_source_rows", staticmethod(fake_list_source_rows))

    await BicAccountingService.export_reconciliation(
        None,  # type: ignore[arg-type]
        user=SimpleNamespace(role="member", org_id=3),
        accounting_year=2026,
        accounting_month=5,
        service_provider="服务商",
    )

    assert calls
    assert calls[0]["service_provider"] == "服务商"
    assert calls[0].get("exact_service_provider") is None


def test_service_provider_filter_uses_fuzzy_multi_value_matching() -> None:
    expr = _service_provider_filter(BicDetail.service_provider, "服务商A，服务商B")

    assert expr is not None
    statement = str(expr.compile(compile_kwargs={"literal_binds": True}))
    assert "LIKE" in statement
    assert "%服务商A%" in statement
    assert "%服务商B%" in statement
    assert " OR " in statement


@pytest.mark.asyncio
async def test_execute_task_marks_source_file_expired_and_preserves_previous_result(monkeypatch: pytest.MonkeyPatch) -> None:
    task = BicTask(
        id=1,
        file_id=2,
        org_id=3,
        user_id=4,
        status="success",
        progress=100,
        processed_rows=6,
        success_rows=6,
        failed_rows=0,
        result_summary={"old": True},
    )
    upload_file = BicUploadFile(
        id=2,
        org_id=3,
        user_id=4,
        original_name="26年02月_bic_抖音旗舰店.xlsx",
        oss_key="oss-key",
        platform_code="douyin",
        shop_name="抖音旗舰店",
        accounting_year=2026,
        accounting_month=2,
        status="processed",
    )
    session = _BicExecuteSession(task=task, upload_file=upload_file)

    monkeypatch.setattr(
        "app.services.bic_accounting_service.oss_service.download_to_temp",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSSObjectUnavailableError("missing")),
    )

    result = await BicAccountingService.execute_task(session, task_id=1)  # type: ignore[arg-type]

    assert session.rollback_count == 1
    assert result.status == "expired"
    assert task.error_message == SOURCE_FILE_UNAVAILABLE_MESSAGE
    assert task.processed_rows == 6
    assert task.success_rows == 6
    assert task.failed_rows == 0
    assert task.result_summary == {"old": True}
    assert upload_file.status == "expired"
    assert upload_file.error_message == SOURCE_FILE_UNAVAILABLE_MESSAGE


def test_bic_detail_workbook_includes_shop_profile_fields() -> None:
    buffer = BicAccountingService._build_detail_workbook(
        [
            {
                "platform_code": "douyin",
                "service_provider": "服务商A",
                "store_short_id": "S001",
                "shop_name": "示例店铺",
                "qic_warehouse": "华东仓",
                "merchant": "示例商户",
                "tax_no": "91310000000000000X",
                "shop_type": "旗舰店",
                "registered_address": "上海市浦东新区张江路 88 号",
                "total_amount": Decimal("123.45"),
            }
        ]
    )
    workbook = load_workbook(buffer)
    sheet = workbook.active

    headers = [sheet.cell(row=1, column=index).value for index in range(1, 12)]
    values = [sheet.cell(row=2, column=index).value for index in range(1, 12)]

    assert headers == ["序号", "平台", "服务商", "店铺id", "店铺名称", "QIC仓", "公司名称", "税号", "抬头类型", "注册地址", "结算金额"]
    assert values == [
        1,
        TransactionAccountingService._format_platform("douyin"),
        "服务商A",
        "S001",
        "示例店铺",
        "华东仓",
        "示例商户",
        "91310000000000000X",
        "旗舰店",
        "上海市浦东新区张江路 88 号",
        123.45,
    ]


def test_bic_reconciliation_workbook_contains_summary_and_shop_source_sheet() -> None:
    buffer = BicAccountingService._build_reconciliation_workbook(
        [
            {
                "shop_id": 5,
                "store_short_id": "S001",
                "service_provider": "服务商A",
                "shop_name": "示例店铺",
                "qic_warehouse": "华东仓",
                "settlement_no": "SETTLE-1",
                "order_code": "ORDER-1",
                "related_order_no": "REL-1",
                "related_waybill_no": "",
                "fee_item": "质检费(通过)",
                "settlement_amount": Decimal("123.45"),
                "billing_params": "金额:123.45(元)",
                "billing_completed_time": "2026-05-01 10:00:00",
                "business_node": "质检通过",
                "business_occurred_time": "2026-05-01 09:59:59",
                "settled_at": "2026-05-01 10:00:01",
                "status": "结算成功",
                "transaction_account": "货款",
                "transaction_flow_no": "SCP-1",
                "remark": "",
                "is_mudaibao": "否",
                "is_child_order": "否",
                "merchant": "示例商户",
                "tax_no": "91310000000000000X",
                "shop_type": "旗舰店",
                "registered_address": "上海市浦东新区张江路 88 号",
            },
            {
                "shop_id": 5,
                "store_short_id": "S001",
                "shop_name": "示例店铺",
                "qic_warehouse": "华东仓",
                "settlement_no": "SETTLE-2",
                "order_code": "ORDER-2",
                "related_order_no": "REL-2",
                "related_waybill_no": "",
                "fee_item": "质检费(通过)",
                "service_provider": "服务商A",
                "settlement_amount": Decimal("1.55"),
                "billing_params": "金额:1.55(元)",
                "billing_completed_time": "2026-05-01 11:00:00",
                "business_node": "质检通过",
                "business_occurred_time": "2026-05-01 10:59:59",
                "settled_at": "2026-05-01 11:00:01",
                "status": "结算成功",
                "transaction_account": "货款",
                "transaction_flow_no": "SCP-2",
                "remark": "",
                "is_mudaibao": "否",
                "is_child_order": "否",
                "merchant": "示例商户",
                "tax_no": "91310000000000000X",
                "shop_type": "旗舰店",
                "registered_address": "上海市浦东新区张江路 88 号",
            }
        ],
        accounting_year=2026,
        accounting_month=5,
        service_provider="服务商A",
    )
    workbook = load_workbook(buffer)
    summary_sheet = workbook["汇总"]
    source_sheet = workbook["明细"]

    summary_headers = [summary_sheet.cell(row=1, column=index).value for index in range(1, 9)]
    summary_values = [summary_sheet.cell(row=2, column=index).value for index in range(1, 9)]
    source_headers = [source_sheet.cell(row=1, column=index).value for index in range(1, 21)]
    source_values = [source_sheet.cell(row=2, column=index).value for index in range(1, 21)]

    assert summary_headers == ["店铺id", "店铺名称", "QIC仓", "公司名称", "税号", "抬头类型", "注册地址", "结算金额"]
    assert summary_values == [
        "S001",
        "示例店铺",
        "华东仓",
        "示例商户",
        "91310000000000000X",
        "旗舰店",
        "上海市浦东新区张江路 88 号",
        125,
    ]
    assert source_headers == BIC_RECONCILIATION_SOURCE_EXPORT_HEADERS
    assert source_values[:9] == ["示例店铺", "SETTLE-1", "ORDER-1", "REL-1", None, "质检费(通过)", "服务商A", "华东仓", 123.45]


def test_bic_reconciliation_workbook_merges_all_rows_into_single_source_sheet() -> None:
    buffer = BicAccountingService._build_reconciliation_workbook(
        [
            {
                "shop_id": 5,
                "shop_name": "同名店铺",
                "qic_warehouse": "华东仓",
                "fee_item": "质检费(通过)",
                "service_provider": "服务商A",
                "settlement_amount": Decimal("10.00"),
                "transaction_flow_no": "FLOW-1",
            },
            {
                "shop_id": 6,
                "shop_name": "同名店铺",
                "qic_warehouse": "华南仓",
                "fee_item": "质检费(通过)",
                "service_provider": "服务商A",
                "settlement_amount": Decimal("20.00"),
                "transaction_flow_no": "FLOW-2",
            },
        ],
        accounting_year=2026,
        accounting_month=5,
        service_provider="服务商A",
    )
    workbook = load_workbook(buffer)

    assert workbook.sheetnames == ["汇总", "明细"]
    assert workbook["明细"].cell(row=2, column=1).value == "同名店铺"
    assert workbook["明细"].cell(row=2, column=17).value == "FLOW-1"
    assert workbook["明细"].cell(row=3, column=1).value == "同名店铺"
    assert workbook["明细"].cell(row=3, column=17).value == "FLOW-2"


def test_bic_detail_rows_group_by_service_provider_and_qic() -> None:
    task = BicTask(id=1, file_id=2, org_id=3, user_id=4, status="success", progress=100)
    upload_file = BicUploadFile(
        id=2,
        org_id=3,
        user_id=4,
        original_name="bic.xlsx",
        oss_key="oss://bic.xlsx",
        file_size=123,
        platform_code="douyin",
        shop_name="示例店铺",
        accounting_year=2026,
        accounting_month=5,
    )

    rows = BicAccountingService._build_detail_rows(
        task=task,
        upload_file=upload_file,
        rows=[
            {"service_provider": "服务商A", "qic_warehouse": "华东仓", "amount": Decimal("1.20"), "raw_row": {}},
            {"service_provider": "服务商A", "qic_warehouse": "华东仓", "amount": Decimal("2.30"), "raw_row": {}},
            {"service_provider": "服务商B", "qic_warehouse": "华东仓", "amount": Decimal("4.50"), "raw_row": {}},
        ],
        platform_code="douyin",
        shop_id=None,
        shop_name="示例店铺",
    )

    assert sorted((row.service_provider, row.qic_warehouse, row.row_count, row.total_amount) for row in rows) == [
        ("服务商A", "华东仓", 2, Decimal("3.50")),
        ("服务商B", "华东仓", 1, Decimal("4.50")),
    ]


def test_bic_source_rows_link_to_detail_rows() -> None:
    task = BicTask(id=1, file_id=2, org_id=3, user_id=4, status="success", progress=100)
    upload_file = BicUploadFile(
        id=2,
        org_id=3,
        user_id=4,
        original_name="bic.xlsx",
        oss_key="oss://bic.xlsx",
        file_size=123,
        platform_code="douyin",
        shop_name="示例店铺",
        accounting_year=2026,
        accounting_month=5,
    )
    detail_rows = BicAccountingService._build_detail_rows(
        task=task,
        upload_file=upload_file,
        rows=[
            {"service_provider": "服务商A", "qic_warehouse": "华东仓", "amount": Decimal("1.20")},
            {"service_provider": "服务商A", "qic_warehouse": "华东仓", "amount": Decimal("2.30")},
            {"service_provider": "服务商B", "qic_warehouse": "华东仓", "amount": Decimal("4.50")},
        ],
        platform_code="douyin",
        shop_id=9,
        shop_name="示例店铺",
    )
    for index, row in enumerate(detail_rows, start=10):
        row.id = index

    source_rows = BicAccountingService._build_source_rows(
        task=task,
        upload_file=upload_file,
        detail_rows=detail_rows,
        rows=[
            {"service_provider": "服务商A", "qic_warehouse": "华东仓", "amount": Decimal("1.20"), "fee_item": "质检费(通过)", "transaction_flow_no": "SCP-1"},
            {
                "service_provider": "服务商A",
                "qic_warehouse": "华东仓",
                "amount": Decimal("2.30"),
                "fee_item": "质检费(通过)",
                "transaction_flow_no": "SCP-2",
                "billing_completed_time": "2026-04-25T19:05:14",
                "business_occurred_time": "2026-04-25T19:04:59",
                "settled_at": "2026-04-25T19:05:15",
            },
            {"service_provider": "服务商B", "qic_warehouse": "华东仓", "amount": Decimal("4.50"), "fee_item": "质检费(通过)", "transaction_flow_no": "SCP-3"},
        ],
        platform_code="douyin",
        shop_id=9,
        shop_name="示例店铺",
    )

    assert sorted((row.service_provider, row.qic_warehouse, row.detail_id, row.settlement_amount) for row in source_rows) == [
        ("服务商A", "华东仓", 10, Decimal("1.20")),
        ("服务商A", "华东仓", 10, Decimal("2.30")),
        ("服务商B", "华东仓", 11, Decimal("4.50")),
    ]
    source_row_with_times = next(row for row in source_rows if row.transaction_flow_no == "SCP-2")
    assert source_row_with_times.billing_completed_time == datetime(2026, 4, 25, 19, 5, 14)
    assert source_row_with_times.business_occurred_time == datetime(2026, 4, 25, 19, 4, 59)
    assert source_row_with_times.settled_at == datetime(2026, 4, 25, 19, 5, 15)


@pytest.mark.asyncio
async def test_list_source_rows_defaults_to_latest_task_scope() -> None:
    session = _BicSourceRowsQuerySession()
    user = SimpleNamespace(role="member", org_id=3)

    rows, total = await BicAccountingService.list_source_rows(
        session,  # type: ignore[arg-type]
        user=user,  # type: ignore[arg-type]
        page=None,
        page_size=None,
    )

    assert rows == []
    assert total == 0
    assert len(session.statements) == 2
    statement_sql = str(session.statements[1].compile(compile_kwargs={"literal_binds": True}))
    assert "fin_bic_source_rows.task_id IN (SELECT max(fin_bic_tasks.id) AS task_id" in statement_sql
    assert "GROUP BY fin_bic_upload_files.org_id" in statement_sql
    assert "fin_bic_upload_files.accounting_month" in statement_sql
    assert "fin_bic_source_rows.shop_id IS NULL" in statement_sql
    assert "EXISTS" in statement_sql
    assert "fin_shops" in statement_sql


@pytest.mark.asyncio
async def test_bic_list_tasks_hides_deleted_shop_uploads() -> None:
    session = _BicSourceRowsQuerySession()
    user = SimpleNamespace(role="member", org_id=3)

    rows, total = await BicAccountingService.list_tasks(
        session,  # type: ignore[arg-type]
        user=user,  # type: ignore[arg-type]
    )

    assert rows == []
    assert total == 0
    statement_sql = "\n".join(
        str(statement.compile(compile_kwargs={"literal_binds": True}))
        for statement in session.statements
    )
    assert "fin_bic_upload_files.shop_id IS NULL" in statement_sql
    assert "EXISTS" in statement_sql
    assert "fin_shops" in statement_sql


@pytest.mark.asyncio
async def test_bic_source_rows_upsert_updates_existing_flow_without_insert() -> None:
    existing = _source_model(
        task_id=1,
        file_id=2,
        detail_id=3,
        amount="1.00",
        flow_no="FLOW-1",
        settlement_no="OLD",
    )
    existing.id = 99
    session = _BicSourceRowsQuerySession()
    session.active_rows = [existing]

    replacement = _source_model(
        task_id=10,
        file_id=20,
        detail_id=30,
        amount="9.99",
        flow_no="FLOW-1",
        settlement_no="NEW",
    )
    inserted = _source_model(
        task_id=10,
        file_id=20,
        detail_id=31,
        amount="2.22",
        flow_no="FLOW-2",
        settlement_no="INSERTED",
    )

    synced_rows = await _upsert_source_rows_by_flow(
        session,  # type: ignore[arg-type]
        platform_code="douyin",
        accounting_period=202605,
        source_rows=[replacement, inserted],
    )

    assert synced_rows == [existing, inserted]
    assert existing.id == 99
    assert existing.task_id == 10
    assert existing.file_id == 20
    assert existing.detail_id == 30
    assert existing.settlement_no == "NEW"
    assert existing.settlement_amount == Decimal("9.99")
    assert inserted in session.active_rows
    assert replacement not in session.active_rows


def test_bic_rows_dedupe_by_transaction_flow_number() -> None:
    rows = _dedupe_bic_rows(
        [
            {"transaction_flow_no": "SCP-1", "amount": Decimal("1.00")},
            {"transaction_flow_no": "SCP-2", "amount": Decimal("2.00")},
            {"transaction_flow_no": "SCP-1", "amount": Decimal("3.00")},
            {"transaction_flow_no": "", "amount": Decimal("4.00")},
        ]
    )

    assert [(row["transaction_flow_no"], row["amount"]) for row in rows] == [
        ("SCP-1", Decimal("3.00")),
        ("SCP-2", Decimal("2.00")),
        ("", Decimal("4.00")),
    ]
