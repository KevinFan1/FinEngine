from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.services.douyin_dongzhang_detail_service import DouyinDongzhangDetailService
from app.services.partition_service import DOUYIN_SOURCE_PARTITION
from app.models.douyin_dongzhang_detail import DouyinDongzhangDetail
from app.services.upload_period_service import get_upload_period_header


def test_douyin_detail_partition_key_uses_source_period_from_upload_period_rule() -> None:
    table_options = DouyinDongzhangDetail.__table_args__[-1]

    assert table_options["postgresql_partition_by"] == "RANGE (source_period)"
    assert DOUYIN_SOURCE_PARTITION.partition_column == "source_period"
    assert DouyinDongzhangDetail.__table__.columns["source_period"].primary_key is True
    assert "summary_period" not in DouyinDongzhangDetail.__table__.columns


def test_douyin_dongzhang_upload_period_rule_uses_transaction_time() -> None:
    assert get_upload_period_header("douyin", "动账") == "动账时间"


def test_serialize_detail_row_formats_money_and_dates() -> None:
    class Row:
        id = 1
        summary_id = 2
        org_id = 3
        shop_id = 4
        shop_name = "抖音旗舰店"
        source_year = 2026
        source_month = 5
        source_period = 202605
        summary_year = 2026
        summary_month = 4
        source_platform_code = "douyin"
        report_platform_code = "douyin"
        period_source = "order_time"
        source_row_number = 9
        matched_compensation = "运费赔付"
        refund_to_compensation = Decimal("10.20")
        cashback = Decimal("3.30")
        order_paid = Decimal("99.50")
        refund_amount = Decimal("18.80")
        gmv = Decimal("80.70")
        platform_income = Decimal("2.10")
        platform_fee_positive = Decimal("4.00")
        return_cost = Decimal("6.60")
        commission_derived = Decimal("7.70")
        bic = Decimal("0")
        insurance_fee = Decimal("0")
        transaction_time = datetime(2026, 5, 2, 10, 0, 0)
        transaction_flow_no = "flow-1"
        transaction_direction = "入账"
        transaction_amount = Decimal("12.30")
        transaction_account = "账户A"
        transaction_scene = "订单结算"
        billing_type = "类型A"
        sub_order_no = "sub-1"
        order_no = "order-1"
        after_sale_no = "after-1"
        order_time = datetime(2026, 4, 20, 12, 0, 0)
        product_id = "prod-1"
        product_code = "CAN123"
        product_name = "商品A"
        author_id = "author-1"
        author_name = "达人A"
        order_type = "普通订单"
        order_paid_amount_raw = Decimal("100.00")
        shipping_fee = Decimal("8.00")
        platform_subsidy_shipping = Decimal("1.00")
        platform_subsidy = Decimal("2.00")
        other_platform_subsidy = Decimal("3.00")
        trade_in_deduction = Decimal("4.00")
        gov_subsidy_platform = Decimal("5.00")
        author_subsidy = Decimal("6.00")
        douyin_pay_subsidy = Decimal("7.00")
        douyin_monthly_subsidy = Decimal("8.00")
        bank_subsidy = Decimal("9.00")
        order_refund_raw = Decimal("10.00")
        platform_fee_raw = Decimal("-11.00")
        commission_raw = Decimal("-12.00")
        provider_commission_raw = Decimal("-13.00")
        channel_share = Decimal("14.00")
        merchant_fee_raw = Decimal("-15.00")
        promotion_fee_raw = Decimal("-16.00")
        other_share = Decimal("17.00")
        is_commission_free = "否"
        commission_free_amount = Decimal("18.00")
        merchant_name = "商户主体A"
        remark = "备注A"

    payload = DouyinDongzhangDetailService.serialize_detail_row(Row(), org_name="组织A", shop_color="#fff")

    assert payload["source_date"] == "2026-05"
    assert payload["source_period"] == 202605
    assert payload["summary_date"] == "2026-04"
    assert payload["org_name"] == "组织A"
    assert payload["platform"] == "douyin"
    assert payload["platform_label"] == "抖音"
    assert payload["transaction_time"] == "2026-05-02 10:00:00"
    assert payload["order_time"] == "2026-04-20 12:00:00"
    assert payload["refund_to_compensation"] == 10.2
    assert payload["platform_fee_raw"] == -11.0
    assert payload["remark"] == "备注A"
    assert payload["major_merchant_name"] == ""
    assert payload["merchant_receipt_subject"] == ""
    assert payload["allocated_bic"] == 0
    assert payload["allocated_insurance_fee"] == 0
    assert payload["live_amount"] == 0


def test_build_export_workbook_writes_header_only() -> None:
    buffer = DouyinDongzhangDetailService.build_export_workbook([], include_header_only=True)

    assert buffer.getbuffer().nbytes > 0


def test_build_export_workbook_header_order_matches_frontend() -> None:
    buffer = DouyinDongzhangDetailService.build_export_workbook([], include_header_only=True)
    workbook = load_workbook(buffer, read_only=True)
    worksheet = workbook["抖音动账源明细"]
    headers = [cell for cell in next(worksheet.iter_rows(values_only=True))]

    assert headers[:6] == [
        "平台",
        "店铺",
        "动账时间",
        "动帐流水号",
        "动账方向",
        "动账金额",
    ]
    assert "调年月(系统的业务年月)" in headers
    assert "商品编码" in headers
    start = headers.index("大商家名称")
    assert headers[start:] == [
        "大商家名称",
        "我方主体",
        "收款商家",
        "分摊BIC",
        "分摊运费险",
        "直播款",
        "商家对账匹配状态",
        "商家对账匹配失败原因",
    ]
    assert "商家收款主体" not in headers


def test_normalize_payload_strips_excel_text_prefix() -> None:
    normalized = DouyinDongzhangDetailService._normalize_payload(
        {
            "source_row_number": 2,
            "summary_year": 2026,
            "summary_month": 4,
            "transaction_time": "2026-04-01 01:17:29",
            "transaction_flow_no": "'2026040101172401977718277000",
            "transaction_direction": "入账",
            "transaction_amount": Decimal("83.16"),
            "transaction_account": "聚合账户",
            "transaction_scene": "货款结算入账",
            "billing_type": "精选联盟",
            "sub_order_no": "'6925040387642719692",
            "order_no": "'6925040387642719692",
            "order_time": "2026-03-22 16:17:29",
        }
    )

    assert normalized["transaction_time"] == datetime(2026, 4, 1, 1, 17, 29)
    assert normalized["order_time"] == datetime(2026, 3, 22, 16, 17, 29)
    assert normalized["transaction_flow_no"] == "2026040101172401977718277000"
    assert normalized["sub_order_no"] == "6925040387642719692"
    assert normalized["order_no"] == "6925040387642719692"
    assert normalized["transaction_amount"] == Decimal("83.16")


def test_normalize_payload_derives_missing_product_code() -> None:
    normalized = DouyinDongzhangDetailService._normalize_payload(
        {
            "source_row_number": 2,
            "summary_year": 2026,
            "summary_month": 4,
            "product_name": "淡水珍珠项链-多样性发一件-约4.5mm-V45054-25（东哥）",
        }
    )

    assert normalized["product_code"] == "V45054"


def test_normalize_payload_keeps_existing_product_code() -> None:
    normalized = DouyinDongzhangDetailService._normalize_payload(
        {
            "source_row_number": 2,
            "summary_year": 2026,
            "summary_month": 4,
            "product_code": "MANUAL123",
            "product_name": "淡水珍珠项链-多样性发一件-约4.5mm-V45054-25（东哥）",
        }
    )

    assert normalized["product_code"] == "MANUAL123"
