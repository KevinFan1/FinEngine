import pytest
from openpyxl import load_workbook

from app.models.shop import Shop
from app.schemas.shop import ShopCreate
from app.services.shop_service import SHOP_IMPORT_HEADERS, ShopService


class _ScalarResult:
    def __init__(self, value):
        self._value = value

    def scalar(self):
        return self._value

    def scalar_one_or_none(self):
        return self._value


class _UploadRaceSession:
    def __init__(self, existing_shop: Shop) -> None:
        self.existing_shop = existing_shop
        self.statements = []

    async def execute(self, stmt):
        self.statements.append(stmt)
        if getattr(stmt, "is_insert", False):
            return _ScalarResult(None)
        if getattr(stmt, "is_select", False):
            if not any(getattr(item, "is_insert", False) for item in self.statements):
                return _ScalarResult(None)
            return _ScalarResult(self.existing_shop)
        raise AssertionError(f"unexpected statement: {stmt!r}")

    def add(self, _instance):
        raise AssertionError("get_or_create_shop must use an atomic insert for upload races")

    async def flush(self):
        raise AssertionError("get_or_create_shop must not rely on select-then-flush")


class _CreateShopRaceSession:
    def __init__(self, inserted_shop: Shop | None) -> None:
        self.inserted_shop = inserted_shop
        self.statements = []

    async def execute(self, stmt):
        self.statements.append(stmt)
        if getattr(stmt, "is_insert", False):
            return _ScalarResult(self.inserted_shop)
        raise AssertionError("create_shop must rely on a single atomic insert")

    def add(self, _instance):
        raise AssertionError("create_shop must not use select-then-add")

    async def flush(self):
        raise AssertionError("create_shop must not use select-then-flush")

    async def refresh(self, _instance):
        raise AssertionError("create_shop should return the inserted row")


class _ListShopsCaptureSession:
    def __init__(self) -> None:
        self.statements = []

    async def execute(self, stmt):
        self.statements.append(stmt)
        if len(self.statements) == 1:
            return _ScalarResult(0)
        return _AllResult([])


class _AllResult:
    def __init__(self, rows):
        self._rows = rows

    def all(self):
        return self._rows


@pytest.mark.asyncio
async def test_get_or_create_shop_uses_atomic_insert_for_upload_races() -> None:
    existing_shop = Shop(
        id=7,
        org_id=1,
        platform_name="douyin",
        shop_name="琢木承珠",
        shop_color="#F97316",
    )
    db = _UploadRaceSession(existing_shop)

    shop = await ShopService.get_or_create_shop(
        db,  # type: ignore[arg-type]
        org_id=1,
        platform_name="douyin",
        shop_name="琢木承珠",
    )

    assert shop is existing_shop
    assert getattr(db.statements[0], "is_insert", False)
    assert any(getattr(item, "is_select", False) for item in db.statements[1:])


@pytest.mark.asyncio
async def test_create_shop_uses_atomic_insert(monkeypatch: pytest.MonkeyPatch) -> None:
    created_shop = Shop(
        id=8,
        org_id=1,
        platform_name="douyin",
        shop_name="并发店铺",
        shop_color="#22C55E",
    )
    db = _CreateShopRaceSession(created_shop)
    audit_calls = []

    async def fake_log(*args, **kwargs):
        audit_calls.append(kwargs)
        return None

    monkeypatch.setattr("app.services.shop_service.AuditService.log", fake_log)

    shop = await ShopService.create_shop(
        db,  # type: ignore[arg-type]
        data=ShopCreate(platform_name="douyin", shop_name="并发店铺"),
        org_id=1,
        operator=type("UserStub", (), {"id": 3, "username": "u3", "display_name": "用户3"})(),
    )

    assert shop is created_shop
    assert len(db.statements) == 1
    assert getattr(db.statements[0], "is_insert", False)
    assert len(audit_calls) == 1


def test_shop_import_template_marks_unique_key_headers() -> None:
    workbook = load_workbook(ShopService.build_import_template())
    sheet = workbook["店铺导入模板"]

    headers = [sheet.cell(row=1, column=index).value for index in range(1, len(SHOP_IMPORT_HEADERS) + 1)]

    assert headers == list(SHOP_IMPORT_HEADERS)
    assert "组织" not in headers
    for column_index in (1, 2):
        comment = sheet.cell(row=1, column=column_index).comment
        assert comment is not None
        assert "唯一" in comment.text


def test_shop_export_workbook_uses_import_headers_and_shop_fields() -> None:
    shop = Shop(
        id=1,
        org_id=2,
        platform_name="douyin",
        shop_name="抖音旗舰店",
        tax_no="91310000X",
        merchant="上海示例商户",
        registered_address="上海市浦东新区",
        legal_person="张三",
        previous_name="旧店铺名",
        store_long_id="LONG-1",
        store_short_id="SHORT-1",
        settlement_period="T+1",
        primary_account="main-account",
        anchor="主播A",
        shop_type="旗舰店",
        purpose="直播",
        former_name="former-a",
    )

    workbook = load_workbook(ShopService.build_export_workbook([(shop, "组织A")]))
    sheet = workbook["店铺信息"]

    headers = [sheet.cell(row=1, column=index).value for index in range(1, len(SHOP_IMPORT_HEADERS) + 1)]
    values = [sheet.cell(row=2, column=index).value for index in range(1, len(SHOP_IMPORT_HEADERS) + 1)]

    assert headers == list(SHOP_IMPORT_HEADERS)
    assert values == [
        "douyin",
        "抖音旗舰店",
        "91310000X",
        "上海示例商户",
        "上海市浦东新区",
        "张三",
        "旧店铺名",
        "LONG-1",
        "SHORT-1",
        "T+1",
        "main-account",
        "主播A",
        "旗舰店",
        "直播",
        "former-a",
    ]


@pytest.mark.asyncio
async def test_list_shops_filters_selected_ids() -> None:
    db = _ListShopsCaptureSession()

    await ShopService.list_shops(
        db,  # type: ignore[arg-type]
        org_id=1,
        ids="3,5",
        page=1,
        page_size=20,
    )

    statement_sql = str(db.statements[1].compile(compile_kwargs={"literal_binds": True}))

    assert "fin_shops.id IN (3, 5)" in statement_sql


@pytest.mark.asyncio
async def test_list_shops_orders_by_updated_at_desc_then_id_desc() -> None:
    db = _ListShopsCaptureSession()

    await ShopService.list_shops(
        db,  # type: ignore[arg-type]
        org_id=1,
        page=1,
        page_size=20,
    )

    statement_sql = str(db.statements[1].compile(compile_kwargs={"literal_binds": True}))

    assert "ORDER BY fin_shops.updated_at DESC, fin_shops.id DESC" in statement_sql


def test_shop_import_header_aliases_keep_chinese_and_english_former_names_distinct() -> None:
    header_map = ShopService._parse_import_headers(("平台", "店铺名", "曾用名", "former_name"))

    assert header_map["previous_name"] == 2
    assert header_map["former_name"] == 3


@pytest.mark.asyncio
async def test_create_shop_reports_duplicate_after_concurrent_insert() -> None:
    db = _CreateShopRaceSession(None)

    with pytest.raises(ValueError, match="已存在同名店铺"):
        await ShopService.create_shop(
            db,  # type: ignore[arg-type]
            data=ShopCreate(platform_name="douyin", shop_name="并发店铺"),
            org_id=1,
            operator=type("UserStub", (), {"id": 3, "username": "u3", "display_name": "用户3"})(),
        )

    assert len(db.statements) == 1
    assert getattr(db.statements[0], "is_insert", False)
