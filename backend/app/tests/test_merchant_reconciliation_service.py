import asyncio
from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from app.models.merchant_reconciliation import MerchantOpeningBalance, MerchantRedSheetPayment, MerchantRedSheetPurchase
from app.services.merchant_reconciliation_exporter import MerchantReconciliationExporter
from app.services.merchant_reconciliation_matcher import MerchantReconciliationMatcher
from app.services.merchant_reconciliation_summary import MerchantReconciliationSummaryBuilder
from app.services.merchant_reconciliation_service import (
    PAYMENT_HEADERS,
    PURCHASE_HEADERS,
    MerchantReconciliationService,
    _ReconciliationLoadContext,
    _RedSheetContext,
)
from app.utils.live_code import extract_live_code


def test_summary_builder_refreshes_bank_statuses() -> None:
    row = MerchantReconciliationSummaryBuilder.empty_summary_row(
        org_id=2,
        org_name="组织A",
        accounting_year=2026,
        accounting_month=4,
        our_subject="我方主体",
        receipt_subject="收款商家",
        adjustment={"paid_flow_amount": Decimal("70.00")},
    )
    row["merchant_payable_net_amount"] = Decimal("100.00")
    row["opening_balance"] = Decimal("10.00")
    row["business_fee_deduction"] = Decimal("5.00")
    row["other_deduction_amount"] = Decimal("2.00")
    row["bank_flow_amount"] = Decimal("33.00")

    MerchantReconciliationSummaryBuilder.refresh_summary_flow_amounts(row)

    assert row["payable_goods_balance"] == Decimal("103.00")
    assert row["unpaid_flow_amount"] == Decimal("33.00")
    assert row["bank_payment_diff"] == Decimal("0.00")
    assert row["bank_status"] == "matched"


def test_summary_builder_allocates_money_by_weight_with_rounding_remainder() -> None:
    allocations = MerchantReconciliationSummaryBuilder.allocate_money_by_weight(
        Decimal("100.00"),
        [
            (("2", "主体A", "商家A"), Decimal("1")),
            (("2", "主体B", "商家B"), Decimal("2")),
            (("2", "主体C", "商家C"), Decimal("3")),
        ],
    )

    assert allocations[("2", "主体A", "商家A")] == Decimal("16.67")
    assert allocations[("2", "主体B", "商家B")] == Decimal("33.33")
    assert allocations[("2", "主体C", "商家C")] == Decimal("50.00")
    assert sum(allocations.values(), Decimal("0.00")) == Decimal("100.00")


def test_exporter_formats_money_dates_and_datetimes() -> None:
    assert MerchantReconciliationExporter.format_export_value(Decimal("12.30"), money=True) == 12.3
    assert MerchantReconciliationExporter.format_export_value(datetime(2026, 4, 1, 9, 8, 7)) == "2026-04-01 09:08:07"


def test_matcher_split_product_codes_supports_plus_chains() -> None:
    assert MerchantReconciliationMatcher.split_product_codes("CAN123+CAN456,T20260001") == [
        "CAN123",
        "CAN456",
        "T20260001",
    ]


def test_red_sheet_template_uses_fixed_month_sheet_names_and_headers() -> None:
    buffer = MerchantReconciliationService.build_red_sheet_template(accounting_year=2026, accounting_month=4)
    workbook = load_workbook(buffer, read_only=True)

    assert "202604货款" in workbook.sheetnames
    assert "202604采购" in workbook.sheetnames
    payment_headers = [cell for cell in next(workbook["202604货款"].iter_rows(values_only=True))]
    purchase_headers = [cell for cell in next(workbook["202604采购"].iter_rows(values_only=True))]
    assert payment_headers == PAYMENT_HEADERS
    assert purchase_headers == PURCHASE_HEADERS
    assert "结算状态" in payment_headers
    assert "备注" not in payment_headers


def test_parse_payment_sheet_stores_remark_as_settlement_status() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(PAYMENT_HEADERS)
    row = [""] * len(PAYMENT_HEADERS)
    row[PAYMENT_HEADERS.index("直播间")] = "直播间A"
    row[PAYMENT_HEADERS.index("直播日期")] = "2026-04-01"
    row[PAYMENT_HEADERS.index("商家")] = "商家A"
    row[PAYMENT_HEADERS.index("收款商家")] = "收款商家A"
    row[PAYMENT_HEADERS.index("结算状态")] = "已结算"
    worksheet.append(row)

    rows = MerchantReconciliationService._parse_payment_sheet(worksheet)

    assert rows[0]["remark"] == "已结算"
    assert rows[0]["raw_row"]["结算状态"] == "已结算"


def test_parse_bank_flow_filename_extracts_year_month() -> None:
    assert MerchantReconciliationService.parse_bank_flow_filename("202601_银行流水_壹韵.xls") == (2026, 1)
    assert MerchantReconciliationService.parse_bank_flow_filename("2026-01_银行流水_壹韵.xls") == (2026, 1)
    assert MerchantReconciliationService.parse_bank_flow_filename("2026年01月_银行流水_邮政.xls") == (2026, 1)
    assert MerchantReconciliationService.parse_bank_flow_filename("银行流水_202601.xls") is None


def test_parse_live_date_from_bank_text_uses_red_sheet_display_format() -> None:
    assert (
        MerchantReconciliationService.parse_live_date_from_bank_text(
            "11月11-20日场次直播款",
            accounting_year=2026,
            accounting_month=1,
        )
        == "11月11-20日"
    )
    assert (
        MerchantReconciliationService.parse_live_date_from_bank_text(
            "2026-01-05直播款",
            accounting_year=2026,
            accounting_month=1,
        )
        == "1月5日"
    )
    assert (
        MerchantReconciliationService.parse_live_date_from_bank_text(
            "2026-01-05转账，3月1-5日直播款",
            accounting_year=2026,
            accounting_month=4,
        )
        == "3月1-5日"
    )


def test_parse_bank_flow_file_supports_standard_export(tmp_path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([
        "账号",
        "账户名称",
        "交易时间",
        "借方发生额（支取）",
        "贷方发生额（收入）",
        "余额",
        "币种",
        "对方户名",
        "对方账号",
        "对方开户机构",
        "记账日期",
        "摘要",
        "备注",
        "账户明细编号-交易流水号",
    ])
    worksheet.append([
        "4401",
        "汕尾市壹韵商贸有限公司",
        "20260105 18:22:21",
        191,
        0,
        445321.58,
        "人民币元",
        "汕尾市城区览幻电子商务店（个人独资）",
        "944153",
        "中国邮政储蓄银行股份有限公司汕尾市分行",
        "20260105",
        "电子转账",
        "11月11-20日直播款",
        "4939",
    ])
    worksheet.append([
        "4401",
        "汕尾市壹韵商贸有限公司",
        "20260106 18:22:21",
        100,
        0,
        445221.58,
        "人民币元",
        "汕尾市城区览幻电子商务店（个人独资）",
        "944153",
        "中国邮政储蓄银行股份有限公司汕尾市分行",
        "20260106",
        "电子转账",
        "11月11-20日货款",
        "4940",
    ])
    path = tmp_path / "202601_银行流水_壹韵.xlsx"
    workbook.save(path)

    rows, bank_name, account_name = MerchantReconciliationService._parse_bank_flow_file(
        file_path=path,
        accounting_year=2026,
        accounting_month=1,
    )

    assert bank_name == "企业网银"
    assert account_name == "汕尾市壹韵商贸有限公司"
    assert len(rows) == 1
    assert rows[0]["counterparty_name"] == "汕尾市城区览幻电子商务店（个人独资）"
    assert rows[0]["transaction_time"].strftime("%Y-%m-%d %H:%M:%S") == "2026-01-05 18:22:21"
    assert rows[0]["live_date"] == "11月11-20日"
    assert rows[0]["flow_amount"] == Decimal("191.00")


def test_parse_bank_flow_file_reads_postal_account_name_from_f2(tmp_path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["中国邮政储蓄银行账户交易明细", "", "", "", "", "", "", "", "", "", "", ""])
    worksheet.append(["账号", "944150013000365212", "", "", "户名", "汕尾市永臻商贸有限公司", "", ""])
    worksheet.append(["收入总金额", "800,000.00", "", "", "支出总金额", "911,912.25", "", ""])
    worksheet.append(["收入总笔数", "1笔", "", "", "支出总笔数", "123笔", "", ""])
    worksheet.append(["起止日期", "2026年01月01日-2026年02月02日", "", "", "数据查询导出时间", "2026年02月02日11:43:13"])
    worksheet.append(["序号", "交易日期", "交易时间", "记账日期", "支出金额", "收入金额", "余额", "对方账号", "对方户名", "对方行名", "用途", "附言", "摘要", "交易流水号"])
    worksheet.append(["1", "2026-01-05", "10:46:05", "2026-01-05", 3600, "0.00", "488,686.32", "944008", "海丰县可塘镇屏苼电子商务商行(个体工商户)", "中国邮政储蓄银行股份有限公司汕尾市分行", "11月21-30日直播款", "11月21-30日直播款", "行内公到公转账", "54947992"])
    worksheet.merge_cells("A1:L1")
    worksheet.merge_cells("F2:H2")
    path = tmp_path / "202601_银行流水_永臻.xlsx"
    workbook.save(path)

    rows, bank_name, account_name = MerchantReconciliationService._parse_bank_flow_file(
        file_path=path,
        accounting_year=2026,
        accounting_month=1,
    )

    assert bank_name == "中国邮政"
    assert account_name == "汕尾市永臻商贸有限公司"
    assert rows[0]["account_name"] == "汕尾市永臻商贸有限公司"
    assert rows[0]["transaction_time"].strftime("%Y-%m-%d %H:%M:%S") == "2026-01-05 10:46:05"
    assert rows[0]["live_date"] == "11月21-30日"


def test_parse_payment_sheet_rejects_unknown_settlement_status() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(PAYMENT_HEADERS)
    row = [""] * len(PAYMENT_HEADERS)
    row[PAYMENT_HEADERS.index("直播间")] = "直播间A"
    row[PAYMENT_HEADERS.index("结算状态")] = "未知状态"
    worksheet.append(row)

    with pytest.raises(ValueError, match="结算状态必须是"):
        MerchantReconciliationService._parse_payment_sheet(worksheet)


def test_red_sheet_payment_payload_exposes_settlement_status() -> None:
    payment = MerchantRedSheetPayment(
        red_sheet_id=1,
        org_id=2,
        shop_id=3,
        accounting_period=202604,
        source_row_number=2,
        remark="已结算",
    )

    payload = MerchantReconciliationService._red_sheet_detail_payload(
        payment,
        org_name="组织A",
        shop_color="#fff",
    )

    assert payload["remark"] == "已结算"
    assert payload["settlement_status"] == "已结算"


def test_build_detail_payload_matches_red_sheet_and_allocates_amounts() -> None:
    class Detail:
        id = 1
        org_id = 2
        shop_id = 3
        shop_name = "抖音旗舰店"
        source_year = 2026
        source_month = 5
        summary_year = 2026
        summary_month = 4
        source_platform_code = "douyin"
        source_row_number = 9
        transaction_time = datetime(2026, 4, 2, 10, 0, 0)
        transaction_flow_no = "flow-1"
        transaction_direction = "入账"
        transaction_amount = Decimal("100")
        transaction_scene = "货款结算入账"
        sub_order_no = "sub-1"
        order_no = "order-1"
        order_time = datetime(2026, 4, 1, 8, 0, 0)
        product_id = "product-1"
        product_code = "CAN123"
        product_name = "商品 CAN123"
        author_name = "达人A"
        merchant_name = "动账商家A"
        gmv = Decimal("100")

    purchase = MerchantRedSheetPurchase(
        red_sheet_id=1,
        org_id=2,
        shop_id=3,
        accounting_period=202604,
        source_row_number=2,
        live_room="直播间A",
        merchant="商家A",
        live_date="2026-04-01",
        live_code="CAN123",
    )
    payment = MerchantRedSheetPayment(
        red_sheet_id=1,
        org_id=2,
        shop_id=3,
        accounting_period=202604,
        source_row_number=2,
        live_room="直播间A",
        live_date="2026-04-01",
        merchant="商家A",
        receipt_subject="收款主体A",
        receipt_merchant="收款商家A",
    )
    context = _RedSheetContext(
        purchases_by_code={"CAN123": purchase},
        payments_by_key={("商家A", "4月1日", "直播间A"): payment},
    )

    payload = MerchantReconciliationService._build_detail_payload(
        Detail(),
        org_name="组织A",
        shop_color="#fff",
        load_context=_ReconciliationLoadContext(
            org_id=2,
            total_gmv=Decimal("200"),
            total_bic=Decimal("20"),
            total_insurance=Decimal("10"),
            red_sheet_context=context,
        ),
    )

    assert payload["match_status"] == "matched"
    assert payload["merchant_name"] == "动账商家A"
    assert payload["major_merchant_name"] == "商家A"
    assert payload["merchant_receipt_subject"] == "收款主体A"
    assert payload["receipt_merchant"] == "收款商家A"
    assert payload["red_sheet_payment_id"] is None
    assert payload["allocated_bic"] == Decimal("10.00")
    assert payload["allocated_insurance_fee"] == Decimal("5.00")
    assert payload["live_amount"] == Decimal("70.00")


def test_build_detail_payload_uses_month_total_gmv_as_allocation_denominator() -> None:
    class Detail:
        id = 1
        org_id = 2
        shop_id = 3
        shop_name = "抖音旗舰店"
        source_year = 2026
        source_month = 5
        summary_year = 2026
        summary_month = 4
        source_platform_code = "douyin"
        source_row_number = 9
        transaction_time = None
        transaction_flow_no = ""
        transaction_direction = ""
        transaction_amount = Decimal("0")
        transaction_scene = ""
        sub_order_no = ""
        order_no = ""
        order_time = None
        product_id = ""
        product_code = "CAN123"
        product_name = "商品 CAN123"
        author_name = ""
        merchant_name = ""
        gmv = Decimal("25")

    payload = MerchantReconciliationService._build_detail_payload(
        Detail(),
        org_name=None,
        shop_color=None,
        load_context=_ReconciliationLoadContext(
            org_id=2,
            total_gmv=Decimal("100"),
            total_bic=Decimal("20"),
            total_insurance=Decimal("8"),
            red_sheet_context=_RedSheetContext(purchases_by_code={}, payments_by_key={}),
        ),
    )

    assert payload["allocated_bic"] == Decimal("5.00")
    assert payload["allocated_insurance_fee"] == Decimal("2.00")


def test_extract_live_code_uses_new_regex_rule() -> None:
    assert extract_live_code("主播场次 ABC123-备用") == "ABC123"
    assert extract_live_code("xx re123 yy") == "RE123"
    assert extract_live_code("V89909-F") == "V89909-F"
    assert extract_live_code("123 V89909-F") == "V89909-F"
    assert extract_live_code("v89909-m") == "V89909-M"
    assert extract_live_code("纯中文无编码") == ""


def test_red_sheet_live_date_is_stored_as_text() -> None:
    assert MerchantReconciliationService._cell_text("2月1-10日") == "2月1-10日"
    assert MerchantReconciliationService._cell_text(datetime(2026, 2, 1, 0, 0, 0)) == "2026-02-01"


def test_match_payment_shop_name_uses_purchase_shop_mapping() -> None:
    purchase_rows = [
        {"source_shop_name": "店铺A", "merchant": "商家A", "live_date": "2026-04-01", "live_room": "直播间A"},
        {"source_shop_name": "店铺B", "merchant": "商家B", "live_date": "2026-04-02", "live_room": "直播间B"},
    ]
    purchase_shop_lookup = MerchantReconciliationService._build_purchase_shop_lookup(purchase_rows)

    matched_shop_name = MerchantReconciliationService._match_payment_shop_name(
        {"merchant": "商家B", "live_date": "2026-04-02", "live_room": "直播间B"},
        purchase_shop_lookup,
    )

    assert matched_shop_name == "店铺B"


def test_unmatched_payment_shop_is_warning_not_import_failure() -> None:
    purchase_rows = [
        {"source_shop_name": "店铺A", "merchant": "商家A", "live_date": "2026-04-01", "live_room": "直播间A"},
    ]
    purchase_shop_lookup = MerchantReconciliationService._build_purchase_shop_lookup(purchase_rows)
    payment_row = {
        "source_row_number": 3,
        "merchant": "恒诺珠宝-B2135",
        "live_date": "2026-04-01",
        "live_room": "东哥直播间",
    }

    assert MerchantReconciliationService._match_payment_shop_name(payment_row, purchase_shop_lookup) == ""

    payment_model = MerchantReconciliationService._payment_model(
        payment_row,
        red_sheet_id=1,
        org_id=2,
        shop_id=None,
        shop_name="",
        accounting_period=202604,
    )

    assert payment_model.shop_id is None
    assert payment_model.shop_name == ""


def test_payment_shop_warning_is_short_and_clear() -> None:
    warning = MerchantReconciliationService._payment_shop_warning(
        {
            "source_row_number": 48,
            "merchant": "恒诺珠宝-B2135",
            "live_room": "东哥直播间",
        }
    )

    assert warning == "货款 sheet 第 48 行未识别到店铺：商家 恒诺珠宝-B2135，直播间 东哥直播间"


def test_split_product_codes_supports_plus_chains() -> None:
    assert MerchantReconciliationService._split_product_codes("CAN123+CAN456,T20260001") == [
        "CAN123",
        "CAN456",
        "T20260001",
    ]


def test_collect_summary_detail_totals_groups_by_dongzhang_merchant_and_receipt_merchant() -> None:
    rows = [
        {
            "org_name": "组织A",
            "shop_name": "店铺A",
            "merchant_name": "动账商家A",
            "our_subject": "红单结算主体A",
            "merchant_receipt_subject": "红单收款主体A",
            "receipt_merchant": "收款商家A",
            "gmv": Decimal("100"),
            "live_amount": Decimal("70"),
            "red_sheet_payment_id": 11,
        },
        {
            "org_name": "组织A",
            "shop_name": "店铺A",
            "merchant_name": "动账商家A",
            "our_subject": "红单结算主体A",
            "merchant_receipt_subject": "红单收款主体A",
            "receipt_merchant": "收款商家A",
            "gmv": Decimal("50"),
            "live_amount": Decimal("35"),
            "red_sheet_payment_id": 11,
        },
    ]

    detail_totals, payment_group_weights = MerchantReconciliationService._collect_summary_detail_totals(rows)

    key = ("组织A", "动账商家A", "收款商家A")
    assert detail_totals[key]["gmv"] == Decimal("150")
    assert detail_totals[key]["merchant_payable_net_amount"] == Decimal("105")
    assert detail_totals[key]["row_count"] == 2
    assert payment_group_weights == {11: {key: Decimal("105")}}


def test_collect_summary_detail_totals_keeps_pre_aggregated_row_count() -> None:
    rows = [
        {
            "org_name": "组织A",
            "shop_name": "店铺A",
            "merchant_name": "动账商家A",
            "receipt_merchant": "收款商家A",
            "gmv": Decimal("300"),
            "live_amount": Decimal("210"),
            "row_count": 3,
            "red_sheet_payment_id": 11,
        },
    ]

    detail_totals, payment_group_weights = MerchantReconciliationService._collect_summary_detail_totals(rows)

    key = ("组织A", "动账商家A", "收款商家A")
    assert detail_totals[key]["gmv"] == Decimal("300")
    assert detail_totals[key]["merchant_payable_net_amount"] == Decimal("210")
    assert detail_totals[key]["row_count"] == 3
    assert payment_group_weights == {11: {key: Decimal("210")}}


def test_build_summary_rows_matches_red_sheet_amounts_to_detail_groups() -> None:
    detail_totals = {
        ("组织A", "动账商家A", "收款商家A"): {
            "gmv": Decimal("150"),
            "merchant_payable_net_amount": Decimal("105"),
            "row_count": 2,
        }
    }
    payment_adjustments = {
        ("组织A", "动账商家A", "收款商家A"): {
            "business_fee_deduction": Decimal("3"),
            "other_deduction_amount": Decimal("2"),
            "payable_goods_balance": Decimal("120"),
        }
    }

    rows = MerchantReconciliationService._build_summary_rows_from_aggregates(
        detail_totals=detail_totals,
        payment_adjustments=payment_adjustments,
        accounting_year=2026,
        accounting_month=4,
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["our_subject"] == "动账商家A"
    assert row["merchant_receipt_subject"] == "收款商家A"
    assert row["receipt_merchant"] == "收款商家A"
    assert row["gmv"] == Decimal("150")
    assert row["merchant_payable_net_amount"] == Decimal("105")
    assert row["business_fee_deduction"] == Decimal("3")
    assert row["other_deduction_amount"] == Decimal("2")
    assert row["payable_goods_balance"] == Decimal("100")
    assert row["unpaid_flow_amount"] == Decimal("100")
    assert row["row_count"] == 2


def test_refresh_summary_flow_amounts_uses_summary_formula() -> None:
    row = {
        "merchant_payable_net_amount": Decimal("105"),
        "opening_balance": Decimal("20"),
        "business_fee_deduction": Decimal("3"),
        "other_deduction_amount": Decimal("2"),
        "paid_flow_amount": Decimal("30"),
        "bank_flow_amount": Decimal("30"),
    }

    MerchantReconciliationService._refresh_summary_flow_amounts(row)

    assert row["payable_goods_balance"] == Decimal("120")
    assert row["unpaid_flow_amount"] == Decimal("90")
    assert row["bank_payment_diff"] == Decimal("60")
    assert row["bank_status"] == "diff"


def test_merge_bank_flow_totals_updates_bank_columns_without_overwriting_paid_amount() -> None:
    row = MerchantReconciliationService._empty_summary_row(
        org_name="组织A",
        accounting_year=2026,
        accounting_month=1,
        our_subject="汕尾市壹韵商贸有限公司",
        receipt_subject="汕尾市城区览幻电子商务店（个人独资）",
        adjustment={
            "payable_goods_balance": Decimal("191"),
            "paid_flow_amount": Decimal("180"),
        },
    )

    MerchantReconciliationService._merge_bank_flow_totals(
        [row],
        {
            (
                "组织A",
                "汕尾市壹韵商贸有限公司",
                "汕尾市城区览幻电子商务店(个人独资)",
            ): Decimal("191"),
        },
    )

    assert row["paid_flow_amount"] == Decimal("180")
    assert row["bank_flow_amount"] == Decimal("191")
    assert row["bank_payment_diff"] == Decimal("-371")
    assert row["bank_status"] == "diff"


def test_merge_opening_balance_rows_updates_formula_and_appends_saved_only_row() -> None:
    row = MerchantReconciliationService._empty_summary_row(
        org_id=2,
        org_name="组织A",
        accounting_year=2026,
        accounting_month=2,
        our_subject="我方主体A",
        receipt_subject="收款商家A",
        adjustment={"business_fee_deduction": Decimal("3")},
    )
    row["merchant_payable_net_amount"] = Decimal("70")
    MerchantReconciliationService._refresh_summary_flow_amounts(row)
    matched_balance = MerchantOpeningBalance(
        org_id=2,
        platform_code="douyin",
        accounting_year=2026,
        accounting_month=2,
        accounting_period=202602,
        our_subject="我方主体A",
        receipt_merchant="收款商家A",
        opening_balance=Decimal("20"),
        created_by=1,
    )
    saved_only_balance = MerchantOpeningBalance(
        org_id=2,
        platform_code="douyin",
        accounting_year=2026,
        accounting_month=2,
        accounting_period=202602,
        our_subject="我方主体B",
        receipt_merchant="收款商家B",
        opening_balance=Decimal("15"),
        created_by=1,
    )

    summary_rows = [row]
    MerchantReconciliationService._merge_opening_balance_rows(
        summary_rows,
        [(matched_balance, "组织A"), (saved_only_balance, "组织A")],
        accounting_year=2026,
        accounting_month=2,
        append_missing=True,
    )

    assert row["opening_balance"] == Decimal("20")
    assert row["payable_goods_balance"] == Decimal("87")
    assert row["unpaid_flow_amount"] == Decimal("87")
    assert len(summary_rows) == 2
    saved_only_row = summary_rows[1]
    assert saved_only_row["our_subject"] == "我方主体B"
    assert saved_only_row["merchant_receipt_subject"] == "收款商家B"
    assert saved_only_row["opening_balance"] == Decimal("15")
    assert saved_only_row["payable_goods_balance"] == Decimal("15")


def test_filter_summary_rows_by_bank_status() -> None:
    rows = [
        {"bank_status": "matched", "key": "a"},
        {"bank_status": "diff", "key": "b"},
        {"bank_status": "pending", "key": "c"},
    ]

    filtered = MerchantReconciliationService._filter_summary_rows_by_bank_status(rows, bank_status="matched")

    assert filtered == [{"bank_status": "matched", "key": "a"}]


def test_allocate_payment_adjustments_by_matched_detail_weight() -> None:
    payment = SimpleNamespace(
        id=11,
        settlement_subject="红单结算主体",
        receipt_subject="红单收款主体",
        receipt_merchant="红单收款商家",
        business_fee_deduction=Decimal("10"),
        deduction_amount=Decimal("1"),
        payable_goods_amount=Decimal("100"),
    )
    rows = [(payment, "组织A")]
    weights = {
        11: {
            ("组织A", "动账商家A", "收款商家A"): Decimal("70"),
            ("组织A", "动账商家B", "收款商家A"): Decimal("30"),
        }
    }

    class Result:
        def all(self):
            return rows

    class DB:
        async def execute(self, _stmt):
            return Result()

    adjustments = asyncio.run(
        MerchantReconciliationService._load_payment_adjustments(
            DB(),
            user=SimpleNamespace(role="admin", org_id=2),
            org_id=None,
            shop_id=None,
            accounting_year=2026,
            accounting_month=4,
            payment_group_weights=weights,
        )
    )

    assert adjustments[("组织A", "动账商家A", "收款商家A")]["payable_goods_balance"] == Decimal("70.00")
    assert adjustments[("组织A", "动账商家B", "收款商家A")]["payable_goods_balance"] == Decimal("30.00")
    assert adjustments[("组织A", "动账商家A", "收款商家A")]["business_fee_deduction"] == Decimal("7.00")
    assert adjustments[("组织A", "动账商家B", "收款商家A")]["business_fee_deduction"] == Decimal("3.00")
    assert adjustments[("组织A", "动账商家A", "收款商家A")]["other_deduction_amount"] == Decimal("0.70")
    assert adjustments[("组织A", "动账商家B", "收款商家A")]["other_deduction_amount"] == Decimal("0.30")
