from datetime import datetime
from decimal import Decimal
from pathlib import Path
import re
from types import SimpleNamespace
from collections.abc import Sequence

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import select
from sqlalchemy.dialects import postgresql

from app.models.reconciliation_checklist import (
    ReconciliationChecklistDetail,
    ReconciliationChecklistOrderKey,
    ReconciliationChecklistPayableBalanceSummaryRow,
    ReconciliationChecklistProductSummaryRow,
    ReconciliationChecklistReceiptSummaryRow,
    ReconciliationChecklistTask,
    ReconciliationChecklistUploadFile,
)
from app.services.partition_maintenance_service import ensure_reconciliation_checklist_partitions_for_year
from app.services.partition_service import (
    RECONCILIATION_CHECKLIST_PAYABLE_BALANCE_SUMMARY_PARTITION,
    partition_name,
)
from app.services.reconciliation_checklist_service import (
    CHECKLIST_EMPTY_SUMMARY_MESSAGE,
    CHECKLIST_FILE_TYPE_INVOICE,
    CHECKLIST_FILE_TYPE_MERCHANT,
    CHECKLIST_FILE_TYPE_SOURCE,
    SOURCE_DETAIL_COMPARE_FIELDS,
    CHECKLIST_TASK_TYPE_INVOICE_EDIT,
    CHECKLIST_TASK_TYPE_MERCHANT_EDIT,
    SOURCE_HEADERS,
    ReconciliationChecklistService,
    _build_header_column_lookup,
    _canonical_header,
    _detect_checklist_type,
    _value_from_row,
    _row_fingerprint,
    _to_int,
)


def _write_workbook(path: Path, sheets: list[tuple[str, list[str], list[list[object]]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


def _source_row(**overrides: object) -> list[object]:
    values: dict[str, object] = {
        "进驻的直播平台": "抖音",
        "结算时间": datetime(2026, 6, 1, 8, 30, 0),
        "子订单号": "SO-1",
        "下单时间": datetime(2026, 5, 31, 20, 0, 0),
        "商品ID": "P-1",
        "商品名称": "商品A",
        "商品数量": 2,
        "达人名称": "达人A",
        "平台补贴": "1.00",
        "达人补贴": "2.00",
        "抖音支付补贴": "3.00",
        "抖音月付营销补贴": "4.00",
        "银行补贴": "5.00",
        "用户实付 （订单金额）": "100.50",
        "平台服务费": "6.00",
        "达人佣金": "7.00",
        "招商服务费": "8.00",
        "商户主体名称": "商户主体A",
        "客服代码": "KF01",
        "收款商家": "收款商家A",
        "直播推广佣金": "9.00",
        "佣金率": "0.10",
        "应付商家净额": "91.50",
        "付款金额": "90.00",
        "付款时间（商家）": datetime(2026, 6, 5, 9, 0, 0),
        "开票时间": datetime(2026, 6, 6, 10, 0, 0),
        "发票号码": "FP-1",
    }
    values.update(overrides)
    return [values[header] for header in SOURCE_HEADERS]


class _PartitionSession:
    def __init__(self) -> None:
        self.statements: list[str] = []
        self.primary_keys_by_table: dict[str, list[str]] = {}
        self.partition_bounds_by_table: dict[tuple[str, str], str] = {}

    async def scalar(self, statement: object, params: dict[str, object] | None = None):
        self.statements.append(str(statement))
        _ = params
        return None

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        sql = str(statement)
        self.statements.append(sql)
        _ = params
        if "SELECT to_regclass" in sql:
            return SimpleNamespace(scalar=lambda: True)
        if "FROM pg_partitioned_table" in sql:
            return SimpleNamespace(scalar=lambda: True)
        if "SELECT obj_description" in sql:
            return SimpleNamespace(scalar=lambda: None)
        if "SELECT a.attname, d.description" in sql:
            return SimpleNamespace(mappings=lambda: [])
        if "FROM pg_index i" in sql:
            table_name = str((params or {}).get("table_name", ""))
            columns = self.primary_keys_by_table.get(table_name, ["id", "org_id", "accounting_period"])
            return SimpleNamespace(scalar=lambda: columns)
        if "FROM pg_inherits i" in sql and "pg_get_expr(child.relpartbound, child.oid, true)" in sql:
            table_name = str((params or {}).get("table_name", ""))
            partition_bound = str((params or {}).get("partition_bound", ""))
            existing_partition = self.partition_bounds_by_table.get((table_name, partition_bound))
            return SimpleNamespace(scalar=lambda: existing_partition)
        return SimpleNamespace(scalar=lambda: None)


class _OrderLookupSession:
    def __init__(self) -> None:
        self.statements: list[str] = []
        self.param_counts: list[int] = []
        self.tuple_param_counts: list[int] = []

    async def execute(self, statement: object):
        compiled = statement.compile(dialect=postgresql.dialect())
        sql = str(compiled)
        self.statements.append(sql)
        expanded_count = 0
        tuple_count = 0
        for value in compiled.params.values():
            if isinstance(value, list):
                if value and isinstance(value[0], tuple):
                    expanded_count += len(value) * len(value[0])
                    tuple_count += len(value) * len(value[0])
                else:
                    expanded_count += len(value)
            else:
                expanded_count += 1
        self.param_counts.append(expanded_count)
        if tuple_count:
            self.tuple_param_counts.append(tuple_count)
        if "FILTER (WHERE" in sql and "fin_reconciliation_checklist_details" in sql:
            return _BatchWriteResult(all_rows=[(2501, 0, 0)])
        empty_result = _BatchWriteResult()
        return SimpleNamespace(
            scalars=empty_result.scalars,
            mappings=empty_result.mappings,
            all=empty_result.all,
            one=empty_result.one,
        )


class _PersistTaskSession:
    def __init__(self) -> None:
        self.flush_count = 0
        self.commit_count = 0

    async def flush(self) -> None:
        self.flush_count += 1

    async def commit(self) -> None:
        self.commit_count += 1


class _RebuildSummarySession:
    def __init__(self) -> None:
        self.statements: list[str] = []

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        self.statements.append(str(statement))
        _ = params
        return SimpleNamespace(rowcount=0)


class _DashboardMetricsResult:
    def __init__(self, rows: list[object]) -> None:
        self._rows = rows

    def all(self):
        return self._rows


class _DashboardMetricsSession:
    def __init__(self) -> None:
        self.scalar_sql: list[str] = []
        self.execute_sql: list[str] = []

    async def scalar(self, statement: object):
        sql = _compile_sql(statement)
        self.scalar_sql.append(sql)
        assert "fin_reconciliation_checklist_details" not in sql
        if "count(*)" in sql and "status IN" in sql:
            return 3
        if "count(*)" in sql and "status = " in sql:
            return 1
        if "count(*)" in sql:
            return 4
        if "sum(fin_reconciliation_checklist_tasks.total_rows)" in sql:
            return 1200
        if "sum(fin_reconciliation_checklist_receipt_summary_rows.total_user_paid_amount)" in sql:
            return Decimal("13579.24")
        if "count(distinct(fin_reconciliation_checklist_receipt_summary_rows.merchant_subject_name))" in sql:
            return 2
        if "count(distinct(fin_reconciliation_checklist_receipt_summary_rows.accounting_period))" in sql:
            return 2
        raise AssertionError(f"unexpected scalar SQL: {sql}")

    async def execute(self, statement: object):
        sql = _compile_sql(statement)
        self.execute_sql.append(sql)
        assert "fin_reconciliation_checklist_details" not in sql
        if "date_part" in sql:
            return _DashboardMetricsResult(
                [
                    SimpleNamespace(month=5, task_count=1),
                    SimpleNamespace(month=6, task_count=2),
                ]
            )
        if "accounting_month" in sql and "GROUP BY" in sql:
            return _DashboardMetricsResult(
                [
                    SimpleNamespace(month=5, total_user_paid_amount=Decimal("100.00")),
                    SimpleNamespace(month=6, total_user_paid_amount=Decimal("300.50")),
                ]
            )
        if "merchant_subject_name" in sql and "LIMIT" in sql:
            return _DashboardMetricsResult(
                [
                    SimpleNamespace(merchant_name="商户主体A", total_user_paid_amount=Decimal("300.50")),
                    SimpleNamespace(merchant_name="商户主体B", total_user_paid_amount=Decimal("100.00")),
                ]
            )
        if "fin_reconciliation_checklist_upload_files" in sql:
            return _DashboardMetricsResult(
                [
                    SimpleNamespace(
                        id=10,
                        original_name="source.xlsx",
                        task_type="source_import",
                        status="success",
                        total_rows=100,
                        success_rows=100,
                        failed_rows=0,
                        inserted_rows=100,
                        finished_at=datetime(2026, 6, 8, 12, 0, 0),
                    )
                ]
            )
        raise AssertionError(f"unexpected execute SQL: {sql}")


class _BatchWriteResult:
    def __init__(
        self,
        *,
        scalar_one_value: object | None = None,
        all_rows: list[tuple[object, ...]] | None = None,
        rowcount: int = 0,
    ) -> None:
        self._scalar_one_value = scalar_one_value
        self._all_rows = all_rows or []
        self.rowcount = rowcount

    def scalar_one(self):
        return self._scalar_one_value

    def one(self):
        return self._all_rows[0] if self._all_rows else ()

    def all(self):
        return self._all_rows

    def scalars(self):
        rows = self._all_rows

        class _Scalars:
            @staticmethod
            def all():
                return rows

        return _Scalars()

    def mappings(self):
        rows = self._all_rows

        class _Mappings:
            @staticmethod
            def all():
                return [row if isinstance(row, dict) else vars(row) for row in rows]

        return _Mappings()


class _BatchWriteSession:
    def __init__(self, *, existing_detail_pairs: list[tuple[object, ...]] | None = None, existing_details: list[object] | None = None) -> None:
        self.statements: list[str] = []
        self.detail_insert_calls = 0
        self.detail_lookup_calls = 0
        self.order_key_insert_calls = 0
        self.update_calls = 0
        self.delete_calls = 0
        self.existing_detail_pairs = existing_detail_pairs or []
        self.existing_details = existing_details or []
        self.commit_count = 0
        self.stage_rows: list[dict[str, object]] = []

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        sql = _compile_sql(statement)
        self.statements.append(sql)
        _ = params

        if sql.lstrip().startswith("CREATE TEMP TABLE"):
            return _BatchWriteResult(rowcount=0)
        if sql.lstrip().startswith("DROP TABLE IF EXISTS"):
            return _BatchWriteResult(rowcount=0)
        if "INSERT INTO tmp_fin_reconciliation_checklist_source_stage" in sql:
            self.stage_rows.extend(list(getattr(statement, "_multi_values", [()])[0]))
            return _BatchWriteResult(rowcount=len(getattr(statement, "_multi_values", [()])[0]))
        if "FILTER (WHERE" in sql and "fin_reconciliation_checklist_details" in sql:
            if self.existing_details:
                changed = 0
                unchanged = len(self.existing_details)
                for stage_row, existing_detail in zip(self.stage_rows, self.existing_details, strict=False):
                    if not self._rows_match(stage_row, existing_detail):
                        changed += 1
                        unchanged -= 1
                inserted = max(len(self.stage_rows) - len(self.existing_details), 0)
                return _BatchWriteResult(all_rows=[(inserted, changed, max(unchanged, 0))])
            inserted = len(self.stage_rows)
            return _BatchWriteResult(all_rows=[(inserted, 0, 0)])
        if "FROM fin_reconciliation_checklist_details JOIN (VALUES" in sql:
            self.detail_lookup_calls += 1
            if self.existing_details:
                return _BatchWriteResult(all_rows=list(self.existing_details))
            return _BatchWriteResult(all_rows=list(self.existing_detail_pairs))
        if "INSERT INTO fin_reconciliation_checklist_details" in sql:
            self.detail_insert_calls += 1
            return _BatchWriteResult(
                scalar_one_value=1000 + self.detail_insert_calls,
                rowcount=2,
            )
        if "INSERT INTO fin_reconciliation_checklist_order_keys" in sql:
            self.order_key_insert_calls += 1
            return _BatchWriteResult(rowcount=2)
        if sql.lstrip().startswith("UPDATE fin_reconciliation_checklist_details SET"):
            self.update_calls += 1
            return _BatchWriteResult(rowcount=2)
        if sql.lstrip().startswith("DELETE FROM fin_reconciliation_checklist_details"):
            self.delete_calls += 1
            return _BatchWriteResult(rowcount=0)
        return _BatchWriteResult(rowcount=0)

    async def commit(self) -> None:
        self.commit_count += 1

    @staticmethod
    def _rows_match(stage_row: dict[str, object], existing_detail: object) -> bool:
        if not stage_row:
            return False
        existing_values = vars(existing_detail)
        return all(existing_values.get(field_name) == stage_row.get(field_name) for field_name in SOURCE_DETAIL_COMPARE_FIELDS)


class _SourceStagingSession:
    def __init__(self, *, counts: tuple[int, int, int]) -> None:
        self.statements: list[str] = []
        self.counts = counts
        self.stage_insert_calls = 0
        self.detail_upsert_calls = 0
        self.commit_count = 0

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        sql = _compile_sql(statement)
        self.statements.append(sql)
        _ = params

        if sql.lstrip().startswith("CREATE TEMP TABLE"):
            return _BatchWriteResult(rowcount=0)
        if sql.lstrip().startswith("DROP TABLE IF EXISTS"):
            return _BatchWriteResult(rowcount=0)
        if "INSERT INTO tmp_fin_reconciliation_checklist_source_stage" in sql:
            self.stage_insert_calls += 1
            return _BatchWriteResult(rowcount=2)
        if "FILTER (WHERE" in sql and "fin_reconciliation_checklist_details" in sql:
            return _BatchWriteResult(all_rows=[self.counts])
        if "INSERT INTO fin_reconciliation_checklist_details" in sql and "SELECT" in sql:
            self.detail_upsert_calls += 1
            return _BatchWriteResult(rowcount=sum(self.counts[:2]))
        return _BatchWriteResult(rowcount=0)

    async def commit(self) -> None:
        self.commit_count += 1


class _CopyCapableSession(_SourceStagingSession):
    def __init__(self, *, counts: tuple[int, int, int], copy_should_fail: bool = False) -> None:
        super().__init__(counts=counts)
        self.copy_should_fail = copy_should_fail
        self.copy_calls = 0
        self.fallback_insert_calls = 0
        self.copy_columns: list[str] = []

    async def _copy_records_to_stage(self, table_name: str, rows: Sequence[dict[str, object]], columns: Sequence[str]) -> bool:
        assert table_name == "tmp_fin_reconciliation_checklist_source_stage"
        self.copy_calls += 1
        self.copy_columns = list(columns)
        if self.copy_should_fail:
            raise RuntimeError("copy failed")
        self.stage_insert_calls += 1
        return True

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        sql = _compile_sql(statement)
        if "INSERT INTO tmp_fin_reconciliation_checklist_source_stage" in sql:
            self.fallback_insert_calls += 1
        return await super().execute(statement, params)


class _ChecklistExecuteSession:
    def __init__(self, *, task: ReconciliationChecklistTask, upload_file: ReconciliationChecklistUploadFile) -> None:
        self.task = task
        self.upload_file = upload_file
        self.commit_count = 0
        self.rollback_count = 0
        self.allow_db_get = True
        self.flush_count = 0

    async def get(self, model: type, _item_id: int):
        if not self.allow_db_get:
            raise AssertionError("db.get called while DB access should be paused")
        if model is ReconciliationChecklistTask:
            return self.task
        if model is ReconciliationChecklistUploadFile:
            return self.upload_file
        raise AssertionError(f"unexpected model: {model!r}")

    async def flush(self) -> None:
        self.flush_count += 1
        return None

    async def commit(self) -> None:
        self.commit_count += 1

    async def rollback(self) -> None:
        self.rollback_count += 1

    async def refresh(self, _instance: object) -> None:
        return None


class _ManualEditQuerySession:
    def __init__(self, rows: list[object]) -> None:
        self.rows = rows
        self.flush_count = 0

    async def execute(self, _statement: object):
        return SimpleNamespace(scalars=lambda: SimpleNamespace(all=lambda: self.rows))

    async def flush(self) -> None:
        self.flush_count += 1


class _ManualEditSaveSession:
    def __init__(
        self,
        *,
        existing_details: list[tuple[int, str]] | None = None,
        existing_summary_values: list[tuple[int, str, str, Decimal | None, Decimal | None]] | None = None,
        detail_rows: list[object] | None = None,
        update_rowcount: int = 0,
    ) -> None:
        self.statements: list[str] = []
        self.flush_count = 0
        self.update_calls = 0
        self.existing_details = existing_details or []
        self.existing_summary_values = existing_summary_values or []
        self.detail_rows = detail_rows or []
        self.update_rowcount = update_rowcount

    async def execute(self, statement: object, params: dict[str, object] | None = None):
        sql = _compile_sql(statement)
        self.statements.append(sql)
        _ = params
        if (
            "row_fingerprint IN" in sql
            or "sub_order_no IN" in sql
            or (
                sql.lstrip().startswith("SELECT")
                and "FROM fin_reconciliation_checklist_details" in sql
                and "merchant_net_amount, fin_reconciliation_checklist_details.payment_amount" not in sql
                and "accounting_period, fin_reconciliation_checklist_details.sub_order_no" not in sql
            )
        ):
            return SimpleNamespace(scalars=lambda: SimpleNamespace(all=lambda: list(self.detail_rows)))
        if "SELECT fin_reconciliation_checklist_details.accounting_period, fin_reconciliation_checklist_details.sub_order_no, fin_reconciliation_checklist_details.receipt_merchant, fin_reconciliation_checklist_details.merchant_net_amount, fin_reconciliation_checklist_details.payment_amount" in sql:
            return SimpleNamespace(all=lambda: list(self.existing_summary_values))
        if "SELECT fin_reconciliation_checklist_details.accounting_period, fin_reconciliation_checklist_details.sub_order_no" in sql:
            return SimpleNamespace(all=lambda: list(self.existing_details))
        if sql.lstrip().startswith("UPDATE fin_reconciliation_checklist_details SET"):
            self.update_calls += 1
            return SimpleNamespace(rowcount=self.update_rowcount)
        return SimpleNamespace(rowcount=0)

    async def flush(self) -> None:
        self.flush_count += 1


class _CreateManualEditTaskSession:
    def __init__(self) -> None:
        self.added: list[object] = []
        self._next_id = 200
        self.info: dict[str, object] = {}
        self.flush_count = 0
        self.objects_by_id: dict[tuple[type[object], int], object] = {}

    def add(self, instance: object) -> None:
        self.added.append(instance)

    async def flush(self) -> None:
        self.flush_count += 1
        for instance in self.added:
            if getattr(instance, "id", None) is None:
                instance.id = self._next_id
                self._next_id += 1
            self.objects_by_id[(type(instance), instance.id)] = instance

    async def get(self, model: type[object], object_id: int) -> object | None:
        return self.objects_by_id.get((model, object_id))

    async def execute(self, _statement: object) -> object:
        return SimpleNamespace(scalars=lambda: SimpleNamespace(first=lambda: None))

    async def refresh(self, _instance: object) -> None:
        return None

    async def run_callbacks(self) -> None:
        from app.core.database import run_after_commit_callbacks

        await run_after_commit_callbacks(self)  # type: ignore[arg-type]


def _compile_sql(statement: object) -> str:
    return str(statement.compile(dialect=postgresql.dialect()))


def _parsed_source_detail_row(*, sub_order_no: str, settlement_time: datetime, source_row_number: int = 2, **overrides: object) -> dict[str, object]:
    values = dict(zip(SOURCE_HEADERS, _source_row(子订单号=sub_order_no, 结算时间=settlement_time, **overrides)))
    return ReconciliationChecklistService._parse_source_row(values, source_row_number)


def _existing_detail_lookup_row(row: dict[str, object], *, org_id: int = 3) -> SimpleNamespace:
    detail = SimpleNamespace(**row)
    detail.org_id = org_id
    detail.row_fingerprint = _row_fingerprint(
        org_id=org_id,
        settlement_time=row["settlement_time"],
        sub_order_no=row["sub_order_no"],
        platform_subsidy=row["platform_subsidy"],
        talent_subsidy=row["talent_subsidy"],
        douyin_pay_subsidy=row["douyin_pay_subsidy"],
        douyin_monthly_pay_subsidy=row["douyin_monthly_pay_subsidy"],
        bank_subsidy=row["bank_subsidy"],
        user_paid_amount=row["user_paid_amount"],
        platform_service_fee=row["platform_service_fee"],
        talent_commission=row["talent_commission"],
        investment_service_fee=row["investment_service_fee"],
    )
    return detail


def test_models_define_order_key_and_hash_partitioned_summary_tables() -> None:
    assert ReconciliationChecklistOrderKey.__tablename__ == "fin_reconciliation_checklist_order_keys"
    order_key_indexes = {index.name: index for index in ReconciliationChecklistOrderKey.__table__.indexes}
    assert [column.name for column in order_key_indexes["uq_fin_reconciliation_checklist_order_key"].columns] == ["org_id", "sub_order_no"]

    assert ReconciliationChecklistDetail.__table_args__[-1]["postgresql_partition_by"] == "RANGE (accounting_period)"
    assert [column.name for column in ReconciliationChecklistDetail.__table__.primary_key.columns] == [
        "id",
        "org_id",
        "accounting_period",
    ]
    detail_columns = ReconciliationChecklistDetail.__table__.columns
    for column_name in [
        "live_platform",
        "settlement_time",
        "sub_order_no",
        "order_time",
        "product_id",
        "product_name",
        "product_quantity",
        "talent_name",
        "merchant_subject_name",
        "receipt_merchant",
        "user_paid_amount",
        "live_commission",
        "merchant_net_amount",
        "merchant_net_balance",
        "payment_amount",
        "merchant_payment_time",
        "invoice_time",
        "invoice_number",
    ]:
        assert column_name in detail_columns

    detail_indexes = {index.name: index for index in ReconciliationChecklistDetail.__table__.indexes}
    assert [column.name for column in detail_indexes["uq_fin_reconciliation_checklist_detail_row_fingerprint"].columns] == [
        "org_id",
        "accounting_period",
        "row_fingerprint",
    ]

    assert ReconciliationChecklistProductSummaryRow.__tablename__ == "fin_reconciliation_checklist_product_summary_rows"
    assert ReconciliationChecklistReceiptSummaryRow.__tablename__ == "fin_reconciliation_checklist_receipt_summary_rows"
    assert ReconciliationChecklistPayableBalanceSummaryRow.__tablename__ == "fin_reconciliation_checklist_payable_balance_summary_rows"
    assert "task_type" in ReconciliationChecklistTask.__table__.columns
    assert ReconciliationChecklistProductSummaryRow.__table_args__[-1]["postgresql_partition_by"] == "RANGE (accounting_period)"
    assert ReconciliationChecklistReceiptSummaryRow.__table_args__[-1]["postgresql_partition_by"] == "RANGE (accounting_period)"
    assert ReconciliationChecklistPayableBalanceSummaryRow.__table_args__[-1]["postgresql_partition_by"] == "RANGE (accounting_period)"
    for model in [
        ReconciliationChecklistProductSummaryRow,
        ReconciliationChecklistReceiptSummaryRow,
        ReconciliationChecklistPayableBalanceSummaryRow,
    ]:
        assert [column.name for column in model.__table__.primary_key.columns] == [
            "id",
            "org_id",
            "accounting_period",
        ]


def test_product_quantity_parse_failure_defaults_to_zero() -> None:
    assert _to_int("") == 0
    assert _to_int("abc") == 0
    assert _to_int("1,234") == 1234


def test_unique_id_treats_null_and_zero_money_as_same_value() -> None:
    settlement_time = datetime(2026, 6, 1, 8, 30, 0)

    zero_unique_id = _row_fingerprint(
        org_id=9,
        settlement_time=settlement_time,
        sub_order_no="SO-1",
        platform_subsidy=Decimal("0"),
        talent_subsidy=Decimal("0"),
        douyin_pay_subsidy=Decimal("0"),
        douyin_monthly_pay_subsidy=Decimal("0"),
        bank_subsidy=Decimal("0"),
        user_paid_amount=Decimal("0"),
        platform_service_fee=Decimal("0"),
        talent_commission=Decimal("0"),
        investment_service_fee=Decimal("0"),
    )
    null_unique_id = _row_fingerprint(
        org_id=9,
        settlement_time=settlement_time,
        sub_order_no="SO-1",
        platform_subsidy=None,
        talent_subsidy=None,
        douyin_pay_subsidy=None,
        douyin_monthly_pay_subsidy=None,
        bank_subsidy=None,
        user_paid_amount=None,
        platform_service_fee=None,
        talent_commission=None,
        investment_service_fee=None,
    )

    assert zero_unique_id == null_unique_id


def test_common_summary_filters_support_selected_row_keys() -> None:
    user = SimpleNamespace(role="superadmin", org_id=1)

    product_filters = ReconciliationChecklistService._common_summary_filters(
        ReconciliationChecklistProductSummaryRow,
        user=user,  # type: ignore[arg-type]
        ids=["1:202606:收款商家A:商户主体A:商品A"],
    )
    receipt_filters = ReconciliationChecklistService._common_summary_filters(
        ReconciliationChecklistReceiptSummaryRow,
        user=user,  # type: ignore[arg-type]
        ids=["1:202606:商户主体A:抖音:收款商家A"],
    )
    payable_filters = ReconciliationChecklistService._common_summary_filters(
        ReconciliationChecklistPayableBalanceSummaryRow,
        user=user,  # type: ignore[arg-type]
        ids=["1:202606:商户主体A:收款商家A"],
    )

    product_sql = _compile_sql(select(ReconciliationChecklistProductSummaryRow).where(*product_filters))
    receipt_sql = _compile_sql(select(ReconciliationChecklistReceiptSummaryRow).where(*receipt_filters))
    payable_sql = _compile_sql(select(ReconciliationChecklistPayableBalanceSummaryRow).where(*payable_filters))

    assert "product_name" in product_sql
    assert "live_platform" in receipt_sql
    assert "receipt_merchant" in payable_sql


@pytest.mark.asyncio
async def test_dashboard_metrics_use_receipt_summary_rows_without_scanning_details() -> None:
    session = _DashboardMetricsSession()

    metrics = await ReconciliationChecklistService.get_dashboard_metrics(
        session,  # type: ignore[arg-type]
        user=SimpleNamespace(role="admin", org_id=1),  # type: ignore[arg-type]
        year=2026,
    )

    sql = "\n".join(session.scalar_sql + session.execute_sql)
    assert "fin_reconciliation_checklist_receipt_summary_rows" in sql
    assert "fin_reconciliation_checklist_details" not in sql
    assert metrics["total_user_paid_amount"] == Decimal("13579.24")
    assert metrics["monthly_user_paid_amounts"][4]["total_user_paid_amount"] == Decimal("100.00")
    assert metrics["top_merchants"][0] == {
        "merchant_name": "商户主体A",
        "total_user_paid_amount": Decimal("300.50"),
    }
    assert metrics["recent_tasks"][0]["original_name"] == "source.xlsx"


def test_checklist_partition_migration_primary_keys_include_org_hash_key() -> None:
    migration_path = (
        Path(__file__).resolve().parents[2]
        / "alembic"
        / "versions"
        / "049_rebuild_reconciliation_checklist_order_base.py"
    )
    source = migration_path.read_text(encoding="utf-8")

    detail_block = _migration_create_table_block(source, "fin_reconciliation_checklist_details")
    assert 'sa.PrimaryKeyConstraint("id", "org_id", "accounting_period")' in detail_block

    summary_loop = source[source.index("for table_name, extra_columns") :]
    assert summary_loop.count('sa.PrimaryKeyConstraint("id", "org_id", "accounting_period")') == 1


def _migration_create_table_block(source: str, table_name: str) -> str:
    match = re.search(
        rf'op\.create_table\(\s*"{re.escape(table_name)}"(?P<body>.*?)\n    \)',
        source,
        re.S,
    )
    assert match, f"create_table block missing for {table_name}"
    return match.group("body")


def test_header_detection_recognizes_three_checklist_types_and_aliases() -> None:
    source_headers = SOURCE_HEADERS.copy()
    source_headers[source_headers.index("商户主体名称")] = "商家主体名称"
    source_headers[source_headers.index("用户实付 （订单金额）")] = "用户实付 (订单金额)"

    assert _detect_checklist_type(source_headers).file_type == CHECKLIST_FILE_TYPE_SOURCE
    assert _detect_checklist_type(["唯一ID", "子订单号", "收款商家", "开票时间", "发票号码"]).file_type == CHECKLIST_FILE_TYPE_INVOICE
    assert _detect_checklist_type(["唯一ID", "子订单号", "收款商家", "应付商家净额", "付款金额", "付款时间（商家）"]).file_type == CHECKLIST_FILE_TYPE_MERCHANT
    assert _canonical_header(" 用户实付    （订单金额） ") == "用户实付 （订单金额）"


def test_header_detection_treats_truncated_payment_time_header_as_source() -> None:
    source_headers = SOURCE_HEADERS.copy()
    source_headers[source_headers.index("付款时间（商家）")] = "付款时间（商家"

    assert _detect_checklist_type(source_headers).file_type == CHECKLIST_FILE_TYPE_SOURCE


def test_header_column_lookup_reads_aliases_directly_from_row() -> None:
    headers = ["系统行定位值", "商家主体名称", "用户实付（订单金额）"]
    row = ["UID-1", "商户主体A", "100.50"]

    lookup = _build_header_column_lookup(headers)

    assert _value_from_row(row, lookup, "唯一ID") == "UID-1"
    assert _value_from_row(row, lookup, "商户主体名称") == "商户主体A"
    assert _value_from_row(row, lookup, "用户实付 （订单金额）") == "100.50"


def test_parse_source_file_does_not_stringify_datetime_cells(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    file_path = tmp_path / "source_datetime_fast_path.xlsx"
    _write_workbook(file_path, [("原始数据-表头", SOURCE_HEADERS, [_source_row(子订单号="SO-1")])])

    def fail_on_datetime(value: object) -> str:
        if isinstance(value, datetime):
            raise AssertionError("datetime cells should not be stringified before parsing")
        if value is None:
            return ""
        return str(value).strip()

    monkeypatch.setattr("app.services.reconciliation_checklist_service.safe_str", fail_on_datetime)

    result = ReconciliationChecklistService.parse_file(str(file_path))

    assert result["failed_rows"] == 0
    assert result["rows"][0]["settlement_time"] == datetime(2026, 6, 1, 8, 30, 0)


def test_parse_source_file_extracts_order_level_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "source.xlsx"
    source_headers_with_balance = SOURCE_HEADERS + ["应付商家净额余额"]
    _write_workbook(
        file_path,
        [
            (
                "原始数据-表头",
                source_headers_with_balance,
                [
                    _source_row(子订单号="SO-1", **{"用户实付 （订单金额）": None}) + ["1.50"],
                    _source_row(子订单号="SO-2", 结算时间="2026/07/02 10:00:00", 商品名称="商品B", **{"用户实付 （订单金额）": None}) + ["2.50"],
                ],
            )
        ],
    )

    result = ReconciliationChecklistService.parse_file(str(file_path))

    assert result["file_type"] == CHECKLIST_FILE_TYPE_SOURCE
    assert result["total_rows"] == 2
    assert result["success_rows"] == 2
    assert result["failed_rows"] == 0
    assert result["fatal_error"] is False
    assert [
        (
            row["sub_order_no"],
            row["accounting_period"],
            row["live_platform"],
            row["merchant_subject_name"],
            row["receipt_merchant"],
            row["user_paid_amount"],
            row["live_commission"],
            row["merchant_net_amount"],
            row["merchant_net_balance"],
        )
        for row in result["rows"]
    ] == [
        ("SO-1", 202606, "抖音", "商户主体A", "收款商家A", Decimal("0.00"), Decimal("9.00"), Decimal("91.50"), Decimal("1.50")),
        ("SO-2", 202607, "抖音", "商户主体A", "收款商家A", Decimal("0.00"), Decimal("9.00"), Decimal("91.50"), Decimal("2.50")),
    ]


def test_parse_source_file_defaults_missing_balance_header_to_zero(tmp_path: Path) -> None:
    file_path = tmp_path / "source_without_balance.xlsx"
    _write_workbook(file_path, [("原始数据-表头", SOURCE_HEADERS, [_source_row(子订单号="SO-1")])])

    result = ReconciliationChecklistService.parse_file(str(file_path))

    assert result["file_type"] == CHECKLIST_FILE_TYPE_SOURCE
    assert result["rows"][0]["merchant_net_balance"] == Decimal("0.00")


def test_parse_source_file_rejects_overflow_commission_rate(tmp_path: Path) -> None:
    file_path = tmp_path / "source_overflow_rate.xlsx"
    _write_workbook(
        file_path,
        [("原始数据-表头", SOURCE_HEADERS, [_source_row(**{"佣金率": "10000"})])],
    )

    result = ReconciliationChecklistService.parse_file(str(file_path))

    assert result["total_rows"] == 1
    assert result["success_rows"] == 0
    assert result["failed_rows"] == 1
    assert result["rows"] == []
    assert any("佣金率超出范围" in str(item) for item in result["errors"])


def test_parse_update_files_and_empty_files_are_successful(tmp_path: Path) -> None:
    invoice_file = tmp_path / "invoice.xlsx"
    merchant_file = tmp_path / "merchant.xlsx"
    header_only = tmp_path / "header_only.xlsx"
    blank = tmp_path / "blank.xlsx"

    _write_workbook(invoice_file, [("发票取数-更新", ["唯一ID", "子订单号", "收款商家", "开票时间", "发票号码"], [["UID-1", "SO-1", "", None, ""]])])
    _write_workbook(merchant_file, [("商家取数-更新", ["唯一ID", "子订单号", "收款商家", "应付商家净额", "付款金额", "付款时间（商家）"], [["UID-1", "SO-1", "", "", "", None]])])
    _write_workbook(header_only, [("原始数据-表头", SOURCE_HEADERS, [])])
    wb = Workbook()
    wb.save(blank)
    wb.close()

    invoice = ReconciliationChecklistService.parse_file(str(invoice_file))
    merchant = ReconciliationChecklistService.parse_file(str(merchant_file))
    empty = ReconciliationChecklistService.parse_file(str(header_only))
    blank_result = ReconciliationChecklistService.parse_file(str(blank))

    assert invoice["file_type"] == CHECKLIST_FILE_TYPE_INVOICE
    assert invoice["rows"][0]["unique_id"] == "UID-1"
    assert invoice["rows"][0]["receipt_merchant"] == ""
    assert invoice["rows"][0]["invoice_time"] is None
    assert merchant["file_type"] == CHECKLIST_FILE_TYPE_MERCHANT
    assert merchant["rows"][0]["unique_id"] == "UID-1"
    assert merchant["rows"][0]["merchant_net_amount"] is None
    assert merchant["rows"][0]["payment_amount"] is None
    assert empty["fatal_error"] is False
    assert empty["success_rows"] == 0
    assert empty["result_message"] == CHECKLIST_EMPTY_SUMMARY_MESSAGE
    assert blank_result["fatal_error"] is False
    assert blank_result["result_message"].startswith("空数据")


def test_parse_merchant_update_file_accepts_legacy_system_locator_header(tmp_path: Path) -> None:
    merchant_file = tmp_path / "merchant-legacy-locator.xlsx"
    payment_time = datetime(2026, 6, 8, 12, 30, 0)
    _write_workbook(
        merchant_file,
        [
            (
                "商家修改",
                ["系统行定位值", "子订单号", "收款商家", "应付商家净额", "付款金额", "付款时间（商家）"],
                [["UID-1", "SO-1", "收款商家B", "", "", payment_time]],
            )
        ],
    )

    result = ReconciliationChecklistService.parse_file(str(merchant_file))

    assert result["file_type"] == CHECKLIST_FILE_TYPE_MERCHANT
    assert result["failed_rows"] == 0
    assert result["rows"][0]["unique_id"] == "UID-1"
    assert result["rows"][0]["merchant_payment_time"] == payment_time


@pytest.mark.parametrize(
    "raw_value",
    ["2026-04-12", "2026/04/12", "2026年04月12", "2026年04月12日"],
)
def test_manual_edit_date_text_variants_parse_to_midnight(raw_value: str) -> None:
    expected = datetime(2026, 4, 12)

    invoice_row = ReconciliationChecklistService._parse_invoice_row(
        {
            "唯一ID": "UID-1",
            "子订单号": "SO-1",
            "收款商家": "收款商家A",
            "开票时间": raw_value,
            "发票号码": "FP-1",
        },
        2,
    )
    merchant_row = ReconciliationChecklistService._parse_merchant_row(
        {
            "唯一ID": "UID-1",
            "子订单号": "SO-1",
            "收款商家": "收款商家A",
            "应付商家净额": "91.50",
            "付款金额": "90.00",
            "付款时间（商家）": raw_value,
        },
        2,
    )

    assert invoice_row["invoice_time"] == expected
    assert merchant_row["merchant_payment_time"] == expected


def test_summarize_errors_groups_same_messages_and_limits_row_samples() -> None:
    errors = [
        "Row 12: 子订单号不能为空",
        "Row 18: 子订单号不能为空",
        "Row 29: 子订单号不能为空",
        "Row 31: 子订单号不能为空",
        "Row 32: 子订单号不能为空",
        "Row 33: 子订单号不能为空",
        "Row 34: 子订单号不能为空",
        "Row 35: 子订单号不能为空",
        "Row 36: 子订单号不能为空",
        "Row 37: 子订单号不能为空",
        "Row 38: 子订单号不能为空",
        "Row 52: 收款商家为空",
        "缺少对账清单必要表头",
    ]

    summary = ReconciliationChecklistService._summarize_errors(errors)

    assert summary[0] == "子订单号不能为空：11行，部分行序号: 12、18、29、31、32、33、34、35、36、37、38"
    assert summary[1] == "收款商家为空：1行，部分行序号: 52"
    assert summary[2] == "缺少对账清单必要表头：1次"


def test_parse_file_error_summary_counts_all_failed_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "source_missing_sub_order.xlsx"
    rows = [_source_row(子订单号="") for _ in range(25)]
    _write_workbook(file_path, [("原始数据-表头", SOURCE_HEADERS, rows)])

    result = ReconciliationChecklistService.parse_file(str(file_path))
    summary = ReconciliationChecklistService._summarize_parse_result_errors(result)

    assert result["failed_rows"] == 25
    assert len(result["errors"]) == 20
    assert summary == [
        "子订单号不能为空：25行，部分行序号: 2、3、4、5、6、7、8、9、10、11、12、13、14、15、16、17、18、19、20、21",
    ]


@pytest.mark.asyncio
async def test_persist_parsed_result_uses_grouped_error_summary() -> None:
    session = _PersistTaskSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")

    summary = await ReconciliationChecklistService.persist_parsed_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        parse_result={
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 4,
            "success_rows": 1,
            "failed_rows": 3,
            "errors": [
                "Row 8: 子订单号不能为空",
                "Row 15: 子订单号不能为空",
                "Row 22: 收款商家为空",
            ],
            "warnings": [],
            "fatal_error": False,
            "rows": [],
            "result_message": "处理完成",
        },
    )

    assert summary["错误明细"] == [
        "子订单号不能为空：2行，部分行序号: 8、15",
        "收款商家为空：1行，部分行序号: 22",
    ]


@pytest.mark.asyncio
async def test_persist_source_rows_batches_detail_and_order_key_upserts(monkeypatch: pytest.MonkeyPatch) -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-2", settlement_time=datetime(2026, 7, 1, 8, 0, 0), source_row_number=3),
    ]
    existing_row = dict(rows[1])
    existing_row["receipt_merchant"] = "旧收款商家"
    session = _BatchWriteSession(existing_details=[_existing_detail_lookup_row(existing_row)])

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert inserted == 1
    assert updated == 1
    assert unchanged == 0
    assert moved_periods == set()
    assert session.detail_insert_calls == 1
    assert session.stage_rows
    assert session.order_key_insert_calls == 0
    assert "RETURNING xmax" not in "\n".join(session.statements)


@pytest.mark.asyncio
async def test_persist_source_rows_skips_upsert_when_existing_rows_are_unchanged() -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-2", settlement_time=datetime(2026, 7, 1, 8, 0, 0), source_row_number=3),
    ]
    session = _BatchWriteSession(existing_details=[_existing_detail_lookup_row(row) for row in rows])

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert inserted == 0
    assert updated == 0
    assert unchanged == 2
    assert moved_periods == set()
    assert session.stage_rows
    assert session.detail_insert_calls == 0
    sql = "\n".join(session.statements)
    assert "CREATE TEMP TABLE" in sql
    assert "tmp_fin_reconciliation_checklist_source_stage" in sql
    assert "FILTER (WHERE" in sql


@pytest.mark.asyncio
async def test_apply_invoice_rows_batches_updates(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _BatchWriteSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="invoice.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家A",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-1",
        },
        {
            "source_row_number": 3,
            "unique_id": "UID-2",
            "sub_order_no": "SO-2",
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 7, 6, 10, 0, 0),
            "invoice_number": "FP-2",
        },
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail_1 = SimpleNamespace(id=101, sub_order_no="SO-1", row_fingerprint="UID-1", accounting_period=202606, settlement_time=datetime(2026, 6, 1, 8, 0, 0), receipt_merchant="")
        detail_2 = SimpleNamespace(id=102, sub_order_no="SO-2", row_fingerprint="UID-2", accounting_period=202607, settlement_time=datetime(2026, 7, 1, 8, 0, 0), receipt_merchant="")
        return {"UID-1": detail_1, "UID-2": detail_2}, {"SO-1": [detail_1], "SO-2": [detail_2]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_invoice_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert updated == 2
    assert failed == 0
    assert errors == []
    assert periods == {202606, 202607}
    assert session.update_calls == 1
    joined_sql = "\n".join(session.statements)
    assert any("FROM (VALUES" in sql for sql in session.statements)
    assert "invoice_time=CAST(reconciliation_invoice_updates.invoice_time AS TIMESTAMP WITHOUT TIME ZONE)" in joined_sql
    assert "is_deleted IS false" in joined_sql


@pytest.mark.asyncio
async def test_apply_merchant_rows_batches_updates(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _BatchWriteSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="merchant.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家A",
            "merchant_net_amount": Decimal("91.50"),
            "payment_amount": Decimal("90.00"),
            "merchant_payment_time": datetime(2026, 6, 6, 10, 0, 0),
        },
        {
            "source_row_number": 3,
            "unique_id": "UID-2",
            "sub_order_no": "SO-2",
            "receipt_merchant": "收款商家B",
            "merchant_net_amount": None,
            "payment_amount": None,
            "merchant_payment_time": datetime(2026, 7, 6, 10, 0, 0),
        },
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail_1 = SimpleNamespace(id=101, sub_order_no="SO-1", row_fingerprint="UID-1", accounting_period=202606, settlement_time=datetime(2026, 6, 1, 8, 0, 0), receipt_merchant="", merchant_net_amount=Decimal("0.00"), payment_amount=None, merchant_payment_time=None)
        detail_2 = SimpleNamespace(id=102, sub_order_no="SO-2", row_fingerprint="UID-2", accounting_period=202607, settlement_time=datetime(2026, 7, 1, 8, 0, 0), receipt_merchant="", merchant_net_amount=Decimal("0.00"), payment_amount=None, merchant_payment_time=None)
        return {"UID-1": detail_1, "UID-2": detail_2}, {"SO-1": [detail_1], "SO-2": [detail_2]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_merchant_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert updated == 2
    assert failed == 0
    assert errors == []
    assert periods == {202606, 202607}
    assert session.update_calls == 1
    joined_sql = "\n".join(session.statements)
    assert any("FROM (VALUES" in sql for sql in session.statements)
    assert "merchant_net_balance" in joined_sql
    assert "merchant_payment_time=CAST(reconciliation_merchant_updates.merchant_payment_time AS TIMESTAMP WITHOUT TIME ZONE)" in joined_sql
    assert "is_deleted IS false" in joined_sql


@pytest.mark.asyncio
async def test_apply_merchant_edit_rows_updates_balance_and_filters_deleted_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(
        existing_details=[(202606, "SO-1")],
        existing_summary_values=[(202606, "SO-1", "收款商家A", Decimal("91.50"), Decimal("90.00"))],
        detail_rows=[
            SimpleNamespace(
                id=101,
                accounting_period=202606,
                sub_order_no="SO-1",
                settlement_time=datetime(2026, 6, 1, 8, 0, 0),
                row_fingerprint="UID-1",
                receipt_merchant="收款商家A",
                merchant_net_amount=Decimal("91.50"),
                payment_amount=Decimal("90.00"),
            )
        ],
        update_rowcount=1,
    )
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家A",
            "merchant_net_amount": Decimal("91.50"),
            "payment_amount": Decimal("90.00"),
            "merchant_payment_time": datetime(2026, 6, 6, 10, 0, 0),
        }
    ]

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_merchant_edit_rows(
        session,  # type: ignore[arg-type]
        org_id=9,
        rows=rows,
    )

    assert updated > 0
    assert failed == 0
    assert errors == []
    assert periods == set()
    assert session.update_calls == 1
    joined_sql = "\n".join(session.statements)
    assert "merchant_net_balance" in joined_sql
    assert (
        "merchant_payment_time=CAST(reconciliation_merchant_manual_updates.merchant_payment_time AS TIMESTAMP WITHOUT TIME ZONE)"
        in joined_sql
    )
    assert "is_deleted IS false" in joined_sql


@pytest.mark.asyncio
async def test_execute_task_marks_empty_file_success(monkeypatch: pytest.MonkeyPatch) -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="queued", progress=0)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="empty.xlsx", oss_key="oss-key", status="uploaded")
    session = _ChecklistExecuteSession(task=task, upload_file=upload_file)

    def fake_parse_file(*_args, **_kwargs) -> dict:
        return {
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 0,
            "success_rows": 0,
            "failed_rows": 0,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [],
            "result_message": CHECKLIST_EMPTY_SUMMARY_MESSAGE,
        }

    async def fake_persist_parsed_result(*_args, **_kwargs) -> dict:
        return {
            "文件类型": CHECKLIST_FILE_TYPE_SOURCE,
            "处理结果": CHECKLIST_EMPTY_SUMMARY_MESSAGE,
            "总行数": 0,
            "成功行数": 0,
            "失败行数": 0,
            "新增行数": 0,
            "更新行数": 0,
        }

    monkeypatch.setattr("app.services.reconciliation_checklist_service.oss_service.download_to_temp", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(ReconciliationChecklistService, "parse_file", staticmethod(fake_parse_file))
    monkeypatch.setattr(ReconciliationChecklistService, "persist_parsed_result", staticmethod(fake_persist_parsed_result))

    result = await ReconciliationChecklistService.execute_task(session, task_id=1)  # type: ignore[arg-type]

    assert result.status == "success"
    assert result.error_message is None
    assert result.result_summary["处理结果"] == CHECKLIST_EMPTY_SUMMARY_MESSAGE
    assert upload_file.status == "processed"
    assert session.commit_count == 3


@pytest.mark.asyncio
async def test_execute_task_masks_internal_exception_and_logs_details(monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture) -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="queued", progress=0)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="large.xlsx", oss_key="oss-key", status="uploaded")
    session = _ChecklistExecuteSession(task=task, upload_file=upload_file)

    def fake_parse_file(*_args, **_kwargs) -> dict:
        return {
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 1,
            "success_rows": 1,
            "failed_rows": 0,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [{"sub_order_no": "SO-1", "accounting_period": 202606}],
            "result_message": "处理完成",
        }

    async def fake_persist_parsed_result(*_args, **_kwargs) -> dict:
        raise RuntimeError("the number of query arguments cannot exceed 32767")

    monkeypatch.setattr("app.services.reconciliation_checklist_service.oss_service.download_to_temp", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(ReconciliationChecklistService, "parse_file", staticmethod(fake_parse_file))
    monkeypatch.setattr(ReconciliationChecklistService, "persist_parsed_result", staticmethod(fake_persist_parsed_result))

    with caplog.at_level("ERROR", logger="finengine.reconciliation_checklist"):
        result = await ReconciliationChecklistService.execute_task(session, task_id=1)  # type: ignore[arg-type]

    assert result.status == "failed"
    assert result.error_message == "处理异常，请联系管理员或稍后重试"
    assert upload_file.error_message == "处理异常，请联系管理员或稍后重试"
    assert "32767" not in (result.error_message or "")
    assert "reconciliation_checklist.task_failed" in caplog.text
    assert "the number of query arguments cannot exceed 32767" in caplog.text


@pytest.mark.asyncio
async def test_execute_task_parses_file_before_reentering_database_work(monkeypatch: pytest.MonkeyPatch) -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="queued", progress=0)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    session = _ChecklistExecuteSession(task=task, upload_file=upload_file)
    parse_called = False

    def fake_parse_file(_file_path: str) -> dict:
        nonlocal parse_called
        parse_called = True
        assert session.allow_db_get is False
        return {
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 0,
            "success_rows": 0,
            "failed_rows": 0,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [],
            "result_message": CHECKLIST_EMPTY_SUMMARY_MESSAGE,
        }

    async def fake_persist_parsed_result(*_args, **kwargs) -> dict:
        assert parse_called is True
        assert session.allow_db_get is True
        assert kwargs["parse_result"]["download_seconds"] >= 0
        return {
            "文件类型": CHECKLIST_FILE_TYPE_SOURCE,
            "处理结果": CHECKLIST_EMPTY_SUMMARY_MESSAGE,
            "总行数": 0,
            "成功行数": 0,
            "失败行数": 0,
            "新增行数": 0,
            "更新行数": 0,
            "文件下载耗时秒": kwargs["parse_result"]["download_seconds"],
        }

    monkeypatch.setattr("app.services.reconciliation_checklist_service.oss_service.download_to_temp", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(ReconciliationChecklistService, "parse_file", staticmethod(fake_parse_file))
    monkeypatch.setattr(ReconciliationChecklistService, "persist_parsed_result", staticmethod(fake_persist_parsed_result))

    result = await ReconciliationChecklistService.execute_task(session, task_id=1)  # type: ignore[arg-type]

    assert result.status == "success"
    assert parse_called is True


@pytest.mark.asyncio
async def test_execute_task_skips_partition_precreate_for_parsed_cross_year_periods(monkeypatch: pytest.MonkeyPatch) -> None:
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="queued", progress=0)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    session = _ChecklistExecuteSession(task=task, upload_file=upload_file)

    async def fail_partition_precreate(*_args, **_kwargs) -> None:
        raise AssertionError("import task must not check or create partitions")

    def fake_parse_file(_file_path: str) -> dict:
        return {
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 2,
            "success_rows": 2,
            "failed_rows": 0,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [
                {"sub_order_no": "SO-1", "accounting_period": 202512},
                {"sub_order_no": "SO-2", "accounting_period": 202602},
            ],
            "result_message": "处理完成",
        }

    async def fake_persist_parsed_result(*_args, **_kwargs) -> dict:
        return {
            "文件类型": CHECKLIST_FILE_TYPE_SOURCE,
            "处理结果": "处理完成",
            "总行数": 2,
            "成功行数": 2,
            "失败行数": 0,
            "新增行数": 2,
            "更新行数": 0,
            "涉及年月": [202512, 202602],
        }

    monkeypatch.setattr("app.services.reconciliation_checklist_service.oss_service.download_to_temp", lambda *_args, **_kwargs: None)
    monkeypatch.setattr("app.services.reconciliation_checklist_service.ensure_reconciliation_checklist_partitions_for_year", fail_partition_precreate, raising=False)
    monkeypatch.setattr("app.services.reconciliation_checklist_service.ensure_reconciliation_checklist_partitions_for_window", fail_partition_precreate, raising=False)
    monkeypatch.setattr(ReconciliationChecklistService, "parse_file", staticmethod(fake_parse_file))
    monkeypatch.setattr(ReconciliationChecklistService, "persist_parsed_result", staticmethod(fake_persist_parsed_result))

    result = await ReconciliationChecklistService.execute_task(session, task_id=1)  # type: ignore[arg-type]

    assert result.status == "success"


@pytest.mark.asyncio
async def test_order_key_lookup_chunks_large_sub_order_lists() -> None:
    session = _OrderLookupSession()
    rows = [{"sub_order_no": f"SO-{index}"} for index in range(70000)]

    result = await ReconciliationChecklistService._lookup_order_keys(session, org_id=3, rows=rows)  # type: ignore[arg-type]

    assert result == {}
    assert len(session.param_counts) > 1
    assert max(session.param_counts) <= 30001


@pytest.mark.asyncio
async def test_persist_source_rows_uses_large_tuple_lookup_batches() -> None:
    session = _OrderLookupSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(
            sub_order_no=f"SO-{index}",
            settlement_time=datetime(2026, 6, 1, 8, 0, 0),
            source_row_number=index + 2,
        )
        for index in range(2501)
    ]

    await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert session.tuple_param_counts == []
    sql = "\n".join(session.statements)
    assert "CREATE TEMP TABLE" in sql
    assert "tmp_fin_reconciliation_checklist_source_stage" in sql


@pytest.mark.asyncio
async def test_persist_parsed_result_reports_unchanged_source_rows(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _PersistTaskSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rebuilt_calls: list[list[int]] = []

    async def fake_persist_source_rows(*_args, **_kwargs):
        return 1, 2, 2, set()

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        assert org_id == 3
        rebuilt_calls.append(sorted(periods))

    monkeypatch.setattr(ReconciliationChecklistService, "_persist_source_rows", staticmethod(fake_persist_source_rows))
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))

    summary = await ReconciliationChecklistService.persist_parsed_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        parse_result={
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 5,
            "success_rows": 5,
            "failed_rows": 0,
            "download_seconds": 4.321,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [{"accounting_period": 202606} for _ in range(5)],
            "result_message": "处理完成",
        },
    )

    assert summary["新增行数"] == 1
    assert summary["更新行数"] == 2
    assert summary["未变更行数"] == 2
    assert summary["文件下载耗时秒"] == 4.321
    assert rebuilt_calls == [[202606]]


@pytest.mark.asyncio
async def test_persist_parsed_result_rebuilds_summary_when_source_rows_are_all_unchanged(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _PersistTaskSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rebuilt_calls: list[list[int]] = []

    async def fake_persist_source_rows(*_args, **_kwargs):
        return 0, 0, 5, set()

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        assert org_id == 3
        rebuilt_calls.append(sorted(periods))

    monkeypatch.setattr(ReconciliationChecklistService, "_persist_source_rows", staticmethod(fake_persist_source_rows))
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))

    summary = await ReconciliationChecklistService.persist_parsed_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        parse_result={
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 5,
            "success_rows": 5,
            "failed_rows": 0,
            "download_seconds": 0.123,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [{"accounting_period": 202606} for _ in range(5)],
            "result_message": "处理完成",
        },
    )

    assert summary["新增行数"] == 0
    assert summary["更新行数"] == 0
    assert summary["未变更行数"] == 5
    assert rebuilt_calls == [[202606]]


@pytest.mark.asyncio
async def test_persist_parsed_result_shards_source_rows_by_accounting_period(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _PersistTaskSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    persisted_periods: list[list[int]] = []
    rebuilt_periods: list[int] = []

    async def fake_persist_source_rows_sharded(*_args, rows, **_kwargs):
        period_list = sorted({int(row["accounting_period"]) for row in rows if row.get("accounting_period")})
        persisted_periods.append(period_list)
        return 3, 2, 1, {202608}

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        assert org_id == 3
        rebuilt_periods.extend(sorted(periods))

    monkeypatch.setattr(
        ReconciliationChecklistService,
        "_persist_source_rows_sharded",
        staticmethod(fake_persist_source_rows_sharded),
    )
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))

    summary = await ReconciliationChecklistService.persist_parsed_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        parse_result={
            "file_type": CHECKLIST_FILE_TYPE_SOURCE,
            "total_rows": 6,
            "success_rows": 6,
            "failed_rows": 0,
            "download_seconds": 1.234,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "rows": [
                {"accounting_period": 202606},
                {"accounting_period": 202606},
                {"accounting_period": 202607},
                {"accounting_period": 202607},
                {"accounting_period": 202608},
                {"accounting_period": 202608},
            ],
            "result_message": "处理完成",
        },
    )

    assert persisted_periods == [[202606, 202607, 202608]]
    assert rebuilt_periods == [202606, 202607, 202608]
    assert summary["新增行数"] == 3
    assert summary["更新行数"] == 2
    assert summary["未变更行数"] == 1
    assert summary["涉及年月"] == [202606, 202607, 202608]


@pytest.mark.asyncio
async def test_persist_source_rows_sharded_splits_large_single_period_batches(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _BatchWriteSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-202606-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-202606-2", settlement_time=datetime(2026, 6, 2, 8, 0, 0), source_row_number=3),
        _parsed_source_detail_row(sub_order_no="SO-202606-3", settlement_time=datetime(2026, 6, 3, 8, 0, 0), source_row_number=4),
    ]
    persisted_chunk_sizes: list[int] = []

    async def fake_persist_source_rows(_db, *, rows, **_kwargs):
        period_values = sorted({int(row["accounting_period"]) for row in rows if row.get("accounting_period")})
        assert len(period_values) == 1
        persisted_chunk_sizes.append(len(list(rows)))
        return 1, 1, 0, set()

    monkeypatch.setattr("app.services.reconciliation_checklist_service.CHECKLIST_SOURCE_PERSIST_SHARD_SIZE", 2)
    monkeypatch.setattr(
        ReconciliationChecklistService,
        "_persist_source_rows",
        staticmethod(fake_persist_source_rows),
    )

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows_sharded(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert persisted_chunk_sizes == [2, 1]
    assert inserted == 2
    assert updated == 2
    assert unchanged == 0
    assert moved_periods == set()
    assert session.commit_count == 2


@pytest.mark.asyncio
async def test_persist_source_rows_uses_staging_sql_path() -> None:
    session = _SourceStagingSession(counts=(2, 1, 1))
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-2", settlement_time=datetime(2026, 6, 2, 8, 0, 0), source_row_number=3),
    ]

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert inserted == 2
    assert updated == 1
    assert unchanged == 1
    assert moved_periods == set()
    assert session.stage_insert_calls == 1
    assert session.detail_upsert_calls == 1
    sql = "\n".join(session.statements)
    assert "CREATE TEMP TABLE" in sql
    assert "tmp_fin_reconciliation_checklist_source_stage" in sql
    assert "JOIN (VALUES" not in sql


@pytest.mark.asyncio
async def test_persist_source_rows_prefers_copy_for_stage_loading() -> None:
    session = _CopyCapableSession(counts=(2, 0, 0))
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-2", settlement_time=datetime(2026, 6, 2, 8, 0, 0), source_row_number=3),
    ]

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert inserted == 2
    assert updated == 0
    assert unchanged == 0
    assert moved_periods == set()
    assert session.copy_calls == 1
    assert session.fallback_insert_calls == 0
    assert "id" not in session.copy_columns
    assert "created_at" not in session.copy_columns
    assert "is_deleted" not in session.copy_columns


@pytest.mark.asyncio
async def test_persist_source_rows_falls_back_to_insert_when_copy_fails() -> None:
    session = _CopyCapableSession(counts=(2, 0, 0), copy_should_fail=True)
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="source.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        _parsed_source_detail_row(sub_order_no="SO-1", settlement_time=datetime(2026, 6, 1, 8, 0, 0), source_row_number=2),
        _parsed_source_detail_row(sub_order_no="SO-2", settlement_time=datetime(2026, 6, 2, 8, 0, 0), source_row_number=3),
    ]

    inserted, updated, unchanged, moved_periods = await ReconciliationChecklistService._persist_source_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert inserted == 2
    assert updated == 0
    assert unchanged == 0
    assert moved_periods == set()
    assert session.copy_calls == 1
    assert session.fallback_insert_calls == 1


@pytest.mark.asyncio
async def test_lookup_existing_receipt_merchants_chunks_large_pair_lists() -> None:
    session = _OrderLookupSession()
    pairs = [(202601 + index % 2, f"SO-{index}") for index in range(40000)]

    result = await ReconciliationChecklistService._lookup_existing_receipt_merchants(
        session,  # type: ignore[arg-type]
        org_id=3,
        pairs=pairs,
    )

    assert result == {}
    assert len(session.param_counts) > 1
    assert max(session.tuple_param_counts) <= 30000


@pytest.mark.asyncio
async def test_lookup_existing_merchant_summary_values_chunks_large_pair_lists() -> None:
    session = _OrderLookupSession()
    pairs = [(202601 + index % 2, f"SO-{index}") for index in range(40000)]

    result = await ReconciliationChecklistService._lookup_existing_merchant_summary_values(
        session,  # type: ignore[arg-type]
        org_id=3,
        pairs=pairs,
    )

    assert result == {}
    assert len(session.param_counts) > 1
    assert max(session.param_counts) <= 30001


def test_manual_edit_normalize_sub_order_numbers_dedupes_and_limits() -> None:
    result = ReconciliationChecklistService._normalize_manual_edit_sub_orders(
        [" SO-1 ", "SO-2,SO-1", "SO-3\nSO-4", "，SO-5,,SO-6"]
    )

    assert result == ["SO-1", "SO-2", "SO-3", "SO-4", "SO-5", "SO-6"]
    with pytest.raises(ValueError, match="请先输入子订单号"):
        ReconciliationChecklistService._normalize_manual_edit_sub_orders(["", " ", ",，"])
    with pytest.raises(ValueError, match="单次最多查询 100 个子订单号"):
        ReconciliationChecklistService._normalize_manual_edit_sub_orders([f"SO-{index}" for index in range(1, 102)])


@pytest.mark.asyncio
async def test_query_invoice_edit_items_returns_matched_and_missing(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _ManualEditQuerySession(
        [
            SimpleNamespace(
                id=22,
                row_fingerprint="UID-2",
                sub_order_no="SO-2",
                settlement_time=datetime(2026, 6, 6, 10, 0, 0),
                platform_subsidy=Decimal("1.11"),
                talent_subsidy=Decimal("2.22"),
                douyin_pay_subsidy=Decimal("3.33"),
                douyin_monthly_pay_subsidy=Decimal("4.44"),
                bank_subsidy=Decimal("5.55"),
                user_paid_amount=Decimal("20.00"),
                platform_service_fee=Decimal("2.00"),
                talent_commission=Decimal("6.66"),
                investment_service_fee=Decimal("7.77"),
                receipt_merchant="收款商家B",
                invoice_time=datetime(2026, 6, 6, 10, 0, 0),
                invoice_number="FP-2",
            ),
            SimpleNamespace(
                id=11,
                row_fingerprint="UID-1",
                sub_order_no="SO-1",
                settlement_time=datetime(2026, 6, 1, 8, 0, 0),
                platform_subsidy=Decimal("0.11"),
                talent_subsidy=Decimal("0.22"),
                douyin_pay_subsidy=Decimal("0.33"),
                douyin_monthly_pay_subsidy=Decimal("0.44"),
                bank_subsidy=Decimal("0.55"),
                user_paid_amount=Decimal("10.00"),
                platform_service_fee=Decimal("1.00"),
                talent_commission=Decimal("0.66"),
                investment_service_fee=Decimal("0.77"),
                receipt_merchant="收款商家A",
                invoice_time=None,
                invoice_number="FP-1",
            ),
        ]
    )
    user = SimpleNamespace(role="admin", org_id=9)

    matched_items, missing_sub_order_nos = await ReconciliationChecklistService.query_invoice_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        sub_order_nos=["SO-1", "SO-2", "SO-3", "SO-1"],
    )

    assert matched_items == [
        {
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "settlement_time": datetime(2026, 6, 1, 8, 0, 0),
            "platform_subsidy": Decimal("0.11"),
            "talent_subsidy": Decimal("0.22"),
            "douyin_pay_subsidy": Decimal("0.33"),
            "douyin_monthly_pay_subsidy": Decimal("0.44"),
            "bank_subsidy": Decimal("0.55"),
            "user_paid_amount": Decimal("10.00"),
            "platform_service_fee": Decimal("1.00"),
            "talent_commission": Decimal("0.66"),
            "investment_service_fee": Decimal("0.77"),
            "receipt_merchant": "收款商家A",
            "invoice_time": None,
            "invoice_number": "FP-1",
        },
        {
            "unique_id": "UID-2",
            "sub_order_no": "SO-2",
            "settlement_time": datetime(2026, 6, 6, 10, 0, 0),
            "platform_subsidy": Decimal("1.11"),
            "talent_subsidy": Decimal("2.22"),
            "douyin_pay_subsidy": Decimal("3.33"),
            "douyin_monthly_pay_subsidy": Decimal("4.44"),
            "bank_subsidy": Decimal("5.55"),
            "user_paid_amount": Decimal("20.00"),
            "platform_service_fee": Decimal("2.00"),
            "talent_commission": Decimal("6.66"),
            "investment_service_fee": Decimal("7.77"),
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-2",
        },
    ]
    assert missing_sub_order_nos == ["SO-3"]


@pytest.mark.asyncio
async def test_query_invoice_edit_items_treats_missing_details_as_missing_sub_orders(
) -> None:
    session = _ManualEditQuerySession(
        [
            SimpleNamespace(
                id=22,
                row_fingerprint="UID-2",
                sub_order_no="SO-2",
                settlement_time=datetime(2026, 6, 6, 10, 0, 0),
                platform_subsidy=Decimal("1.11"),
                talent_subsidy=Decimal("2.22"),
                douyin_pay_subsidy=Decimal("3.33"),
                douyin_monthly_pay_subsidy=Decimal("4.44"),
                bank_subsidy=Decimal("5.55"),
                user_paid_amount=Decimal("20.00"),
                platform_service_fee=Decimal("2.00"),
                talent_commission=Decimal("6.66"),
                investment_service_fee=Decimal("7.77"),
                receipt_merchant="收款商家B",
                invoice_time=datetime(2026, 6, 6, 10, 0, 0),
                invoice_number="FP-2",
            ),
        ]
    )

    matched_items, missing_sub_order_nos = await ReconciliationChecklistService.query_invoice_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=SimpleNamespace(role="admin", org_id=9),  # type: ignore[arg-type]
        sub_order_nos=["SO-1", "SO-2", "SO-3"],
    )

    assert matched_items == [
        {
            "unique_id": "UID-2",
            "sub_order_no": "SO-2",
            "settlement_time": datetime(2026, 6, 6, 10, 0, 0),
            "platform_subsidy": Decimal("1.11"),
            "talent_subsidy": Decimal("2.22"),
            "douyin_pay_subsidy": Decimal("3.33"),
            "douyin_monthly_pay_subsidy": Decimal("4.44"),
            "bank_subsidy": Decimal("5.55"),
            "user_paid_amount": Decimal("20.00"),
            "platform_service_fee": Decimal("2.00"),
            "talent_commission": Decimal("6.66"),
            "investment_service_fee": Decimal("7.77"),
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-2",
        }
    ]
    assert missing_sub_order_nos == ["SO-1", "SO-3"]


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("query_method_name", "expected_field"),
    [
        ("query_invoice_edit_items", "invoice_number"),
        ("query_merchant_edit_items", "merchant_net_amount"),
    ],
)
async def test_manual_edit_queries_validate_org_scope_before_lookup(
    query_method_name: str,
    expected_field: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    query_method = getattr(ReconciliationChecklistService, query_method_name)

    with pytest.raises(ValueError, match="数据不存在或无权访问"):
        await query_method(
            _ManualEditQuerySession([]),  # type: ignore[arg-type]
            org_id=9,
            user=SimpleNamespace(role="admin", org_id=8),  # type: ignore[arg-type]
            sub_order_nos=["SO-1"],
        )


@pytest.mark.asyncio
async def test_query_merchant_edit_items_returns_matched_and_missing(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _ManualEditQuerySession(
        [
            SimpleNamespace(
                id=22,
                row_fingerprint="UID-2",
                sub_order_no="SO-2",
                settlement_time=datetime(2026, 6, 6, 10, 0, 0),
                platform_subsidy=Decimal("1.11"),
                talent_subsidy=Decimal("2.22"),
                douyin_pay_subsidy=Decimal("3.33"),
                douyin_monthly_pay_subsidy=Decimal("4.44"),
                bank_subsidy=Decimal("5.55"),
                user_paid_amount=Decimal("20.00"),
                platform_service_fee=Decimal("2.00"),
                talent_commission=Decimal("6.66"),
                investment_service_fee=Decimal("7.77"),
                receipt_merchant="收款商家B",
                merchant_net_amount=None,
                payment_amount=None,
                merchant_payment_time=None,
            ),
            SimpleNamespace(
                id=11,
                row_fingerprint="UID-1",
                sub_order_no="SO-1",
                settlement_time=datetime(2026, 6, 1, 8, 0, 0),
                platform_subsidy=Decimal("0.11"),
                talent_subsidy=Decimal("0.22"),
                douyin_pay_subsidy=Decimal("0.33"),
                douyin_monthly_pay_subsidy=Decimal("0.44"),
                bank_subsidy=Decimal("0.55"),
                user_paid_amount=Decimal("10.00"),
                platform_service_fee=Decimal("1.00"),
                talent_commission=Decimal("0.66"),
                investment_service_fee=Decimal("0.77"),
                receipt_merchant="收款商家A",
                merchant_net_amount=Decimal("91.50"),
                payment_amount=Decimal("90.00"),
                merchant_payment_time=datetime(2026, 6, 5, 9, 0, 0),
            ),
        ]
    )
    user = SimpleNamespace(role="admin", org_id=9)

    matched_items, missing_sub_order_nos = await ReconciliationChecklistService.query_merchant_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        sub_order_nos=["SO-1", "SO-2", "SO-3", "SO-1"],
    )

    assert matched_items == [
        {
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "settlement_time": datetime(2026, 6, 1, 8, 0, 0),
            "platform_subsidy": Decimal("0.11"),
            "talent_subsidy": Decimal("0.22"),
            "douyin_pay_subsidy": Decimal("0.33"),
            "douyin_monthly_pay_subsidy": Decimal("0.44"),
            "bank_subsidy": Decimal("0.55"),
            "user_paid_amount": Decimal("10.00"),
            "platform_service_fee": Decimal("1.00"),
            "talent_commission": Decimal("0.66"),
            "investment_service_fee": Decimal("0.77"),
            "receipt_merchant": "收款商家A",
            "merchant_net_amount": Decimal("91.50"),
            "payment_amount": Decimal("90.00"),
            "merchant_payment_time": datetime(2026, 6, 5, 9, 0, 0),
        },
        {
            "unique_id": "UID-2",
            "sub_order_no": "SO-2",
            "settlement_time": datetime(2026, 6, 6, 10, 0, 0),
            "platform_subsidy": Decimal("1.11"),
            "talent_subsidy": Decimal("2.22"),
            "douyin_pay_subsidy": Decimal("3.33"),
            "douyin_monthly_pay_subsidy": Decimal("4.44"),
            "bank_subsidy": Decimal("5.55"),
            "user_paid_amount": Decimal("20.00"),
            "platform_service_fee": Decimal("2.00"),
            "talent_commission": Decimal("6.66"),
            "investment_service_fee": Decimal("7.77"),
            "receipt_merchant": "收款商家B",
            "merchant_net_amount": None,
            "payment_amount": None,
            "merchant_payment_time": None,
        },
    ]
    assert missing_sub_order_nos == ["SO-3"]


@pytest.mark.asyncio
async def test_query_merchant_edit_items_backfills_missing_unique_id() -> None:
    detail = SimpleNamespace(
        id=11,
        row_fingerprint="",
        sub_order_no="SO-1",
        settlement_time=datetime(2026, 6, 1, 8, 0, 0),
        platform_subsidy=Decimal("0.11"),
        talent_subsidy=Decimal("0.22"),
        douyin_pay_subsidy=Decimal("0.33"),
        douyin_monthly_pay_subsidy=Decimal("0.44"),
        bank_subsidy=Decimal("0.55"),
        user_paid_amount=Decimal("10.00"),
        platform_service_fee=Decimal("1.00"),
        talent_commission=Decimal("0.66"),
        investment_service_fee=Decimal("0.77"),
        receipt_merchant="收款商家A",
        merchant_net_amount=Decimal("91.50"),
        payment_amount=Decimal("90.00"),
        merchant_payment_time=datetime(2026, 6, 5, 9, 0, 0),
    )
    session = _ManualEditQuerySession([detail])
    expected_unique_id = _row_fingerprint(
        org_id=9,
        settlement_time=detail.settlement_time,
        sub_order_no=detail.sub_order_no,
        platform_subsidy=detail.platform_subsidy,
        talent_subsidy=detail.talent_subsidy,
        douyin_pay_subsidy=detail.douyin_pay_subsidy,
        douyin_monthly_pay_subsidy=detail.douyin_monthly_pay_subsidy,
        bank_subsidy=detail.bank_subsidy,
        user_paid_amount=detail.user_paid_amount,
        platform_service_fee=detail.platform_service_fee,
        talent_commission=detail.talent_commission,
        investment_service_fee=detail.investment_service_fee,
    )

    matched_items, missing_sub_order_nos = await ReconciliationChecklistService.query_merchant_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=SimpleNamespace(role="admin", org_id=9),  # type: ignore[arg-type]
        sub_order_nos=["SO-1"],
    )

    assert matched_items[0]["unique_id"] == expected_unique_id
    assert detail.row_fingerprint == expected_unique_id
    assert session.flush_count == 1
    assert missing_sub_order_nos == []


@pytest.mark.asyncio
async def test_query_invoice_edit_items_returns_multiple_rows_for_same_sub_order() -> None:
    session = _ManualEditQuerySession(
        [
            SimpleNamespace(
                id=12,
                row_fingerprint="UID-2",
                sub_order_no="SO-1",
                settlement_time=datetime(2026, 6, 2, 8, 0, 0),
                platform_subsidy=Decimal("1.11"),
                talent_subsidy=Decimal("2.22"),
                douyin_pay_subsidy=Decimal("3.33"),
                douyin_monthly_pay_subsidy=Decimal("4.44"),
                bank_subsidy=Decimal("5.55"),
                user_paid_amount=Decimal("20.00"),
                platform_service_fee=Decimal("2.00"),
                talent_commission=Decimal("6.66"),
                investment_service_fee=Decimal("7.77"),
                receipt_merchant="收款商家B",
                invoice_time=None,
                invoice_number="FP-2",
            ),
            SimpleNamespace(
                id=11,
                row_fingerprint="UID-1",
                sub_order_no="SO-1",
                settlement_time=datetime(2026, 6, 1, 8, 0, 0),
                platform_subsidy=Decimal("0.11"),
                talent_subsidy=Decimal("0.22"),
                douyin_pay_subsidy=Decimal("0.33"),
                douyin_monthly_pay_subsidy=Decimal("0.44"),
                bank_subsidy=Decimal("0.55"),
                user_paid_amount=Decimal("10.00"),
                platform_service_fee=Decimal("1.00"),
                talent_commission=Decimal("0.66"),
                investment_service_fee=Decimal("0.77"),
                receipt_merchant="收款商家A",
                invoice_time=None,
                invoice_number="FP-1",
            ),
        ]
    )

    matched_items, missing_sub_order_nos = await ReconciliationChecklistService.query_invoice_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=SimpleNamespace(role="admin", org_id=9),  # type: ignore[arg-type]
        sub_order_nos=["SO-1"],
    )

    assert [item["unique_id"] for item in matched_items] == ["UID-1", "UID-2"]
    assert missing_sub_order_nos == []


@pytest.mark.asyncio
async def test_save_invoice_edit_items_reuses_existing_parse_rules(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _ManualEditSaveSession()
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)
    parsed_rows: list[dict[str, object]] = []
    rebuilt_calls: list[tuple[int, list[int]]] = []
    audit_calls: list[dict[str, object]] = []

    def fake_parse_invoice_row(values: dict[str, object], row_number: int) -> dict[str, object]:
        assert row_number == 2
        return {
            "source_row_number": row_number,
            "unique_id": str(values.get("唯一ID") or ""),
            "row_fingerprint": str(values.get("唯一ID") or ""),
            "sub_order_no": str(values.get("子订单号") or ""),
            "receipt_merchant": f"parsed:{str(values.get('收款商家') or '')}",
            "invoice_time": values.get("开票时间"),
            "invoice_number": f"parsed:{str(values.get('发票号码') or '')}",
        }

    async def fake_apply_invoice_edit_rows(_db, *, org_id: int, rows):
        assert org_id == 9
        parsed_rows.extend(list(rows))
        return 1, 0, [], {202606}

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        rebuilt_calls.append((org_id, sorted(periods)))

    async def fake_audit_log(_db, **kwargs):
        audit_calls.append(kwargs)

    monkeypatch.setattr(ReconciliationChecklistService, "_parse_invoice_row", staticmethod(fake_parse_invoice_row))
    monkeypatch.setattr(ReconciliationChecklistService, "_apply_invoice_edit_rows", staticmethod(fake_apply_invoice_edit_rows))
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    result = await ReconciliationChecklistService.save_invoice_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        items=[
            {
                "unique_id": "UID-1",
                "sub_order_no": "SO-1",
                "receipt_merchant": "收款商家A",
                "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
                "invoice_number": "FP-1",
            }
        ],
        ip="127.0.0.1",
        user_agent="pytest",
    )

    assert parsed_rows == [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "row_fingerprint": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "parsed:收款商家A",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "parsed:FP-1",
        }
    ]
    assert result == {
        "success_count": 1,
        "failed_count": 0,
        "unchanged_count": 0,
        "missing_sub_order_nos": [],
        "affected_periods": [202606],
        "error_messages": [],
    }
    assert rebuilt_calls == [(9, [202606])]
    assert session.flush_count == 1
    assert len(audit_calls) == 1
    assert audit_calls[0]["module"] == "reconciliation_checklist"
    assert audit_calls[0]["action"] == "invoice_edit"
    assert audit_calls[0]["extra_data"]["success_count"] == 1
    assert audit_calls[0]["extra_data"]["failed_count"] == 0
    assert audit_calls[0]["extra_data"]["unchanged_count"] == 0
    assert audit_calls[0]["extra_data"]["affected_periods"] == [202606]


@pytest.mark.asyncio
async def test_save_invoice_edit_items_skips_summary_rebuild_when_only_invoice_fields_change(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession()
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)
    rebuilt_calls: list[tuple[int, list[int]]] = []

    async def fake_apply_invoice_edit_rows(_db, *, org_id: int, rows):
        assert org_id == 9
        _ = list(rows)
        return 1, 0, [], set()

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        rebuilt_calls.append((org_id, sorted(periods)))

    async def fake_audit_log(_db, **_kwargs):
        return None

    monkeypatch.setattr(ReconciliationChecklistService, "_apply_invoice_edit_rows", staticmethod(fake_apply_invoice_edit_rows))
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    result = await ReconciliationChecklistService.save_invoice_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        items=[
            {
                "unique_id": "UID-1",
                "sub_order_no": "SO-1",
                "receipt_merchant": "收款商家A",
                "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
                "invoice_number": "FP-1",
            }
        ],
        ip="127.0.0.1",
        user_agent="pytest",
    )

    assert result == {
        "success_count": 1,
        "failed_count": 0,
        "unchanged_count": 0,
        "missing_sub_order_nos": [],
        "affected_periods": [],
        "error_messages": [],
    }
    assert rebuilt_calls == []
    assert session.flush_count == 1


@pytest.mark.asyncio
async def test_apply_invoice_edit_rows_marks_period_when_receipt_merchant_changes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(update_rowcount=1)
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "row_fingerprint": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-1",
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail = SimpleNamespace(
            id=101,
            accounting_period=202606,
            settlement_time=datetime(2026, 6, 1, 8, 0, 0),
            row_fingerprint="UID-1",
            receipt_merchant="收款商家A",
            merchant_net_amount=Decimal("91.50"),
            payment_amount=Decimal("90.00"),
        )
        return {"UID-1": detail}, {"SO-1": [detail]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_invoice_edit_rows(
        session,  # type: ignore[arg-type]
        org_id=9,
        rows=rows,
    )

    assert updated > 0
    assert failed == 0
    assert errors == []
    assert periods == {202606}
    assert (
        "invoice_time=CAST(reconciliation_invoice_manual_updates.invoice_time AS TIMESTAMP WITHOUT TIME ZONE)"
        in "\n".join(session.statements)
    )


@pytest.mark.asyncio
async def test_apply_invoice_rows_marks_period_when_receipt_merchant_changes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _BatchWriteSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=9, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=9, user_id=4, original_name="invoice.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-1",
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail = SimpleNamespace(
            id=101,
            accounting_period=202606,
            settlement_time=datetime(2026, 6, 1, 8, 0, 0),
            row_fingerprint="UID-1",
            receipt_merchant="收款商家A",
        )
        return {"UID-1": detail}, {}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_invoice_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert updated > 0
    assert failed == 0
    assert errors == []
    assert periods == {202606}


@pytest.mark.asyncio
async def test_save_invoice_edit_items_reports_missing_when_order_key_exists_but_detail_is_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(existing_details=[])
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)
    rebuilt_calls: list[tuple[int, list[int]]] = []
    audit_calls: list[dict[str, object]] = []

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        rebuilt_calls.append((org_id, sorted(periods)))

    async def fake_audit_log(_db, **kwargs):
        audit_calls.append(kwargs)

    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    with pytest.raises(ValueError, match="唯一ID不能为空"):
        await ReconciliationChecklistService.save_invoice_edit_items(
            session,  # type: ignore[arg-type]
            org_id=9,
            user=user,  # type: ignore[arg-type]
            items=[
                {
                    "sub_order_no": "SO-404",
                    "receipt_merchant": "收款商家A",
                    "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
                    "invoice_number": "FP-404",
                }
            ],
            ip="127.0.0.1",
            user_agent="pytest",
        )
    assert session.update_calls == 0
    assert rebuilt_calls == []
    assert audit_calls == []


@pytest.mark.asyncio
async def test_save_merchant_edit_items_reports_missing_sub_orders_without_updating(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession()
    user = SimpleNamespace(id=8, username="tester", display_name="测试用户", role="admin", org_id=9)
    rebuilt_calls: list[tuple[int, list[int]]] = []
    audit_calls: list[dict[str, object]] = []

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        rebuilt_calls.append((org_id, sorted(periods)))

    async def fake_audit_log(_db, **kwargs):
        audit_calls.append(kwargs)

    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    result = await ReconciliationChecklistService.save_merchant_edit_items(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        items=[
            {
                "unique_id": "UID-404",
                "sub_order_no": "SO-404",
                "receipt_merchant": "收款商家A",
                "merchant_net_amount": None,
                "payment_amount": Decimal("90.00"),
                "merchant_payment_time": datetime(2026, 6, 5, 9, 0, 0),
            },
            {
                "unique_id": "UID-405",
                "sub_order_no": "SO-405",
                "receipt_merchant": "收款商家B",
                "merchant_net_amount": Decimal("10.00"),
                "payment_amount": None,
                "merchant_payment_time": None,
            },
        ],
        ip="127.0.0.1",
        user_agent="pytest",
    )

    assert result == {
        "success_count": 0,
        "failed_count": 2,
        "unchanged_count": 0,
        "missing_sub_order_nos": [],
        "affected_periods": [],
        "error_messages": ["Row 2: 唯一ID不存在", "Row 3: 唯一ID不存在"],
    }
    assert session.update_calls == 0
    assert rebuilt_calls == []
    assert len(audit_calls) == 1
    assert audit_calls[0]["module"] == "reconciliation_checklist"
    assert audit_calls[0]["action"] == "merchant_edit"
    assert audit_calls[0]["extra_data"]["missing_sub_order_nos"] == []
    assert audit_calls[0]["extra_data"]["affected_periods"] == []
    assert audit_calls[0]["extra_data"]["unchanged_count"] == 0


@pytest.mark.asyncio
async def test_save_merchant_edit_items_treats_blank_net_amount_as_zero(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession()
    captured_rows: list[dict[str, object]] = []

    async def fake_apply_merchant_edit_rows(_db, *, org_id: int, rows):
        assert org_id == 3
        captured_rows.extend(list(rows))
        return 1, 0, [], {202606}

    async def fake_rebuild(*_args, **_kwargs):
        return None

    async def fake_audit_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(
        ReconciliationChecklistService,
        "_apply_merchant_edit_rows",
        staticmethod(fake_apply_merchant_edit_rows),
    )
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    await ReconciliationChecklistService.save_merchant_edit_items(
        session,  # type: ignore[arg-type]
        user=SimpleNamespace(id=9, username="qa", display_name="质检", role="org_admin", org_id=3),
        org_id=3,
        items=[
            {
                "unique_id": "UID-1",
                "sub_order_no": "SO-1",
                "receipt_merchant": "收款商家A",
                "merchant_net_amount": None,
                "payment_amount": None,
                "merchant_payment_time": None,
            }
        ],
    )

    assert captured_rows[0]["merchant_net_amount"] is None


@pytest.mark.asyncio
async def test_save_merchant_edit_items_skips_summary_rebuild_when_only_payment_time_changes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession()
    rebuilt_calls: list[tuple[int, list[int]]] = []

    async def fake_apply_merchant_edit_rows(_db, *, org_id: int, rows):
        assert org_id == 3
        _ = list(rows)
        return 1, 0, [], set()

    async def fake_rebuild(*_args, **_kwargs):
        rebuilt_calls.append((3, [202606]))

    async def fake_audit_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(
        ReconciliationChecklistService,
        "_apply_merchant_edit_rows",
        staticmethod(fake_apply_merchant_edit_rows),
    )
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    result = await ReconciliationChecklistService.save_merchant_edit_items(
        session,  # type: ignore[arg-type]
        user=SimpleNamespace(id=9, username="qa", display_name="质检", role="org_admin", org_id=3),
        org_id=3,
        items=[
            {
                "unique_id": "UID-1",
                "sub_order_no": "SO-1",
                "receipt_merchant": "收款商家A",
                "merchant_net_amount": "10.00",
                "payment_amount": "8.00",
                "merchant_payment_time": datetime(2026, 6, 6, 10, 0, 0),
            }
        ],
    )

    assert result == {
        "success_count": 1,
        "failed_count": 0,
        "unchanged_count": 0,
        "missing_sub_order_nos": [],
        "affected_periods": [],
        "error_messages": [],
    }
    assert rebuilt_calls == []
    assert session.flush_count == 1


@pytest.mark.asyncio
async def test_apply_merchant_edit_rows_marks_period_when_summary_fields_change(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(update_rowcount=1)
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "row_fingerprint": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "merchant_net_amount": Decimal("12.00"),
            "payment_amount": Decimal("9.00"),
            "merchant_payment_time": datetime(2026, 6, 6, 10, 0, 0),
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail = SimpleNamespace(
            id=101,
            accounting_period=202606,
            settlement_time=datetime(2026, 6, 1, 8, 0, 0),
            row_fingerprint="UID-1",
            receipt_merchant="收款商家A",
            merchant_net_amount=Decimal("91.50"),
            payment_amount=Decimal("90.00"),
        )
        return {"UID-1": detail}, {"SO-1": [detail]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_merchant_edit_rows(
        session,  # type: ignore[arg-type]
        org_id=9,
        rows=rows,
    )

    assert updated > 0
    assert failed == 0
    assert errors == []
    assert periods == {202606}


@pytest.mark.asyncio
async def test_apply_invoice_edit_rows_requires_unique_id_when_sub_order_has_multiple_matches(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(update_rowcount=0)
    rows = [
        {
            "source_row_number": 2,
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-1",
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail_a = SimpleNamespace(id=101, accounting_period=202606, settlement_time=datetime(2026, 6, 1, 8, 0, 0), row_fingerprint="UID-1")
        detail_b = SimpleNamespace(id=102, accounting_period=202606, settlement_time=datetime(2026, 6, 2, 8, 0, 0), row_fingerprint="UID-2")
        return {}, {"SO-1": [detail_a, detail_b]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_invoice_edit_rows(
        session,  # type: ignore[arg-type]
        org_id=9,
        rows=rows,
    )

    assert updated == 0
    assert failed == 1
    assert errors == ["Row 2: 唯一ID不能为空"]
    assert periods == set()


@pytest.mark.asyncio
async def test_apply_invoice_edit_rows_requires_unique_id_even_when_sub_order_is_unique(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _ManualEditSaveSession(update_rowcount=1)
    rows = [
        {
            "source_row_number": 2,
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "invoice_time": datetime(2026, 6, 6, 10, 0, 0),
            "invoice_number": "FP-1",
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail = SimpleNamespace(id=101, accounting_period=202606, settlement_time=datetime(2026, 6, 1, 8, 0, 0), row_fingerprint="UID-1")
        return {}, {"SO-1": [detail]}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_invoice_edit_rows(
        session,  # type: ignore[arg-type]
        org_id=9,
        rows=rows,
    )

    assert updated == 0
    assert failed == 1
    assert errors == ["Row 2: 唯一ID不能为空"]
    assert periods == set()


@pytest.mark.asyncio
async def test_create_manual_edit_upload_task_creates_new_invoice_task(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _CreateManualEditTaskSession()
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)
    delay_calls: list[int] = []

    async def fake_audit_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(
        "app.services.reconciliation_checklist_service.oss_service.upload_file",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("backend must not proxy manual edit uploads to OSS")),
    )
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)
    monkeypatch.setattr(
        "app.tasks.reconciliation_checklist.run_reconciliation_checklist_task",
        SimpleNamespace(delay=lambda task_id: delay_calls.append(task_id) or SimpleNamespace(id=f"manual-{task_id}")),
    )

    task = await ReconciliationChecklistService.create_manual_edit_upload_task_from_oss(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        original_name="invoice-edit.xlsx",
        oss_key="user-upload/reconciliation-checklist/manual-edits/9/200/invoice-edit.xlsx",
        file_size=1234,
        file_hash="abc123",
        task_type=CHECKLIST_TASK_TYPE_INVOICE_EDIT,
    )

    upload_file = next(item for item in session.added if isinstance(item, ReconciliationChecklistUploadFile))
    assert task.task_type == CHECKLIST_TASK_TYPE_INVOICE_EDIT
    assert task.file_id == upload_file.id
    assert upload_file.source_upload_file_id is None
    assert upload_file.original_name == "invoice-edit.xlsx"
    assert upload_file.oss_key == "user-upload/reconciliation-checklist/manual-edits/9/200/invoice-edit.xlsx"
    assert upload_file.file_size == 1234
    assert upload_file.file_hash == "abc123"
    assert delay_calls == []

    await session.run_callbacks()

    assert delay_calls == [task.id]
    assert task.celery_task_id == f"manual-{task.id}"


@pytest.mark.asyncio
async def test_manual_edit_upload_callback_reuses_initialized_upload_file(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _CreateManualEditTaskSession()
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)

    async def fake_audit_log(*_args, **_kwargs):
        return None

    monkeypatch.setattr(
        "app.services.reconciliation_checklist_service.oss_service.upload_file",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("backend must not proxy manual edit uploads to OSS")),
    )
    monkeypatch.setattr("app.services.reconciliation_checklist_service.AuditService.log", fake_audit_log)

    upload_file = await ReconciliationChecklistService.init_manual_edit_upload(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        original_name="merchant-edit.xlsx",
        file_size=4096,
        task_type=CHECKLIST_TASK_TYPE_MERCHANT_EDIT,
    )

    task = await ReconciliationChecklistService.create_manual_edit_upload_task_from_oss(
        session,  # type: ignore[arg-type]
        org_id=9,
        user=user,  # type: ignore[arg-type]
        file_id=upload_file.id,
        original_name=upload_file.original_name,
        oss_key=f"{upload_file.oss_key}.uploaded",
        file_size=4096,
        file_hash="def456",
        task_type=CHECKLIST_TASK_TYPE_MERCHANT_EDIT,
    )

    assert upload_file.status == "uploaded"
    assert upload_file.oss_key.endswith(".uploaded")
    assert upload_file.file_hash == "def456"
    assert task.file_id == upload_file.id
    assert task.task_type == CHECKLIST_TASK_TYPE_MERCHANT_EDIT


@pytest.mark.asyncio
async def test_create_manual_edit_upload_task_rejects_wrong_manual_task_type() -> None:
    session = _CreateManualEditTaskSession()
    user = SimpleNamespace(id=7, username="tester", display_name="测试用户", role="admin", org_id=9)

    with pytest.raises(ValueError, match="不支持的修改任务类型"):
        await ReconciliationChecklistService.create_manual_edit_upload_task_from_oss(
            session,  # type: ignore[arg-type]
            org_id=9,
            user=user,  # type: ignore[arg-type]
            original_name="merchant-edit.xlsx",
            oss_key="user-upload/reconciliation-checklist/manual-edits/9/200/merchant-edit.xlsx",
            file_size=1234,
            file_hash=None,
            task_type="source_import",
        )


@pytest.mark.asyncio
async def test_apply_merchant_rows_marks_period_when_receipt_merchant_changes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = _BatchWriteSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=9, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=9, user_id=4, original_name="merchant.xlsx", oss_key="oss-key", status="uploaded")
    rows = [
        {
            "source_row_number": 2,
            "unique_id": "UID-1",
            "sub_order_no": "SO-1",
            "receipt_merchant": "收款商家B",
            "merchant_net_amount": Decimal("91.50"),
            "payment_amount": Decimal("90.00"),
            "merchant_payment_time": datetime(2026, 6, 6, 10, 0, 0),
        }
    ]

    async def fake_resolve(*_args, **_kwargs):
        detail = SimpleNamespace(
            id=101,
            accounting_period=202606,
            settlement_time=datetime(2026, 6, 1, 8, 0, 0),
            row_fingerprint="UID-1",
            receipt_merchant="收款商家A",
            merchant_net_amount=Decimal("91.50"),
            payment_amount=Decimal("90.00"),
            merchant_payment_time=None,
        )
        return {"UID-1": detail}, {}

    monkeypatch.setattr(ReconciliationChecklistService, "_resolve_target_details", staticmethod(fake_resolve))

    updated, failed, errors, periods = await ReconciliationChecklistService._apply_merchant_rows(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        rows=rows,
    )

    assert updated > 0
    assert failed == 0
    assert errors == []
    assert periods == {202606}


@pytest.mark.asyncio
async def test_reconciliation_partitions_create_month_and_org_hash_subpartitions() -> None:
    session = _PartitionSession()

    result = await ensure_reconciliation_checklist_partitions_for_year(session, year=2026)  # type: ignore[arg-type]

    sql = "\n".join(session.statements)
    assert result == {"start_period": 202601, "end_period": 202612, "year": 2026}
    assert "CREATE TABLE IF NOT EXISTS fin_reconciliation_checklist_details_202601" in sql
    assert "PARTITION BY HASH (org_id)" in sql
    assert "fin_reconciliation_checklist_details_202601_h00" in sql
    assert "FOR VALUES WITH (MODULUS 16, REMAINDER 0)" in sql
    assert "fin_reconciliation_checklist_product_summary_rows_202601_h15" in sql
    assert "fin_reconciliation_checklist_receipt_summary_rows_202601_h15" in sql
    assert "fin_rcl_payable_balance_summary_202601_h15" in sql


def test_payable_balance_partition_prefix_stays_within_postgres_identifier_limit() -> None:
    partition = partition_name(RECONCILIATION_CHECKLIST_PAYABLE_BALANCE_SUMMARY_PARTITION, 202602)
    assert partition == "fin_rcl_payable_balance_summary_202602"
    assert len(partition) <= 58
    assert len(f"{partition}_h15") <= 63


@pytest.mark.asyncio
async def test_reconciliation_partition_maintenance_rebuilds_truncated_legacy_month_partition() -> None:
    session = _PartitionSession()
    session.partition_bounds_by_table[
        (
            "fin_reconciliation_checklist_payable_balance_summary_rows",
            "FOR VALUES FROM (202602) TO (202603)",
        )
    ] = "fin_reconciliation_checklist_payable_balance_summary_rows_20260"

    await ensure_reconciliation_checklist_partitions_for_year(session, year=2026)  # type: ignore[arg-type]

    sql = "\n".join(session.statements)
    assert 'DROP TABLE IF EXISTS "fin_reconciliation_checklist_payable_balance_summary_rows_20260" CASCADE' in sql
    assert "CREATE TABLE IF NOT EXISTS fin_rcl_payable_balance_summary_202602" in sql
    assert "CREATE TABLE IF NOT EXISTS fin_rcl_payable_balance_summary_202602_h15" in sql


@pytest.mark.asyncio
async def test_reconciliation_partition_maintenance_repairs_legacy_parent_primary_keys() -> None:
    session = _PartitionSession()
    for table_name in [
        "fin_reconciliation_checklist_details",
        "fin_reconciliation_checklist_product_summary_rows",
        "fin_reconciliation_checklist_receipt_summary_rows",
        "fin_reconciliation_checklist_payable_balance_summary_rows",
    ]:
        session.primary_keys_by_table[table_name] = ["id", "accounting_period"]

    await ensure_reconciliation_checklist_partitions_for_year(session, year=2026)  # type: ignore[arg-type]

    sql = "\n".join(session.statements)
    assert 'ALTER TABLE "fin_reconciliation_checklist_details" DROP CONSTRAINT IF EXISTS "fin_reconciliation_checklist_details_pkey"' in sql
    assert 'ALTER TABLE "fin_reconciliation_checklist_details" ADD PRIMARY KEY (id, org_id, accounting_period)' in sql
    assert 'ALTER TABLE "fin_reconciliation_checklist_product_summary_rows" ADD PRIMARY KEY (id, org_id, accounting_period)' in sql
    assert 'ALTER TABLE "fin_reconciliation_checklist_receipt_summary_rows" ADD PRIMARY KEY (id, org_id, accounting_period)' in sql
    assert 'ALTER TABLE "fin_reconciliation_checklist_payable_balance_summary_rows" ADD PRIMARY KEY (id, org_id, accounting_period)' in sql
    assert sql.index('ALTER TABLE "fin_reconciliation_checklist_details" ADD PRIMARY KEY') < sql.index("CREATE TABLE IF NOT EXISTS fin_reconciliation_checklist_details_202601")


@pytest.mark.asyncio
async def test_persist_task_result_does_not_check_partitions_for_update_periods(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _PersistTaskSession()
    task = ReconciliationChecklistTask(id=1, file_id=2, org_id=3, user_id=4, status="running", progress=5)
    upload_file = ReconciliationChecklistUploadFile(id=2, org_id=3, user_id=4, original_name="invoice.xlsx", oss_key="oss-key", status="uploaded")
    rebuilt: list[int] = []

    parse_result = {
        "file_type": CHECKLIST_FILE_TYPE_INVOICE,
        "total_rows": 1,
        "success_rows": 1,
        "failed_rows": 0,
        "errors": [],
        "warnings": [],
        "fatal_error": False,
        "rows": [
            {
                "source_row_number": 2,
                "sub_order_no": "SO-202602",
                "receipt_merchant": "收款商家A",
                "invoice_time": None,
                "invoice_number": "FP-1",
            }
        ],
        "result_message": "处理完成",
    }

    def fake_parse_file(*_args, **_kwargs):
        return parse_result

    async def fail_ensure_month_partition(*_args, **_kwargs):
        raise AssertionError("import persistence must not check or create partitions")

    async def fake_apply_invoice_rows(*_args, **_kwargs):
        return 1, 0, [], {202602}

    async def fake_rebuild_summaries(_db, *, org_id: int, periods):
        assert org_id == 3
        rebuilt.extend(sorted(periods))

    monkeypatch.setattr(ReconciliationChecklistService, "parse_file", staticmethod(fake_parse_file))
    monkeypatch.setattr("app.services.reconciliation_checklist_service.ensure_month_partition", fail_ensure_month_partition, raising=False)
    monkeypatch.setattr(ReconciliationChecklistService, "_apply_invoice_rows", staticmethod(fake_apply_invoice_rows))
    monkeypatch.setattr(ReconciliationChecklistService, "_rebuild_summaries", staticmethod(fake_rebuild_summaries))

    summary = await ReconciliationChecklistService.persist_task_result(
        session,  # type: ignore[arg-type]
        task=task,
        upload_file=upload_file,
        file_path="invoice.xlsx",
    )

    assert summary["更新行数"] == 1
    assert summary["涉及年月"] == [202602]
    assert rebuilt == [202602]


@pytest.mark.asyncio
async def test_rebuild_summaries_does_not_check_target_period_partitions(monkeypatch: pytest.MonkeyPatch) -> None:
    session = _RebuildSummarySession()

    async def fail_ensure_month_partition(*_args, **_kwargs):
        raise AssertionError("summary rebuild must not check or create partitions")

    monkeypatch.setattr("app.services.reconciliation_checklist_service.ensure_month_partition", fail_ensure_month_partition, raising=False)

    await ReconciliationChecklistService._rebuild_summaries(
        session,  # type: ignore[arg-type]
        org_id=1,
        periods={202602},
    )

    sql = "\n".join(session.statements)
    assert "DELETE FROM fin_reconciliation_checklist_product_summary_rows" in sql
    assert "INSERT INTO fin_reconciliation_checklist_product_summary_rows" in sql


def test_summary_export_specs_are_registered() -> None:
    from app.services.export_job_service import EXPORT_SPECS

    assert "reconciliation_checklist.product_summary" in EXPORT_SPECS
    assert "reconciliation_checklist.receipt_summary" in EXPORT_SPECS
    assert "reconciliation_checklist.payable_balance_summary" in EXPORT_SPECS
    assert EXPORT_SPECS["reconciliation_checklist.summary"].module == "reconciliation_checklist"


@pytest.mark.asyncio
async def test_export_product_summary_file_uses_reference_layout(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    output_path = tmp_path / "product_summary.xlsx"
    rows = [
        {
            "key": "1:202606:收款商家A:商户主体A:商品A",
            "org_id": 1,
            "org_name": "组织A",
            "accounting_year": 2026,
            "accounting_month": 6,
            "accounting_period": 202606,
            "receipt_merchant": "收款商家A",
            "merchant_subject_name": "商户主体A",
            "product_name": "商品A",
            "product_quantity": 2,
            "total_user_paid_amount": Decimal("100.50"),
            "total_live_commission": Decimal("9.00"),
            "total_merchant_net_amount": Decimal("91.50"),
        }
    ]

    async def fake_list_product_summary(*_args, **_kwargs):
        return rows, 1

    monkeypatch.setattr(ReconciliationChecklistService, "list_product_summary", staticmethod(fake_list_product_summary))

    row_count = await ReconciliationChecklistService.export_product_summary_to_file(
        SimpleNamespace(),  # type: ignore[arg-type]
        user=SimpleNamespace(role="admin", org_id=1),  # type: ignore[arg-type]
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    try:
        ws = workbook.active
        assert row_count == 1
        assert ws["A1"].value == "直播运营费用明细"
        assert ws["A3"].value == "收款商家"
        assert ws["C3"].value == "商户主体名称"
        assert ws["B6"].value == "商品A"
        assert ws["A7"].value == "总计"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_product_summary_current_page_forwards_pagination_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output_path = tmp_path / "product_summary_current_page.xlsx"
    captured_kwargs: dict[str, object] = {}
    rows = [
        {
            "key": "1:202606:收款商家A:商户主体A:商品A",
            "org_id": 1,
            "org_name": "组织A",
            "accounting_year": 2026,
            "accounting_month": 6,
            "accounting_period": 202606,
            "receipt_merchant": "收款商家A",
            "merchant_subject_name": "商户主体A",
            "product_name": "商品A",
            "product_quantity": 2,
            "total_user_paid_amount": Decimal("100.50"),
            "total_live_commission": Decimal("9.00"),
            "total_merchant_net_amount": Decimal("91.50"),
        }
    ]

    async def fake_list_product_summary(*_args, **kwargs):
        captured_kwargs.update(kwargs)
        return rows, 1

    monkeypatch.setattr(ReconciliationChecklistService, "list_product_summary", staticmethod(fake_list_product_summary))

    row_count = await ReconciliationChecklistService.export_product_summary_to_file(
        SimpleNamespace(),  # type: ignore[arg-type]
        user=SimpleNamespace(role="admin", org_id=1),  # type: ignore[arg-type]
        output_path=output_path,
        page=2,
        page_size=50,
    )

    workbook = load_workbook(output_path)
    try:
        ws = workbook.active
        assert row_count == 1
        assert captured_kwargs["page"] == 2
        assert captured_kwargs["page_size"] == 50
        assert ws["B6"].value == "商品A"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_receipt_summary_current_page_forwards_pagination_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output_path = tmp_path / "receipt_summary_current_page.xlsx"
    captured_kwargs: dict[str, object] = {}
    rows = [
        {
            "key": "1:202606:商户主体A:抖音:收款商家A",
            "org_id": 1,
            "org_name": "组织A",
            "accounting_year": 2026,
            "accounting_month": 6,
            "accounting_period": 202606,
            "merchant_subject_name": "商户主体A",
            "live_platform": "抖音",
            "receipt_merchant": "收款商家A",
            "order_count": 2,
            "total_user_paid_amount": Decimal("100.50"),
            "total_live_commission": Decimal("9.00"),
            "total_merchant_net_amount": Decimal("91.50"),
        }
    ]

    async def fake_list_receipt_summary(*_args, **kwargs):
        captured_kwargs.update(kwargs)
        return rows, 1

    monkeypatch.setattr(ReconciliationChecklistService, "list_receipt_summary", staticmethod(fake_list_receipt_summary))

    row_count = await ReconciliationChecklistService.export_receipt_summary_to_file(
        SimpleNamespace(),  # type: ignore[arg-type]
        user=SimpleNamespace(role="admin", org_id=1),  # type: ignore[arg-type]
        output_path=output_path,
        page=3,
        page_size=20,
    )

    workbook = load_workbook(output_path)
    try:
        ws = workbook.active
        assert row_count == 1
        assert captured_kwargs["page"] == 3
        assert captured_kwargs["page_size"] == 20
        assert ws["A1"].value == "商户主体名称"
        assert ws["A6"].value == 1
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_payable_balance_summary_current_page_forwards_pagination_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output_path = tmp_path / "payable_balance_summary.xlsx"
    captured_kwargs: dict[str, object] = {}
    rows = [
        {
            "key": "1:202606:商户主体A:收款商家A",
            "org_id": 1,
            "org_name": "组织A",
            "accounting_year": 2026,
            "accounting_month": 6,
            "accounting_period": 202606,
            "merchant_subject_name": "商户主体A",
            "receipt_merchant": "收款商家A",
            "total_user_paid_amount": Decimal("100.50"),
            "total_merchant_net_amount": Decimal("91.50"),
            "total_payment_amount": Decimal("90.00"),
            "total_merchant_net_balance": Decimal("1.50"),
        }
    ]

    async def fake_list_payable_balance_summary(*_args, **kwargs):
        captured_kwargs.update(kwargs)
        return rows, 1

    monkeypatch.setattr(
        ReconciliationChecklistService,
        "list_payable_balance_summary",
        staticmethod(fake_list_payable_balance_summary),
    )

    row_count = await ReconciliationChecklistService.export_payable_balance_summary_to_file(
        SimpleNamespace(),  # type: ignore[arg-type]
        user=SimpleNamespace(role="admin", org_id=1),  # type: ignore[arg-type]
        output_path=output_path,
        page=2,
        page_size=50,
    )

    workbook = load_workbook(output_path)
    try:
        ws = workbook.active
        assert row_count == 1
        assert captured_kwargs["page"] == 2
        assert captured_kwargs["page_size"] == 50
        assert ws["A1"].value == "商户主体名称"
        assert ws["A2"].value == "商户主体A"
        assert ws["G3"].value == float(Decimal("1.50"))
    finally:
        workbook.close()
