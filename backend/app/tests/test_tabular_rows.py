from pathlib import Path

import pytest
from openpyxl import Workbook

import app.tasks.processors.base as base_module
from app.tasks.processors.base import TabularFileOpenError, open_tabular_rows


def _read_rows(path: Path) -> list[list[object]]:
    with open_tabular_rows(str(path)) as rows:
        assert rows is not None
        return [list(row) for row in rows]


def test_open_tabular_rows_reads_csv_even_when_suffix_is_xlsx(tmp_path: Path) -> None:
    file_path = tmp_path / "wrong_suffix.xlsx"
    file_path.write_text("日期,金额\n2026-05-01,12.30\n", encoding="utf-8-sig")

    assert _read_rows(file_path) == [["日期", "金额"], ["2026-05-01", "12.30"]]


def test_open_tabular_rows_reads_html_table_even_when_suffix_is_xlsx(tmp_path: Path) -> None:
    file_path = tmp_path / "html_export.xlsx"
    file_path.write_text(
        "<html><body><table><tr><th>日期</th><th>金额</th></tr><tr><td>2026-05-01</td><td>12.30</td></tr></table></body></html>",
        encoding="gb18030",
    )

    assert _read_rows(file_path) == [["日期", "金额"], ["2026-05-01", "12.30"]]


def test_open_tabular_rows_rejects_non_tabular_xlsx_with_clear_error(tmp_path: Path) -> None:
    file_path = tmp_path / "bad.xlsx"
    file_path.write_text("this is not a table", encoding="utf-8")

    try:
        _read_rows(file_path)
    except TabularFileOpenError as exc:
        assert "不是 xlsx 压缩包" in str(exc)
    else:
        raise AssertionError("TabularFileOpenError was not raised")


def test_open_tabular_rows_prefers_calamine_for_xlsx(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    file_path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "金额"])
    sheet.append(["2026-05-01", 12.3])
    workbook.save(file_path)
    workbook.close()

    calls = {"calamine": 0, "openpyxl": 0}

    def fake_iter_calamine_rows(_file_path: str):
        calls["calamine"] += 1
        yield ("日期", "金额")
        yield ("2026-05-01", 12.3)

    def fail_load_workbook(*args, **kwargs):
        calls["openpyxl"] += 1
        raise AssertionError("openpyxl should not run when calamine succeeds")

    monkeypatch.setattr(base_module, "iter_calamine_rows", fake_iter_calamine_rows)
    monkeypatch.setattr(base_module, "load_workbook", fail_load_workbook)

    assert _read_rows(file_path) == [["日期", "金额"], ["2026-05-01", 12.3]]
    assert calls == {"calamine": 1, "openpyxl": 0}


def test_open_tabular_rows_falls_back_to_openpyxl_when_calamine_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    file_path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "金额"])
    sheet.append(["2026-05-01", 12.3])
    workbook.save(file_path)
    workbook.close()

    def fail_iter_calamine_rows(_file_path: str):
        raise ValueError("calamine failed")
        yield

    monkeypatch.setattr(base_module, "iter_calamine_rows", fail_iter_calamine_rows)

    assert _read_rows(file_path) == [["日期", "金额"], ["2026-05-01", 12.3]]
