"""Base classes for platform-specific tabular file processors.

The template here owns file IO, header lookup, row iteration, error accounting,
and monthly aggregation. Platform processors only implement date selection and
field formulas.
"""

import csv
from html.parser import HTMLParser
import logging
import re
from zipfile import BadZipFile
from abc import ABC, abstractmethod
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from xlrd import XL_CELL_DATE, XL_CELL_NUMBER, XLRDError, open_workbook, xldate_as_datetime

from app.utils.money import ZERO_MONEY, safe_decimal

logger = logging.getLogger("finengine.processor")

DATE_FORMATS: tuple[str, ...] = (
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y年%m月%d日 %H:%M:%S",
    "%Y年%m月%d日 %H:%M",
    "%Y年%m月%d日",
    "%Y.%m.%d %H:%M:%S",
    "%Y.%m.%d %H:%M",
    "%Y.%m.%d",
)

FINANCIAL_SUMMARY_FIELDS: tuple[str, ...] = (
    "order_paid_amount",
    "refund_amount",
    "gmv",
    "platform_income",
    "platform_fee",
    "return_cost",
    "commission",
    "merchant_fee",
    "promotion_fee",
    "provider_commission",
    "donation_fee",
)
POSITIVE_SUMMARY_FIELDS: frozenset[str] = frozenset(
    (
        "refund_amount",
        "commission",
        "merchant_fee",
        "promotion_fee",
        "provider_commission",
    )
)
CSV_ENCODINGS: tuple[str, ...] = ("utf-8-sig", "utf-8", "gb18030")
HEADER_SCAN_LIMIT = 20
HEADER_NORMALIZE_PATTERN = re.compile(r"[\s　\uFEFF\u200B-\u200D]+")
ZIP_SIGNATURES: tuple[bytes, ...] = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
OLE2_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
TEXT_SAMPLE_SIZE = 4096


class TabularFileOpenError(ValueError):
    """Raised when an uploaded file cannot be opened as a supported table."""


def parse_datetime(value: object) -> datetime | None:
    """Try multiple date formats to parse a datetime value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value

    s = str(value).strip()
    if not s:
        return None

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def safe_str(value: object) -> str:
    """Convert a cell value to stripped string, returning '' for None."""
    if value is None:
        return ""
    return str(value).strip()


def canonical_remark(value: str):
    if not value:
        return ""

    return "".join(re.findall(r"[\u4e00-\u9fff]+", value))


def canonical_header(value: object) -> str:
    """Normalize small header variants shared by frontend and backend matching."""
    header = HEADER_NORMALIZE_PATTERN.sub("", safe_str(value))
    return header.replace("帐", "账").replace("（", "(").replace("）", ")").lower()


def normalize_positive_summary_fields(values: dict[str, Decimal]) -> dict[str, Decimal]:
    """Normalize fee/refund fields that should display as positive totals."""
    for field in POSITIVE_SUMMARY_FIELDS:
        if field in values:
            values[field] = abs(safe_decimal(values[field]))
    return values


def detect_csv_encoding(file_path: str) -> str:
    """Detect common encodings used by Chinese marketplace CSV exports."""
    path = Path(file_path)
    with path.open("rb") as file:
        sample = file.read(64 * 1024)

    for encoding in CSV_ENCODINGS:
        try:
            sample.decode(encoding)
        except UnicodeDecodeError:
            continue
        return encoding
    return "gb18030"


def detect_csv_dialect(file_path: str, encoding: str) -> type[csv.Dialect] | csv.Dialect:
    """Best-effort delimiter detection; default to normal comma CSV."""
    path = Path(file_path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as file:
        sample = file.read(8192)

    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return csv.excel


def _read_file_signature(file_path: str, size: int = TEXT_SAMPLE_SIZE) -> bytes:
    path = Path(file_path)
    with path.open("rb") as file:
        return file.read(size)


def _has_zip_signature(sample: bytes) -> bool:
    return sample.startswith(ZIP_SIGNATURES)


def _has_ole2_signature(sample: bytes) -> bool:
    return sample.startswith(OLE2_SIGNATURE)


def _sample_is_binary(sample: bytes) -> bool:
    if not sample:
        return False
    if b"\x00" in sample:
        return True
    control_bytes = sum(1 for byte in sample if byte < 32 and byte not in b"\t\n\r\f\b")
    return control_bytes / len(sample) > 0.30


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        _ = attrs
        if tag.lower() == "tr":
            self._current_row = []
        elif tag.lower() in {"td", "th"} and self._current_row is not None:
            self._current_cell = []

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        lower_tag = tag.lower()
        if lower_tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            self._current_row.append("".join(self._current_cell).strip())
            self._current_cell = None
        elif lower_tag == "tr" and self._current_row is not None:
            if any(cell for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None


def _looks_like_html_table(file_path: str) -> bool:
    sample = _read_file_signature(file_path, 2048).lstrip().lower()
    return sample.startswith((b"<html", b"<!doctype html", b"<table")) or b"<table" in sample[:512]


def _looks_like_csv_text(file_path: str) -> bool:
    sample = _read_file_signature(file_path)
    if not sample or _sample_is_binary(sample):
        return False

    encoding = detect_csv_encoding(file_path)
    text = sample.decode(encoding, errors="replace").lstrip("\ufeff")
    if "<table" in text[:512].lower():
        return False

    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return False
    first_lines = lines[: min(10, len(lines))]
    delimiters = (",", "\t", ";")
    return any(any(delimiter in line for delimiter in delimiters) for line in first_lines)


def read_html_table_rows(file_path: str) -> list[list[str]]:
    encoding = detect_csv_encoding(file_path)
    path = Path(file_path)
    with path.open("r", encoding=encoding, errors="replace") as file:
        content = file.read()

    parser = _HtmlTableParser()
    parser.feed(content)
    return parser.rows


def iter_xls_rows(file_path: str) -> Iterator[tuple[object, ...]]:
    workbook = open_workbook(file_path)
    if workbook.nsheets == 0:
        return
    sheet = workbook.sheet_by_index(0)
    for row_index in range(sheet.nrows):
        values = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            value = cell.value
            if cell.ctype == XL_CELL_DATE:
                value = xldate_as_datetime(value, workbook.datemode)
            elif cell.ctype == XL_CELL_NUMBER and isinstance(value, float) and value.is_integer():
                value = int(value)
            values.append(value)
        yield tuple(values)


def _tabular_open_error(file_path: str, reason: str) -> TabularFileOpenError:
    suffix = Path(file_path).suffix.lower() or "无后缀"
    return TabularFileOpenError(f"无法打开表格文件：文件内容不是支持的 Excel/CSV 格式（后缀 {suffix}，{reason}）")


@contextmanager
def open_tabular_rows(file_path: str) -> Iterator[Iterable[tuple[object, ...] | list[str]] | None]:
    """Yield rows from Excel/CSV files with the same row shape."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    sample = _read_file_signature(file_path)

    if _looks_like_html_table(file_path):
        yield read_html_table_rows(file_path)
        return

    if suffix == ".xls" or _has_ole2_signature(sample):
        try:
            yield iter_xls_rows(file_path)
        except XLRDError as exc:
            raise _tabular_open_error(file_path, "xls 解析失败") from exc
        return

    if _has_zip_signature(sample):
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as exc:
            raise _tabular_open_error(file_path, "xlsx 解析失败") from exc
        try:
            ws = wb.active
            if ws is None:
                yield None
                return
            ws.reset_dimensions()
            yield ws.iter_rows(values_only=True)
        finally:
            wb.close()
        return

    if suffix == ".csv" or _looks_like_csv_text(file_path):
        encoding = detect_csv_encoding(file_path)
        dialect = detect_csv_dialect(file_path, encoding)
        with path.open("r", encoding=encoding, errors="replace", newline="") as file:
            yield csv.reader(file, dialect=dialect)
        return

    if suffix in {".xlsx", ".xlsm"}:
        raise _tabular_open_error(file_path, "不是 xlsx 压缩包")

    raise _tabular_open_error(file_path, "无法识别文件内容")


@dataclass(frozen=True)
class GroupKey:
    """Hashable composite key for (shop, year, month)."""

    shop: str
    year: int
    month: int


class ProcessingResult(dict):
    """Result dict compatible with the existing Celery ingestion flow."""

    @classmethod
    def empty(cls, error: str | None = None) -> "ProcessingResult":
        result = cls(
            total_rows=0,
            success_rows=0,
            failed_rows=0,
            errors=[],
            groups={},
        )
        if error:
            result["errors"].append(error)
        return result


class BasePlatformProcessor(ABC):
    """Base interface for platform processors."""

    def process(
        self,
        file_path: str,
        shop_name: str,
        type_code: str = "动账",
        category_dict: dict[str, list[str]] | None = None,
    ) -> dict:
        lower_type = type_code.lower()
        normalized_type = lower_type if lower_type in {"bic", "gmv"} else type_code
        handler = self.get_type_processor(normalized_type)
        if handler is None:
            return ProcessingResult.empty(error=f"{self.platform_label}暂不支持 [{type_code}] 类型文件")
        try:
            return handler(file_path=file_path, shop_name=shop_name, category_dict=category_dict)
        except TabularFileOpenError as exc:
            logger.warning(
                "表格文件打开失败 platform=%s type=%s file=%s error=%s",
                self.platform_label,
                normalized_type,
                file_path,
                exc,
            )
            return ProcessingResult.empty(error=str(exc))

    @property
    @abstractmethod
    def platform_label(self) -> str:
        """Human-readable platform label used in error messages."""

    def get_type_processor(self, type_code: str):
        if type_code == "动账":
            return self._process_financial_summary
        return None


class FinancialSummaryStrategy(ABC):
    """Platform strategy for row formulas that write FinancialSummary fields."""

    fields: tuple[str, ...] = FINANCIAL_SUMMARY_FIELDS

    @property
    @abstractmethod
    def required_headers(self) -> tuple[str, ...]:
        """Headers required by this strategy."""

    @abstractmethod
    def compute_year_month(self, vals: dict[str, object]) -> tuple[int, int] | None:
        """Return the aggregation month for a row."""

    @abstractmethod
    def compute_values(
        self,
        vals: dict[str, object],
        category_dict: dict[str, list[str]] | None = None,
    ) -> dict[str, Decimal]:
        """Return FinancialSummary-compatible values for a row."""

    def empty_agg(self) -> dict[str, Decimal]:
        return {field: ZERO_MONEY for field in self.fields}

    def finalize_agg(self, agg: dict[str, Decimal]) -> dict[str, Decimal]:
        return normalize_positive_summary_fields(agg)


class FinancialSummaryExcelProcessorMixin(BasePlatformProcessor):
    """Template method implementation for summary-field tabular processors."""

    summary_strategy: FinancialSummaryStrategy

    def _process_financial_summary(
        self,
        *,
        file_path: str,
        shop_name: str,
        category_dict: dict[str, list[str]] | None = None,
    ) -> dict:
        return self._process_financial_summary_with_strategy(
            file_path=file_path,
            shop_name=shop_name,
            category_dict=category_dict,
            strategy=self.summary_strategy,
        )

    def _process_financial_summary_with_strategy(
        self,
        *,
        file_path: str,
        shop_name: str,
        category_dict: dict[str, list[str]] | None = None,
        strategy: FinancialSummaryStrategy,
    ) -> dict:
        result = ProcessingResult.empty()

        with open_tabular_rows(file_path) as rows:
            if rows is None:
                result["errors"].append("无法打开表格文件")
                return result

            row_iter = iter(rows)
            header_result = self._find_header_row(row_iter, strategy.required_headers)
            if header_result is None:
                result["errors"].append("无法读取表头")
                self._log_header_compare(
                    file_path=file_path,
                    type_code=self._current_type_code(strategy),
                    actual_headers=[],
                    required_headers=strategy.required_headers,
                    missing_headers=list(strategy.required_headers),
                    header_row_number=None,
                )
                return result

            headers, _header_row_number = header_result
            col_idx = self._build_col_idx(headers, strategy.required_headers)
            missing = self._missing_headers(col_idx, strategy.required_headers)
            if missing:
                result["errors"].append(f"缺少必要表头: {', '.join(missing)}")
                self._log_header_compare(
                    file_path=file_path,
                    type_code=self._current_type_code(strategy),
                    actual_headers=headers,
                    required_headers=strategy.required_headers,
                    missing_headers=missing,
                    header_row_number=_header_row_number,
                )
                return result

            groups: dict[GroupKey, dict[str, Decimal]] = {}

            for row in row_iter:
                result["total_rows"] += 1
                try:
                    vals = self._row_to_values(row, col_idx)
                    year_month = strategy.compute_year_month(vals)
                    if year_month is None:
                        result["failed_rows"] += 1
                        result["errors"].append(f"Row {result['total_rows'] + 1}: 无法解析归属年月")
                        continue

                    key = GroupKey(shop=shop_name, year=year_month[0], month=year_month[1])
                    agg = groups.setdefault(key, strategy.empty_agg())
                    row_values = strategy.compute_values(vals, category_dict)
                    for field, value in row_values.items():
                        agg[field] = agg.get(field, ZERO_MONEY) + safe_decimal(value)

                    result["success_rows"] += 1
                except Exception as e:
                    result["failed_rows"] += 1
                    result["errors"].append(f"Row {result['total_rows'] + 1}: {e}")

        result["groups"] = self._serialize_groups({key: strategy.finalize_agg(agg) for key, agg in groups.items()})
        return result

    @staticmethod
    def _read_headers(row: Iterable[object]) -> list[str]:
        return [safe_str(h) for h in row]

    @classmethod
    def _find_header_row(
        cls,
        row_iter: Iterator[tuple[object, ...] | list[str]],
        required_headers: Iterable[str],
    ) -> tuple[list[str], int] | None:
        last_headers: list[str] = []
        for row_number, row in enumerate(row_iter, start=1):
            headers = cls._read_headers(row)
            if any(headers):
                last_headers = headers
                col_idx = cls._build_col_idx(headers, required_headers)
                if not cls._missing_headers(col_idx, required_headers):
                    return headers, row_number
            if row_number >= HEADER_SCAN_LIMIT:
                break
        if last_headers:
            return last_headers, min(HEADER_SCAN_LIMIT, row_number)
        return None

    @staticmethod
    def _build_col_idx(headers: Iterable[str], expected_headers: Iterable[str] | None = None) -> dict[str, int]:
        col_idx: dict[str, int] = {}
        canonical_to_name = {canonical_header(h): h for h in expected_headers or []}
        for idx, header in enumerate(headers):
            canonical = canonical_header(header)
            if not canonical:
                continue
            col_idx[canonical_to_name.get(canonical, header)] = idx
        return col_idx

    @staticmethod
    def _missing_headers(col_idx: dict[str, int], required_headers: Iterable[str]) -> list[str]:
        return [h for h in required_headers if h not in col_idx]

    @staticmethod
    def _row_to_values(row: tuple[object, ...] | list[str], col_idx: dict[str, int]) -> dict[str, object]:
        return {h: row[idx] if idx < len(row) else None for h, idx in col_idx.items()}

    @staticmethod
    def _serialize_groups(groups: dict[GroupKey, dict[str, Decimal]]) -> dict[str, dict[str, Decimal]]:
        return {f"{k.shop}|{k.year}|{k.month}": v for k, v in groups.items()}

    def _current_type_code(self, strategy: FinancialSummaryStrategy) -> str:
        if strategy is self.summary_strategy:
            return "动账"
        return strategy.__class__.__name__

    def _log_header_compare(
        self,
        *,
        file_path: str,
        type_code: str,
        actual_headers: Iterable[str],
        required_headers: Iterable[str],
        missing_headers: Iterable[str],
        header_row_number: int | None,
    ) -> None:
        actual = [safe_str(h) for h in actual_headers]
        required = [safe_str(h) for h in required_headers]
        missing = [safe_str(h) for h in missing_headers]
        logger.warning(
            "文件表头校验失败 platform=%s type=%s file=%s header_row=%s actual_count=%s required_count=%s missing=%s actual_headers=%s required_headers=%s",
            self.platform_label,
            type_code,
            file_path,
            header_row_number,
            len(actual),
            len(required),
            missing,
            actual,
            required,
        )


class SimpleMonthlySumProcessorMixin(BasePlatformProcessor):
    """Template for files that aggregate one amount column by one date column."""

    def _process_simple_monthly_sum(
        self,
        *,
        file_path: str,
        shop_name: str,
        date_header: str,
        amount_header: str,
        output_key: str,
    ) -> dict:
        result = ProcessingResult.empty()

        with open_tabular_rows(file_path) as rows:
            if rows is None:
                result["errors"].append("无法打开表格文件")
                return result

            row_iter = iter(rows)
            required_headers = (date_header, amount_header)
            header_result = FinancialSummaryExcelProcessorMixin._find_header_row(row_iter, required_headers)
            if header_result is None:
                result["errors"].append("无法读取表头")
                self._log_header_compare(
                    file_path=file_path,
                    type_code=output_key,
                    actual_headers=[],
                    required_headers=required_headers,
                    missing_headers=required_headers,
                    header_row_number=None,
                )
                return result

            headers, _header_row_number = header_result
            col_idx = FinancialSummaryExcelProcessorMixin._build_col_idx(headers, required_headers)
            missing = FinancialSummaryExcelProcessorMixin._missing_headers(col_idx, required_headers)
            if missing:
                result["errors"].append(f"缺少必要表头: {', '.join(missing)}")
                self._log_header_compare(
                    file_path=file_path,
                    type_code=output_key,
                    actual_headers=headers,
                    required_headers=required_headers,
                    missing_headers=missing,
                    header_row_number=_header_row_number,
                )
                return result

            groups: dict[GroupKey, dict[str, Decimal]] = {}

            for row in row_iter:
                result["total_rows"] += 1
                try:
                    date_value = row[col_idx[date_header]] if col_idx[date_header] < len(row) else None
                    dt = parse_datetime(date_value)
                    if dt is None:
                        result["failed_rows"] += 1
                        result["errors"].append(f"Row {result['total_rows'] + 1}: 无法解析{date_header}")
                        continue

                    amount = row[col_idx[amount_header]] if col_idx[amount_header] < len(row) else None
                    key = GroupKey(shop=shop_name, year=dt.year, month=dt.month)
                    agg = groups.setdefault(key, {output_key: ZERO_MONEY})
                    agg[output_key] += safe_decimal(amount)
                    result["success_rows"] += 1
                except Exception as e:
                    result["failed_rows"] += 1
                    result["errors"].append(f"Row {result['total_rows'] + 1}: {e}")

        result["groups"] = FinancialSummaryExcelProcessorMixin._serialize_groups(groups)
        return result
