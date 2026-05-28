"""Summary service — query and export financial summaries."""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.organization import Organization
from app.models.shop import Shop
from app.models.summary import FinancialSummary
from app.services.douyin_dongzhang_detail_service import (
    DONGZHANG_DETAIL_EXPORT_COLUMNS,
    MONEY_FIELDS as DONGZHANG_DETAIL_MONEY_FIELDS,
    DouyinDongzhangDetailService,
)
from app.services.summary_adjustment_service import SummaryAdjustmentService
from app.services.shop_visibility import active_shop_filter

# Export header aligned with actual business Excel.
EXPORT_HEADERS = [
    "核算年月",
    "归属年月",
    "平台",
    "店铺",
    "订单实付金额",
    "退款金额",
    "实收GMV",
    "平台其他收入",
    "平台服务费",
    "退货费用及其他费用",
    "达人佣金",
    "招商服务费",
    "站外推广费",
    "服务商佣金",
    "支付捐赠费用",
    "运费险",
    "BIC",
]

REPORT_EXPORT_HEADERS = [
    "核算年月",
    "平台",
    "店铺",
    "订单实付金额",
    "退款金额",
    "实收GMV",
    "平台其他收入",
    "平台服务费",
    "退货费用及其他费用",
    "达人佣金",
    "招商服务费",
    "站外推广费",
    "服务商佣金",
    "支付捐赠费用",
    "运费险",
    "BIC",
]

SUMMARY_MONEY_FIELDS = (
    "order_paid_amount",
    "refund_amount",
    "gmv",
    "platform_income",
    "platform_fee",
    "return_cost",
    "commission",
    "merchant_fee",
    "promotion_fee",
    "provider_commission",
    "donation_fee",
    "insurance_fee",
    "bic",
)


def split_filter_values(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def split_int_filter_values(value: str | None) -> list[int]:
    if not value:
        return []
    values: list[int] = []
    for item in value.split(","):
        item = item.strip()
        if item:
            values.append(int(item))
    return values


def single_filter_value(value: str | None) -> str | None:
    values = split_filter_values(value)
    return values[0] if len(values) == 1 else None


def single_int_filter_value(value: str | None) -> int | None:
    values = split_int_filter_values(value)
    return values[0] if len(values) == 1 else None


def month_period_value(year: int | None, month: int | None) -> int | None:
    if year is None or month is None:
        return None
    return int(year) * 100 + int(month)


def build_month_range_filters(
    period_expr,
    *,
    start_year: int | None = None,
    start_month: int | None = None,
    end_year: int | None = None,
    end_month: int | None = None,
) -> list:
    filters = []
    start_period = month_period_value(start_year, start_month)
    end_period = month_period_value(end_year, end_month)
    if start_period is not None:
        filters.append(period_expr >= start_period)
    if end_period is not None:
        filters.append(period_expr <= end_period)
    return filters


PLATFORM_LABELS = {
    "douyin": "抖音",
    "抖店": "抖音",
    "kuaishou": "快手",
    "快手": "快手",
    "xiaohongshu": "小红书",
    "小红书": "小红书",
    "weixin_video": "微信小店",
    "视频号": "微信小店",
    "微信小店": "微信小店",
    "tmall": "天猫",
    "天猫": "天猫",
    "taobao": "淘宝",
    "淘宝": "淘宝",
    "alipay": "支付宝",
    "支付宝": "支付宝",
    "qianniu": "千牛",
    "千牛": "千牛",
    "miniprogram": "小程序",
    "小程序": "小程序",
}

EXCEL_EXPORT_ROW_LIMIT = 20000


def _ensure_export_row_limit(label: str, row_count: int) -> None:
    if row_count > EXCEL_EXPORT_ROW_LIMIT:
        raise ValueError(f"{label}导出数据量 {row_count} 行，超过系统上限 {EXCEL_EXPORT_ROW_LIMIT} 行，请缩小筛选范围后再导出")


def _write_only_header_row(sheet, headers: list[str]) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for label in headers:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = Font(bold=True)
        cells.append(cell)
    return cells


class SummaryService:
    @staticmethod
    async def list_summaries(
        db: AsyncSession,
        *,
        org_ids: list[int] | None,
        summary_year: int | None = None,
        summary_month: int | None = None,
        summary_start_year: int | None = None,
        summary_start_month: int | None = None,
        summary_end_year: int | None = None,
        summary_end_month: int | None = None,
        source_year: int | None = None,
        source_month: int | None = None,
        source_start_year: int | None = None,
        source_start_month: int | None = None,
        source_end_year: int | None = None,
        source_end_month: int | None = None,
        platform_name: str | None = None,
        report_platform_name: str | None = None,
        shop_ids: str | None = None,
        shop_name: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict], int]:
        """Query financial summaries with filters."""
        stmt = (
            select(
                FinancialSummary,
                Shop.shop_color.label("shop_color"),
                Organization.name.label("org_name"),
            )
            .outerjoin(Shop, FinancialSummary.shop_id == Shop.id)
            .outerjoin(Organization, FinancialSummary.org_id == Organization.id)
            .where(
                FinancialSummary.is_deleted.is_(False),
                active_shop_filter(FinancialSummary.shop_id),
            )
        )
        count_stmt = (
            select(func.count())
            .select_from(FinancialSummary)
            .where(
                FinancialSummary.is_deleted.is_(False),
                active_shop_filter(FinancialSummary.shop_id),
            )
        )
        if org_ids is not None:
            stmt = stmt.where(FinancialSummary.org_id.in_(org_ids))
            count_stmt = count_stmt.where(FinancialSummary.org_id.in_(org_ids))

        if summary_year is not None:
            stmt = stmt.where(FinancialSummary.summary_year == summary_year)
            count_stmt = count_stmt.where(FinancialSummary.summary_year == summary_year)

        if summary_month is not None:
            stmt = stmt.where(FinancialSummary.summary_month == summary_month)
            count_stmt = count_stmt.where(FinancialSummary.summary_month == summary_month)

        if source_year is not None:
            stmt = stmt.where(FinancialSummary.source_year == source_year)
            count_stmt = count_stmt.where(FinancialSummary.source_year == source_year)

        if source_month is not None:
            stmt = stmt.where(FinancialSummary.source_month == source_month)
            count_stmt = count_stmt.where(FinancialSummary.source_month == source_month)

        summary_period = FinancialSummary.summary_year * 100 + FinancialSummary.summary_month
        for period_filter in build_month_range_filters(
            summary_period,
            start_year=summary_start_year,
            start_month=summary_start_month,
            end_year=summary_end_year,
            end_month=summary_end_month,
        ):
            stmt = stmt.where(period_filter)
            count_stmt = count_stmt.where(period_filter)

        source_period = FinancialSummary.source_year * 100 + FinancialSummary.source_month
        for period_filter in build_month_range_filters(
            source_period,
            start_year=source_start_year,
            start_month=source_start_month,
            end_year=source_end_year,
            end_month=source_end_month,
        ):
            stmt = stmt.where(period_filter)
            count_stmt = count_stmt.where(period_filter)

        platform_names = split_filter_values(platform_name)
        if platform_names:
            stmt = stmt.where(FinancialSummary.source_platform_code.in_(platform_names))
            count_stmt = count_stmt.where(FinancialSummary.source_platform_code.in_(platform_names))

        report_platform_names = split_filter_values(report_platform_name)
        if report_platform_names:
            stmt = stmt.where(FinancialSummary.report_platform_code.in_(report_platform_names))
            count_stmt = count_stmt.where(FinancialSummary.report_platform_code.in_(report_platform_names))

        shop_id_list = split_int_filter_values(shop_ids)
        if shop_id_list:
            stmt = stmt.where(FinancialSummary.shop_id.in_(shop_id_list))
            count_stmt = count_stmt.where(FinancialSummary.shop_id.in_(shop_id_list))

        shop_names = split_filter_values(shop_name)
        if shop_names:
            stmt = stmt.where(FinancialSummary.shop_name.in_(shop_names))
            count_stmt = count_stmt.where(FinancialSummary.shop_name.in_(shop_names))

        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            keyword_cond = or_(
                FinancialSummary.shop_name.ilike(like_pattern),
                FinancialSummary.source_platform_code.ilike(like_pattern),
                FinancialSummary.report_platform_code.ilike(like_pattern),
            )
            stmt = stmt.where(keyword_cond)
            count_stmt = count_stmt.where(keyword_cond)

        stmt = stmt.order_by(
            FinancialSummary.updated_at.desc(),
            FinancialSummary.id.desc(),
        )

        total_result = await db.execute(count_stmt)
        total = total_result.scalar() or 0

        stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(stmt)
        summaries = [
            {
                "summary": summary,
                "shop_color": shop_color,
                "org_name": org_name,
            }
            for summary, shop_color, org_name in result.all()
        ]

        return summaries, total

    @staticmethod
    async def export_summaries(
        db: AsyncSession,
        *,
        org_ids: list[int] | None,
        summary_year: int | None = None,
        summary_month: int | None = None,
        summary_start_year: int | None = None,
        summary_start_month: int | None = None,
        summary_end_year: int | None = None,
        summary_end_month: int | None = None,
        source_year: int | None = None,
        source_month: int | None = None,
        source_start_year: int | None = None,
        source_start_month: int | None = None,
        source_end_year: int | None = None,
        source_end_month: int | None = None,
        platform_name: str | None = None,
        report_platform_name: str | None = None,
        shop_ids: str | None = None,
        shop_name: str | None = None,
        keyword: str | None = None,
        ids: list[int] | None = None,
        page: int | None = None,
        page_size: int | None = None,
        include_dongzhang_details: bool = False,
    ) -> io.BytesIO:
        """Export financial summaries to an Excel file (in-memory BytesIO)."""
        # Query all matching rows (no pagination)
        stmt = select(FinancialSummary).where(
            FinancialSummary.is_deleted.is_(False),
            active_shop_filter(FinancialSummary.shop_id),
        )
        if org_ids is not None:
            stmt = stmt.where(FinancialSummary.org_id.in_(org_ids))

        if ids is not None:
            if not ids:
                rows = []
                return SummaryService._build_summary_workbook(rows)
            stmt = stmt.where(FinancialSummary.id.in_(ids))

        if summary_year is not None:
            stmt = stmt.where(FinancialSummary.summary_year == summary_year)
        if summary_month is not None:
            stmt = stmt.where(FinancialSummary.summary_month == summary_month)
        if source_year is not None:
            stmt = stmt.where(FinancialSummary.source_year == source_year)
        if source_month is not None:
            stmt = stmt.where(FinancialSummary.source_month == source_month)
        summary_period = FinancialSummary.summary_year * 100 + FinancialSummary.summary_month
        for period_filter in build_month_range_filters(
            summary_period,
            start_year=summary_start_year,
            start_month=summary_start_month,
            end_year=summary_end_year,
            end_month=summary_end_month,
        ):
            stmt = stmt.where(period_filter)
        source_period = FinancialSummary.source_year * 100 + FinancialSummary.source_month
        for period_filter in build_month_range_filters(
            source_period,
            start_year=source_start_year,
            start_month=source_start_month,
            end_year=source_end_year,
            end_month=source_end_month,
        ):
            stmt = stmt.where(period_filter)
        platform_names = split_filter_values(platform_name)
        if platform_names:
            stmt = stmt.where(FinancialSummary.source_platform_code.in_(platform_names))
        report_platform_names = split_filter_values(report_platform_name)
        if report_platform_names:
            stmt = stmt.where(FinancialSummary.report_platform_code.in_(report_platform_names))
        shop_id_list = split_int_filter_values(shop_ids)
        if shop_id_list:
            stmt = stmt.where(FinancialSummary.shop_id.in_(shop_id_list))
        shop_names = split_filter_values(shop_name)
        if shop_names:
            stmt = stmt.where(FinancialSummary.shop_name.in_(shop_names))
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            stmt = stmt.where(
                or_(
                    FinancialSummary.shop_name.ilike(like_pattern),
                    FinancialSummary.source_platform_code.ilike(like_pattern),
                    FinancialSummary.report_platform_code.ilike(like_pattern),
                )
            )

        stmt = stmt.order_by(
            FinancialSummary.updated_at.desc(),
            FinancialSummary.id.desc(),
        )

        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        else:
            count_stmt = select(func.count()).select_from(stmt.subquery())
            total = (await db.execute(count_stmt)).scalar() or 0
            _ensure_export_row_limit("财务汇总", total)

        result = await db.execute(stmt)
        rows = list(result.scalars().all())

        detail_rows = []
        if include_dongzhang_details:
            detail_rows = await DouyinDongzhangDetailService.load_export_rows_for_summary_ids(
                db,
                summary_ids=[row.id for row in rows],
            )

        return SummaryService._build_summary_workbook(rows, detail_rows=detail_rows)

    @staticmethod
    async def list_report_summaries(
        db: AsyncSession,
        *,
        org_ids: list[int] | None,
        source_year: int | None = None,
        source_month: int | None = None,
        source_start_year: int | None = None,
        source_start_month: int | None = None,
        source_end_year: int | None = None,
        source_end_month: int | None = None,
        platform_name: str | None = None,
        report_platform_name: str | None = None,
        shop_ids: str | None = None,
        shop_name: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict], int]:
        """Aggregate financial summaries by accounting year/month."""
        shop_id = single_int_filter_value(shop_ids)
        filters = SummaryService._build_report_filters(
            org_ids=org_ids,
            source_year=source_year,
            source_month=source_month,
            source_start_year=source_start_year,
            source_start_month=source_start_month,
            source_end_year=source_end_year,
            source_end_month=source_end_month,
            platform_name=report_platform_name or platform_name,
            shop_ids=shop_ids,
            shop_name=shop_name,
            keyword=keyword,
        )
        group_cols = (
            FinancialSummary.org_id,
            FinancialSummary.source_year,
            FinancialSummary.source_month,
            FinancialSummary.report_platform_code,
            FinancialSummary.shop_id,
            FinancialSummary.shop_name,
        )

        grouped_stmt = select(*group_cols).where(*filters).group_by(*group_cols).subquery()
        total_result = await db.execute(select(func.count()).select_from(grouped_stmt))
        total = total_result.scalar() or 0

        metric_cols = [func.coalesce(func.sum(getattr(FinancialSummary, field)), 0).label(field) for field in SUMMARY_MONEY_FIELDS]
        stmt = (
            select(
                FinancialSummary.org_id.label("org_id"),
                func.max(Organization.name).label("org_name"),
                FinancialSummary.source_year.label("source_year"),
                FinancialSummary.source_month.label("source_month"),
                FinancialSummary.report_platform_code.label("platform_name"),
                FinancialSummary.report_platform_code.label("report_platform_code"),
                FinancialSummary.shop_id.label("shop_id"),
                FinancialSummary.shop_name.label("shop_name"),
                func.max(Shop.shop_color).label("shop_color"),
                func.count(FinancialSummary.id).label("summary_count"),
                *metric_cols,
            )
            .outerjoin(Shop, FinancialSummary.shop_id == Shop.id)
            .outerjoin(Organization, FinancialSummary.org_id == Organization.id)
            .where(*filters)
            .group_by(*group_cols)
            .order_by(
                func.max(FinancialSummary.updated_at).desc(),
                FinancialSummary.source_year.desc(),
                FinancialSummary.source_month.desc(),
                FinancialSummary.report_platform_code,
                FinancialSummary.shop_name,
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        result = await db.execute(stmt)
        rows = [dict(row) for row in result.mappings().all()]
        rows = await SummaryService._apply_report_adjustments(
            db,
            org_ids=org_ids,
            rows=rows,
            source_year=source_year,
            source_month=source_month,
            source_start_year=source_start_year,
            source_start_month=source_start_month,
            source_end_year=source_end_year,
            source_end_month=source_end_month,
            platform_name=single_filter_value(report_platform_name or platform_name),
            shop_id=shop_id,
            shop_name=single_filter_value(shop_name),
        )
        return rows, total

    @staticmethod
    async def export_report_summaries(
        db: AsyncSession,
        *,
        org_ids: list[int] | None,
        source_year: int | None = None,
        source_month: int | None = None,
        source_start_year: int | None = None,
        source_start_month: int | None = None,
        source_end_year: int | None = None,
        source_end_month: int | None = None,
        platform_name: str | None = None,
        report_platform_name: str | None = None,
        shop_ids: str | None = None,
        shop_name: str | None = None,
        keyword: str | None = None,
        ids: list[str] | None = None,
        page: int | None = None,
        page_size: int | None = None,
        include_dongzhang_details: bool = False,
    ) -> io.BytesIO:
        """Export accounting-month aggregated summaries to Excel."""
        shop_id = single_int_filter_value(shop_ids)
        filters = SummaryService._build_report_filters(
            org_ids=org_ids,
            source_year=source_year,
            source_month=source_month,
            source_start_year=source_start_year,
            source_start_month=source_start_month,
            source_end_year=source_end_year,
            source_end_month=source_end_month,
            platform_name=report_platform_name or platform_name,
            shop_ids=shop_ids,
            shop_name=shop_name,
            keyword=keyword,
        )
        group_cols = (
            FinancialSummary.org_id,
            FinancialSummary.source_year,
            FinancialSummary.source_month,
            FinancialSummary.report_platform_code,
            FinancialSummary.shop_id,
            FinancialSummary.shop_name,
        )
        metric_cols = [func.coalesce(func.sum(getattr(FinancialSummary, field)), 0).label(field) for field in SUMMARY_MONEY_FIELDS]
        stmt = (
            select(
                FinancialSummary.org_id.label("org_id"),
                func.max(Organization.name).label("org_name"),
                FinancialSummary.source_year.label("source_year"),
                FinancialSummary.source_month.label("source_month"),
                FinancialSummary.report_platform_code.label("platform_name"),
                FinancialSummary.report_platform_code.label("report_platform_code"),
                FinancialSummary.shop_id.label("shop_id"),
                FinancialSummary.shop_name.label("shop_name"),
                func.max(Shop.shop_color).label("shop_color"),
                func.count(FinancialSummary.id).label("summary_count"),
                *metric_cols,
            )
            .outerjoin(Shop, FinancialSummary.shop_id == Shop.id)
            .outerjoin(Organization, FinancialSummary.org_id == Organization.id)
            .where(*filters)
            .group_by(*group_cols)
            .order_by(
                func.max(FinancialSummary.updated_at).desc(),
                FinancialSummary.source_year.desc(),
                FinancialSummary.source_month.desc(),
                FinancialSummary.report_platform_code,
                FinancialSummary.shop_name,
            )
        )

        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        else:
            grouped_subquery = stmt.subquery()
            count_stmt = select(func.count()).select_from(grouped_subquery)
            total = (await db.execute(count_stmt)).scalar() or 0
            _ensure_export_row_limit("汇总报表", total)

        result = await db.execute(stmt)
        rows = [dict(row) for row in result.mappings().all()]
        rows = await SummaryService._apply_report_adjustments(
            db,
            org_ids=org_ids,
            rows=rows,
            source_year=source_year,
            source_month=source_month,
            source_start_year=source_start_year,
            source_start_month=source_start_month,
            source_end_year=source_end_year,
            source_end_month=source_end_month,
            platform_name=single_filter_value(report_platform_name or platform_name),
            shop_id=shop_id,
            shop_name=single_filter_value(shop_name),
        )
        if ids is not None:
            selected_ids = set(ids)
            rows = [row for row in rows if SummaryService.report_row_id(row) in selected_ids]
        detail_rows = []
        if include_dongzhang_details:
            detail_rows = await DouyinDongzhangDetailService.load_export_rows_for_report_rows(
                db,
                rows=rows,
            )
        return SummaryService._build_report_workbook(rows, detail_rows=detail_rows)

    @staticmethod
    def _build_summary_workbook(
        rows: list[FinancialSummary],
        *,
        detail_rows: list[dict[str, object]] | None = None,
    ) -> io.BytesIO:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="财务汇总")
        ws.append(_write_only_header_row(ws, EXPORT_HEADERS))
        for s in rows:
            ws.append(
                [
                    SummaryService._month_label(s.source_year, s.source_month),
                    f"{s.summary_year}{s.summary_month:02d}",
                    SummaryService._platform_label(s.source_platform_code),
                    s.shop_name,
                    float(s.order_paid_amount or 0),
                    float(s.refund_amount or 0),
                    float(s.gmv or 0),
                    float(s.platform_income or 0),
                    float(s.platform_fee or 0),
                    float(s.return_cost or 0),
                    float(s.commission or 0),
                    float(s.merchant_fee or 0),
                    float(s.promotion_fee or 0),
                    float(s.provider_commission or 0),
                    float(s.donation_fee or 0),
                    float(s.insurance_fee or 0),
                    float(s.bic or 0),
                ]
            )
        if detail_rows is not None:
            SummaryService._append_dongzhang_detail_sheet(wb, detail_rows)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_report_workbook(
        rows: list[dict],
        *,
        detail_rows: list[dict[str, object]] | None = None,
    ) -> io.BytesIO:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="汇总报表")
        ws.append(_write_only_header_row(ws, REPORT_EXPORT_HEADERS))
        for row in rows:
            ws.append(
                [
                    SummaryService._month_label(int(row.get("source_year") or 0), int(row.get("source_month") or 0)),
                    SummaryService._platform_label(str(row.get("platform_name") or "")),
                    row.get("shop_name") or "",
                    float(row.get("order_paid_amount") or 0),
                    float(row.get("refund_amount") or 0),
                    float(row.get("original_gmv") or 0),
                    float(row.get("platform_income") or 0),
                    float(row.get("platform_fee") or 0),
                    float(row.get("original_return_cost") or 0),
                    float(row.get("commission") or 0),
                    float(row.get("merchant_fee") or 0),
                    float(row.get("promotion_fee") or 0),
                    float(row.get("provider_commission") or 0),
                    float(row.get("donation_fee") or 0),
                    float(row.get("insurance_fee") or 0),
                    float(row.get("bic") or 0),
                ]
            )
        if detail_rows is not None:
            SummaryService._append_dongzhang_detail_sheet(wb, detail_rows)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _append_dongzhang_detail_sheet(wb: Workbook, detail_rows: list[dict[str, object]]) -> None:
        ws = wb.create_sheet(title="Douyin动账源明细")
        ws.append(_write_only_header_row(ws, [label for _field, label in DONGZHANG_DETAIL_EXPORT_COLUMNS]))
        for row in detail_rows:
            ws.append(
                [
                    DouyinDongzhangDetailService._format_export_value(
                        row.get(field),
                        money=(field in DONGZHANG_DETAIL_MONEY_FIELDS),
                    )
                    for field, _label in DONGZHANG_DETAIL_EXPORT_COLUMNS
                ]
            )

    @staticmethod
    def _platform_label(platform_name: str) -> str:
        return PLATFORM_LABELS.get(platform_name, platform_name)

    @staticmethod
    def report_row_id(row: dict) -> str:
        return f"{int(row.get('org_id') or 0)}-{int(row.get('source_year') or 0)}-{int(row.get('source_month') or 0)}-{row.get('platform_name') or ''}-{int(row.get('shop_id') or 0)}"

    @staticmethod
    def _month_label(year: int | None, month: int | None) -> str:
        if not year or not month:
            return "未解析"
        return f"{int(year)}{int(month):02d}"

    @staticmethod
    async def _apply_report_adjustments(
        db: AsyncSession,
        *,
        org_ids: list[int] | None,
        rows: list[dict],
        source_year: int | None,
        source_month: int | None,
        source_start_year: int | None,
        source_start_month: int | None,
        source_end_year: int | None,
        source_end_month: int | None,
        platform_name: str | None,
        shop_id: int | None,
        shop_name: str | None,
    ) -> list[dict]:
        if not rows:
            return rows

        adjustment_sums = await SummaryAdjustmentService.list_adjustment_sums(
            db,
            org_ids=org_ids,
            source_year=source_year,
            source_month=source_month,
            source_start_year=source_start_year,
            source_start_month=source_start_month,
            source_end_year=source_end_year,
            source_end_month=source_end_month,
            platform_name=platform_name,
            shop_id=shop_id,
            shop_name=shop_name,
        )
        for row in rows:
            key = (
                int(row.get("org_id") or 0),
                int(row.get("source_year") or 0),
                int(row.get("source_month") or 0),
                str(row.get("platform_name") or ""),
                int(row.get("shop_id") or 0),
                str(row.get("shop_name") or ""),
            )
            adjustments = adjustment_sums.get(key, {})
            original_gmv = Decimal(str(row.get("gmv") or 0))
            gmv_adjustment = adjustments.get("gmv", Decimal("0"))
            original_return_cost = Decimal(str(row.get("return_cost") or 0))
            return_cost_adjustment = adjustments.get("return_cost", Decimal("0"))

            row["original_gmv"] = original_gmv
            row["gmv_adjustment"] = gmv_adjustment
            row["gmv"] = original_gmv + gmv_adjustment
            row["original_return_cost"] = original_return_cost
            row["return_cost_adjustment"] = return_cost_adjustment
            row["return_cost"] = original_return_cost + return_cost_adjustment
        return rows

    @staticmethod
    def _build_report_filters(
        *,
        org_ids: list[int] | None,
        source_year: int | None,
        source_month: int | None,
        source_start_year: int | None = None,
        source_start_month: int | None = None,
        source_end_year: int | None = None,
        source_end_month: int | None = None,
        platform_name: str | None,
        shop_ids: str | None,
        shop_name: str | None,
        keyword: str | None = None,
    ):
        filters = [FinancialSummary.is_deleted.is_(False), active_shop_filter(FinancialSummary.shop_id)]
        if org_ids is not None:
            filters.append(FinancialSummary.org_id.in_(org_ids))
        if source_year is not None:
            filters.append(FinancialSummary.source_year == source_year)
        if source_month is not None:
            filters.append(FinancialSummary.source_month == source_month)
        source_period = FinancialSummary.source_year * 100 + FinancialSummary.source_month
        filters.extend(
            build_month_range_filters(
                source_period,
                start_year=source_start_year,
                start_month=source_start_month,
                end_year=source_end_year,
                end_month=source_end_month,
            )
        )
        platform_names = split_filter_values(platform_name)
        if platform_names:
            filters.append(FinancialSummary.report_platform_code.in_(platform_names))
        shop_id_list = split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(FinancialSummary.shop_id.in_(shop_id_list))
        shop_names = split_filter_values(shop_name)
        if shop_names:
            filters.append(FinancialSummary.shop_name.in_(shop_names))
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    FinancialSummary.shop_name.ilike(like_pattern),
                    FinancialSummary.source_platform_code.ilike(like_pattern),
                    FinancialSummary.report_platform_code.ilike(like_pattern),
                )
            )
        return filters
