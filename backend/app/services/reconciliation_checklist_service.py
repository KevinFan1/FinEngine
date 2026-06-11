"""Service layer for reconciliation checklist imports and reports."""

from __future__ import annotations

import hashlib
import logging
import re
import tempfile
from collections import defaultdict
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, Side
from sqlalchemy import cast, column, delete, func, or_, select, table, text, tuple_, update, values
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import register_after_commit
from app.models.organization import Organization
from app.models.reconciliation_checklist import (
    ReconciliationChecklistDetail,
    ReconciliationChecklistOrderKey,
    ReconciliationChecklistPayableBalanceSummaryRow,
    ReconciliationChecklistProductSummaryRow,
    ReconciliationChecklistReceiptSummaryRow,
    ReconciliationChecklistTask,
    ReconciliationChecklistUploadFile,
)
from app.models.upload import UploadFile
from app.models.user import User
from app.services.audit_service import AuditService
from app.services.oss_service import SOURCE_FILE_UNAVAILABLE_MESSAGE, is_oss_object_unavailable_error, oss_service
from app.tasks.processors.base import open_tabular_rows, parse_datetime, safe_str
from app.utils.money import safe_decimal
from app.utils.oss_paths import checklist_manual_edit_dir, current_oss_period, sanitize_oss_filename
from app.utils.query_filters import resolve_org_ids

logger = logging.getLogger("finengine.reconciliation_checklist")

CHECKLIST_FILE_TYPE_PREFIX = "对账清单"
CHECKLIST_FILE_TYPE_SOURCE = "对账清单-原始数据"
CHECKLIST_FILE_TYPE_INVOICE = "对账清单-发票更新"
CHECKLIST_FILE_TYPE_MERCHANT = "对账清单-商家更新"
CHECKLIST_COMPAT_FILE_TYPE = "对账清单"
CHECKLIST_FILE_TYPES = {
    CHECKLIST_COMPAT_FILE_TYPE,
    CHECKLIST_FILE_TYPE_SOURCE,
    CHECKLIST_FILE_TYPE_INVOICE,
    CHECKLIST_FILE_TYPE_MERCHANT,
}
CHECKLIST_EMPTY_SUMMARY_MESSAGE = "空数据"
CHECKLIST_INTERNAL_ERROR_MESSAGE = "处理异常，请联系管理员或稍后重试"
CHECKLIST_ERROR_SAMPLE_LIMIT = 20
CHECKLIST_ERROR_ROW_SAMPLE_LIMIT = 20
CHECKLIST_BATCH_MAX_QUERY_ARGS = 30000
CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE = CHECKLIST_BATCH_MAX_QUERY_ARGS
CHECKLIST_TUPLE_IN_LOOKUP_CHUNK_SIZE = CHECKLIST_BATCH_MAX_QUERY_ARGS // 2
CHECKLIST_SOURCE_PERSIST_SHARD_SIZE = 5000
CHECKLIST_SOURCE_STAGE_TABLE_NAME = "tmp_fin_reconciliation_checklist_source_stage"
CHECKLIST_MANUAL_EDIT_MAX_SUB_ORDERS = 100
CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT = 10
CHECKLIST_COMMISSION_RATE_MAX = Decimal("9999.999999")
CHECKLIST_PROCESSED_TASK_STATUSES = ("success", "partial_success")
CHECKLIST_RESULT_TASK_STATUSES = ("success", "partial_success", "failed", "expired")
CHECKLIST_EXPORT_SIDE = Side(style="thin", color="000000")
CHECKLIST_EXPORT_BORDER = Border(
    left=CHECKLIST_EXPORT_SIDE,
    right=CHECKLIST_EXPORT_SIDE,
    top=CHECKLIST_EXPORT_SIDE,
    bottom=CHECKLIST_EXPORT_SIDE,
)
CHECKLIST_EXPORT_NO_BORDER = Border()
CHECKLIST_EXPORT_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CHECKLIST_EXPORT_LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
CHECKLIST_EXPORT_RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")
CHECKLIST_EXPORT_HEADER_FONT = Font(bold=True)
CHECKLIST_EXPORT_MONEY_FORMAT = "0.00"
CHECKLIST_PRODUCT_SUMMARY_TITLE = "直播运营费用明细"

SOURCE_HEADERS = [
    "进驻的直播平台",
    "结算时间",
    "子订单号",
    "下单时间",
    "商品ID",
    "商品名称",
    "商品数量",
    "达人名称",
    "平台补贴",
    "达人补贴",
    "抖音支付补贴",
    "抖音月付营销补贴",
    "银行补贴",
    "用户实付 （订单金额）",
    "平台服务费",
    "达人佣金",
    "招商服务费",
    "商户主体名称",
    "客服代码",
    "收款商家",
    "直播推广佣金",
    "佣金率",
    "应付商家净额",
    "付款金额",
    "付款时间（商家）",
    "开票时间",
    "发票号码",
]
SOURCE_OPTIONAL_HEADERS = ["应付商家净额余额"]
INVOICE_UPDATE_HEADERS = ["唯一ID", "子订单号", "收款商家", "开票时间", "发票号码"]
MERCHANT_UPDATE_HEADERS = ["唯一ID", "子订单号", "收款商家", "应付商家净额", "付款金额", "付款时间（商家）"]

CHECKLIST_HEADERS = SOURCE_HEADERS
CHECKLIST_FILE_TYPE = CHECKLIST_COMPAT_FILE_TYPE
CHECKLIST_TASK_TYPE_SOURCE_IMPORT = "source_import"
CHECKLIST_TASK_TYPE_INVOICE_EDIT = "invoice_edit"
CHECKLIST_TASK_TYPE_MERCHANT_EDIT = "merchant_edit"
CHECKLIST_MANUAL_EDIT_TASK_TYPES = {
    CHECKLIST_TASK_TYPE_INVOICE_EDIT: CHECKLIST_FILE_TYPE_INVOICE,
    CHECKLIST_TASK_TYPE_MERCHANT_EDIT: CHECKLIST_FILE_TYPE_MERCHANT,
}

HEADER_ALIASES = {
    "商家主体名称": "商户主体名称",
    "用户实付(订单金额)": "用户实付 （订单金额）",
    "用户实付（订单金额）": "用户实付 （订单金额）",
    "用户实付订单金额": "用户实付 （订单金额）",
    "系统行定位值": "唯一ID",
    "系统唯一ID": "唯一ID",
    "付款时间(商家": "付款时间(商家)",
    "付款时间（商家": "付款时间(商家)",
}
SOURCE_FIELD_MAP = {
    "进驻的直播平台": "live_platform",
    "结算时间": "settlement_time",
    "子订单号": "sub_order_no",
    "下单时间": "order_time",
    "商品ID": "product_id",
    "商品名称": "product_name",
    "商品数量": "product_quantity",
    "达人名称": "talent_name",
    "平台补贴": "platform_subsidy",
    "达人补贴": "talent_subsidy",
    "抖音支付补贴": "douyin_pay_subsidy",
    "抖音月付营销补贴": "douyin_monthly_pay_subsidy",
    "银行补贴": "bank_subsidy",
    "用户实付 （订单金额）": "user_paid_amount",
    "平台服务费": "platform_service_fee",
    "达人佣金": "talent_commission",
    "招商服务费": "investment_service_fee",
    "商户主体名称": "merchant_subject_name",
    "客服代码": "customer_service_code",
    "收款商家": "receipt_merchant",
    "直播推广佣金": "live_commission",
    "佣金率": "commission_rate",
    "应付商家净额": "merchant_net_amount",
    "应付商家净额余额": "merchant_net_balance",
    "付款金额": "payment_amount",
    "付款时间（商家）": "merchant_payment_time",
    "开票时间": "invoice_time",
    "发票号码": "invoice_number",
}
MONEY_FIELDS = {
    "platform_subsidy",
    "talent_subsidy",
    "douyin_pay_subsidy",
    "douyin_monthly_pay_subsidy",
    "bank_subsidy",
    "user_paid_amount",
    "platform_service_fee",
    "talent_commission",
    "investment_service_fee",
    "live_commission",
    "merchant_net_amount",
    "merchant_net_balance",
}
OPTION_KINDS = {"merchant_subject", "receipt_merchant", "live_platform", "product_name"}


@dataclass(frozen=True)
class HeaderSpec:
    file_type: str
    headers: list[str]


@dataclass
class ErrorSummaryGroup:
    count: int = 0
    rows: list[int] = field(default_factory=list)


HEADER_SPECS = [
    HeaderSpec(CHECKLIST_FILE_TYPE_SOURCE, SOURCE_HEADERS),
    HeaderSpec(CHECKLIST_FILE_TYPE_MERCHANT, MERCHANT_UPDATE_HEADERS),
    HeaderSpec(CHECKLIST_FILE_TYPE_INVOICE, INVOICE_UPDATE_HEADERS),
]


def _canonical_header(value: object) -> str:
    text_value = safe_str(value).replace("帐", "账")
    text_value = "".join(text_value.split())
    text_value = text_value.replace("（", "(").replace("）", ")")
    return HEADER_ALIASES.get(text_value, text_value)


def _normalized_required_header(header: str) -> str:
    return _canonical_header(header)


def _build_header_column_lookup(headers: Sequence[object]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalized_required_header(safe_str(header))
        if normalized:
            lookup.setdefault(normalized, index)
    return lookup


def _value_from_row(row: Sequence[object], lookup: dict[str, int], header: str) -> object:
    index = lookup.get(header)
    if index is None:
        index = lookup.get(_normalized_required_header(header))
    if index is None or index >= len(row):
        return None
    return row[index]


@dataclass(frozen=True)
class _RowValueAccessor:
    row: Sequence[object]
    lookup: dict[str, int]

    def get(self, header: str) -> object:
        return _value_from_row(self.row, self.lookup, header)


def _header_value(values: dict[str, object] | _RowValueAccessor, header: str) -> object:
    if isinstance(values, _RowValueAccessor):
        return values.get(header)
    if header in values:
        return values.get(header)
    canonical = _canonical_header(header)
    if canonical in values:
        return values.get(canonical)
    normalized = _normalized_required_header(header)
    if normalized in values:
        return values.get(normalized)
    return None


def _accounting_period(year: int, month: int) -> int:
    if month < 1 or month > 12:
        raise ValueError("月份必须在 1-12 之间")
    return int(year) * 100 + int(month)


def _money(value: object) -> Decimal:
    return safe_decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _nullable_money(value: object) -> Decimal | None:
    if safe_str(value) == "":
        return None
    return _money(value)


def _nullable_rate(value: object) -> Decimal | None:
    if safe_str(value) == "":
        return None
    try:
        parsed = Decimal(str(value).replace("%", "").replace(",", "").strip()).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("佣金率无法解析") from exc
    if abs(parsed) > CHECKLIST_COMMISSION_RATE_MAX:
        raise ValueError("佣金率超出范围")
    return parsed


def _nullable_datetime(value: object) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if safe_str(value) == "":
        return None
    return parse_datetime(value)


def _to_int(value: object) -> int:
    text_value = safe_str(value).replace(",", "")
    if text_value == "":
        return 0
    try:
        return int(Decimal(text_value))
    except (InvalidOperation, ValueError):
        return 0


def _fingerprint_part(value: object) -> str:
    if isinstance(value, Decimal):
        return format(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")
    return safe_str(value)


def _fingerprint_datetime_part(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return safe_str(value)


def _fingerprint_money_part(value: object) -> str:
    return format(safe_decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")


def _row_fingerprint(
    *,
    org_id: int,
    settlement_time: object,
    sub_order_no: object,
    platform_subsidy: object,
    talent_subsidy: object,
    douyin_pay_subsidy: object,
    douyin_monthly_pay_subsidy: object,
    bank_subsidy: object,
    user_paid_amount: object,
    platform_service_fee: object,
    talent_commission: object,
    investment_service_fee: object,
) -> str:
    payload = "|".join(
        [
            str(int(org_id)),
            _fingerprint_datetime_part(settlement_time),
            _fingerprint_part(sub_order_no),
            _fingerprint_money_part(platform_subsidy),
            _fingerprint_money_part(talent_subsidy),
            _fingerprint_money_part(douyin_pay_subsidy),
            _fingerprint_money_part(douyin_monthly_pay_subsidy),
            _fingerprint_money_part(bank_subsidy),
            _fingerprint_money_part(user_paid_amount),
            _fingerprint_money_part(platform_service_fee),
            _fingerprint_money_part(talent_commission),
            _fingerprint_money_part(investment_service_fee),
        ]
    )
    return hashlib.md5(payload.encode("utf-8")).hexdigest()


def _merchant_net_balance_value(
    merchant_net_amount: object,
    payment_amount: object,
) -> Decimal:
    net_amount = safe_decimal(merchant_net_amount)
    paid_amount = safe_decimal(payment_amount)
    return (net_amount - paid_amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _split_filter_values(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.replace("，", ",").split(",") if item.strip()]


def _product_summary_key(row: dict[str, object]) -> str:
    return ":".join(
        [
            str(row["org_id"]),
            str(row["accounting_period"]),
            safe_str(row.get("receipt_merchant")),
            safe_str(row.get("merchant_subject_name")),
            safe_str(row.get("product_name")),
        ]
    )


def _receipt_summary_key(row: dict[str, object]) -> str:
    return ":".join(
        [
            str(row["org_id"]),
            str(row["accounting_period"]),
            safe_str(row.get("merchant_subject_name")),
            safe_str(row.get("live_platform")),
            safe_str(row.get("receipt_merchant")),
        ]
    )


def _payable_balance_summary_key(row: dict[str, object]) -> str:
    return ":".join(
        [
            str(row["org_id"]),
            str(row["accounting_period"]),
            safe_str(row.get("merchant_subject_name")),
            safe_str(row.get("receipt_merchant")),
        ]
    )


def _detect_checklist_type(headers: Sequence[object]) -> HeaderSpec:
    canonical_headers = {_canonical_header(header) for header in headers if _canonical_header(header)}
    best_spec: HeaderSpec | None = None
    best_missing: list[str] = []
    for spec in HEADER_SPECS:
        missing = [header for header in spec.headers if _normalized_required_header(header) not in canonical_headers]
        if not missing:
            return spec
        if best_spec is None or len(missing) < len(best_missing):
            best_spec = spec
            best_missing = missing
    missing_text = "、".join(best_missing) if best_missing else "必要表头"
    raise ValueError(f"缺少对账清单必要表头: {missing_text}")


def _capture_result_state(task: ReconciliationChecklistTask) -> dict[str, object]:
    return {
        "total_rows": task.total_rows,
        "success_rows": task.success_rows,
        "failed_rows": task.failed_rows,
        "inserted_rows": task.inserted_rows,
        "updated_rows": task.updated_rows,
        "result_summary": task.result_summary,
    }


def _restore_result_state(task: ReconciliationChecklistTask, state: dict[str, object]) -> None:
    task.total_rows = int(state.get("total_rows") or 0)
    task.success_rows = int(state.get("success_rows") or 0)
    task.failed_rows = int(state.get("failed_rows") or 0)
    task.inserted_rows = int(state.get("inserted_rows") or 0)
    task.updated_rows = int(state.get("updated_rows") or 0)
    task.result_summary = state.get("result_summary")  # type: ignore[assignment]


def _period_filters(
    column,
    *,
    accounting_year: int | None = None,
    accounting_month: int | None = None,
    accounting_start_year: int | None = None,
    accounting_start_month: int | None = None,
    accounting_end_year: int | None = None,
    accounting_end_month: int | None = None,
) -> list:
    filters = []
    if accounting_year is not None:
        filters.append(column >= accounting_year * 100 + 1)
        filters.append(column <= accounting_year * 100 + 12)
    if accounting_year is not None and accounting_month is not None:
        filters = [column == _accounting_period(accounting_year, accounting_month)]
    if accounting_start_year is not None and accounting_start_month is not None:
        filters.append(column >= _accounting_period(accounting_start_year, accounting_start_month))
    if accounting_end_year is not None and accounting_end_month is not None:
        filters.append(column <= _accounting_period(accounting_end_year, accounting_end_month))
    return filters


def _apply_export_sheet_style(sheet) -> None:
    for column_name, width in {"A": 10, "B": 36, "C": 18, "D": 18, "E": 18}.items():
        sheet.column_dimensions[column_name].width = width


def _export_cell(
    sheet,
    value: object,
    *,
    bold: bool = False,
    font_size: int | None = None,
    border: Border = CHECKLIST_EXPORT_BORDER,
    alignment: Alignment = CHECKLIST_EXPORT_ALIGNMENT,
    number_format: str | None = None,
) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.border = border
    cell.alignment = alignment
    if bold or font_size:
        cell.font = Font(bold=bold, size=font_size)
    if number_format:
        cell.number_format = number_format
    return cell


def _export_row(sheet, values: Iterable[object], *, bold: bool = False, font_size: int | None = None) -> list[WriteOnlyCell]:
    return [_export_cell(sheet, value, bold=bold, font_size=font_size) for value in values]


def _set_sheet_cell(
    sheet,
    row: int,
    column: int,
    value: object,
    *,
    bold: bool = False,
    font_size: int | None = None,
    border: Border = CHECKLIST_EXPORT_BORDER,
    alignment: Alignment = CHECKLIST_EXPORT_ALIGNMENT,
    number_format: str | None = None,
):
    cell = sheet.cell(row=row, column=column, value=value)
    cell.border = border
    cell.alignment = alignment
    if bold or font_size:
        cell.font = Font(bold=bold, size=font_size)
    if number_format:
        cell.number_format = number_format
    return cell


def _is_zero_money_export_row(*values: object) -> bool:
    return all(safe_decimal(value) == Decimal("0.00") for value in values)


def _sheet_title(*parts: object) -> str:
    raw = "_".join(safe_str(part) for part in parts if safe_str(part)) or "对账清单"
    for char in r"[]:*?/\\":
        raw = raw.replace(char, "_")
    return raw[:31]


def _chunk_records(records: Sequence[Any], *, bind_count_per_record: int | None = None) -> list[Sequence[Any]]:
    if not records:
        return []
    per_record = bind_count_per_record if bind_count_per_record and bind_count_per_record > 0 else max(len(records[0]), 1)
    chunk_size = max(CHECKLIST_BATCH_MAX_QUERY_ARGS // per_record, 1)
    return [records[start : start + chunk_size] for start in range(0, len(records), chunk_size)]


def _chunk_tuple_lookup_records(records: Sequence[Any]) -> list[Sequence[Any]]:
    if not records:
        return []
    return [records[start : start + CHECKLIST_TUPLE_IN_LOOKUP_CHUNK_SIZE] for start in range(0, len(records), CHECKLIST_TUPLE_IN_LOOKUP_CHUNK_SIZE)]


def _source_stage_table():
    return table(
        CHECKLIST_SOURCE_STAGE_TABLE_NAME,
        *(column(table_column.name, table_column.type) for table_column in ReconciliationChecklistDetail.__table__.columns),
    )


def _copy_stage_column_names(rows: Sequence[dict[str, object]]) -> list[str]:
    if not rows:
        return []
    return list(rows[0].keys())


def _format_seconds(value: float) -> str:
    return f"{value:.3f}"


SOURCE_DETAIL_COMPARE_FIELDS = (
    "org_id",
    "accounting_year",
    "accounting_month",
    "accounting_period",
    "live_platform",
    "settlement_time",
    "sub_order_no",
    "order_time",
    "product_id",
    "product_name",
    "product_quantity",
    "talent_name",
    "platform_subsidy",
    "talent_subsidy",
    "douyin_pay_subsidy",
    "douyin_monthly_pay_subsidy",
    "bank_subsidy",
    "user_paid_amount",
    "platform_service_fee",
    "talent_commission",
    "investment_service_fee",
    "merchant_subject_name",
    "customer_service_code",
    "receipt_merchant",
    "live_commission",
    "commission_rate",
    "merchant_net_amount",
    "payment_amount",
    "merchant_net_balance",
    "merchant_payment_time",
    "invoice_time",
    "invoice_number",
)


def _source_detail_key(row: dict[str, object]) -> tuple[int, str]:
    return int(row["accounting_period"]), safe_str(row.get("row_fingerprint"))


class ReconciliationChecklistService:
    ERROR_ROW_PATTERN = re.compile(r"^Row\s+(?P<row>\d+):\s*(?P<message>.+)$")

    @staticmethod
    def _typed_detail_update_value(updates: object, field_name: str) -> object:
        value = getattr(updates.c, field_name)
        if field_name in {"invoice_time", "merchant_payment_time"}:
            return cast(value, ReconciliationChecklistDetail.__table__.c[field_name].type)
        return value

    @staticmethod
    def validate_org_scope(*, org_id: int, user: User) -> None:
        if user.role != "superadmin" and org_id != user.org_id:
            raise ValueError("数据不存在或无权访问")

    @staticmethod
    def _normalize_manual_edit_sub_orders(raw_value: str | Sequence[str]) -> list[str]:
        values = [raw_value] if isinstance(raw_value, str) else list(raw_value)
        normalized: list[str] = []
        for value in values:
            parts = re.split(r"[\s,，]+", safe_str(value))
            normalized.extend(part for part in parts if part)
        deduped = list(dict.fromkeys(normalized))
        if not deduped:
            raise ValueError("请先输入子订单号")
        if len(deduped) > CHECKLIST_MANUAL_EDIT_MAX_SUB_ORDERS:
            raise ValueError(f"单次最多查询 {CHECKLIST_MANUAL_EDIT_MAX_SUB_ORDERS} 个子订单号")
        return deduped

    @staticmethod
    async def _query_manual_edit_details(
        db: AsyncSession,
        *,
        org_id: int,
        sub_order_nos: Sequence[str],
    ) -> tuple[list[ReconciliationChecklistDetail], list[str]]:
        normalized = ReconciliationChecklistService._normalize_manual_edit_sub_orders(sub_order_nos)
        result = await db.execute(
            select(ReconciliationChecklistDetail).where(
                ReconciliationChecklistDetail.org_id == org_id,
                ReconciliationChecklistDetail.is_deleted.is_(False),
                ReconciliationChecklistDetail.sub_order_no.in_(normalized),
            )
        )
        rows = result.scalars().all()
        grouped: dict[str, list[ReconciliationChecklistDetail]] = defaultdict(list)
        for item in rows:
            grouped[safe_str(item.sub_order_no)].append(item)
        matched_items: list[ReconciliationChecklistDetail] = []
        for sub_order_no in normalized:
            items = grouped.get(sub_order_no, [])
            items.sort(key=lambda item: (item.settlement_time or datetime.min, safe_str(item.row_fingerprint), item.id or 0))
            matched_items.extend(items)
        if ReconciliationChecklistService._backfill_missing_unique_ids(org_id=org_id, details=matched_items):
            await db.flush()
        missing_sub_orders = [sub_order_no for sub_order_no in normalized if not grouped.get(sub_order_no)]
        return matched_items, missing_sub_orders

    @staticmethod
    def _backfill_missing_unique_ids(*, org_id: int, details: Iterable[ReconciliationChecklistDetail]) -> bool:
        changed = False
        for detail in details:
            if safe_str(detail.row_fingerprint):
                continue
            detail.row_fingerprint = _row_fingerprint(
                org_id=org_id,
                settlement_time=detail.settlement_time,
                sub_order_no=detail.sub_order_no,
                platform_subsidy=detail.platform_subsidy,
                talent_subsidy=detail.talent_subsidy,
                douyin_pay_subsidy=detail.douyin_pay_subsidy,
                douyin_monthly_pay_subsidy=detail.douyin_monthly_pay_subsidy,
                bank_subsidy=detail.bank_subsidy,
                user_paid_amount=detail.user_paid_amount,
                platform_service_fee=detail.platform_service_fee,
                talent_commission=detail.talent_commission,
                investment_service_fee=detail.investment_service_fee,
            )
            changed = True
        return changed

    @staticmethod
    async def query_invoice_edit_items(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        sub_order_nos: Sequence[str],
    ) -> tuple[list[dict[str, object]], list[str]]:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        matched_rows, missing_sub_order_nos = await ReconciliationChecklistService._query_manual_edit_details(
            db,
            org_id=org_id,
            sub_order_nos=sub_order_nos,
        )
        matched_items = [
            {
                "unique_id": safe_str(item.row_fingerprint),
                "sub_order_no": safe_str(item.sub_order_no),
                "settlement_time": item.settlement_time,
                "platform_subsidy": item.platform_subsidy,
                "talent_subsidy": item.talent_subsidy,
                "douyin_pay_subsidy": item.douyin_pay_subsidy,
                "douyin_monthly_pay_subsidy": item.douyin_monthly_pay_subsidy,
                "bank_subsidy": item.bank_subsidy,
                "user_paid_amount": item.user_paid_amount,
                "platform_service_fee": item.platform_service_fee,
                "talent_commission": item.talent_commission,
                "investment_service_fee": item.investment_service_fee,
                "receipt_merchant": safe_str(item.receipt_merchant),
                "invoice_time": item.invoice_time,
                "invoice_number": safe_str(item.invoice_number),
            }
            for item in matched_rows
        ]
        return matched_items, missing_sub_order_nos

    @staticmethod
    async def query_merchant_edit_items(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        sub_order_nos: Sequence[str],
    ) -> tuple[list[dict[str, object]], list[str]]:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        matched_rows, missing_sub_order_nos = await ReconciliationChecklistService._query_manual_edit_details(
            db,
            org_id=org_id,
            sub_order_nos=sub_order_nos,
        )
        matched_items = [
            {
                "unique_id": safe_str(item.row_fingerprint),
                "sub_order_no": safe_str(item.sub_order_no),
                "settlement_time": item.settlement_time,
                "platform_subsidy": item.platform_subsidy,
                "talent_subsidy": item.talent_subsidy,
                "douyin_pay_subsidy": item.douyin_pay_subsidy,
                "douyin_monthly_pay_subsidy": item.douyin_monthly_pay_subsidy,
                "bank_subsidy": item.bank_subsidy,
                "user_paid_amount": item.user_paid_amount,
                "platform_service_fee": item.platform_service_fee,
                "talent_commission": item.talent_commission,
                "investment_service_fee": item.investment_service_fee,
                "receipt_merchant": safe_str(item.receipt_merchant),
                "merchant_net_amount": item.merchant_net_amount,
                "payment_amount": item.payment_amount,
                "merchant_payment_time": item.merchant_payment_time,
            }
            for item in matched_rows
        ]
        return matched_items, missing_sub_order_nos

    @staticmethod
    def _extract_missing_sub_orders_from_errors(
        parsed_rows: Iterable[dict[str, object]],
        errors: Iterable[str],
    ) -> list[str]:
        row_to_sub_order = {int(row.get("source_row_number") or 0): safe_str(row.get("sub_order_no")) for row in parsed_rows if safe_str(row.get("sub_order_no"))}
        missing_sub_orders: list[str] = []
        for error in errors:
            match = ReconciliationChecklistService.ERROR_ROW_PATTERN.match(safe_str(error))
            if not match:
                continue
            if "子订单号不存在" not in match.group("message"):
                continue
            sub_order_no = row_to_sub_order.get(int(match.group("row")))
            if sub_order_no:
                missing_sub_orders.append(sub_order_no)
        return list(dict.fromkeys(missing_sub_orders))

    @staticmethod
    def _add_error_summary(
        error_summary: dict[str, ErrorSummaryGroup],
        message: str,
        *,
        row_number: int | None = None,
    ) -> None:
        normalized_message = safe_str(message)
        if not normalized_message:
            return
        group = error_summary.setdefault(normalized_message, ErrorSummaryGroup())
        group.count += 1
        if row_number is not None and len(group.rows) < CHECKLIST_ERROR_ROW_SAMPLE_LIMIT:
            group.rows.append(row_number)

    @staticmethod
    def _record_parse_error(result: dict, row_number: int, exc: Exception) -> None:
        message = safe_str(exc)
        result["failed_rows"] += 1
        ReconciliationChecklistService._add_error_summary(
            result.setdefault("error_summary", {}),
            message,
            row_number=row_number,
        )
        if len(result["errors"]) < CHECKLIST_ERROR_SAMPLE_LIMIT:
            result["errors"].append(f"Row {row_number}: {message}")

    @staticmethod
    def _add_error_text_to_summary(parse_result: dict, error: object) -> None:
        message_text = safe_str(error)
        if not message_text:
            return
        match = ReconciliationChecklistService.ERROR_ROW_PATTERN.match(message_text)
        if match:
            ReconciliationChecklistService._add_error_summary(
                parse_result.setdefault("error_summary", {}),
                safe_str(match.group("message")),
                row_number=int(match.group("row")),
            )
            return
        ReconciliationChecklistService._add_error_summary(
            parse_result.setdefault("error_summary", {}),
            message_text,
        )

    @staticmethod
    def _summarize_parse_result_errors(parse_result: dict) -> list[str]:
        error_summary = parse_result.get("error_summary")
        if isinstance(error_summary, dict) and error_summary:
            summarized: list[str] = []
            for message, group in error_summary.items():
                if isinstance(group, ErrorSummaryGroup):
                    count = group.count
                    rows = group.rows
                elif isinstance(group, dict):
                    count = int(group.get("count") or 0)
                    rows = [int(row) for row in group.get("rows", []) if int(row) > 0]
                else:
                    continue
                if rows:
                    row_text = "、".join(str(item) for item in rows[:CHECKLIST_ERROR_ROW_SAMPLE_LIMIT])
                    summarized.append(f"{message}：{count}行，部分行序号: {row_text}")
                else:
                    summarized.append(f"{message}：{count}次")
                if len(summarized) >= CHECKLIST_ERROR_SAMPLE_LIMIT:
                    break
            return summarized
        return ReconciliationChecklistService._summarize_errors(parse_result.get("errors", []))

    @staticmethod
    async def save_invoice_edit_items(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        items: Sequence[dict[str, object]],
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> dict[str, object]:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        if not items:
            raise ValueError("请先查询并修改数据")
        if len(items) > CHECKLIST_MANUAL_EDIT_MAX_SUB_ORDERS:
            raise ValueError("单次最多保存 100 条记录")

        parsed_rows = [
            ReconciliationChecklistService._parse_invoice_row(
                {
                    "唯一ID": item.get("unique_id"),
                    "子订单号": item.get("sub_order_no"),
                    "收款商家": item.get("receipt_merchant"),
                    "开票时间": item.get("invoice_time"),
                    "发票号码": item.get("invoice_number"),
                },
                index + 2,
            )
            for index, item in enumerate(items)
        ]
        success_count, failed_count, errors, affected_periods = await ReconciliationChecklistService._apply_invoice_edit_rows(
            db,
            org_id=org_id,
            rows=parsed_rows,
        )
        unchanged_count = max(len(parsed_rows) - failed_count - success_count, 0)
        if affected_periods:
            await ReconciliationChecklistService._rebuild_summaries(db, org_id=org_id, periods=affected_periods)
        await db.flush()

        missing_sub_order_nos = ReconciliationChecklistService._extract_missing_sub_orders_from_errors(parsed_rows, errors)
        affected_period_list = sorted(int(period) for period in affected_periods)
        sample_sub_order_nos = [safe_str(row.get("sub_order_no")) for row in parsed_rows[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT] if safe_str(row.get("sub_order_no"))]
        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=org_id,
            module="reconciliation_checklist",
            action="invoice_edit",
            description=f"保存发票人工编辑，共 {len(parsed_rows)} 条",
            target_type="reconciliation_checklist_manual_edit",
            target_id=None,
            target_name="invoice_edit",
            ip=ip,
            user_agent=user_agent,
            extra_data={
                "success_count": success_count,
                "failed_count": failed_count,
                "unchanged_count": unchanged_count,
                "missing_count": len(missing_sub_order_nos),
                "missing_sub_order_nos": missing_sub_order_nos[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT],
                "sample_sub_order_nos": sample_sub_order_nos,
                "affected_periods": affected_period_list,
                "error_messages": errors[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT],
            },
        )
        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "unchanged_count": unchanged_count,
            "missing_sub_order_nos": missing_sub_order_nos,
            "affected_periods": affected_period_list,
            "error_messages": errors,
        }

    @staticmethod
    async def save_merchant_edit_items(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        items: Sequence[dict[str, object]],
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> dict[str, object]:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        if not items:
            raise ValueError("请先查询并修改数据")
        if len(items) > CHECKLIST_MANUAL_EDIT_MAX_SUB_ORDERS:
            raise ValueError("单次最多保存 100 条记录")

        parsed_rows = [
            ReconciliationChecklistService._parse_merchant_row(
                {
                    "唯一ID": item.get("unique_id"),
                    "子订单号": item.get("sub_order_no"),
                    "收款商家": item.get("receipt_merchant"),
                    "应付商家净额": item.get("merchant_net_amount"),
                    "付款金额": item.get("payment_amount"),
                    "付款时间（商家）": item.get("merchant_payment_time"),
                },
                index + 2,
            )
            for index, item in enumerate(items)
        ]
        success_count, failed_count, errors, affected_periods = await ReconciliationChecklistService._apply_merchant_edit_rows(
            db,
            org_id=org_id,
            rows=parsed_rows,
        )
        unchanged_count = max(len(parsed_rows) - failed_count - success_count, 0)
        if affected_periods:
            await ReconciliationChecklistService._rebuild_summaries(db, org_id=org_id, periods=affected_periods)
        await db.flush()

        missing_sub_order_nos = ReconciliationChecklistService._extract_missing_sub_orders_from_errors(parsed_rows, errors)
        affected_period_list = sorted(int(period) for period in affected_periods)
        sample_sub_order_nos = [safe_str(row.get("sub_order_no")) for row in parsed_rows[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT] if safe_str(row.get("sub_order_no"))]
        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=org_id,
            module="reconciliation_checklist",
            action="merchant_edit",
            description=f"保存商家人工编辑，共 {len(parsed_rows)} 条",
            target_type="reconciliation_checklist_manual_edit",
            target_id=None,
            target_name="merchant_edit",
            ip=ip,
            user_agent=user_agent,
            extra_data={
                "success_count": success_count,
                "failed_count": failed_count,
                "unchanged_count": unchanged_count,
                "missing_count": len(missing_sub_order_nos),
                "missing_sub_order_nos": missing_sub_order_nos[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT],
                "sample_sub_order_nos": sample_sub_order_nos,
                "affected_periods": affected_period_list,
                "error_messages": errors[:CHECKLIST_MANUAL_EDIT_SAMPLE_LIMIT],
            },
        )
        return {
            "success_count": success_count,
            "failed_count": failed_count,
            "unchanged_count": unchanged_count,
            "missing_sub_order_nos": missing_sub_order_nos,
            "affected_periods": affected_period_list,
            "error_messages": errors,
        }

    @staticmethod
    async def create_from_shared_upload(
        db: AsyncSession,
        *,
        upload_file: UploadFile,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> ReconciliationChecklistTask:
        if (upload_file.parsed_type or "").strip() not in CHECKLIST_FILE_TYPES:
            raise ValueError("仅对账清单文件可创建对账清单任务")

        existing_stmt = select(ReconciliationChecklistUploadFile).where(
            ReconciliationChecklistUploadFile.source_upload_file_id == upload_file.id,
            ReconciliationChecklistUploadFile.is_deleted.is_(False),
        )
        existing = (await db.execute(existing_stmt)).scalar_one_or_none()
        checklist_upload_file = existing
        if existing is not None:
            task_stmt = (
                select(ReconciliationChecklistTask)
                .where(
                    ReconciliationChecklistTask.file_id == existing.id,
                    ReconciliationChecklistTask.is_deleted.is_(False),
                )
                .order_by(ReconciliationChecklistTask.id.desc())
            )
            current_task = (await db.execute(task_stmt)).scalars().first()
            if current_task is not None:
                return current_task

        if checklist_upload_file is None:
            checklist_upload_file = ReconciliationChecklistUploadFile(
                org_id=upload_file.org_id,
                user_id=user.id,
                source_upload_file_id=upload_file.id,
                original_name=upload_file.original_name,
                oss_key=upload_file.oss_key,
                file_size=upload_file.file_size,
                file_hash=upload_file.file_hash,
                status="uploaded",
            )
            db.add(checklist_upload_file)
            await db.flush()

        task = ReconciliationChecklistTask(
            file_id=checklist_upload_file.id,
            org_id=checklist_upload_file.org_id,
            user_id=user.id,
            status="queued",
            progress=0,
        )
        db.add(task)
        await db.flush()

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=checklist_upload_file.org_id,
            module="reconciliation_checklist",
            action="upload_file",
            description=f"上传对账清单文件 [{checklist_upload_file.original_name}]",
            target_type="reconciliation_checklist_upload_file",
            target_id=checklist_upload_file.id,
            target_name=checklist_upload_file.original_name,
            ip=ip,
            user_agent=user_agent,
            extra_data={"task_id": task.id, "source_upload_file_id": upload_file.id},
        )
        ReconciliationChecklistService.dispatch_task_after_commit(db, task=task, upload_file=checklist_upload_file)
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def init_manual_edit_upload(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        original_name: str,
        file_size: int,
        task_type: str,
    ) -> ReconciliationChecklistUploadFile:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        if task_type not in CHECKLIST_MANUAL_EDIT_TASK_TYPES:
            raise ValueError("不支持的修改任务类型")

        upload_file = ReconciliationChecklistUploadFile(
            org_id=org_id,
            user_id=user.id,
            source_upload_file_id=None,
            original_name=original_name,
            file_size=file_size,
            status="initialized",
        )
        db.add(upload_file)
        await db.flush()
        upload_file.oss_key = (
            f"upload/{checklist_manual_edit_dir(task_type)}/{current_oss_period(datetime.now())}/"
            f"{upload_file.id}_{sanitize_oss_filename(original_name)}"
        )
        await db.flush()
        await db.refresh(upload_file)
        return upload_file

    @staticmethod
    async def create_manual_edit_upload_task_from_oss(
        db: AsyncSession,
        *,
        org_id: int,
        user: User,
        original_name: str,
        oss_key: str,
        file_size: int,
        file_hash: str | None,
        task_type: str,
        file_id: int | None = None,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> ReconciliationChecklistTask:
        ReconciliationChecklistService.validate_org_scope(org_id=org_id, user=user)
        if task_type not in CHECKLIST_MANUAL_EDIT_TASK_TYPES:
            raise ValueError("不支持的修改任务类型")

        checklist_upload_file: ReconciliationChecklistUploadFile | None = None
        if file_id is not None:
            candidate = await db.get(ReconciliationChecklistUploadFile, file_id)
            if candidate is None or candidate.is_deleted:
                raise ValueError("上传文件不存在")
            ReconciliationChecklistService.validate_org_scope(org_id=candidate.org_id, user=user)
            if candidate.org_id != org_id:
                raise ValueError("上传文件组织与当前组织不一致")
            checklist_upload_file = candidate

            existing_stmt = (
                select(ReconciliationChecklistTask)
                .where(
                    ReconciliationChecklistTask.file_id == checklist_upload_file.id,
                    ReconciliationChecklistTask.is_deleted.is_(False),
                )
                .order_by(ReconciliationChecklistTask.id.desc())
            )
            existing_task = (await db.execute(existing_stmt)).scalars().first()
            if existing_task is not None:
                return existing_task

            checklist_upload_file.original_name = original_name
            checklist_upload_file.oss_key = oss_key
            checklist_upload_file.file_size = file_size
            checklist_upload_file.file_hash = file_hash
            checklist_upload_file.status = "uploaded"
            checklist_upload_file.error_message = None

        if checklist_upload_file is None:
            checklist_upload_file = ReconciliationChecklistUploadFile(
                org_id=org_id,
                user_id=user.id,
                source_upload_file_id=None,
                original_name=original_name,
                oss_key=oss_key,
                file_size=file_size,
                file_hash=file_hash,
                status="uploaded",
            )
            db.add(checklist_upload_file)
            await db.flush()

        task = ReconciliationChecklistTask(
            file_id=checklist_upload_file.id,
            org_id=org_id,
            user_id=user.id,
            task_type=task_type,
            status="queued",
            progress=0,
        )
        db.add(task)
        await db.flush()

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=org_id,
            module="reconciliation_checklist",
            action=task_type,
            description=f"上传对账清单人工修改文件 [{original_name}]",
            target_type="reconciliation_checklist_manual_edit_task",
            target_id=task.id,
            target_name=original_name,
            ip=ip,
            user_agent=user_agent,
            extra_data={"task_id": task.id, "file_id": checklist_upload_file.id, "task_type": task_type},
        )
        ReconciliationChecklistService.dispatch_task_after_commit(db, task=task, upload_file=checklist_upload_file)
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    def dispatch_task_after_commit(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
    ) -> None:
        async def dispatch() -> None:
            try:
                from app.tasks.reconciliation_checklist import run_reconciliation_checklist_task

                async_result = run_reconciliation_checklist_task.delay(task.id)
                task.celery_task_id = async_result.id
            except Exception as exc:
                task.status = "failed"
                task.progress = 100
                task.error_message = f"对账清单任务投递失败: {exc}"
                upload_file.status = "failed"
                upload_file.error_message = task.error_message
            await db.flush()

        register_after_commit(db, dispatch)

    @staticmethod
    async def rerun_task(
        db: AsyncSession,
        *,
        task_id: int,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> ReconciliationChecklistTask | None:
        task = await db.get(ReconciliationChecklistTask, task_id)
        if task is None or task.is_deleted:
            return None
        ReconciliationChecklistService.validate_org_scope(org_id=task.org_id, user=user)
        if task.status == "expired":
            raise ValueError("源文件已过期或不存在，不能重新导入，请重新上传文件")
        upload_file = await db.get(ReconciliationChecklistUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("对账清单上传文件不存在")

        task.status = "queued"
        task.progress = 0
        task.celery_task_id = None
        task.error_message = None
        task.started_at = None
        task.finished_at = None
        task.updated_at = datetime.now(timezone.utc)
        upload_file.status = "uploaded"
        upload_file.error_message = None

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=task.org_id,
            module="reconciliation_checklist",
            action="task_rerun",
            description=f"重跑对账清单任务 [{task.id}]",
            target_type="reconciliation_checklist_task",
            target_id=task.id,
            ip=ip,
            user_agent=user_agent,
        )
        ReconciliationChecklistService.dispatch_task_after_commit(db, task=task, upload_file=upload_file)
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def execute_task(db: AsyncSession, *, task_id: int) -> ReconciliationChecklistTask:
        task_started = perf_counter()
        task = await db.get(ReconciliationChecklistTask, task_id)
        if task is None or task.is_deleted:
            raise ValueError("对账清单任务不存在")
        upload_file = await db.get(ReconciliationChecklistUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("对账清单上传文件不存在")

        previous_state = _capture_result_state(task)
        task.status = "running"
        task.progress = 5
        task.started_at = datetime.now(timezone.utc)
        task.finished_at = None
        task.error_message = None
        await db.flush()
        await db.commit()

        try:
            suffix = Path(upload_file.original_name).suffix or ".xlsx"
            with tempfile.NamedTemporaryFile(suffix=suffix) as tmp:
                download_started = perf_counter()
                oss_service.download_to_temp(upload_file.oss_key, tmp.name)
                download_seconds = perf_counter() - download_started
                await db.commit()
                try:
                    if hasattr(db, "allow_db_get"):
                        setattr(db, "allow_db_get", False)
                    parse_started = perf_counter()
                    parse_result = ReconciliationChecklistService.parse_file(tmp.name)
                    parse_seconds = perf_counter() - parse_started
                finally:
                    if hasattr(db, "allow_db_get"):
                        setattr(db, "allow_db_get", True)
                parse_result["parse_seconds"] = parse_seconds
                parse_result["download_seconds"] = download_seconds
                task_type = safe_str(getattr(task, "task_type", "")) or CHECKLIST_TASK_TYPE_SOURCE_IMPORT
                if task_type in CHECKLIST_MANUAL_EDIT_TASK_TYPES:
                    expected_file_type = CHECKLIST_MANUAL_EDIT_TASK_TYPES[task_type]
                    if parse_result.get("file_type") != expected_file_type and not parse_result.get("fatal_error"):
                        parse_result["fatal_error"] = True
                        parse_result["failed_rows"] = max(1, int(parse_result.get("failed_rows") or 0))
                        parse_result["errors"].append("上传文件类型与任务类型不一致")
                        ReconciliationChecklistService._add_error_summary(
                            parse_result.setdefault("error_summary", {}),
                            "上传文件类型与任务类型不一致",
                        )
                summary = await ReconciliationChecklistService.persist_parsed_result(
                    db,
                    task=task,
                    upload_file=upload_file,
                    parse_result=parse_result,
                )

            task.total_rows = int(summary.get("总行数", 0))
            task.success_rows = int(summary.get("成功行数", 0))
            task.failed_rows = int(summary.get("失败行数", 0))
            task.inserted_rows = int(summary.get("新增行数", 0))
            task.updated_rows = int(summary.get("更新行数", 0))
            task.result_summary = summary
            task.progress = 100
            task.status = (
                "failed"
                if summary.get("文件解析失败")
                else (
                    "partial_success"
                    if task.failed_rows > 0 and task.success_rows > 0
                    else ("failed" if task.failed_rows > 0 and task.success_rows == 0 and task.total_rows > 0 else "success")
                )
            )
            error_messages = summary.get("错误明细")
            task.error_message = "\n".join(map(str, error_messages)) if isinstance(error_messages, list) and errors_are_fatal(summary) else None
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "failed" if task.status == "failed" else "processed"
            upload_file.error_message = task.error_message
        except Exception as exc:
            await db.rollback()
            task = await db.get(ReconciliationChecklistTask, task_id)
            if task is None or task.is_deleted:
                raise ValueError("对账清单任务不存在") from exc
            upload_file = await db.get(ReconciliationChecklistUploadFile, task.file_id)
            if upload_file is None or upload_file.is_deleted:
                raise ValueError("对账清单上传文件不存在") from exc
            is_expired = is_oss_object_unavailable_error(exc)
            if not is_expired:
                logger.exception(
                    "reconciliation_checklist.task_failed task_id=%s file_id=%s org_id=%s",
                    task.id,
                    task.file_id,
                    task.org_id,
                )
            _restore_result_state(task, previous_state)
            task.status = "expired" if is_expired else "failed"
            task.progress = 100
            task.error_message = SOURCE_FILE_UNAVAILABLE_MESSAGE if is_expired else CHECKLIST_INTERNAL_ERROR_MESSAGE
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "expired" if is_expired else "failed"
            upload_file.error_message = task.error_message

        logger.info(
            "对账清单任务阶段总览 task_id=%s file_id=%s 任务状态=%s 总行数=%s 新增行数=%s 更新行数=%s 未变更行数=%s 文件下载耗时秒=%s 解析耗时秒=%s 明细入库耗时秒=%s 汇总重建耗时秒=%s 任务总耗时秒=%s",
            task.id,
            task.file_id,
            task.status,
            task.total_rows,
            task.inserted_rows,
            task.updated_rows,
            int((task.result_summary or {}).get("未变更行数") or 0),
            _format_seconds(float((task.result_summary or {}).get("文件下载耗时秒") or 0)),
            _format_seconds(float((task.result_summary or {}).get("解析耗时秒") or 0)),
            _format_seconds(float((task.result_summary or {}).get("明细入库耗时秒") or 0)),
            _format_seconds(float((task.result_summary or {}).get("汇总重建耗时秒") or 0)),
            _format_seconds(perf_counter() - task_started),
        )
        await db.flush()
        await db.commit()
        await db.refresh(task)
        return task

    @staticmethod
    async def persist_task_result(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        file_path: str,
    ) -> dict:
        parse_result = ReconciliationChecklistService.parse_file(file_path)
        return await ReconciliationChecklistService.persist_parsed_result(
            db,
            task=task,
            upload_file=upload_file,
            parse_result=parse_result,
        )

    @staticmethod
    async def persist_parsed_result(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        parse_result: dict,
    ) -> dict:
        inserted_rows = 0
        updated_rows = 0
        unchanged_rows = 0
        persist_seconds = 0.0
        summary_seconds = 0.0
        affected_periods: set[int] = set()
        should_rebuild_summaries = False
        if not parse_result.get("fatal_error") and parse_result.get("rows"):
            affected_periods.update({int(row["accounting_period"]) for row in parse_result["rows"] if row.get("accounting_period")})
            if parse_result["file_type"] == CHECKLIST_FILE_TYPE_SOURCE:
                persist_started = perf_counter()
                inserted_rows, updated_rows, unchanged_rows, moved_periods = await ReconciliationChecklistService._persist_source_rows_sharded(
                    db,
                    task=task,
                    upload_file=upload_file,
                    rows=parse_result["rows"],
                )
                persist_seconds += perf_counter() - persist_started
                affected_periods.update(moved_periods)
                should_rebuild_summaries = bool(affected_periods)
            elif parse_result["file_type"] == CHECKLIST_FILE_TYPE_INVOICE:
                persist_started = perf_counter()
                updated_rows, failed_rows, errors, periods = await ReconciliationChecklistService._apply_invoice_rows(
                    db, task=task, upload_file=upload_file, rows=parse_result["rows"]
                )
                persist_seconds += perf_counter() - persist_started
                parse_result["failed_rows"] = int(parse_result["failed_rows"]) + failed_rows
                parse_result["success_rows"] = max(0, int(parse_result["success_rows"]) - failed_rows)
                parse_result["errors"].extend(errors)
                for error in errors:
                    ReconciliationChecklistService._add_error_text_to_summary(parse_result, error)
                affected_periods.update(periods)
                should_rebuild_summaries = bool(periods)
            elif parse_result["file_type"] == CHECKLIST_FILE_TYPE_MERCHANT:
                persist_started = perf_counter()
                updated_rows, failed_rows, errors, periods = await ReconciliationChecklistService._apply_merchant_rows(
                    db, task=task, upload_file=upload_file, rows=parse_result["rows"]
                )
                persist_seconds += perf_counter() - persist_started
                parse_result["failed_rows"] = int(parse_result["failed_rows"]) + failed_rows
                parse_result["success_rows"] = max(0, int(parse_result["success_rows"]) - failed_rows)
                parse_result["errors"].extend(errors)
                for error in errors:
                    ReconciliationChecklistService._add_error_text_to_summary(parse_result, error)
                affected_periods.update(periods)
                should_rebuild_summaries = bool(periods)
            if should_rebuild_summaries and affected_periods:
                summary_started = perf_counter()
                await ReconciliationChecklistService._rebuild_summaries(db, org_id=task.org_id, periods=affected_periods)
                summary_seconds += perf_counter() - summary_started
            await db.flush()

        logger.info(
            "对账清单导入阶段耗时 task_id=%s file_id=%s 文件类型=%s 总行数=%s 新增行数=%s 更新行数=%s 未变更行数=%s 解析耗时秒=%s 明细入库耗时秒=%s 汇总重建耗时秒=%s 涉及年月=%s",
            task.id,
            upload_file.id,
            parse_result.get("file_type") or upload_file.original_name,
            parse_result.get("total_rows", 0),
            inserted_rows,
            updated_rows,
            unchanged_rows,
            _format_seconds(float(parse_result.get("parse_seconds") or 0.0)),
            _format_seconds(persist_seconds),
            _format_seconds(summary_seconds),
            sorted(affected_periods),
        )

        summary = {
            "文件类型": parse_result.get("file_type") or upload_file.original_name,
            "处理结果": parse_result.get("result_message") or "处理完成",
            "总行数": parse_result.get("total_rows", 0),
            "成功行数": parse_result.get("success_rows", 0),
            "失败行数": parse_result.get("failed_rows", 0),
            "新增行数": inserted_rows,
            "更新行数": updated_rows,
            "未变更行数": unchanged_rows,
            "涉及年月": sorted(affected_periods),
            "文件下载耗时秒": round(float(parse_result.get("download_seconds") or 0), 3),
            "解析耗时秒": round(float(parse_result.get("parse_seconds") or 0), 3),
            "明细入库耗时秒": round(persist_seconds, 3),
            "汇总重建耗时秒": round(summary_seconds, 3),
        }
        errors = ReconciliationChecklistService._summarize_parse_result_errors(parse_result)
        if errors:
            summary["错误明细"] = errors
        if parse_result.get("fatal_error"):
            summary["文件解析失败"] = "是"
        return summary

    @staticmethod
    def _summarize_errors(errors: Sequence[object]) -> list[str]:
        grouped_rows: dict[str, list[int]] = {}
        grouped_counts: dict[str, int] = {}
        ordered_messages: list[str] = []

        for raw_error in errors:
            message_text = safe_str(raw_error)
            if not message_text:
                continue
            match = ReconciliationChecklistService.ERROR_ROW_PATTERN.match(message_text)
            if match:
                normalized_message = safe_str(match.group("message"))
                row_number = int(match.group("row"))
                if normalized_message not in grouped_rows:
                    grouped_rows[normalized_message] = []
                    grouped_counts[normalized_message] = 0
                    ordered_messages.append(normalized_message)
                grouped_rows[normalized_message].append(row_number)
                grouped_counts[normalized_message] += 1
                continue

            if message_text not in grouped_counts:
                grouped_rows[message_text] = []
                grouped_counts[message_text] = 0
                ordered_messages.append(message_text)
            grouped_counts[message_text] += 1

        summarized: list[str] = []
        for message in ordered_messages:
            row_numbers = grouped_rows.get(message) or []
            count = grouped_counts.get(message, 0)
            if row_numbers:
                row_text = "、".join(str(item) for item in row_numbers[:CHECKLIST_ERROR_ROW_SAMPLE_LIMIT])
                summarized.append(f"{message}：{count}行，部分行序号: {row_text}")
            else:
                summarized.append(f"{message}：{count}次")
            if len(summarized) >= CHECKLIST_ERROR_SAMPLE_LIMIT:
                break
        return summarized

    @staticmethod
    def parse_file(file_path: str) -> dict:
        result = {
            "file_type": None,
            "total_rows": 0,
            "success_rows": 0,
            "failed_rows": 0,
            "errors": [],
            "error_summary": {},
            "warnings": [],
            "fatal_error": False,
            "rows": [],
            "result_message": "",
        }

        def mark_empty(message: str = CHECKLIST_EMPTY_SUMMARY_MESSAGE) -> dict:
            result["result_message"] = message
            result["fatal_error"] = False
            result["total_rows"] = 0
            result["success_rows"] = 0
            result["failed_rows"] = 0
            return result

        sheet_rows = ReconciliationChecklistService._iter_candidate_sheet_rows(file_path)
        if not sheet_rows:
            return mark_empty("空数据：未发现可处理行")

        header_info = ReconciliationChecklistService._find_header_row(sheet_rows)
        if header_info is None:
            if ReconciliationChecklistService._rows_are_empty(sheet_rows):
                return mark_empty("空数据：未发现可处理行")
            result["errors"].append("缺少对账清单必要表头")
            result["failed_rows"] = 1
            result["fatal_error"] = True
            return result

        spec, canonical_headers, header_row_index, rows = header_info
        result["file_type"] = spec.file_type
        known_headers = list(spec.headers)
        if spec.file_type == CHECKLIST_FILE_TYPE_SOURCE:
            known_headers.extend(SOURCE_OPTIONAL_HEADERS)
        all_column_lookup = _build_header_column_lookup(canonical_headers)
        column_lookup = {
            normalized_header: index
            for header in known_headers
            if (normalized_header := _normalized_required_header(header)) in all_column_lookup
            for index in (all_column_lookup[normalized_header],)
        }
        for row_index in range(header_row_index + 1, len(rows)):
            row = rows[row_index]
            source_index = row_index + 1
            if not any(safe_str(value) for value in row):
                continue
            result["total_rows"] += 1
            try:
                values = _RowValueAccessor(row=row, lookup=column_lookup)
                if spec.file_type == CHECKLIST_FILE_TYPE_SOURCE:
                    parsed = ReconciliationChecklistService._parse_source_row(values, source_index)
                elif spec.file_type == CHECKLIST_FILE_TYPE_INVOICE:
                    parsed = ReconciliationChecklistService._parse_invoice_row(values, source_index)
                else:
                    parsed = ReconciliationChecklistService._parse_merchant_row(values, source_index)
                result["rows"].append(parsed)
                result["success_rows"] += 1
            except Exception as exc:
                ReconciliationChecklistService._record_parse_error(result, source_index, exc)
        if result["total_rows"] == 0:
            return mark_empty(CHECKLIST_EMPTY_SUMMARY_MESSAGE)
        return result

    @staticmethod
    def _iter_candidate_sheet_rows(file_path: str) -> list[list[object]]:
        path = Path(file_path)
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
            except Exception:
                wb = None
            if wb is not None:
                try:
                    fallback_rows: list[list[object]] = []
                    for ws in wb.worksheets:
                        ws.reset_dimensions()
                        rows = [list(row) for row in ws.iter_rows(values_only=True)]
                        if not rows:
                            continue
                        if ReconciliationChecklistService._find_header_row([rows]) is not None:
                            return rows
                        if not fallback_rows and not ReconciliationChecklistService._rows_are_empty([rows]):
                            fallback_rows = rows
                    return fallback_rows
                finally:
                    wb.close()
        with open_tabular_rows(file_path) as rows:
            if rows is None:
                return []
            return [list(row) for row in rows]

    @staticmethod
    def _rows_are_empty(sheet_rows: list[list[object]] | list[list[list[object]]]) -> bool:
        rows: Iterable
        if sheet_rows and isinstance(sheet_rows[0], list) and sheet_rows[0] and isinstance(sheet_rows[0][0], list):
            rows = (row for sheet in sheet_rows for row in sheet)  # type: ignore[assignment]
        else:
            rows = sheet_rows
        return not any(any(safe_str(value) for value in row) for row in rows)

    @staticmethod
    def _find_header_row(sheet_rows: list[list[object]] | list[list[list[object]]]):
        candidate_sheets: list[list[list[object]]]
        if sheet_rows and isinstance(sheet_rows[0], list) and sheet_rows[0] and isinstance(sheet_rows[0][0], list):
            candidate_sheets = sheet_rows  # type: ignore[assignment]
        else:
            candidate_sheets = [sheet_rows]  # type: ignore[list-item]
        for rows in candidate_sheets:
            for index, row in enumerate(rows[:20]):
                canonical_headers = [_canonical_header(value) for value in row]
                try:
                    spec = _detect_checklist_type(canonical_headers)
                except ValueError:
                    continue
                return spec, canonical_headers, index, rows
        return None

    @staticmethod
    def _parse_source_row(values: dict[str, object] | _RowValueAccessor, row_number: int) -> dict[str, object]:
        settlement_time = parse_datetime(_header_value(values, "结算时间"))
        if settlement_time is None:
            raise ValueError("结算时间无法解析")
        sub_order_no = safe_str(_header_value(values, "子订单号"))
        if not sub_order_no:
            raise ValueError("子订单号不能为空")
        live_platform = safe_str(_header_value(values, "进驻的直播平台"))
        merchant_subject_name = safe_str(_header_value(values, "商户主体名称"))
        receipt_merchant = safe_str(_header_value(values, "收款商家"))
        if not live_platform:
            raise ValueError("进驻的直播平台不能为空")
        if not merchant_subject_name:
            raise ValueError("商户主体名称不能为空")
        period = _accounting_period(settlement_time.year, settlement_time.month)
        parsed: dict[str, object] = {
            "source_row_number": row_number,
            "accounting_year": settlement_time.year,
            "accounting_month": settlement_time.month,
            "accounting_period": period,
            "settlement_time": settlement_time,
            "sub_order_no": sub_order_no,
            "live_platform": live_platform,
            "merchant_subject_name": merchant_subject_name,
            "receipt_merchant": receipt_merchant,
        }
        for header, field in SOURCE_FIELD_MAP.items():
            if field in parsed:
                continue
            value = _header_value(values, header)
            if field in MONEY_FIELDS:
                parsed[field] = _money(value)
            elif field in {"order_time", "merchant_payment_time", "invoice_time"}:
                parsed[field] = _nullable_datetime(value)
            elif field == "product_quantity":
                parsed[field] = _to_int(value)
            elif field == "commission_rate":
                parsed[field] = _nullable_rate(value)
            elif field == "payment_amount":
                parsed[field] = _nullable_money(value)
            else:
                parsed[field] = safe_str(value)
        parsed["row_fingerprint"] = _row_fingerprint(
            org_id=0,
            settlement_time=settlement_time,
            sub_order_no=sub_order_no,
            platform_subsidy=parsed.get("platform_subsidy"),
            talent_subsidy=parsed.get("talent_subsidy"),
            douyin_pay_subsidy=parsed.get("douyin_pay_subsidy"),
            douyin_monthly_pay_subsidy=parsed.get("douyin_monthly_pay_subsidy"),
            bank_subsidy=parsed.get("bank_subsidy"),
            user_paid_amount=parsed.get("user_paid_amount"),
            platform_service_fee=parsed.get("platform_service_fee"),
            talent_commission=parsed.get("talent_commission"),
            investment_service_fee=parsed.get("investment_service_fee"),
        )
        return parsed

    @staticmethod
    def _parse_invoice_row(values: dict[str, object] | _RowValueAccessor, row_number: int) -> dict[str, object]:
        unique_id = safe_str(_header_value(values, "唯一ID")) or safe_str(_header_value(values, "系统行定位值"))
        if not unique_id:
            raise ValueError("唯一ID不能为空")
        return {
            "source_row_number": row_number,
            "unique_id": unique_id,
            "row_fingerprint": unique_id,
            "sub_order_no": safe_str(_header_value(values, "子订单号")),
            "receipt_merchant": safe_str(_header_value(values, "收款商家")),
            "invoice_time": _nullable_datetime(_header_value(values, "开票时间")),
            "invoice_number": safe_str(_header_value(values, "发票号码")),
        }

    @staticmethod
    def _parse_merchant_row(values: dict[str, object] | _RowValueAccessor, row_number: int) -> dict[str, object]:
        unique_id = safe_str(_header_value(values, "唯一ID")) or safe_str(_header_value(values, "系统行定位值"))
        if not unique_id:
            raise ValueError("唯一ID不能为空")
        return {
            "source_row_number": row_number,
            "unique_id": unique_id,
            "row_fingerprint": unique_id,
            "sub_order_no": safe_str(_header_value(values, "子订单号")),
            "receipt_merchant": safe_str(_header_value(values, "收款商家")),
            "merchant_net_amount": _nullable_money(_header_value(values, "应付商家净额")),
            "payment_amount": _nullable_money(_header_value(values, "付款金额")),
            "merchant_payment_time": _nullable_datetime(_header_value(values, "付款时间（商家）")),
        }

    @staticmethod
    def _dedupe_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
        positions: dict[tuple[str, str], int] = {}
        deduped: list[dict[str, object]] = []
        for row in rows:
            unique_id = safe_str(row.get("unique_id")) or safe_str(row.get("row_fingerprint"))
            key = ("unique_id", unique_id) if unique_id else ("sub_order_no", safe_str(row.get("sub_order_no")))
            if key in positions:
                deduped[positions[key]] = row
            else:
                positions[key] = len(deduped)
                deduped.append(row)
        return deduped

    @staticmethod
    async def _lookup_details_by_unique_ids(
        db: AsyncSession,
        *,
        org_id: int,
        unique_ids: Sequence[str],
    ) -> dict[str, ReconciliationChecklistDetail]:
        detail_map: dict[str, ReconciliationChecklistDetail] = {}
        unique_values = [value for value in dict.fromkeys(safe_str(item) for item in unique_ids) if value]
        if not unique_values:
            return detail_map
        for start in range(0, len(unique_values), CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE):
            chunk = unique_values[start : start + CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE]
            result = await db.execute(
                select(ReconciliationChecklistDetail).where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.row_fingerprint.in_(chunk),
                )
            )
            detail_map.update({safe_str(item.row_fingerprint): item for item in result.scalars().all()})
        return detail_map

    @staticmethod
    async def _lookup_details_by_sub_order_numbers(
        db: AsyncSession,
        *,
        org_id: int,
        sub_order_numbers: Sequence[str],
    ) -> dict[str, list[ReconciliationChecklistDetail]]:
        grouped: dict[str, list[ReconciliationChecklistDetail]] = defaultdict(list)
        unique_numbers = [value for value in dict.fromkeys(safe_str(item) for item in sub_order_numbers) if value]
        if not unique_numbers:
            return grouped
        for start in range(0, len(unique_numbers), CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE):
            chunk = unique_numbers[start : start + CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE]
            result = await db.execute(
                select(ReconciliationChecklistDetail).where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.sub_order_no.in_(chunk),
                )
            )
            for item in result.scalars().all():
                grouped[safe_str(item.sub_order_no)].append(item)
        for items in grouped.values():
            items.sort(key=lambda item: (item.settlement_time or datetime.min, safe_str(item.row_fingerprint), item.id or 0))
        return grouped

    @staticmethod
    async def _resolve_target_details(
        db: AsyncSession,
        *,
        org_id: int,
        rows: Sequence[dict[str, object]],
        require_unique_id: bool = False,
    ) -> tuple[dict[str, ReconciliationChecklistDetail], dict[str, list[ReconciliationChecklistDetail]]]:
        unique_id_map = await ReconciliationChecklistService._lookup_details_by_unique_ids(
            db,
            org_id=org_id,
            unique_ids=[safe_str(row.get("unique_id")) or safe_str(row.get("row_fingerprint")) for row in rows],
        )
        if require_unique_id:
            return unique_id_map, {}
        sub_order_map = await ReconciliationChecklistService._lookup_details_by_sub_order_numbers(
            db,
            org_id=org_id,
            sub_order_numbers=[safe_str(row.get("sub_order_no")) for row in rows],
        )
        return unique_id_map, sub_order_map

    @staticmethod
    def _resolve_target_detail(
        *,
        row: dict[str, object],
        unique_id_map: dict[str, ReconciliationChecklistDetail],
        sub_order_map: dict[str, list[ReconciliationChecklistDetail]],
        require_unique_id: bool = False,
    ) -> tuple[ReconciliationChecklistDetail | None, str | None]:
        unique_id = safe_str(row.get("unique_id")) or safe_str(row.get("row_fingerprint"))
        sub_order_no = safe_str(row.get("sub_order_no"))
        if require_unique_id and not unique_id:
            return None, "唯一ID不能为空"
        if unique_id:
            detail = unique_id_map.get(unique_id)
            if detail is None:
                return None, "唯一ID不存在"
            return detail, None
        matches = sub_order_map.get(sub_order_no, [])
        if not matches:
            return None, "子订单号不存在"
        if len(matches) > 1:
            return None, "子订单号存在多条记录，请重新查询后使用最新模板"
        return matches[0], None

    @staticmethod
    async def _fetch_order_keys_by_sub_order_numbers(
        db: AsyncSession,
        *,
        org_id: int,
        sub_order_numbers: Sequence[str],
    ) -> dict[str, ReconciliationChecklistOrderKey]:
        key_map: dict[str, ReconciliationChecklistOrderKey] = {}
        unique_numbers = [value for value in dict.fromkeys(safe_str(item) for item in sub_order_numbers) if value]
        for start in range(0, len(unique_numbers), CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE):
            chunk = unique_numbers[start : start + CHECKLIST_ORDER_KEY_LOOKUP_CHUNK_SIZE]
            result = await db.execute(
                select(ReconciliationChecklistOrderKey).where(
                    ReconciliationChecklistOrderKey.org_id == org_id,
                    ReconciliationChecklistOrderKey.sub_order_no.in_(chunk),
                )
            )
            key_map.update({item.sub_order_no: item for item in result.scalars().all()})
        return key_map

    @staticmethod
    async def _persist_source_rows(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        rows: Iterable[dict[str, object]],
        accounting_period: int | None = None,
    ) -> tuple[int, int, int, set[int]]:
        deduped = ReconciliationChecklistService._dedupe_rows(rows)
        if not deduped:
            return 0, 0, 0, set()
        shard_period = accounting_period
        if shard_period is None:
            shard_periods = {int(row.get("accounting_period") or 0) for row in deduped if int(row.get("accounting_period") or 0) > 0}
            shard_period = next(iter(shard_periods)) if len(shard_periods) == 1 else None
        moved_periods: set[int] = set()
        detail_rows = [ReconciliationChecklistService._detail_values(task=task, upload_file=upload_file, row=row) for row in deduped]
        stage_table = _source_stage_table()
        stage_prepare_started = perf_counter()
        await db.execute(text(f"DROP TABLE IF EXISTS {CHECKLIST_SOURCE_STAGE_TABLE_NAME}"))
        await db.execute(
            text(
                f"""
                CREATE TEMP TABLE {CHECKLIST_SOURCE_STAGE_TABLE_NAME}
                (LIKE {ReconciliationChecklistDetail.__tablename__} INCLUDING DEFAULTS)
                ON COMMIT DROP
                """
            )
        )
        stage_prepare_seconds = perf_counter() - stage_prepare_started

        loaded_by_copy = False
        stage_load_started = perf_counter()
        try:
            loaded_by_copy = await ReconciliationChecklistService._copy_stage_rows_with_savepoint(
                db,
                table_name=CHECKLIST_SOURCE_STAGE_TABLE_NAME,
                rows=detail_rows,
            )
        except Exception:
            logger.warning(
                "reconciliation_checklist.stage_copy_failed org_id=%s task_id=%s file_id=%s; falling back to insert-values",
                task.org_id,
                task.id,
                upload_file.id,
                exc_info=True,
            )

        if not loaded_by_copy:
            for chunk in _chunk_records(detail_rows):
                await db.execute(insert(stage_table).values(list(chunk)))
        stage_load_seconds = perf_counter() - stage_load_started

        join_condition = (
            (ReconciliationChecklistDetail.org_id == stage_table.c.org_id)
            & (ReconciliationChecklistDetail.accounting_period == stage_table.c.accounting_period)
            & (ReconciliationChecklistDetail.row_fingerprint == stage_table.c.row_fingerprint)
            & ReconciliationChecklistDetail.is_deleted.is_(False)
        )
        if shard_period is not None:
            join_condition = join_condition & (ReconciliationChecklistDetail.accounting_period == shard_period)
        business_changes = or_(*(getattr(ReconciliationChecklistDetail, field_name).is_distinct_from(stage_table.c[field_name]) for field_name in SOURCE_DETAIL_COMPARE_FIELDS))
        joined_stage = stage_table.outerjoin(ReconciliationChecklistDetail, join_condition)

        compare_started = perf_counter()
        counts_result = await db.execute(
            select(
                func.count().filter(ReconciliationChecklistDetail.id.is_(None)),
                func.count().filter(ReconciliationChecklistDetail.id.is_not(None) & business_changes),
                func.count().filter(ReconciliationChecklistDetail.id.is_not(None) & ~business_changes),
            ).select_from(joined_stage)
        )
        insert_count, update_count, unchanged_count = tuple(int(value or 0) for value in counts_result.one())
        compare_seconds = perf_counter() - compare_started
        if insert_count == 0 and update_count == 0:
            logger.info(
                "对账清单分片入库统计 task_id=%s file_id=%s 分片账期=%s 分片行数=%s 临时表准备耗时秒=%s 分片装载耗时秒=%s 分片比对耗时秒=%s 分片写入耗时秒=%s 新增行数=%s 更新行数=%s 未变更行数=%s COPY装载=%s",
                task.id,
                upload_file.id,
                shard_period,
                len(detail_rows),
                _format_seconds(stage_prepare_seconds),
                _format_seconds(stage_load_seconds),
                _format_seconds(compare_seconds),
                _format_seconds(0.0),
                insert_count,
                update_count,
                unchanged_count,
                "是" if loaded_by_copy else "否",
            )
            return insert_count, update_count, unchanged_count, moved_periods

        detail_column_names = list(detail_rows[0].keys())
        source_select = (
            select(*(stage_table.c[column_name] for column_name in detail_column_names))
            .select_from(joined_stage)
            .where(ReconciliationChecklistDetail.id.is_(None) | business_changes)
        )
        stmt = insert(ReconciliationChecklistDetail).from_select(detail_column_names, source_select)
        update_values = {column_name: stmt.excluded[column_name] for column_name in detail_column_names if column_name not in {"id", "created_at"}}
        update_values["updated_at"] = func.now()
        stmt = stmt.on_conflict_do_update(
            index_elements=["org_id", "accounting_period", "row_fingerprint"],
            index_where=text("is_deleted = false AND row_fingerprint <> ''"),
            set_=update_values,
        )
        write_started = perf_counter()
        await db.execute(stmt)
        write_seconds = perf_counter() - write_started

        logger.info(
            "对账清单分片入库统计 task_id=%s file_id=%s 分片账期=%s 分片行数=%s 临时表准备耗时秒=%s 分片装载耗时秒=%s 分片比对耗时秒=%s 分片写入耗时秒=%s 新增行数=%s 更新行数=%s 未变更行数=%s COPY装载=%s",
            task.id,
            upload_file.id,
            shard_period,
            len(detail_rows),
            _format_seconds(stage_prepare_seconds),
            _format_seconds(stage_load_seconds),
            _format_seconds(compare_seconds),
            _format_seconds(write_seconds),
            insert_count,
            update_count,
            unchanged_count,
            "是" if loaded_by_copy else "否",
        )

        return insert_count, update_count, unchanged_count, moved_periods

    @staticmethod
    async def _copy_stage_rows(
        db: AsyncSession,
        *,
        table_name: str,
        rows: Sequence[dict[str, object]],
    ) -> bool:
        if not rows:
            return True

        copy_hook = getattr(db, "_copy_records_to_stage", None)
        if callable(copy_hook):
            return bool(await copy_hook(table_name, rows, _copy_stage_column_names(rows)))

        session_connection = await db.connection()
        raw_connection = await session_connection.get_raw_connection()
        driver_connection = getattr(raw_connection, "driver_connection", None)
        if driver_connection is None or not hasattr(driver_connection, "copy_records_to_table"):
            return False

        column_names = _copy_stage_column_names(rows)
        records = [tuple(row.get(column_name) for column_name in column_names) for row in rows]
        await driver_connection.copy_records_to_table(
            table_name,
            records=records,
            columns=column_names,
        )
        return True

    @staticmethod
    async def _copy_stage_rows_with_savepoint(
        db: AsyncSession,
        *,
        table_name: str,
        rows: Sequence[dict[str, object]],
    ) -> bool:
        begin_nested = getattr(db, "begin_nested", None)
        if not callable(begin_nested):
            return await ReconciliationChecklistService._copy_stage_rows(
                db,
                table_name=table_name,
                rows=rows,
            )

        nested_transaction = await begin_nested()
        try:
            loaded = await ReconciliationChecklistService._copy_stage_rows(
                db,
                table_name=table_name,
                rows=rows,
            )
        except Exception:
            await nested_transaction.rollback()
            raise

        await nested_transaction.commit()
        return loaded

    @staticmethod
    async def _persist_source_rows_sharded(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        rows: Iterable[dict[str, object]],
    ) -> tuple[int, int, int, set[int]]:
        deduped = ReconciliationChecklistService._dedupe_rows(rows)
        if not deduped:
            return 0, 0, 0, set()

        rows_by_period: dict[int, list[dict[str, object]]] = defaultdict(list)
        rows_without_period: list[dict[str, object]] = []
        for row in deduped:
            period = int(row.get("accounting_period") or 0)
            if period > 0:
                rows_by_period[period].append(row)
            else:
                rows_without_period.append(row)

        ordered_shards: list[tuple[int | None, list[dict[str, object]]]] = []
        for period, period_rows in sorted(rows_by_period.items(), key=lambda item: item[0]):
            ordered_shards.extend(
                [(period, period_rows[start : start + CHECKLIST_SOURCE_PERSIST_SHARD_SIZE]) for start in range(0, len(period_rows), CHECKLIST_SOURCE_PERSIST_SHARD_SIZE)]
            )
        if rows_without_period:
            ordered_shards.extend(
                [
                    (None, rows_without_period[start : start + CHECKLIST_SOURCE_PERSIST_SHARD_SIZE])
                    for start in range(0, len(rows_without_period), CHECKLIST_SOURCE_PERSIST_SHARD_SIZE)
                ]
            )

        inserted_total = 0
        updated_total = 0
        unchanged_total = 0
        moved_periods: set[int] = set()

        for shard_period, shard_rows in ordered_shards:
            shard_started = perf_counter()
            inserted, updated, unchanged, shard_moved_periods = await ReconciliationChecklistService._persist_source_rows(
                db,
                task=task,
                upload_file=upload_file,
                rows=shard_rows,
                accounting_period=shard_period,
            )
            inserted_total += inserted
            updated_total += updated
            unchanged_total += unchanged
            moved_periods.update(shard_moved_periods)
            await db.commit()
            logger.info(
                "对账清单分片提交完成 task_id=%s file_id=%s 分片账期=%s 分片行数=%s 新增行数=%s 更新行数=%s 未变更行数=%s 分片总耗时秒=%s",
                task.id,
                upload_file.id,
                shard_period,
                len(shard_rows),
                inserted,
                updated,
                unchanged,
                _format_seconds(perf_counter() - shard_started),
            )

        return inserted_total, updated_total, unchanged_total, moved_periods

    @staticmethod
    def _source_detail_has_business_changes(existing_detail: dict[str, object], row: dict[str, object]) -> bool:
        return any(existing_detail.get(field_name) != row.get(field_name) for field_name in SOURCE_DETAIL_COMPARE_FIELDS)

    @staticmethod
    def _detail_values(*, task: ReconciliationChecklistTask, upload_file: ReconciliationChecklistUploadFile, row: dict[str, object]) -> dict[str, object]:
        fields = [
            "accounting_year",
            "accounting_month",
            "accounting_period",
            "source_row_number",
            "live_platform",
            "settlement_time",
            "sub_order_no",
            "order_time",
            "product_id",
            "product_name",
            "product_quantity",
            "talent_name",
            "platform_subsidy",
            "talent_subsidy",
            "douyin_pay_subsidy",
            "douyin_monthly_pay_subsidy",
            "bank_subsidy",
            "user_paid_amount",
            "platform_service_fee",
            "talent_commission",
            "investment_service_fee",
            "merchant_subject_name",
            "customer_service_code",
            "receipt_merchant",
            "live_commission",
            "commission_rate",
            "merchant_net_amount",
            "payment_amount",
            "merchant_payment_time",
            "invoice_time",
            "invoice_number",
            "merchant_net_balance",
        ]
        values = {field: row.get(field) for field in fields}
        values.update({"task_id": task.id, "file_id": upload_file.id, "org_id": task.org_id})
        values["row_fingerprint"] = _row_fingerprint(
            org_id=task.org_id,
            settlement_time=values.get("settlement_time"),
            sub_order_no=values.get("sub_order_no"),
            platform_subsidy=values.get("platform_subsidy"),
            talent_subsidy=values.get("talent_subsidy"),
            douyin_pay_subsidy=values.get("douyin_pay_subsidy"),
            douyin_monthly_pay_subsidy=values.get("douyin_monthly_pay_subsidy"),
            bank_subsidy=values.get("bank_subsidy"),
            user_paid_amount=values.get("user_paid_amount"),
            platform_service_fee=values.get("platform_service_fee"),
            talent_commission=values.get("talent_commission"),
            investment_service_fee=values.get("investment_service_fee"),
        )
        return values

    @staticmethod
    async def _lookup_order_keys(db: AsyncSession, *, org_id: int, rows: Iterable[dict[str, object]]) -> dict[str, ReconciliationChecklistOrderKey]:
        sub_order_numbers = [safe_str(row["sub_order_no"]) for row in rows]
        if not sub_order_numbers:
            return {}
        return await ReconciliationChecklistService._fetch_order_keys_by_sub_order_numbers(
            db,
            org_id=org_id,
            sub_order_numbers=sub_order_numbers,
        )

    @staticmethod
    async def _lookup_existing_detail_pairs(
        db: AsyncSession,
        *,
        org_id: int,
        pairs: Iterable[tuple[int, str]],
    ) -> set[tuple[int, str]]:
        pair_list = [(int(period), safe_str(sub_order_no)) for period, sub_order_no in pairs if int(period) > 0 and safe_str(sub_order_no)]
        if not pair_list:
            return set()
        detail_pairs: set[tuple[int, str]] = set()
        for chunk in _chunk_tuple_lookup_records(pair_list):
            result = await db.execute(
                select(
                    ReconciliationChecklistDetail.accounting_period,
                    ReconciliationChecklistDetail.sub_order_no,
                ).where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    tuple_(
                        ReconciliationChecklistDetail.accounting_period,
                        ReconciliationChecklistDetail.sub_order_no,
                    ).in_(chunk),
                )
            )
            detail_pairs.update((int(accounting_period), safe_str(sub_order_no)) for accounting_period, sub_order_no in result.all())
        return detail_pairs

    @staticmethod
    async def _lookup_existing_receipt_merchants(
        db: AsyncSession,
        *,
        org_id: int,
        pairs: Iterable[tuple[int, str]],
    ) -> dict[tuple[int, str], str]:
        pair_list = [(int(period), safe_str(sub_order_no)) for period, sub_order_no in pairs if int(period) > 0 and safe_str(sub_order_no)]
        if not pair_list:
            return {}
        receipt_merchant_map: dict[tuple[int, str], str] = {}
        for chunk in _chunk_tuple_lookup_records(pair_list):
            result = await db.execute(
                select(
                    ReconciliationChecklistDetail.accounting_period,
                    ReconciliationChecklistDetail.sub_order_no,
                    ReconciliationChecklistDetail.receipt_merchant,
                ).where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    tuple_(
                        ReconciliationChecklistDetail.accounting_period,
                        ReconciliationChecklistDetail.sub_order_no,
                    ).in_(chunk),
                )
            )
            receipt_merchant_map.update(
                {(int(accounting_period), safe_str(sub_order_no)): safe_str(receipt_merchant) for accounting_period, sub_order_no, receipt_merchant in result.all()}
            )
        return receipt_merchant_map

    @staticmethod
    async def _lookup_existing_merchant_summary_values(
        db: AsyncSession,
        *,
        org_id: int,
        pairs: Iterable[tuple[int, str]],
    ) -> dict[tuple[int, str], tuple[str, Decimal, Decimal | None]]:
        pair_list = [(int(period), safe_str(sub_order_no)) for period, sub_order_no in pairs if int(period) > 0 and safe_str(sub_order_no)]
        if not pair_list:
            return {}
        summary_value_map: dict[tuple[int, str], tuple[str, Decimal, Decimal | None]] = {}
        for chunk in _chunk_tuple_lookup_records(pair_list):
            result = await db.execute(
                select(
                    ReconciliationChecklistDetail.accounting_period,
                    ReconciliationChecklistDetail.sub_order_no,
                    ReconciliationChecklistDetail.receipt_merchant,
                    ReconciliationChecklistDetail.merchant_net_amount,
                    ReconciliationChecklistDetail.payment_amount,
                ).where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    tuple_(
                        ReconciliationChecklistDetail.accounting_period,
                        ReconciliationChecklistDetail.sub_order_no,
                    ).in_(chunk),
                )
            )
            summary_value_map.update(
                {
                    (int(accounting_period), safe_str(sub_order_no)): (
                        safe_str(receipt_merchant),
                        merchant_net_amount if merchant_net_amount is not None else Decimal("0.00"),
                        payment_amount,
                    )
                    for accounting_period, sub_order_no, receipt_merchant, merchant_net_amount, payment_amount in result.all()
                }
            )
        return summary_value_map

    @staticmethod
    async def _apply_invoice_rows(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        rows: Iterable[dict[str, object]],
    ) -> tuple[int, int, list[str], set[int]]:
        row_list = ReconciliationChecklistService._dedupe_rows(rows)
        unique_id_map, sub_order_map = await ReconciliationChecklistService._resolve_target_details(
            db,
            org_id=task.org_id,
            rows=row_list,
            require_unique_id=True,
        )
        updated = 0
        failed = 0
        errors: list[str] = []
        periods: set[int] = set()
        matched_rows: list[dict[str, object]] = []
        for row in row_list:
            detail, error_message = ReconciliationChecklistService._resolve_target_detail(
                row=row,
                unique_id_map=unique_id_map,
                sub_order_map=sub_order_map,
                require_unique_id=True,
            )
            if detail is None:
                failed += 1
                errors.append(f"Row {row.get('source_row_number')}: {error_message}")
                continue
            next_receipt_merchant = safe_str(row.get("receipt_merchant"))
            if safe_str(detail.receipt_merchant) != next_receipt_merchant:
                periods.add(int(detail.accounting_period))
            matched_rows.append(
                {
                    "detail_id": int(detail.id),
                    "accounting_period": int(detail.accounting_period),
                    "receipt_merchant": next_receipt_merchant,
                    "invoice_time": row.get("invoice_time"),
                    "invoice_number": row.get("invoice_number"),
                }
            )

        update_columns = (
            column("detail_id", ReconciliationChecklistDetail.id.type),
            column("accounting_period", ReconciliationChecklistDetail.accounting_period.type),
            column("receipt_merchant", ReconciliationChecklistDetail.receipt_merchant.type),
            column("invoice_time", ReconciliationChecklistDetail.invoice_time.type),
            column("invoice_number", ReconciliationChecklistDetail.invoice_number.type),
        )
        for chunk in _chunk_records(matched_rows):
            update_rows = [
                (
                    int(row["detail_id"]),
                    int(row["accounting_period"]),
                    row.get("receipt_merchant"),
                    row.get("invoice_time"),
                    row.get("invoice_number"),
                )
                for row in chunk
            ]
            updates = values(*update_columns, name="reconciliation_invoice_updates").data(update_rows).alias("reconciliation_invoice_updates")
            result = await db.execute(
                update(ReconciliationChecklistDetail)
                .where(
                    ReconciliationChecklistDetail.org_id == task.org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.id == updates.c.detail_id,
                    ReconciliationChecklistDetail.accounting_period == updates.c.accounting_period,
                )
                .values(
                    task_id=task.id,
                    file_id=upload_file.id,
                    receipt_merchant=updates.c.receipt_merchant,
                    invoice_time=ReconciliationChecklistService._typed_detail_update_value(updates, "invoice_time"),
                    invoice_number=updates.c.invoice_number,
                    updated_at=func.now(),
                )
            )
            updated += int(getattr(result, "rowcount", 0) or 0)
        return updated, failed, errors, periods

    @staticmethod
    async def _apply_invoice_edit_rows(
        db: AsyncSession,
        *,
        org_id: int,
        rows: Iterable[dict[str, object]],
    ) -> tuple[int, int, list[str], set[int]]:
        row_list = ReconciliationChecklistService._dedupe_rows(rows)
        unique_id_map, sub_order_map = await ReconciliationChecklistService._resolve_target_details(
            db,
            org_id=org_id,
            rows=row_list,
            require_unique_id=True,
        )
        updated = 0
        failed = 0
        errors: list[str] = []
        periods: set[int] = set()
        matched_rows: list[dict[str, object]] = []
        for row in row_list:
            detail, error_message = ReconciliationChecklistService._resolve_target_detail(
                row=row,
                unique_id_map=unique_id_map,
                sub_order_map=sub_order_map,
                require_unique_id=True,
            )
            if detail is None:
                failed += 1
                errors.append(f"Row {row.get('source_row_number')}: {error_message}")
                continue
            next_receipt_merchant = safe_str(row.get("receipt_merchant"))
            next_invoice_time = row.get("invoice_time")
            next_invoice_number = safe_str(row.get("invoice_number"))
            if safe_str(detail.receipt_merchant) != next_receipt_merchant:
                periods.add(int(detail.accounting_period))
            if safe_str(detail.receipt_merchant) == next_receipt_merchant and detail.invoice_time == next_invoice_time and safe_str(detail.invoice_number) == next_invoice_number:
                continue
            matched_rows.append(
                {
                    "detail_id": int(detail.id),
                    "accounting_period": int(detail.accounting_period),
                    "receipt_merchant": next_receipt_merchant,
                    "invoice_time": next_invoice_time,
                    "invoice_number": next_invoice_number,
                }
            )

        update_columns = (
            column("detail_id", ReconciliationChecklistDetail.id.type),
            column("accounting_period", ReconciliationChecklistDetail.accounting_period.type),
            column("receipt_merchant", ReconciliationChecklistDetail.receipt_merchant.type),
            column("invoice_time", ReconciliationChecklistDetail.invoice_time.type),
            column("invoice_number", ReconciliationChecklistDetail.invoice_number.type),
        )
        for chunk in _chunk_records(matched_rows):
            update_rows = [
                (
                    int(row["detail_id"]),
                    int(row["accounting_period"]),
                    row.get("receipt_merchant"),
                    row.get("invoice_time"),
                    row.get("invoice_number"),
                )
                for row in chunk
            ]
            updates = values(*update_columns, name="reconciliation_invoice_manual_updates").data(update_rows).alias("reconciliation_invoice_manual_updates")
            await db.execute(
                update(ReconciliationChecklistDetail)
                .where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.id == updates.c.detail_id,
                    ReconciliationChecklistDetail.accounting_period == updates.c.accounting_period,
                )
                .values(
                    receipt_merchant=updates.c.receipt_merchant,
                    invoice_time=ReconciliationChecklistService._typed_detail_update_value(updates, "invoice_time"),
                    invoice_number=updates.c.invoice_number,
                    updated_at=func.now(),
                )
            )
            updated += len(update_rows)
        return updated, failed, errors, periods

    @staticmethod
    async def _apply_merchant_rows(
        db: AsyncSession,
        *,
        task: ReconciliationChecklistTask,
        upload_file: ReconciliationChecklistUploadFile,
        rows: Iterable[dict[str, object]],
    ) -> tuple[int, int, list[str], set[int]]:
        row_list = ReconciliationChecklistService._dedupe_rows(rows)
        unique_id_map, sub_order_map = await ReconciliationChecklistService._resolve_target_details(
            db,
            org_id=task.org_id,
            rows=row_list,
            require_unique_id=True,
        )
        updated = 0
        failed = 0
        errors: list[str] = []
        periods: set[int] = set()
        matched_rows: list[dict[str, object]] = []
        for row in row_list:
            detail, error_message = ReconciliationChecklistService._resolve_target_detail(
                row=row,
                unique_id_map=unique_id_map,
                sub_order_map=sub_order_map,
                require_unique_id=True,
            )
            if detail is None:
                failed += 1
                errors.append(f"Row {row.get('source_row_number')}: {error_message}")
                continue
            next_receipt_merchant = safe_str(row.get("receipt_merchant"))
            next_merchant_net_amount = row.get("merchant_net_amount") or Decimal("0.00")
            next_payment_amount = row.get("payment_amount")
            next_merchant_payment_time = row.get("merchant_payment_time")
            summary_fields_changed = (
                safe_str(detail.receipt_merchant) != next_receipt_merchant
                or (detail.merchant_net_amount if detail.merchant_net_amount is not None else Decimal("0.00")) != next_merchant_net_amount
                or detail.payment_amount != next_payment_amount
            )
            if summary_fields_changed:
                periods.add(int(detail.accounting_period))
            if not summary_fields_changed and detail.merchant_payment_time == next_merchant_payment_time:
                continue
            matched_rows.append(
                {
                    "detail_id": int(detail.id),
                    "accounting_period": int(detail.accounting_period),
                    "receipt_merchant": next_receipt_merchant,
                    "merchant_net_amount": next_merchant_net_amount,
                    "payment_amount": next_payment_amount,
                    "merchant_net_balance": _merchant_net_balance_value(
                        next_merchant_net_amount,
                        next_payment_amount,
                    ),
                    "merchant_payment_time": next_merchant_payment_time,
                }
            )

        update_columns = (
            column("detail_id", ReconciliationChecklistDetail.id.type),
            column("accounting_period", ReconciliationChecklistDetail.accounting_period.type),
            column("receipt_merchant", ReconciliationChecklistDetail.receipt_merchant.type),
            column("merchant_net_amount", ReconciliationChecklistDetail.merchant_net_amount.type),
            column("payment_amount", ReconciliationChecklistDetail.payment_amount.type),
            column("merchant_net_balance", ReconciliationChecklistDetail.merchant_net_balance.type),
            column("merchant_payment_time", ReconciliationChecklistDetail.merchant_payment_time.type),
        )
        for chunk in _chunk_records(matched_rows):
            update_rows = [
                (
                    int(row["detail_id"]),
                    int(row["accounting_period"]),
                    row.get("receipt_merchant"),
                    row.get("merchant_net_amount") or Decimal("0.00"),
                    row.get("payment_amount"),
                    row.get("merchant_net_balance"),
                    row.get("merchant_payment_time"),
                )
                for row in chunk
            ]
            updates = values(*update_columns, name="reconciliation_merchant_updates").data(update_rows).alias("reconciliation_merchant_updates")
            result = await db.execute(
                update(ReconciliationChecklistDetail)
                .where(
                    ReconciliationChecklistDetail.org_id == task.org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.id == updates.c.detail_id,
                    ReconciliationChecklistDetail.accounting_period == updates.c.accounting_period,
                )
                .values(
                    task_id=task.id,
                    file_id=upload_file.id,
                    receipt_merchant=updates.c.receipt_merchant,
                    merchant_net_amount=updates.c.merchant_net_amount,
                    payment_amount=updates.c.payment_amount,
                    merchant_net_balance=updates.c.merchant_net_balance,
                    merchant_payment_time=ReconciliationChecklistService._typed_detail_update_value(updates, "merchant_payment_time"),
                    updated_at=func.now(),
                )
            )
            updated += int(getattr(result, "rowcount", 0) or 0)
        return updated, failed, errors, periods

    @staticmethod
    async def _apply_merchant_edit_rows(
        db: AsyncSession,
        *,
        org_id: int,
        rows: Iterable[dict[str, object]],
    ) -> tuple[int, int, list[str], set[int]]:
        row_list = ReconciliationChecklistService._dedupe_rows(rows)
        unique_id_map, sub_order_map = await ReconciliationChecklistService._resolve_target_details(
            db,
            org_id=org_id,
            rows=row_list,
            require_unique_id=True,
        )
        updated = 0
        failed = 0
        errors: list[str] = []
        periods: set[int] = set()
        matched_rows: list[dict[str, object]] = []
        for row in row_list:
            detail, error_message = ReconciliationChecklistService._resolve_target_detail(
                row=row,
                unique_id_map=unique_id_map,
                sub_order_map=sub_order_map,
                require_unique_id=True,
            )
            if detail is None:
                failed += 1
                errors.append(f"Row {row.get('source_row_number')}: {error_message}")
                continue
            next_receipt_merchant = safe_str(row.get("receipt_merchant"))
            next_merchant_net_amount = row.get("merchant_net_amount") or Decimal("0.00")
            next_payment_amount = row.get("payment_amount")
            if (
                safe_str(detail.receipt_merchant) != next_receipt_merchant
                or (detail.merchant_net_amount if detail.merchant_net_amount is not None else Decimal("0.00")) != next_merchant_net_amount
                or detail.payment_amount != next_payment_amount
            ):
                periods.add(int(detail.accounting_period))
            matched_rows.append(
                {
                    "detail_id": int(detail.id),
                    "accounting_period": int(detail.accounting_period),
                    "receipt_merchant": next_receipt_merchant,
                    "merchant_net_amount": next_merchant_net_amount,
                    "payment_amount": next_payment_amount,
                    "merchant_net_balance": _merchant_net_balance_value(
                        next_merchant_net_amount,
                        next_payment_amount,
                    ),
                    "merchant_payment_time": row.get("merchant_payment_time"),
                }
            )

        update_columns = (
            column("detail_id", ReconciliationChecklistDetail.id.type),
            column("accounting_period", ReconciliationChecklistDetail.accounting_period.type),
            column("receipt_merchant", ReconciliationChecklistDetail.receipt_merchant.type),
            column("merchant_net_amount", ReconciliationChecklistDetail.merchant_net_amount.type),
            column("payment_amount", ReconciliationChecklistDetail.payment_amount.type),
            column("merchant_net_balance", ReconciliationChecklistDetail.merchant_net_balance.type),
            column("merchant_payment_time", ReconciliationChecklistDetail.merchant_payment_time.type),
        )
        for chunk in _chunk_records(matched_rows):
            update_rows = [
                (
                    int(row["detail_id"]),
                    int(row["accounting_period"]),
                    row.get("receipt_merchant"),
                    row.get("merchant_net_amount") or Decimal("0.00"),
                    row.get("payment_amount"),
                    row.get("merchant_net_balance"),
                    row.get("merchant_payment_time"),
                )
                for row in chunk
            ]
            updates = values(*update_columns, name="reconciliation_merchant_manual_updates").data(update_rows).alias("reconciliation_merchant_manual_updates")
            await db.execute(
                update(ReconciliationChecklistDetail)
                .where(
                    ReconciliationChecklistDetail.org_id == org_id,
                    ReconciliationChecklistDetail.is_deleted.is_(False),
                    ReconciliationChecklistDetail.id == updates.c.detail_id,
                    ReconciliationChecklistDetail.accounting_period == updates.c.accounting_period,
                )
                .values(
                    receipt_merchant=updates.c.receipt_merchant,
                    merchant_net_amount=updates.c.merchant_net_amount,
                    payment_amount=updates.c.payment_amount,
                    merchant_net_balance=updates.c.merchant_net_balance,
                    merchant_payment_time=ReconciliationChecklistService._typed_detail_update_value(updates, "merchant_payment_time"),
                    updated_at=func.now(),
                )
            )
            updated += len(update_rows)
        return updated, failed, errors, periods

    @staticmethod
    async def _rebuild_summaries(db: AsyncSession, *, org_id: int, periods: Iterable[int]) -> None:
        period_list = sorted({int(period) for period in periods if int(period) > 0})
        if not period_list:
            return
        await db.execute(
            delete(ReconciliationChecklistProductSummaryRow).where(
                ReconciliationChecklistProductSummaryRow.org_id == org_id, ReconciliationChecklistProductSummaryRow.accounting_period.in_(period_list)
            )
        )
        await db.execute(
            delete(ReconciliationChecklistReceiptSummaryRow).where(
                ReconciliationChecklistReceiptSummaryRow.org_id == org_id, ReconciliationChecklistReceiptSummaryRow.accounting_period.in_(period_list)
            )
        )
        await db.execute(
            delete(ReconciliationChecklistPayableBalanceSummaryRow).where(
                ReconciliationChecklistPayableBalanceSummaryRow.org_id == org_id, ReconciliationChecklistPayableBalanceSummaryRow.accounting_period.in_(period_list)
            )
        )

        product_columns = (
            "org_id",
            "accounting_year",
            "accounting_month",
            "accounting_period",
            "receipt_merchant",
            "merchant_subject_name",
            "product_name",
            "product_quantity",
            "total_user_paid_amount",
            "total_live_commission",
            "total_merchant_net_amount",
        )
        product_select = (
            select(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.receipt_merchant,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.product_name,
                func.sum(ReconciliationChecklistDetail.product_quantity),
                func.coalesce(func.sum(ReconciliationChecklistDetail.user_paid_amount), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.live_commission), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.merchant_net_amount), Decimal("0.00")),
            )
            .where(
                ReconciliationChecklistDetail.org_id == org_id,
                ReconciliationChecklistDetail.accounting_period.in_(period_list),
                ReconciliationChecklistDetail.is_deleted.is_(False),
            )
            .group_by(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.receipt_merchant,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.product_name,
            )
        )
        await db.execute(insert(ReconciliationChecklistProductSummaryRow).from_select(product_columns, product_select))

        receipt_columns = (
            "org_id",
            "accounting_year",
            "accounting_month",
            "accounting_period",
            "merchant_subject_name",
            "live_platform",
            "receipt_merchant",
            "order_count",
            "total_user_paid_amount",
            "total_live_commission",
            "total_merchant_net_amount",
        )
        receipt_select = (
            select(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.live_platform,
                ReconciliationChecklistDetail.receipt_merchant,
                func.count(ReconciliationChecklistDetail.id),
                func.coalesce(func.sum(ReconciliationChecklistDetail.user_paid_amount), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.live_commission), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.merchant_net_amount), Decimal("0.00")),
            )
            .where(
                ReconciliationChecklistDetail.org_id == org_id,
                ReconciliationChecklistDetail.accounting_period.in_(period_list),
                ReconciliationChecklistDetail.is_deleted.is_(False),
            )
            .group_by(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.live_platform,
                ReconciliationChecklistDetail.receipt_merchant,
            )
        )
        await db.execute(insert(ReconciliationChecklistReceiptSummaryRow).from_select(receipt_columns, receipt_select))

        payable_balance_columns = (
            "org_id",
            "accounting_year",
            "accounting_month",
            "accounting_period",
            "merchant_subject_name",
            "receipt_merchant",
            "total_user_paid_amount",
            "total_merchant_net_amount",
            "total_payment_amount",
            "total_merchant_net_balance",
        )
        payable_balance_select = (
            select(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.receipt_merchant,
                func.coalesce(func.sum(ReconciliationChecklistDetail.user_paid_amount), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.merchant_net_amount), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.payment_amount), Decimal("0.00")),
                func.coalesce(func.sum(ReconciliationChecklistDetail.merchant_net_balance), Decimal("0.00")),
            )
            .where(
                ReconciliationChecklistDetail.org_id == org_id,
                ReconciliationChecklistDetail.accounting_period.in_(period_list),
                ReconciliationChecklistDetail.is_deleted.is_(False),
            )
            .group_by(
                ReconciliationChecklistDetail.org_id,
                ReconciliationChecklistDetail.accounting_year,
                ReconciliationChecklistDetail.accounting_month,
                ReconciliationChecklistDetail.accounting_period,
                ReconciliationChecklistDetail.merchant_subject_name,
                ReconciliationChecklistDetail.receipt_merchant,
            )
        )
        await db.execute(insert(ReconciliationChecklistPayableBalanceSummaryRow).from_select(payable_balance_columns, payable_balance_select))

    @staticmethod
    def _common_summary_filters(
        source_model,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        merchant_subject_name: str | None = None,
        receipt_merchant: str | None = None,
        live_platform: str | None = None,
        product_name: str | None = None,
        keyword: str | None = None,
        ids: Sequence[str] | None = None,
    ) -> list:
        filters = []
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(source_model.org_id.in_(org_ids))
        filters.extend(
            _period_filters(
                source_model.accounting_period,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                accounting_start_year=accounting_start_year,
                accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year,
                accounting_end_month=accounting_end_month,
            )
        )
        for column_name, value in (
            ("merchant_subject_name", merchant_subject_name),
            ("receipt_merchant", receipt_merchant),
            ("live_platform", live_platform),
            ("product_name", product_name),
        ):
            if hasattr(source_model, column_name):
                values = _split_filter_values(value)
                if values:
                    filters.append(getattr(source_model, column_name).in_(values))
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            columns = [getattr(source_model, name, None) for name in ("merchant_subject_name", "receipt_merchant", "live_platform", "product_name")]
            filters.append(or_(*(column.ilike(like_pattern) for column in columns if column is not None)))
        if ids:
            if hasattr(source_model, "product_name"):
                key_field_names = ("org_id", "accounting_period", "receipt_merchant", "merchant_subject_name", "product_name")
            elif hasattr(source_model, "live_platform"):
                key_field_names = ("org_id", "accounting_period", "merchant_subject_name", "live_platform", "receipt_merchant")
            else:
                key_field_names = ("org_id", "accounting_period", "merchant_subject_name", "receipt_merchant")

            parsed_keys: list[tuple[object, ...]] = []
            for raw_id in ids:
                parts = str(raw_id).split(":")
                if len(parts) != len(key_field_names):
                    continue
                current_key: list[object] = []
                valid = True
                for field_name, part in zip(key_field_names, parts, strict=False):
                    if field_name in {"org_id", "accounting_period"}:
                        try:
                            current_key.append(int(part))
                        except ValueError:
                            valid = False
                            break
                    else:
                        current_key.append(part)
                if valid:
                    parsed_keys.append(tuple(current_key))

            if not parsed_keys:
                filters.append(text("1 = 0"))
            else:
                key_columns = tuple(getattr(source_model, field_name) for field_name in key_field_names)
                filters.append(tuple_(*key_columns).in_(parsed_keys))
        return filters

    @staticmethod
    async def list_product_summary(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        merchant_subject_name: str | None = None,
        receipt_merchant: str | None = None,
        product_name: str | None = None,
        keyword: str | None = None,
        ids: Sequence[str] | None = None,
        page: int | None = 1,
        page_size: int | None = 20,
        **_ignored,
    ) -> tuple[list[dict[str, object]], int]:
        model = ReconciliationChecklistProductSummaryRow
        filters = ReconciliationChecklistService._common_summary_filters(
            model,
            user=user,
            org_id=org_id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            accounting_start_year=accounting_start_year,
            accounting_start_month=accounting_start_month,
            accounting_end_year=accounting_end_year,
            accounting_end_month=accounting_end_month,
            merchant_subject_name=merchant_subject_name,
            receipt_merchant=receipt_merchant,
            product_name=product_name,
            keyword=keyword,
            ids=ids,
        )
        stmt = (
            select(
                model.org_id,
                Organization.name.label("org_name"),
                model.accounting_year,
                model.accounting_month,
                model.accounting_period,
                model.receipt_merchant,
                model.merchant_subject_name,
                model.product_name,
                model.product_quantity,
                model.total_user_paid_amount,
                model.total_live_commission,
                model.total_merchant_net_amount,
            )
            .outerjoin(Organization, Organization.id == model.org_id)
            .where(*filters)
            .order_by(model.accounting_year.desc(), model.accounting_month.desc(), model.receipt_merchant, model.merchant_subject_name, model.product_name)
        )
        total = 0
        if page is not None and page_size is not None:
            total = (await db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery()))) or 0
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        rows = []
        for row in (await db.execute(stmt)).mappings().all():
            item = dict(row)
            item["key"] = _product_summary_key(item)
            rows.append(item)
        return rows, total

    @staticmethod
    async def list_receipt_summary(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        merchant_subject_name: str | None = None,
        receipt_merchant: str | None = None,
        live_platform: str | None = None,
        keyword: str | None = None,
        ids: Sequence[str] | None = None,
        page: int | None = 1,
        page_size: int | None = 20,
        **_ignored,
    ) -> tuple[list[dict[str, object]], int]:
        model = ReconciliationChecklistReceiptSummaryRow
        filters = ReconciliationChecklistService._common_summary_filters(
            model,
            user=user,
            org_id=org_id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            accounting_start_year=accounting_start_year,
            accounting_start_month=accounting_start_month,
            accounting_end_year=accounting_end_year,
            accounting_end_month=accounting_end_month,
            merchant_subject_name=merchant_subject_name,
            receipt_merchant=receipt_merchant,
            live_platform=live_platform,
            keyword=keyword,
            ids=ids,
        )
        stmt = (
            select(
                model.org_id,
                Organization.name.label("org_name"),
                model.accounting_year,
                model.accounting_month,
                model.accounting_period,
                model.merchant_subject_name,
                model.live_platform,
                model.receipt_merchant,
                model.order_count,
                model.total_user_paid_amount,
                model.total_live_commission,
                model.total_merchant_net_amount,
            )
            .outerjoin(Organization, Organization.id == model.org_id)
            .where(*filters)
            .order_by(model.accounting_year.desc(), model.accounting_month.desc(), model.merchant_subject_name, model.live_platform, model.receipt_merchant)
        )
        total = 0
        if page is not None and page_size is not None:
            total = (await db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery()))) or 0
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        rows = []
        for row in (await db.execute(stmt)).mappings().all():
            item = dict(row)
            item["key"] = _receipt_summary_key(item)
            rows.append(item)
        return rows, total

    @staticmethod
    async def list_payable_balance_summary(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        merchant_subject_name: str | None = None,
        receipt_merchant: str | None = None,
        keyword: str | None = None,
        ids: Sequence[str] | None = None,
        page: int | None = 1,
        page_size: int | None = 20,
        **_ignored,
    ) -> tuple[list[dict[str, object]], int]:
        model = ReconciliationChecklistPayableBalanceSummaryRow
        filters = ReconciliationChecklistService._common_summary_filters(
            model,
            user=user,
            org_id=org_id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            accounting_start_year=accounting_start_year,
            accounting_start_month=accounting_start_month,
            accounting_end_year=accounting_end_year,
            accounting_end_month=accounting_end_month,
            merchant_subject_name=merchant_subject_name,
            receipt_merchant=receipt_merchant,
            keyword=keyword,
            ids=ids,
        )
        stmt = (
            select(
                model.org_id,
                Organization.name.label("org_name"),
                model.accounting_year,
                model.accounting_month,
                model.accounting_period,
                model.merchant_subject_name,
                model.receipt_merchant,
                model.total_user_paid_amount,
                model.total_merchant_net_amount,
                model.total_payment_amount,
                model.total_merchant_net_balance,
            )
            .outerjoin(Organization, Organization.id == model.org_id)
            .where(*filters)
            .order_by(model.accounting_year.desc(), model.accounting_month.desc(), model.merchant_subject_name, model.receipt_merchant)
        )
        total = 0
        if page is not None and page_size is not None:
            total = (await db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery()))) or 0
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        rows = []
        for row in (await db.execute(stmt)).mappings().all():
            item = dict(row)
            item["key"] = _payable_balance_summary_key(item)
            rows.append(item)
        return rows, total

    @staticmethod
    async def list_options(
        db: AsyncSession,
        *,
        user: User,
        kind: str,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        limit: int = 50,
    ) -> list[dict[str, object]]:
        if kind not in OPTION_KINDS:
            raise ValueError("不支持的筛选类型")
        column = {
            "merchant_subject": ReconciliationChecklistDetail.merchant_subject_name,
            "receipt_merchant": ReconciliationChecklistDetail.receipt_merchant,
            "live_platform": ReconciliationChecklistDetail.live_platform,
            "product_name": ReconciliationChecklistDetail.product_name,
        }[kind]
        filters = [ReconciliationChecklistDetail.is_deleted.is_(False)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(ReconciliationChecklistDetail.org_id.in_(org_ids))
        filters.extend(_period_filters(ReconciliationChecklistDetail.accounting_period, accounting_year=accounting_year, accounting_month=accounting_month))
        if keyword:
            filters.append(column.ilike(f"%{keyword.strip()}%"))
        stmt = select(column.label("value")).where(*filters, column != "").group_by(column).order_by(column).limit(min(max(limit, 1), 100))
        return [{"label": row.value, "value": row.value} for row in (await db.execute(stmt)).all()]

    @staticmethod
    async def list_tasks(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        status: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[tuple[ReconciliationChecklistTask, ReconciliationChecklistUploadFile, str | None]], int]:
        filters = [ReconciliationChecklistTask.is_deleted.is_(False), ReconciliationChecklistUploadFile.is_deleted.is_(False)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(ReconciliationChecklistTask.org_id.in_(org_ids))
        status_values = _split_filter_values(status)
        if status_values:
            filters.append(ReconciliationChecklistTask.status.in_(status_values))
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(or_(ReconciliationChecklistUploadFile.original_name.ilike(like_pattern), ReconciliationChecklistTask.error_message.ilike(like_pattern)))
        stmt = (
            select(ReconciliationChecklistTask, ReconciliationChecklistUploadFile, Organization.name.label("org_name"))
            .join(ReconciliationChecklistUploadFile, ReconciliationChecklistUploadFile.id == ReconciliationChecklistTask.file_id)
            .outerjoin(Organization, Organization.id == ReconciliationChecklistTask.org_id)
            .where(*filters)
            .order_by(ReconciliationChecklistTask.updated_at.desc(), ReconciliationChecklistTask.id.desc())
        )
        total = (await db.scalar(select(func.count()).select_from(stmt.order_by(None).subquery()))) or 0
        result = await db.execute(stmt.offset((page - 1) * page_size).limit(page_size))
        return list(result.all()), total

    @staticmethod
    async def get_dashboard_metrics(db: AsyncSession, *, user: User, year: int, org_id: str | int | None = None) -> dict[str, object]:
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        task_base_filters = [ReconciliationChecklistTask.is_deleted.is_(False)]
        task_filters = [*task_base_filters, ReconciliationChecklistTask.status.in_(CHECKLIST_PROCESSED_TASK_STATUSES)]
        failed_task_filters = [*task_base_filters, ReconciliationChecklistTask.status == "failed"]
        recent_task_filters = [*task_base_filters, ReconciliationChecklistTask.status.in_(CHECKLIST_RESULT_TASK_STATUSES)]
        summary_model = ReconciliationChecklistReceiptSummaryRow
        summary_filters = [
            summary_model.accounting_period >= year * 100 + 1,
            summary_model.accounting_period <= year * 100 + 12,
        ]
        if org_ids is not None:
            task_base_filters.append(ReconciliationChecklistTask.org_id.in_(org_ids))
            task_filters.append(ReconciliationChecklistTask.org_id.in_(org_ids))
            failed_task_filters.append(ReconciliationChecklistTask.org_id.in_(org_ids))
            recent_task_filters.append(ReconciliationChecklistTask.org_id.in_(org_ids))
            summary_filters.append(summary_model.org_id.in_(org_ids))

        processed_task_count = await db.scalar(select(func.count()).select_from(ReconciliationChecklistTask).where(*task_filters))
        total_task_count = await db.scalar(select(func.count()).select_from(ReconciliationChecklistTask).where(*task_base_filters))
        failed_task_count = await db.scalar(select(func.count()).select_from(ReconciliationChecklistTask).where(*failed_task_filters))
        total_rows = await db.scalar(select(func.coalesce(func.sum(ReconciliationChecklistTask.total_rows), 0)).where(*task_filters))
        total_user_paid_amount = await db.scalar(select(func.coalesce(func.sum(summary_model.total_user_paid_amount), Decimal("0.00"))).where(*summary_filters))
        merchant_count = await db.scalar(select(func.count(func.distinct(summary_model.merchant_subject_name))).where(*summary_filters, summary_model.merchant_subject_name != ""))
        covered_month_count = await db.scalar(select(func.count(func.distinct(summary_model.accounting_period))).where(*summary_filters))

        month_start = datetime(year, 1, 1, tzinfo=timezone.utc)
        next_year_start = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        task_month_expr = func.date_part("month", ReconciliationChecklistTask.finished_at).label("month")
        monthly_task_rows = await db.execute(
            select(
                task_month_expr,
                func.count().label("task_count"),
            )
            .where(
                *task_filters,
                ReconciliationChecklistTask.finished_at >= month_start,
                ReconciliationChecklistTask.finished_at < next_year_start,
            )
            .group_by(task_month_expr)
            .order_by(task_month_expr)
        )
        monthly_task_counts = [{"month": month, "task_count": 0} for month in range(1, 13)]
        for row in monthly_task_rows.all():
            month = int(row.month or 0)
            if 1 <= month <= 12:
                monthly_task_counts[month - 1]["task_count"] = int(row.task_count or 0)

        monthly_paid_rows = await db.execute(
            select(
                summary_model.accounting_month.label("month"),
                func.coalesce(func.sum(summary_model.total_user_paid_amount), Decimal("0.00")).label("total_user_paid_amount"),
            )
            .where(*summary_filters)
            .group_by(summary_model.accounting_month)
            .order_by(summary_model.accounting_month)
        )
        monthly_user_paid_amounts = [{"month": month, "total_user_paid_amount": Decimal("0.00")} for month in range(1, 13)]
        for row in monthly_paid_rows.all():
            month = int(row.month or 0)
            if 1 <= month <= 12:
                monthly_user_paid_amounts[month - 1]["total_user_paid_amount"] = row.total_user_paid_amount or Decimal("0.00")

        merchant_paid_expr = func.coalesce(func.sum(summary_model.total_user_paid_amount), Decimal("0.00")).label("total_user_paid_amount")
        top_merchant_rows = await db.execute(
            select(
                summary_model.merchant_subject_name.label("merchant_name"),
                merchant_paid_expr,
            )
            .where(*summary_filters, summary_model.merchant_subject_name != "")
            .group_by(summary_model.merchant_subject_name)
            .order_by(merchant_paid_expr.desc(), summary_model.merchant_subject_name)
            .limit(5)
        )
        top_merchants = [
            {
                "merchant_name": row.merchant_name,
                "total_user_paid_amount": row.total_user_paid_amount or Decimal("0.00"),
            }
            for row in top_merchant_rows.all()
        ]

        recent_task_rows = await db.execute(
            select(
                ReconciliationChecklistTask.id,
                ReconciliationChecklistUploadFile.original_name,
                ReconciliationChecklistTask.task_type,
                ReconciliationChecklistTask.status,
                ReconciliationChecklistTask.total_rows,
                ReconciliationChecklistTask.success_rows,
                ReconciliationChecklistTask.failed_rows,
                ReconciliationChecklistTask.inserted_rows,
                ReconciliationChecklistTask.finished_at,
            )
            .join(ReconciliationChecklistUploadFile, ReconciliationChecklistUploadFile.id == ReconciliationChecklistTask.file_id)
            .where(*recent_task_filters, ReconciliationChecklistUploadFile.is_deleted.is_(False))
            .order_by(ReconciliationChecklistTask.finished_at.desc().nullslast(), ReconciliationChecklistTask.id.desc())
            .limit(5)
        )
        recent_tasks = [
            {
                "id": row.id,
                "original_name": row.original_name,
                "task_type": row.task_type,
                "status": row.status,
                "total_rows": row.total_rows,
                "success_rows": row.success_rows,
                "failed_rows": row.failed_rows,
                "inserted_rows": row.inserted_rows,
                "finished_at": row.finished_at,
            }
            for row in recent_task_rows.all()
        ]

        total_task_count_int = int(total_task_count or 0)
        processed_task_count_int = int(processed_task_count or 0)
        completion_rate = Decimal("0.00")
        if total_task_count_int:
            completion_rate = (Decimal(processed_task_count_int) * Decimal("100") / Decimal(total_task_count_int)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return {
            "processed_task_count": processed_task_count_int,
            "total_task_count": total_task_count_int,
            "failed_task_count": int(failed_task_count or 0),
            "total_rows": int(total_rows or 0),
            "total_user_paid_amount": total_user_paid_amount or Decimal("0.00"),
            "merchant_count": int(merchant_count or 0),
            "covered_month_count": int(covered_month_count or 0),
            "completion_rate": completion_rate,
            "year": year,
            "monthly_task_counts": monthly_task_counts,
            "monthly_user_paid_amounts": monthly_user_paid_amounts,
            "top_merchants": top_merchants,
            "recent_tasks": recent_tasks,
        }

    @staticmethod
    async def export_product_summary_to_file(
        db: AsyncSession,
        *,
        user: User,
        output_path: Path,
        **params,
    ) -> int:
        query_params = {"page": None, "page_size": None, **params}
        rows, _ = await ReconciliationChecklistService.list_product_summary(db, user=user, **query_params)
        if not rows:
            raise ValueError("未找到符合条件的对账清单数据")
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        grouped: dict[tuple[int, int, str, str], list[dict[str, object]]] = {}
        for row in rows:
            grouped.setdefault((int(row["accounting_year"]), int(row["accounting_month"]), safe_str(row["receipt_merchant"]), safe_str(row["merchant_subject_name"])), []).append(
                row
            )
        row_count = 0
        for (year, month, receipt_merchant, merchant_subject), group_rows in grouped.items():
            sheet = workbook.create_sheet(_sheet_title(year, month, receipt_merchant, merchant_subject))
            for column_name, width in {"A": 14, "B": 48, "C": 22, "D": 22, "E": 22}.items():
                sheet.column_dimensions[column_name].width = width

            sheet.merge_cells("A1:E1")
            _set_sheet_cell(
                sheet,
                1,
                1,
                CHECKLIST_PRODUCT_SUMMARY_TITLE,
                bold=True,
                font_size=16,
                border=CHECKLIST_EXPORT_NO_BORDER,
            )
            for column in range(2, 6):
                _set_sheet_cell(sheet, 1, column, None, border=CHECKLIST_EXPORT_NO_BORDER)

            for column in range(1, 6):
                _set_sheet_cell(sheet, 2, column, None, bold=True, border=CHECKLIST_EXPORT_NO_BORDER)
            _set_sheet_cell(sheet, 2, 4, "结算时间", bold=True, border=CHECKLIST_EXPORT_NO_BORDER)
            _set_sheet_cell(sheet, 2, 5, f"{year}年{month}月", bold=True, border=CHECKLIST_EXPORT_NO_BORDER)

            _set_sheet_cell(sheet, 3, 1, "收款商家", bold=True)
            _set_sheet_cell(sheet, 3, 2, receipt_merchant, bold=True)
            _set_sheet_cell(sheet, 3, 3, "商户主体名称", bold=True)
            sheet.merge_cells("D3:E3")
            _set_sheet_cell(sheet, 3, 4, merchant_subject, bold=True)
            _set_sheet_cell(sheet, 3, 5, None, bold=True)

            for column in range(1, 6):
                _set_sheet_cell(sheet, 4, column, None, border=CHECKLIST_EXPORT_NO_BORDER)

            headers = ["序号", "商品名称", "用户实付（订单金额）", "直播推广佣金", "应付商家净额"]
            for column, header in enumerate(headers, start=1):
                _set_sheet_cell(sheet, 5, column, header, bold=True)

            total_user_paid = Decimal("0.00")
            total_commission = Decimal("0.00")
            total_net = Decimal("0.00")
            current_row = 6
            display_index = 1
            for row in group_rows:
                user_paid = _money(row.get("total_user_paid_amount"))
                commission = _money(row.get("total_live_commission"))
                net_amount = _money(row.get("total_merchant_net_amount"))
                if _is_zero_money_export_row(user_paid, commission, net_amount):
                    continue
                total_user_paid += user_paid
                total_commission += commission
                total_net += net_amount
                _set_sheet_cell(sheet, current_row, 1, display_index)
                _set_sheet_cell(sheet, current_row, 2, row.get("product_name") or "", alignment=CHECKLIST_EXPORT_LEFT_ALIGNMENT)
                _set_sheet_cell(
                    sheet,
                    current_row,
                    3,
                    float(user_paid),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                _set_sheet_cell(
                    sheet,
                    current_row,
                    4,
                    float(commission),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                _set_sheet_cell(
                    sheet,
                    current_row,
                    5,
                    float(net_amount),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                current_row += 1
                display_index += 1
                row_count += 1
            _set_sheet_cell(sheet, current_row, 1, "总计", bold=True)
            _set_sheet_cell(sheet, current_row, 2, "", bold=True)
            _set_sheet_cell(
                sheet,
                current_row,
                3,
                float(total_user_paid),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
            _set_sheet_cell(
                sheet,
                current_row,
                4,
                float(total_commission),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
            _set_sheet_cell(
                sheet,
                current_row,
                5,
                float(total_net),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
        workbook.save(output_path)
        workbook.close()
        return row_count

    @staticmethod
    async def export_receipt_summary_to_file(
        db: AsyncSession,
        *,
        user: User,
        output_path: Path,
        **params,
    ) -> int:
        query_params = {"page": None, "page_size": None, **params}
        rows, _ = await ReconciliationChecklistService.list_receipt_summary(db, user=user, **query_params)
        if not rows:
            raise ValueError("未找到符合条件的对账清单数据")
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        grouped: dict[tuple[int, int, str, str], list[dict[str, object]]] = {}
        for row in rows:
            grouped.setdefault((int(row["accounting_year"]), int(row["accounting_month"]), safe_str(row["merchant_subject_name"]), safe_str(row["live_platform"])), []).append(row)
        row_count = 0
        for (year, month, merchant_subject, live_platform), group_rows in grouped.items():
            sheet = workbook.create_sheet(_sheet_title(year, month, merchant_subject, live_platform))
            for column_name, width in {"A": 14, "B": 48, "C": 22, "D": 22, "E": 22}.items():
                sheet.column_dimensions[column_name].width = width

            summary_rows = [
                ("商户主体名称", merchant_subject),
                ("进驻的直播平台", live_platform),
                ("结算时间", f"{year}年{month}月"),
            ]
            for row_index, (label, value) in enumerate(summary_rows, start=1):
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
                sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=5)
                _set_sheet_cell(sheet, row_index, 1, label, bold=True)
                _set_sheet_cell(sheet, row_index, 2, None, bold=True)
                _set_sheet_cell(sheet, row_index, 3, value, bold=True)
                _set_sheet_cell(sheet, row_index, 4, None, bold=True)
                _set_sheet_cell(sheet, row_index, 5, None, bold=True)

            for column in range(1, 6):
                _set_sheet_cell(sheet, 4, column, None, border=CHECKLIST_EXPORT_NO_BORDER)

            headers = ["序号", "收款商家", "用户实付 （订单金额）", "直播推广佣金", "应付商家净额"]
            for column, header in enumerate(headers, start=1):
                _set_sheet_cell(sheet, 5, column, header, bold=True)

            total_user_paid = Decimal("0.00")
            total_commission = Decimal("0.00")
            total_net = Decimal("0.00")
            current_row = 6
            display_index = 1
            for row in group_rows:
                user_paid = _money(row.get("total_user_paid_amount"))
                commission = _money(row.get("total_live_commission"))
                net_amount = _money(row.get("total_merchant_net_amount"))
                if _is_zero_money_export_row(user_paid, commission, net_amount):
                    continue
                total_user_paid += user_paid
                total_commission += commission
                total_net += net_amount
                _set_sheet_cell(sheet, current_row, 1, display_index)
                _set_sheet_cell(sheet, current_row, 2, row.get("receipt_merchant") or "", alignment=CHECKLIST_EXPORT_LEFT_ALIGNMENT)
                _set_sheet_cell(
                    sheet,
                    current_row,
                    3,
                    float(user_paid),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                _set_sheet_cell(
                    sheet,
                    current_row,
                    4,
                    float(commission),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                _set_sheet_cell(
                    sheet,
                    current_row,
                    5,
                    float(net_amount),
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                )
                current_row += 1
                display_index += 1
                row_count += 1
            _set_sheet_cell(sheet, current_row, 1, "总计", bold=True)
            _set_sheet_cell(sheet, current_row, 2, "", bold=True)
            _set_sheet_cell(
                sheet,
                current_row,
                3,
                float(total_user_paid),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
            _set_sheet_cell(
                sheet,
                current_row,
                4,
                float(total_commission),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
            _set_sheet_cell(
                sheet,
                current_row,
                5,
                float(total_net),
                bold=True,
                alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
            )
        workbook.save(output_path)
        workbook.close()
        return row_count

    @staticmethod
    async def export_payable_balance_summary_to_file(
        db: AsyncSession,
        *,
        user: User,
        output_path: Path,
        **params,
    ) -> int:
        query_params = {"page": None, "page_size": None, **params}
        rows, _ = await ReconciliationChecklistService.list_payable_balance_summary(db, user=user, **query_params)
        if not rows:
            raise ValueError("未找到符合条件的对账清单数据")
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("商家应付余额明细")
        for column_name, width in {"A": 34, "B": 24, "C": 40, "D": 24, "E": 24, "F": 24, "G": 26}.items():
            sheet.column_dimensions[column_name].width = width
        headers = ["商户主体名称", "结算时间（年月）", "收款商家", "用户实付 （订单金额）", "应付商家净额", "付款金额", "应付商家净额余额"]
        sheet.append(_export_row(sheet, headers, bold=True))
        row_count = 0
        total_user_paid = Decimal("0.00")
        total_net = Decimal("0.00")
        total_payment = Decimal("0.00")
        total_balance = Decimal("0.00")
        for row in rows:
            user_paid = safe_decimal(row.get("total_user_paid_amount"))
            net_amount = safe_decimal(row.get("total_merchant_net_amount"))
            payment_amount = safe_decimal(row.get("total_payment_amount"))
            balance_amount = safe_decimal(row.get("total_merchant_net_balance"))
            total_user_paid += user_paid
            total_net += net_amount
            total_payment += payment_amount
            total_balance += balance_amount
            sheet.append(
                [
                    _export_cell(sheet, row.get("merchant_subject_name") or ""),
                    _export_cell(sheet, f"{int(row['accounting_year'])}年{int(row['accounting_month'])}月"),
                    _export_cell(sheet, row.get("receipt_merchant") or "", alignment=CHECKLIST_EXPORT_LEFT_ALIGNMENT),
                    _export_cell(
                        sheet,
                        float(user_paid),
                        alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                        number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                    ),
                    _export_cell(
                        sheet,
                        float(net_amount),
                        alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                        number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                    ),
                    _export_cell(
                        sheet,
                        float(payment_amount),
                        alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                        number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                    ),
                    _export_cell(
                        sheet,
                        float(balance_amount),
                        alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                        number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                    ),
                ]
            )
            row_count += 1
        sheet.append(
            [
                _export_cell(sheet, "总计", bold=True),
                _export_cell(sheet, "", bold=True),
                _export_cell(sheet, "", bold=True),
                _export_cell(
                    sheet,
                    float(total_user_paid),
                    bold=True,
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                ),
                _export_cell(
                    sheet,
                    float(total_net),
                    bold=True,
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                ),
                _export_cell(
                    sheet,
                    float(total_payment),
                    bold=True,
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                ),
                _export_cell(
                    sheet,
                    float(total_balance),
                    bold=True,
                    alignment=CHECKLIST_EXPORT_RIGHT_ALIGNMENT,
                    number_format=CHECKLIST_EXPORT_MONEY_FORMAT,
                ),
            ]
        )
        workbook.save(output_path)
        workbook.close()
        return row_count

    @staticmethod
    async def export_summary_to_file(db: AsyncSession, *, user: User, output_path: Path, **params) -> int:
        return await ReconciliationChecklistService.export_product_summary_to_file(db, user=user, output_path=output_path, **params)

    @staticmethod
    async def list_summary(db: AsyncSession, *, user: User, **params) -> tuple[list[dict[str, object]], int]:
        return await ReconciliationChecklistService.list_product_summary(db, user=user, **params)

    @staticmethod
    async def list_summary_details(db: AsyncSession, *, user: User, **params) -> tuple[list[dict[str, object]], int]:
        return await ReconciliationChecklistService.list_product_summary(db, user=user, **params)

    @staticmethod
    async def list_entity_options(
        db: AsyncSession,
        *,
        user: User,
        entity_type: str,
        accounting_year: int,
        accounting_month: int,
        org_id: str | int | None = None,
        keyword: str | None = None,
        limit: int = 50,
        **_kwargs,
    ) -> list[dict[str, object]]:
        kind = {
            "merchant": "merchant_subject",
            "receipt_merchant": "receipt_merchant",
            "live_promoter": "live_platform",
        }.get(entity_type, "merchant_subject")
        rows = await ReconciliationChecklistService.list_options(
            db, user=user, kind=kind, org_id=org_id, accounting_year=accounting_year, accounting_month=accounting_month, keyword=keyword, limit=limit
        )
        return [
            {
                "id": index + 1,
                "org_id": user.org_id or 0,
                "parent_id": None,
                "platform_code": "",
                "entity_type": entity_type,
                "name": row["value"],
            }
            for index, row in enumerate(rows)
        ]


def errors_are_fatal(summary: dict[str, object]) -> bool:
    return bool(summary.get("文件解析失败"))
