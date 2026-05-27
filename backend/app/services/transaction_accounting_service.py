"""Service layer for the independent transaction-accounting flow."""

import io
import logging
import re
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.core.database import register_after_commit
from app.models.cash_flow import CashFlowItem
from app.models.organization import Organization
from app.models.transaction_accounting import (
    TransactionCategory,
    TransactionDetail,
    TransactionMajorCategory,
    TransactionRule,
    TransactionSubject,
    TransactionSummaryRow,
    TransactionTask,
    TransactionUploadFile,
)
from app.models.user import User
from app.schemas.transaction_accounting import (
    TransactionCategoryCreate,
    TransactionCategoryUpdate,
    TransactionMajorCategoryCreate,
    TransactionMajorCategoryUpdate,
    TransactionRuleCreate,
    TransactionRuleUpdate,
    TransactionSubjectCreate,
    TransactionSubjectUpdate,
    TransactionUploadCallback,
    TransactionUploadInit,
)
from app.services.audit_service import AuditService
from app.services.oss_service import SOURCE_FILE_UNAVAILABLE_MESSAGE, is_oss_object_unavailable_error, oss_service
from app.services.shop_service import ShopService
from app.services.transaction_rule_engine import TransactionEvaluationResult, TransactionRuleCandidate, evaluate_transaction_row_matches
from app.tasks.processors.base import FinancialSummaryExcelProcessorMixin, open_tabular_rows, parse_datetime, safe_str
from app.tasks.processors.douyin import DouyinProcessor
from app.utils.query_filters import datetime_range_filters, resolve_org_ids

TRANSACTION_FILENAME_PATTERN = re.compile(
    r"^(?P<year>\d{2}|\d{4})年(?P<month>\d{1,2})月[ _](?P<type>动账)[ _](?P<shop>.+)\.(?:xlsx|xlsm|xls|csv)$",
    re.IGNORECASE,
)

DETAIL_EXPORT_HEADERS = [
    "序号",
    "业务年月",
    "平台",
    "店铺",
    "总分类",
    "科目",
    "重分类",
    "汇总数值",
]

SUMMARY_EXPORT_HEADERS = [
    "序号",
    "核算年月",
    "业务年月",
    "平台",
    "店铺",
    "总分类",
    "科目",
    "汇总金额",
]

EXCEL_EXPORT_ROW_LIMIT = 20000


def _ensure_export_row_limit(label: str, row_count: int) -> None:
    if row_count > EXCEL_EXPORT_ROW_LIMIT:
        raise ValueError(f"{label}导出数据量 {row_count} 行，超过系统上限 {EXCEL_EXPORT_ROW_LIMIT} 行，请缩小筛选范围后再导出")


TRANSACTION_PLATFORM_LABELS = {
    "douyin": "抖音",
    "kuaishou": "快手",
    "xiaohongshu": "小红书",
    "weixin_video": "微信小店",
    "weixinvideo": "微信小店",
    "taobao": "淘宝",
    "tmall": "天猫",
    "alipay": "支付宝",
    "qianniu": "千牛",
    "miniprogram": "小程序",
    "mini_program": "小程序",
    "抖店": "抖音",
    "抖音": "抖音",
    "快手": "快手",
}

TRANSACTION_RESULT_TASK_STATUSES = ("success", "partial_success", "failed", "expired")
TRANSACTION_ERROR_SAMPLE_LIMIT = 20
logger = logging.getLogger("finengine.transaction_accounting")


def _capture_transaction_result_state(task: TransactionTask) -> dict[str, object]:
    return {
        "total_rows": task.total_rows,
        "matched_rows": task.matched_rows,
        "unmatched_rows": task.unmatched_rows,
        "failed_rows": task.failed_rows,
        "result_summary": task.result_summary,
    }


def _restore_transaction_result_state(task: TransactionTask, state: dict[str, object]) -> None:
    task.total_rows = int(state.get("total_rows") or 0)
    task.matched_rows = int(state.get("matched_rows") or 0)
    task.unmatched_rows = int(state.get("unmatched_rows") or 0)
    task.failed_rows = int(state.get("failed_rows") or 0)
    task.result_summary = state.get("result_summary")  # type: ignore[assignment]


@dataclass(frozen=True)
class PageResult:
    items: list
    total: int


@dataclass(frozen=True)
class TransactionSummaryReportRow:
    id: int
    task_id: int
    file_id: int
    org_id: int
    org_name: str | None
    shop_id: int | None
    major_category_id: int | None
    major_category_name: str | None
    subject_id: int
    category_id: int
    subject_name: str
    category_name: str
    transaction_direction: str | None
    platform_code: str | None
    shop_name: str | None
    upload_accounting_year: int | None
    upload_accounting_month: int | None
    accounting_year: int | None
    accounting_month: int | None
    cash_flow_group_name: str | None
    row_count: int
    total_amount: Decimal
    created_at: datetime


@dataclass(frozen=True)
class AnnualSummaryAmount:
    cash_flow_item_id: int
    accounting_month: int
    total_amount: Decimal


@dataclass
class TransactionAnnualSummaryReportRow:
    code: str
    name: str
    parent_code: str | None
    level: int
    item_type: str
    flow_section: str
    flow_direction: str | None
    summary_method: str
    months: dict[str, Decimal]
    total_amount: Decimal


@dataclass(frozen=True)
class TransactionAnnualSummaryReport:
    year: int
    months: list[str]
    rows: list[TransactionAnnualSummaryReportRow]


class TransactionAccountingService:
    @staticmethod
    def resolve_org_id(*, user: User, requested_org_id: int | None = None) -> int:
        if user.role == "superadmin":
            if requested_org_id is None:
                raise ValueError("请选择组织")
            return requested_org_id
        if user.org_id is None:
            raise ValueError("当前账号未绑定组织")
        return user.org_id

    @staticmethod
    def validate_org_scope(*, org_id: int, user: User) -> None:
        if user.role != "superadmin" and org_id != user.org_id:
            raise ValueError("数据不存在或无权访问")

    @staticmethod
    async def create_from_shared_upload(
        db: AsyncSession,
        *,
        upload_file,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> TransactionTask:
        if (upload_file.parsed_type or "").lower() != "动账":
            raise ValueError("仅动账文件可创建资金任务")
        source_platform = (upload_file.source_platform_code or upload_file.detected_platform or "").lower()
        if source_platform != "douyin":
            raise ValueError("动账资金核算 v1 仅支持抖音动账")

        shop_id = upload_file.shop_id
        if shop_id is None and upload_file.parsed_shop:
            shop = await ShopService.get_or_create_shop(
                db,
                org_id=upload_file.org_id,
                platform_name=source_platform,
                shop_name=upload_file.parsed_shop,
            )
            shop_id = shop.id

        existing_stmt = select(TransactionUploadFile).where(
            TransactionUploadFile.source_upload_file_id == upload_file.id,
            TransactionUploadFile.is_deleted.is_(False),
        )
        existing = (await db.execute(existing_stmt)).scalar_one_or_none()
        derived_upload_file = existing
        if derived_upload_file is not None and shop_id is not None and derived_upload_file.shop_id is None:
            derived_upload_file.shop_id = shop_id
        if existing is not None:
            task_stmt = (
                select(TransactionTask)
                .where(
                    TransactionTask.file_id == existing.id,
                    TransactionTask.is_deleted.is_(False),
                )
                .order_by(TransactionTask.id.desc())
            )
            current_task = (await db.execute(task_stmt)).scalars().first()
            if current_task is not None:
                return current_task

        if derived_upload_file is None:
            derived_upload_file = TransactionUploadFile(
                org_id=upload_file.org_id,
                user_id=user.id,
                shop_id=shop_id,
                source_upload_file_id=upload_file.id,
                original_name=upload_file.original_name,
                oss_key=upload_file.oss_key,
                file_size=upload_file.file_size,
                file_hash=upload_file.file_hash,
                platform_code=source_platform,
                shop_name=upload_file.parsed_shop,
                accounting_year=upload_file.parsed_year,
                accounting_month=upload_file.parsed_month,
                status="uploaded",
            )
            db.add(derived_upload_file)
            await db.flush()

        task = TransactionTask(
            file_id=derived_upload_file.id,
            org_id=derived_upload_file.org_id,
            user_id=user.id,
            status="queued",
            progress=0,
        )
        db.add(task)
        await db.flush()

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=derived_upload_file.org_id,
            module="transaction_accounting",
            action="upload_file",
            description=f"从统一上传派生动账资金任务 [{derived_upload_file.original_name}]",
            target_type="transaction_upload_file",
            target_id=derived_upload_file.id,
            target_name=derived_upload_file.original_name,
            ip=ip,
            user_agent=user_agent,
            extra_data={
                "task_id": task.id,
                "file_size": derived_upload_file.file_size,
                "source_upload_file_id": upload_file.id,
            },
        )

        TransactionAccountingService.dispatch_task_after_commit(
            db,
            task=task,
            upload_file=derived_upload_file,
        )
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    def dispatch_task_after_commit(
        db: AsyncSession,
        *,
        task: TransactionTask,
        upload_file: TransactionUploadFile | None,
    ) -> None:
        async def dispatch() -> None:
            try:
                from app.tasks.transaction_accounting import run_transaction_accounting_task

                async_result = run_transaction_accounting_task.delay(task.id)
                task.celery_task_id = async_result.id
            except Exception as exc:
                task.status = "failed"
                task.progress = 100
                task.error_message = f"资金任务投递失败: {exc}"
                task.finished_at = datetime.now(timezone.utc)
                if upload_file is not None:
                    upload_file.status = "failed"
                    upload_file.error_message = task.error_message
            await db.flush()

        register_after_commit(db, dispatch)

    @staticmethod
    async def list_major_categories(db: AsyncSession, *, user: User) -> list[TransactionMajorCategory]:
        _ = user
        stmt = (
            select(TransactionMajorCategory)
            .where(TransactionMajorCategory.is_deleted.is_(False))
            .order_by(TransactionMajorCategory.sort_order, TransactionMajorCategory.id)
        )
        result = await db.execute(stmt)
        return list(result.scalars().all())

    @staticmethod
    async def create_major_category(
        db: AsyncSession,
        *,
        data: TransactionMajorCategoryCreate,
        user: User,
    ) -> TransactionMajorCategory:
        _ = user
        item = TransactionMajorCategory(name=data.name, sort_order=data.sort_order, status=data.status)
        db.add(item)
        await db.flush()
        await db.refresh(item)
        return item

    @staticmethod
    async def update_major_category(
        db: AsyncSession,
        *,
        major_category_id: int,
        data: TransactionMajorCategoryUpdate,
        user: User,
    ) -> TransactionMajorCategory | None:
        item = await db.get(TransactionMajorCategory, major_category_id)
        if item is None or item.is_deleted:
            return None
        _ = user
        values = data.model_dump(exclude_unset=True)
        for key, value in values.items():
            setattr(item, key, value)
        await db.flush()
        await db.refresh(item)
        return item

    @staticmethod
    async def delete_major_category(db: AsyncSession, *, major_category_id: int, user: User) -> bool:
        item = await db.get(TransactionMajorCategory, major_category_id)
        if item is None or item.is_deleted:
            return False
        _ = user
        subject_count = (
            await db.execute(
                select(func.count())
                .select_from(TransactionSubject)
                .where(
                    TransactionSubject.is_deleted.is_(False),
                    TransactionSubject.major_category_id == major_category_id,
                )
            )
        ).scalar_one()
        if subject_count:
            raise ValueError("大分类下仍有科目，不能删除")
        item.is_deleted = True
        item.deleted_at = datetime.now(timezone.utc)
        await db.flush()
        return True

    @staticmethod
    async def list_cash_flow_items(db: AsyncSession, *, user: User) -> list[CashFlowItem]:
        _ = user
        parent = aliased(CashFlowItem)
        stmt = (
            select(CashFlowItem, parent.name.label("parent_name"))
            .outerjoin(parent, parent.id == CashFlowItem.parent_id)
            .where(
                CashFlowItem.is_deleted.is_(False),
                CashFlowItem.status == 1,
                CashFlowItem.level == 2,
            )
            .order_by(parent.sort_order, parent.id, CashFlowItem.sort_order, CashFlowItem.id)
        )
        result = await db.execute(stmt)
        items: list[CashFlowItem] = []
        for item, parent_name in result.all():
            item.parent_name = parent_name
            items.append(item)
        return items

    @staticmethod
    async def list_subjects(db: AsyncSession, *, user: User) -> list[TransactionSubject]:
        _ = user
        major = aliased(TransactionMajorCategory)
        cash_item = aliased(CashFlowItem)
        stmt = (
            select(
                TransactionSubject,
                major.name.label("major_category_name"),
                cash_item.name.label("cash_flow_item_name"),
            )
            .outerjoin(major, major.id == TransactionSubject.major_category_id)
            .outerjoin(cash_item, cash_item.id == TransactionSubject.cash_flow_item_id)
            .where(TransactionSubject.is_deleted.is_(False))
            .order_by(
                major.sort_order.nullslast(),
                major.id.nullslast(),
                TransactionSubject.sort_order,
                TransactionSubject.id,
            )
        )
        result = await db.execute(stmt)
        subjects: list[TransactionSubject] = []
        for subject, major_category_name, cash_flow_item_name in result.all():
            subject.major_category_name = major_category_name
            subject.cash_flow_item_name = cash_flow_item_name
            subjects.append(subject)
        return subjects

    @staticmethod
    async def create_subject(db: AsyncSession, *, data: TransactionSubjectCreate, user: User) -> TransactionSubject:
        _ = user
        await TransactionAccountingService._validate_subject_links(
            db,
            major_category_id=data.major_category_id,
            cash_flow_item_id=data.cash_flow_item_id,
        )
        subject = TransactionSubject(
            name=data.name,
            account_type=data.account_type,
            major_category_id=data.major_category_id,
            cash_flow_item_id=data.cash_flow_item_id,
            sort_order=data.sort_order,
            status=data.status,
        )
        db.add(subject)
        await db.flush()
        return await TransactionAccountingService._load_subject_with_labels(db, subject.id)

    @staticmethod
    async def update_subject(db: AsyncSession, *, subject_id: int, data: TransactionSubjectUpdate, user: User) -> TransactionSubject | None:
        subject = await db.get(TransactionSubject, subject_id)
        if subject is None or subject.is_deleted:
            return None
        _ = user
        values = data.model_dump(exclude_unset=True)
        major_category_id = values["major_category_id"] if "major_category_id" in values else subject.major_category_id
        cash_flow_item_id = values["cash_flow_item_id"] if "cash_flow_item_id" in values else subject.cash_flow_item_id
        await TransactionAccountingService._validate_subject_links(
            db,
            major_category_id=major_category_id,
            cash_flow_item_id=cash_flow_item_id,
        )
        for key, value in values.items():
            setattr(subject, key, value)
        await db.flush()
        return await TransactionAccountingService._load_subject_with_labels(db, subject.id)

    @staticmethod
    async def delete_subject(db: AsyncSession, *, subject_id: int, user: User) -> bool:
        subject = await db.get(TransactionSubject, subject_id)
        if subject is None or subject.is_deleted:
            return False
        _ = user
        category_count = (
            await db.execute(
                select(func.count())
                .select_from(TransactionCategory)
                .where(
                    TransactionCategory.is_deleted.is_(False),
                    TransactionCategory.subject_id == subject_id,
                )
            )
        ).scalar_one()
        if category_count:
            raise ValueError("科目下仍有重分类，不能删除")
        subject.is_deleted = True
        subject.deleted_at = datetime.now(timezone.utc)
        await db.flush()
        return True

    @staticmethod
    async def list_categories(db: AsyncSession, *, user: User, subject_id: int | None = None) -> list[TransactionCategory]:
        _ = user
        stmt = select(TransactionCategory).where(TransactionCategory.is_deleted.is_(False))
        if subject_id is not None:
            stmt = stmt.where(TransactionCategory.subject_id == subject_id)
        stmt = stmt.order_by(TransactionCategory.sort_order, TransactionCategory.id)
        result = await db.execute(stmt)
        return list(result.scalars().all())

    @staticmethod
    async def create_category(db: AsyncSession, *, data: TransactionCategoryCreate, user: User) -> TransactionCategory:
        subject = await db.get(TransactionSubject, data.subject_id)
        if subject is None or subject.is_deleted:
            raise ValueError("科目不存在")
        _ = user
        category = TransactionCategory(
            subject_id=data.subject_id,
            name=data.name,
            sort_order=data.sort_order,
            status=data.status,
        )
        db.add(category)
        await db.flush()
        await db.refresh(category)
        return category

    @staticmethod
    async def update_category(db: AsyncSession, *, category_id: int, data: TransactionCategoryUpdate, user: User) -> TransactionCategory | None:
        category = await db.get(TransactionCategory, category_id)
        if category is None or category.is_deleted:
            return None
        _ = user
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(category, key, value)
        await db.flush()
        await db.refresh(category)
        return category

    @staticmethod
    async def delete_category(db: AsyncSession, *, category_id: int, user: User) -> bool:
        category = await db.get(TransactionCategory, category_id)
        if category is None or category.is_deleted:
            return False
        _ = user
        rule_count = (
            await db.execute(
                select(func.count())
                .select_from(TransactionRule)
                .where(
                    TransactionRule.is_deleted.is_(False),
                    TransactionRule.category_id == category_id,
                )
            )
        ).scalar_one()
        if rule_count:
            raise ValueError("重分类下仍有规则，不能删除")
        category.is_deleted = True
        category.deleted_at = datetime.now(timezone.utc)
        await db.flush()
        return True

    @staticmethod
    async def list_rules(db: AsyncSession, *, user: User) -> list[tuple[TransactionRule, str | None, str | None]]:
        _ = user
        stmt = (
            select(TransactionRule, TransactionSubject.name, TransactionCategory.name)
            .join(TransactionSubject, TransactionSubject.id == TransactionRule.subject_id)
            .join(TransactionCategory, TransactionCategory.id == TransactionRule.category_id)
            .where(TransactionRule.is_deleted.is_(False))
            .order_by(TransactionRule.priority, TransactionRule.id)
        )
        result = await db.execute(stmt)
        return list(result.all())

    @staticmethod
    async def create_rule(db: AsyncSession, *, data: TransactionRuleCreate, user: User) -> TransactionRule:
        category = await TransactionAccountingService._get_category_for_rule(db, category_id=data.category_id, subject_id=data.subject_id, user=user)
        _ = category
        values = data.model_dump()
        values["remark_pattern"] = TransactionAccountingService._normalize_rule_match_config(
            match_type=values["match_type"],
            remark_pattern=values["remark_pattern"],
        )
        values["remark_exclude_pattern"] = TransactionAccountingService._normalize_rule_exclude_config(
            values.get("remark_exclude_pattern")
        )
        rule = TransactionRule(**values)
        db.add(rule)
        await db.flush()
        await db.refresh(rule)
        return rule

    @staticmethod
    async def update_rule(db: AsyncSession, *, rule_id: int, data: TransactionRuleUpdate, user: User) -> TransactionRule | None:
        rule = await db.get(TransactionRule, rule_id)
        if rule is None or rule.is_deleted:
            return None
        _ = user
        values = data.model_dump(exclude_unset=True)
        subject_id = int(values.get("subject_id") or rule.subject_id)
        category_id = int(values.get("category_id") or rule.category_id)
        if "subject_id" in values or "category_id" in values:
            await TransactionAccountingService._get_category_for_rule(db, category_id=category_id, subject_id=subject_id, user=user)
        match_type = str(values.get("match_type") or rule.match_type)
        remark_pattern_value = values["remark_pattern"] if "remark_pattern" in values and values["remark_pattern"] is not None else rule.remark_pattern
        values["remark_pattern"] = TransactionAccountingService._normalize_rule_match_config(
            match_type=match_type,
            remark_pattern=str(remark_pattern_value),
        )
        exclude_pattern_value = (
            values["remark_exclude_pattern"]
            if "remark_exclude_pattern" in values and values["remark_exclude_pattern"] is not None
            else rule.remark_exclude_pattern
        )
        values["remark_exclude_pattern"] = TransactionAccountingService._normalize_rule_exclude_config(
            str(exclude_pattern_value)
        )
        for key, value in values.items():
            setattr(rule, key, value)
        await db.flush()
        await db.refresh(rule)
        return rule

    @staticmethod
    async def delete_rule(db: AsyncSession, *, rule_id: int, user: User) -> bool:
        rule = await db.get(TransactionRule, rule_id)
        if rule is None or rule.is_deleted:
            return False
        _ = user
        rule.is_deleted = True
        rule.deleted_at = datetime.now(timezone.utc)
        await db.flush()
        return True

    @staticmethod
    async def init_upload(db: AsyncSession, *, data: TransactionUploadInit, user: User, org_id: int | None = None) -> TransactionUploadFile:
        scoped_org_id = TransactionAccountingService.resolve_org_id(user=user, requested_org_id=org_id)
        parsed = TransactionAccountingService.parse_transaction_filename(data.original_name)
        platform_code = data.platform_code or "douyin"
        if platform_code != "douyin":
            raise ValueError("动账核算 v1 仅支持抖音动账")
        shop = await ShopService.get_or_create_shop(
            db,
            org_id=scoped_org_id,
            platform_name=platform_code,
            shop_name=data.shop_name or parsed["shop"],
        )
        upload_file = TransactionUploadFile(
            org_id=scoped_org_id,
            user_id=user.id,
            shop_id=shop.id,
            original_name=data.original_name,
            file_size=data.file_size,
            platform_code=platform_code,
            shop_name=shop.shop_name,
            accounting_year=data.accounting_year or parsed["year"],
            accounting_month=data.accounting_month or parsed["month"],
            status="initialized",
        )
        db.add(upload_file)
        await db.flush()
        upload_file.oss_key = f"user-upload/transaction-accounting/{scoped_org_id}/{upload_file.id}/{Path(data.original_name).name}"
        await db.flush()
        await db.refresh(upload_file)
        return upload_file

    @staticmethod
    def parse_transaction_filename(filename: str) -> dict[str, int | str]:
        match = TRANSACTION_FILENAME_PATTERN.match(Path(filename).name)
        if not match:
            raise ValueError("文件名不符合规则：请使用 26年02月_动账_店铺名称.xlsx")
        year = int(match.group("year"))
        if year < 100:
            year += 2000
        month = int(match.group("month"))
        if month < 1 or month > 12:
            raise ValueError("文件名月份不正确")
        shop = match.group("shop").strip()
        if not shop:
            raise ValueError("文件名缺少店铺名称")
        return {"year": year, "month": month, "shop": shop}

    @staticmethod
    async def upload_callback(
        db: AsyncSession,
        *,
        data: TransactionUploadCallback,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> TransactionTask:
        upload_file = await db.get(TransactionUploadFile, data.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("上传文件不存在")
        TransactionAccountingService.validate_org_scope(org_id=upload_file.org_id, user=user)
        existing_task = await TransactionAccountingService._get_latest_task_for_upload_file(db, file_id=upload_file.id)
        if existing_task is not None:
            return existing_task
        upload_file.oss_key = data.oss_key
        upload_file.file_size = data.file_size
        upload_file.file_hash = data.file_hash
        upload_file.status = "uploaded"

        task = TransactionTask(
            file_id=upload_file.id,
            org_id=upload_file.org_id,
            user_id=user.id,
            status="queued",
            progress=0,
        )
        db.add(task)
        await db.flush()

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=upload_file.org_id,
            module="transaction_accounting",
            action="upload_file",
            description=f"上传动账核算文件 [{upload_file.original_name}]",
            target_type="transaction_upload_file",
            target_id=upload_file.id,
            target_name=upload_file.original_name,
            ip=ip,
            user_agent=user_agent,
            extra_data={"task_id": task.id, "file_size": data.file_size},
        )

        TransactionAccountingService.dispatch_task_after_commit(
            db,
            task=task,
            upload_file=upload_file,
        )
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def rerun_task(db: AsyncSession, *, task_id: int, user: User, ip: str | None = None, user_agent: str | None = None) -> TransactionTask | None:
        task = await db.get(TransactionTask, task_id)
        if task is None or task.is_deleted:
            return None
        TransactionAccountingService.validate_org_scope(org_id=task.org_id, user=user)
        if task.status == "expired":
            raise ValueError("源文件已过期或不存在，不能重新统计，请重新上传文件")
        upload_file = await db.get(TransactionUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("动账上传文件不存在")

        task.status = "queued"
        task.progress = 0
        task.celery_task_id = None
        task.error_message = None
        task.started_at = None
        task.finished_at = None
        upload_file.status = "uploaded"
        upload_file.error_message = None

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=task.org_id,
            module="transaction_accounting",
            action="task_rerun",
            description=f"重跑动账核算任务 [{task.id}]",
            target_type="transaction_task",
            target_id=task.id,
            ip=ip,
            user_agent=user_agent,
            extra_data={"task_id": task.id, "file_id": upload_file.id},
        )

        TransactionAccountingService.dispatch_task_after_commit(
            db,
            task=task,
            upload_file=upload_file,
        )
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def list_tasks(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        status: str | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        keyword: str | None = None,
        created_start_time: datetime | None = None,
        created_end_time: datetime | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[tuple[TransactionTask, TransactionUploadFile, str | None]], int]:
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        filters = [TransactionTask.is_deleted.is_(False), TransactionUploadFile.is_deleted.is_(False)]
        if org_ids is not None:
            filters.append(TransactionTask.org_id.in_(org_ids))
        if status:
            filters.append(TransactionTask.status.in_(TransactionAccountingService._split_filter_values(status)))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(TransactionUploadFile.platform_code.in_(platform_codes))
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(TransactionUploadFile.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(TransactionUploadFile.shop_id.in_(shop_id_list))
        if accounting_year is not None:
            filters.append(TransactionUploadFile.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(TransactionUploadFile.accounting_month == accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionUploadFile.accounting_year * 100 + TransactionUploadFile.accounting_month,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    TransactionUploadFile.original_name.ilike(like_pattern),
                    TransactionUploadFile.shop_name.ilike(like_pattern),
                    TransactionUploadFile.platform_code.ilike(like_pattern),
                    TransactionTask.error_message.ilike(like_pattern),
                )
            )
        filters.extend(
            datetime_range_filters(
                TransactionTask.created_at,
                start_time=created_start_time,
                end_time=created_end_time,
            )
        )
        stmt = (
            select(TransactionTask, TransactionUploadFile, Organization.name.label("org_name"))
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionTask.file_id)
            .outerjoin(Organization, Organization.id == TransactionTask.org_id)
            .where(*filters)
            .order_by(TransactionTask.id.desc())
        )
        count_stmt = select(func.count()).select_from(TransactionTask).join(TransactionUploadFile, TransactionUploadFile.id == TransactionTask.file_id).where(*filters)
        total = (await db.execute(count_stmt)).scalar() or 0
        result = await db.execute(stmt.offset((page - 1) * page_size).limit(page_size))
        return list(result.all()), total

    @staticmethod
    async def list_details(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        task_id: int | None = None,
        status: str | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
        ids: list[int] | None = None,
        page: int | None = 1,
        page_size: int | None = 50,
    ) -> tuple[list[TransactionDetail], int]:
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if ids is not None and not ids:
            return [], 0
        filters = [
            TransactionDetail.is_deleted.is_(False),
            TransactionUploadFile.is_deleted.is_(False),
            TransactionTask.is_deleted.is_(False),
        ]
        if org_ids is not None:
            filters.append(TransactionDetail.org_id.in_(org_ids))
        resolved_major_category_id = func.coalesce(
            TransactionDetail.major_category_id,
            TransactionSubject.major_category_id,
        )
        resolved_major_category_name = func.coalesce(
            TransactionMajorCategory.name,
            TransactionDetail.major_category_name,
        )
        if ids is not None:
            filters.append(TransactionDetail.id.in_(ids))
        if task_id is not None:
            filters.append(TransactionDetail.task_id == task_id)
        else:
            filters.append(TransactionDetail.task_id.in_(TransactionAccountingService._latest_result_task_ids_select()))
        if status:
            filters.append(TransactionDetail.status == status)
        else:
            filters.extend(
                [
                    TransactionDetail.status == "matched",
                    TransactionDetail.subject_id.is_not(None),
                    TransactionDetail.category_id.is_not(None),
                ]
            )
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(TransactionDetail.platform_code.in_(platform_codes))
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(TransactionDetail.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(TransactionDetail.shop_id.in_(shop_id_list))
        major_category_ids = TransactionAccountingService._split_int_filter_values(major_category_id)
        if major_category_ids:
            filters.append(resolved_major_category_id.in_(major_category_ids))
        subject_ids = TransactionAccountingService._split_int_filter_values(subject_id)
        if subject_ids:
            filters.append(TransactionDetail.subject_id.in_(subject_ids))
        category_ids = TransactionAccountingService._split_int_filter_values(category_id)
        if category_ids:
            filters.append(TransactionDetail.category_id.in_(category_ids))
        if transaction_direction:
            filters.append(TransactionDetail.transaction_direction == transaction_direction)
        if accounting_year is not None:
            filters.append(TransactionDetail.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(TransactionDetail.accounting_month == accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionDetail.accounting_year * 100 + TransactionDetail.accounting_month,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )
        if upload_accounting_year is not None:
            filters.append(TransactionUploadFile.accounting_year == upload_accounting_year)
        if upload_accounting_month is not None:
            filters.append(TransactionUploadFile.accounting_month == upload_accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionUploadFile.accounting_year * 100 + TransactionUploadFile.accounting_month,
                start_year=upload_accounting_start_year,
                start_month=upload_accounting_start_month,
                end_year=upload_accounting_end_year,
                end_month=upload_accounting_end_month,
            )
        )
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    resolved_major_category_name.ilike(like_pattern),
                    TransactionDetail.shop_name.ilike(like_pattern),
                    TransactionDetail.platform_code.ilike(like_pattern),
                    TransactionDetail.remark.ilike(like_pattern),
                    TransactionDetail.amount_field.ilike(like_pattern),
                    TransactionDetail.error_message.ilike(like_pattern),
                )
            )
        stmt = (
            select(
                TransactionDetail,
                Organization.name.label("org_name"),
                TransactionSubject.name,
                resolved_major_category_id.label("resolved_major_category_id"),
                resolved_major_category_name.label("resolved_major_category_name"),
                TransactionCategory.name,
                TransactionUploadFile.accounting_year,
                TransactionUploadFile.accounting_month,
            )
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionDetail.file_id)
            .join(TransactionTask, TransactionTask.id == TransactionDetail.task_id)
            .outerjoin(TransactionSubject, TransactionSubject.id == TransactionDetail.subject_id)
            .outerjoin(TransactionMajorCategory, TransactionMajorCategory.id == resolved_major_category_id)
            .outerjoin(TransactionCategory, TransactionCategory.id == TransactionDetail.category_id)
            .outerjoin(Organization, Organization.id == TransactionDetail.org_id)
            .where(*filters)
            .order_by(
                TransactionUploadFile.accounting_year.desc().nullslast(),
                TransactionUploadFile.accounting_month.desc().nullslast(),
                TransactionDetail.accounting_year.desc().nullslast(),
                TransactionDetail.accounting_month.desc().nullslast(),
                TransactionDetail.platform_code,
                TransactionDetail.shop_id,
                TransactionDetail.shop_name,
                resolved_major_category_name,
                TransactionSubject.name,
                TransactionCategory.name,
            )
        )
        count_stmt = (
            select(func.count())
            .select_from(TransactionDetail)
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionDetail.file_id)
            .join(TransactionTask, TransactionTask.id == TransactionDetail.task_id)
            .outerjoin(TransactionSubject, TransactionSubject.id == TransactionDetail.subject_id)
            .outerjoin(TransactionMajorCategory, TransactionMajorCategory.id == resolved_major_category_id)
            .where(*filters)
        )
        total = (await db.execute(count_stmt)).scalar() or 0
        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(stmt)
        rows = []
        for (
            detail,
            org_name,
            subject_name,
            resolved_major_category_id_value,
            resolved_major_category_name_value,
            category_name,
            upload_year,
            upload_month,
        ) in result.all():
            detail.org_name = org_name
            detail.major_category_id = resolved_major_category_id_value
            detail.major_category_name = resolved_major_category_name_value
            detail.subject_name = subject_name
            detail.category_name = category_name
            detail.upload_accounting_year = upload_year
            detail.upload_accounting_month = upload_month
            detail.cash_flow_group_name = detail.major_category_name
            detail.total_amount = detail.calculated_amount
            rows.append(detail)
        return rows, total

    @staticmethod
    async def list_summaries(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
        ids: list[int] | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ) -> tuple[list[TransactionSummaryReportRow], int]:
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if ids is not None and not ids:
            return [], 0
        filters = [TransactionSummaryRow.is_deleted.is_(False), TransactionUploadFile.is_deleted.is_(False)]
        if org_ids is not None:
            filters.append(TransactionSummaryRow.org_id.in_(org_ids))
        if task_id is not None:
            filters.append(TransactionSummaryRow.task_id == task_id)
        else:
            filters.append(TransactionSummaryRow.task_id.in_(TransactionAccountingService._latest_result_task_ids_select()))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(TransactionSummaryRow.platform_code.in_(platform_codes))
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(TransactionSummaryRow.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(TransactionSummaryRow.shop_id.in_(shop_id_list))
        major_category_ids = TransactionAccountingService._split_int_filter_values(major_category_id)
        if major_category_ids:
            filters.append(TransactionSummaryRow.major_category_id.in_(major_category_ids))
        subject_ids = TransactionAccountingService._split_int_filter_values(subject_id)
        if subject_ids:
            filters.append(TransactionSummaryRow.subject_id.in_(subject_ids))
        category_ids = TransactionAccountingService._split_int_filter_values(category_id)
        if category_ids:
            filters.append(TransactionSummaryRow.category_id.in_(category_ids))
        if transaction_direction:
            filters.append(TransactionSummaryRow.transaction_direction == transaction_direction)
        if accounting_year is not None:
            filters.append(TransactionSummaryRow.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(TransactionSummaryRow.accounting_month == accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionSummaryRow.accounting_year * 100 + TransactionSummaryRow.accounting_month,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )
        if upload_accounting_year is not None:
            filters.append(TransactionUploadFile.accounting_year == upload_accounting_year)
        if upload_accounting_month is not None:
            filters.append(TransactionUploadFile.accounting_month == upload_accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionUploadFile.accounting_year * 100 + TransactionUploadFile.accounting_month,
                start_year=upload_accounting_start_year,
                start_month=upload_accounting_start_month,
                end_year=upload_accounting_end_year,
                end_month=upload_accounting_end_month,
            )
        )
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    TransactionSummaryRow.major_category_name.ilike(like_pattern),
                    TransactionSummaryRow.subject_name.ilike(like_pattern),
                    TransactionSummaryRow.category_name.ilike(like_pattern),
                    TransactionSummaryRow.shop_name.ilike(like_pattern),
                    TransactionSummaryRow.platform_code.ilike(like_pattern),
                )
            )
        grouped_fields = [
            TransactionSummaryRow.org_id,
            TransactionSummaryRow.platform_code,
            TransactionSummaryRow.shop_id,
            TransactionSummaryRow.shop_name,
            TransactionUploadFile.accounting_year,
            TransactionUploadFile.accounting_month,
            TransactionSummaryRow.accounting_year,
            TransactionSummaryRow.accounting_month,
            TransactionSummaryRow.major_category_id,
            TransactionSummaryRow.major_category_name,
            TransactionSummaryRow.subject_id,
            TransactionSummaryRow.subject_name,
        ]
        id_expr = func.min(TransactionSummaryRow.id)
        stmt = (
            select(
                id_expr.label("id"),
                func.min(TransactionSummaryRow.task_id).label("task_id"),
                func.min(TransactionSummaryRow.file_id).label("file_id"),
                TransactionSummaryRow.org_id.label("org_id"),
                func.max(Organization.name).label("org_name"),
                TransactionSummaryRow.shop_id.label("shop_id"),
                TransactionSummaryRow.major_category_id.label("major_category_id"),
                TransactionSummaryRow.major_category_name.label("major_category_name"),
                TransactionSummaryRow.subject_id.label("subject_id"),
                func.min(TransactionSummaryRow.category_id).label("category_id"),
                TransactionSummaryRow.subject_name.label("subject_name"),
                func.min(TransactionSummaryRow.category_name).label("category_name"),
                TransactionSummaryRow.platform_code.label("platform_code"),
                TransactionSummaryRow.shop_name.label("shop_name"),
                TransactionUploadFile.accounting_year.label("upload_accounting_year"),
                TransactionUploadFile.accounting_month.label("upload_accounting_month"),
                TransactionSummaryRow.accounting_year.label("accounting_year"),
                TransactionSummaryRow.accounting_month.label("accounting_month"),
                func.sum(TransactionSummaryRow.row_count).label("row_count"),
                func.sum(TransactionSummaryRow.total_amount).label("total_amount"),
                func.min(TransactionSummaryRow.created_at).label("created_at"),
            )
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionSummaryRow.file_id)
            .outerjoin(Organization, Organization.id == TransactionSummaryRow.org_id)
            .where(*filters)
            .group_by(*grouped_fields)
            .order_by(
                TransactionUploadFile.accounting_year.desc().nullslast(),
                TransactionUploadFile.accounting_month.desc().nullslast(),
                TransactionSummaryRow.accounting_year.desc().nullslast(),
                TransactionSummaryRow.accounting_month.desc().nullslast(),
                TransactionSummaryRow.platform_code,
                TransactionSummaryRow.shop_id,
                TransactionSummaryRow.shop_name,
                TransactionSummaryRow.major_category_name,
                TransactionSummaryRow.subject_name,
            )
        )
        if ids is not None:
            stmt = stmt.having(id_expr.in_(ids))
        count_subquery = (
            select(*grouped_fields)
            .select_from(TransactionSummaryRow)
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionSummaryRow.file_id)
            .where(*filters)
            .group_by(*grouped_fields)
        )
        if ids is not None:
            count_subquery = count_subquery.having(id_expr.in_(ids))
        count_subquery = count_subquery.subquery()
        count_stmt = select(func.count()).select_from(count_subquery)
        total = (await db.execute(count_stmt)).scalar() or 0
        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(stmt)
        rows = [
            TransactionSummaryReportRow(
                id=row.id,
                task_id=row.task_id,
                file_id=row.file_id,
                org_id=row.org_id,
                org_name=row.org_name,
                shop_id=row.shop_id,
                major_category_id=row.major_category_id,
                major_category_name=row.major_category_name,
                subject_id=row.subject_id,
                category_id=row.category_id,
                subject_name=row.subject_name,
                category_name=row.category_name,
                transaction_direction=None,
                platform_code=row.platform_code,
                shop_name=row.shop_name,
                upload_accounting_year=row.upload_accounting_year,
                upload_accounting_month=row.upload_accounting_month,
                accounting_year=row.accounting_year,
                accounting_month=row.accounting_month,
                cash_flow_group_name=row.major_category_name,
                row_count=int(row.row_count or 0),
                total_amount=row.total_amount or Decimal("0"),
                created_at=row.created_at,
            )
            for row in result.all()
        ]
        return rows, total

    @staticmethod
    async def get_annual_summary(
        db: AsyncSession,
        *,
        user: User,
        year: int,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
    ) -> TransactionAnnualSummaryReport:
        cash_items = await TransactionAccountingService._load_cash_flow_items(db)
        amounts = await TransactionAccountingService._load_annual_summary_amounts(
            db,
            user=user,
            year=year,
            org_id=org_id,
            task_id=task_id,
            platform_code=platform_code,
            shop_name=shop_name,
            shop_ids=shop_ids,
            major_category_id=major_category_id,
            subject_id=subject_id,
            category_id=category_id,
            transaction_direction=transaction_direction,
            upload_accounting_year=upload_accounting_year,
            upload_accounting_month=upload_accounting_month,
            upload_accounting_start_year=upload_accounting_start_year,
            upload_accounting_start_month=upload_accounting_start_month,
            upload_accounting_end_year=upload_accounting_end_year,
            upload_accounting_end_month=upload_accounting_end_month,
            keyword=keyword,
        )
        rows = TransactionAccountingService._build_annual_report_rows(year=year, cash_items=cash_items, amounts=amounts)
        return TransactionAnnualSummaryReport(
            year=year,
            months=TransactionAccountingService._annual_month_keys(year),
            rows=rows,
        )

    @staticmethod
    async def export_annual_summary(
        db: AsyncSession,
        *,
        user: User,
        year: int,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
    ) -> io.BytesIO:
        report = await TransactionAccountingService.get_annual_summary(
            db,
            user=user,
            year=year,
            org_id=org_id,
            task_id=task_id,
            platform_code=platform_code,
            shop_name=shop_name,
            shop_ids=shop_ids,
            major_category_id=major_category_id,
            subject_id=subject_id,
            category_id=category_id,
            transaction_direction=transaction_direction,
            upload_accounting_year=upload_accounting_year,
            upload_accounting_month=upload_accounting_month,
            upload_accounting_start_year=upload_accounting_start_year,
            upload_accounting_start_month=upload_accounting_start_month,
            upload_accounting_end_year=upload_accounting_end_year,
            upload_accounting_end_month=upload_accounting_end_month,
            keyword=keyword,
        )
        return TransactionAccountingService._build_annual_summary_workbook(year=report.year, rows=report.rows)

    @staticmethod
    async def export_details(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        task_id: int | None = None,
        status: str | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
        ids: list[int] | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ) -> io.BytesIO:
        if page is not None and page_size is not None:
            rows, _ = await TransactionAccountingService.list_details(
                db, user=user, org_id=org_id, task_id=task_id, status=status,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=page, page_size=page_size,
            )
        else:
            _, total = await TransactionAccountingService.list_details(
                db, user=user, org_id=org_id, task_id=task_id, status=status,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=1, page_size=1,
            )
            _ensure_export_row_limit("动账明细", total)
            rows, _ = await TransactionAccountingService.list_details(
                db, user=user, org_id=org_id, task_id=task_id, status=status,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=None, page_size=None,
            )
        return TransactionAccountingService._build_detail_workbook(rows)

    @staticmethod
    async def export_summaries(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
        ids: list[int] | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ) -> io.BytesIO:
        if page is not None and page_size is not None:
            rows, _ = await TransactionAccountingService.list_summaries(
                db, user=user, org_id=org_id, task_id=task_id,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=page, page_size=page_size,
            )
        else:
            _, total = await TransactionAccountingService.list_summaries(
                db, user=user, org_id=org_id, task_id=task_id,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=1, page_size=1,
            )
            _ensure_export_row_limit("动账汇总", total)
            rows, _ = await TransactionAccountingService.list_summaries(
                db, user=user, org_id=org_id, task_id=task_id,
                platform_code=platform_code, shop_name=shop_name, shop_ids=shop_ids,
                major_category_id=major_category_id, subject_id=subject_id,
                category_id=category_id, transaction_direction=transaction_direction,
                accounting_year=accounting_year, accounting_month=accounting_month,
                accounting_start_year=accounting_start_year, accounting_start_month=accounting_start_month,
                accounting_end_year=accounting_end_year, accounting_end_month=accounting_end_month,
                upload_accounting_year=upload_accounting_year, upload_accounting_month=upload_accounting_month,
                upload_accounting_start_year=upload_accounting_start_year, upload_accounting_start_month=upload_accounting_start_month,
                upload_accounting_end_year=upload_accounting_end_year, upload_accounting_end_month=upload_accounting_end_month,
                keyword=keyword, ids=ids, page=None, page_size=None,
            )
        return TransactionAccountingService._build_summary_workbook(rows)

    @staticmethod
    async def execute_task(db: AsyncSession, *, task_id: int) -> TransactionTask:
        task = await db.get(TransactionTask, task_id)
        if task is None or task.is_deleted:
            raise ValueError("动账核算任务不存在")
        upload_file = await db.get(TransactionUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("动账上传文件不存在")

        previous_result_state = _capture_transaction_result_state(task)
        task.status = "processing"
        task.progress = 5
        task.started_at = datetime.now(timezone.utc)
        task.finished_at = None
        task.error_message = None
        await db.flush()

        try:
            header, data_rows = TransactionAccountingService._load_douyin_dongzhang_rows_from_oss(upload_file)
            rules = await TransactionAccountingService._load_rule_candidates(db, platform_code=upload_file.platform_code)
            if not rules:
                raise ValueError("未配置启用的动账核算规则")

            summary_accumulator: dict[tuple[int, int, int | None, int | None], dict[str, object]] = {}
            rule_accumulator: dict[tuple[int, int | None, int | None], dict[str, object]] = {}
            error_messages: list[str] = []
            counts = defaultdict(int)
            rule_map = {rule.id: rule for rule in rules}
            for offset, row in enumerate(data_rows, start=2):
                direction_field = TransactionAccountingService._first_existing_field(row, [rule.direction_field for rule in rules], default="动账方向")
                scene_field = TransactionAccountingService._first_existing_field(row, [rule.scene_field for rule in rules], default="动账场景")
                remark_field = TransactionAccountingService._first_existing_field(row, [rule.remark_field for rule in rules], default="备注")
                detail_year, detail_month, period_error = TransactionAccountingService._resolve_transaction_period(row)
                evaluations = evaluate_transaction_row_matches(
                    row=row,
                    row_number=offset,
                    rules=rules,
                    direction_field=direction_field,
                    scene_field=scene_field,
                    remark_field=remark_field,
                )
                for evaluation in evaluations:
                    detail_evaluation = TransactionAccountingService._period_failed_result(evaluation=evaluation, message=period_error) if period_error else evaluation
                    counts[detail_evaluation.status] += 1
                    matched_rule = rule_map.get(detail_evaluation.rule_id) if detail_evaluation.rule_id else None
                    if detail_evaluation.status == "matched" and matched_rule is not None:
                        key = (
                            matched_rule.subject_id,
                            matched_rule.category_id,
                            detail_year,
                            detail_month,
                        )
                        bucket = summary_accumulator.setdefault(key, {"amount": Decimal("0"), "count": 0, "first_row": offset})
                        bucket["amount"] = bucket["amount"] + detail_evaluation.calculated_amount
                        bucket["count"] = int(bucket["count"]) + 1
                        bucket["first_row"] = min(int(bucket["first_row"]), offset)
                        rule_key = (matched_rule.id, detail_year, detail_month)
                        rule_bucket = rule_accumulator.setdefault(rule_key, {"amount": Decimal("0"), "count": 0, "first_row": offset})
                        rule_bucket["amount"] = rule_bucket["amount"] + detail_evaluation.calculated_amount
                        rule_bucket["count"] = int(rule_bucket["count"]) + 1
                        rule_bucket["first_row"] = min(int(rule_bucket["first_row"]), offset)
                        continue

                    if detail_evaluation.status in {"unmatched", "failed"}:
                        error_message = TransactionAccountingService._build_transaction_row_error_message(
                            row=row,
                            evaluation=detail_evaluation,
                            direction_field=direction_field,
                            scene_field=scene_field,
                            remark_field=remark_field,
                            period_error=period_error,
                        )
                        if len(error_messages) < TRANSACTION_ERROR_SAMPLE_LIMIT:
                            error_messages.append(error_message)

            TransactionAccountingService._log_rule_result_summaries(
                task=task,
                upload_file=upload_file,
                rule_map=rule_map,
                rule_accumulator=rule_accumulator,
            )
            await db.execute(delete(TransactionDetail).where(TransactionDetail.task_id == task.id))
            await db.execute(delete(TransactionSummaryRow).where(TransactionSummaryRow.task_id == task.id))
            await TransactionAccountingService._create_aggregated_result_rows(
                db,
                task=task,
                upload_file=upload_file,
                summary_accumulator=summary_accumulator,
            )
            await db.flush()
            task.total_rows = len(data_rows)
            task.matched_rows = counts["matched"]
            task.unmatched_rows = counts["unmatched"]
            task.failed_rows = counts["failed"]
            task.progress = 100
            task.status = "success" if task.unmatched_rows == 0 and task.failed_rows == 0 else "partial_success"
            task.result_summary = {
                "总行数": task.total_rows,
                "匹配明细数": task.matched_rows,
                "未匹配行数": task.unmatched_rows,
                "失败行数": task.failed_rows,
                "汇总分组数": len(summary_accumulator),
            }
            if error_messages:
                hidden_error_count = task.unmatched_rows + task.failed_rows - len(error_messages)
                task.result_summary["错误明细"] = error_messages
                if hidden_error_count > 0:
                    task.result_summary["错误明细截断提示"] = f"另有 {hidden_error_count} 条错误未展示，请查看任务错误信息或后台日志"
                task.error_message = "\n".join(error_messages)
                if hidden_error_count > 0:
                    task.error_message += f"\n另有 {hidden_error_count} 条错误未展示，请查看任务错误信息或后台日志"
            else:
                task.error_message = None
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "processed"
        except Exception as exc:
            await db.rollback()
            task = await db.get(TransactionTask, task_id)
            if task is None or task.is_deleted:
                raise ValueError("动账核算任务不存在") from exc
            upload_file = await db.get(TransactionUploadFile, task.file_id)
            if upload_file is None or upload_file.is_deleted:
                raise ValueError("动账上传文件不存在") from exc
            is_expired = is_oss_object_unavailable_error(exc)
            _restore_transaction_result_state(task, previous_result_state)
            task.status = "expired" if is_expired else "failed"
            task.progress = 100
            task.error_message = SOURCE_FILE_UNAVAILABLE_MESSAGE if is_expired else str(exc)
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "expired" if is_expired else "failed"
            upload_file.error_message = task.error_message
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def _load_subject_with_labels(db: AsyncSession, subject_id: int) -> TransactionSubject:
        major = aliased(TransactionMajorCategory)
        cash_item = aliased(CashFlowItem)
        stmt = (
            select(
                TransactionSubject,
                major.name.label("major_category_name"),
                cash_item.name.label("cash_flow_item_name"),
            )
            .outerjoin(major, major.id == TransactionSubject.major_category_id)
            .outerjoin(cash_item, cash_item.id == TransactionSubject.cash_flow_item_id)
            .where(
                TransactionSubject.id == subject_id,
                TransactionSubject.is_deleted.is_(False),
            )
        )
        result = await db.execute(stmt)
        row = result.first()
        if row is None:
            raise ValueError("科目不存在")
        subject, major_category_name, cash_flow_item_name = row
        subject.major_category_name = major_category_name
        subject.cash_flow_item_name = cash_flow_item_name
        return subject

    @staticmethod
    async def _validate_subject_links(
        db: AsyncSession,
        *,
        major_category_id: int | None,
        cash_flow_item_id: int | None,
    ) -> None:
        if major_category_id is not None:
            major_category = await db.get(TransactionMajorCategory, major_category_id)
            if major_category is None or major_category.is_deleted:
                raise ValueError("大分类不存在")
        if cash_flow_item_id is not None:
            cash_flow_item = await db.get(CashFlowItem, cash_flow_item_id)
            if (
                cash_flow_item is None
                or getattr(cash_flow_item, "is_deleted", False)
                or getattr(cash_flow_item, "status", 0) != 1
                or cash_flow_item.level != 2
            ):
                raise ValueError("现金流项目不存在")

    @staticmethod
    async def _get_category_for_rule(db: AsyncSession, *, category_id: int, subject_id: int, user: User) -> TransactionCategory:
        result = await db.execute(
            select(TransactionCategory)
            .join(TransactionSubject, TransactionSubject.id == TransactionCategory.subject_id)
            .where(
                TransactionCategory.id == category_id,
                TransactionCategory.subject_id == subject_id,
                TransactionCategory.is_deleted.is_(False),
                TransactionSubject.is_deleted.is_(False),
            )
        )
        category = result.scalar_one_or_none()
        if category is None:
            raise ValueError("科目或重分类不存在")
        _ = user
        return category

    @staticmethod
    def _normalize_rule_match_config(*, match_type: str, remark_pattern: str) -> str:
        if match_type == "none":
            return ""
        if match_type not in {"exact", "contains", "not_contains"}:
            raise ValueError("备注匹配方式不支持")
        if not safe_str(remark_pattern):
            raise ValueError("备注条件不能为空")
        return remark_pattern

    @staticmethod
    def _normalize_rule_exclude_config(remark_exclude_pattern: str | None) -> str:
        return safe_str(remark_exclude_pattern).strip()

    @staticmethod
    async def _load_rule_candidates(db: AsyncSession, *, platform_code: str | None) -> list[TransactionRuleCandidate]:
        result = await db.execute(
            select(TransactionRule)
            .where(
                TransactionRule.status == 1,
                TransactionRule.is_deleted.is_(False),
                or_(TransactionRule.platform_code.is_(None), TransactionRule.platform_code == platform_code),
            )
            .order_by(TransactionRule.priority, TransactionRule.id)
        )
        return [
            TransactionRuleCandidate(
                id=rule.id,
                subject_id=rule.subject_id,
                category_id=rule.category_id,
                transaction_direction=rule.transaction_direction,
                direction_field=rule.direction_field,
                scene_field="动账场景",
                remark_field=rule.remark_field,
                match_type=rule.match_type,
                transaction_scene=rule.transaction_scene,
                remark_pattern=rule.remark_pattern,
                remark_exclude_pattern=rule.remark_exclude_pattern,
                amount_field=rule.amount_field,
                result_direction=rule.result_direction,
                priority=rule.priority,
            )
            for rule in result.scalars().all()
        ]

    @staticmethod
    def _load_douyin_dongzhang_rows_from_oss(upload_file: TransactionUploadFile) -> tuple[list[str], list[dict[str, object]]]:
        if upload_file.platform_code != "douyin":
            raise ValueError("动账核算 v1 仅支持抖音动账")
        suffix = Path(upload_file.original_name).suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=suffix) as tmp:
            oss_service.download_to_temp(upload_file.oss_key, tmp.name)
            with open_tabular_rows(tmp.name) as rows:
                if rows is None:
                    raise ValueError("无法打开表格文件")
                row_iter = iter(rows)
                required_headers = DouyinProcessor.summary_strategy.required_headers
                header_result = FinancialSummaryExcelProcessorMixin._find_header_row(row_iter, required_headers)
                if header_result is None:
                    raise ValueError("无法读取抖音动账表头")
                header, _header_row_number = header_result
                col_idx = FinancialSummaryExcelProcessorMixin._build_col_idx(header, required_headers)
                missing = FinancialSummaryExcelProcessorMixin._missing_headers(col_idx, required_headers)
                if missing:
                    raise ValueError(f"缺少抖音动账必要表头: {', '.join(missing)}")
                data_rows = [FinancialSummaryExcelProcessorMixin._row_to_values(row, col_idx) for row in row_iter if any(safe_str(cell) for cell in row)]
                return header, data_rows

    @staticmethod
    def _rows_to_dicts(header: list[str], rows: list[list[object]]) -> list[dict[str, object]]:
        normalized_header = [safe_str(value) for value in header]
        data: list[dict[str, object]] = []
        for row in rows:
            item = {}
            for index, field in enumerate(normalized_header):
                if not field:
                    continue
                item[field] = row[index] if index < len(row) else None
            if item:
                data.append(item)
        return data

    @staticmethod
    def _first_existing_field(row: dict[str, object], candidates: list[str], *, default: str) -> str:
        for candidate in candidates:
            if candidate in row:
                return candidate
        return default

    @staticmethod
    def _resolve_transaction_period(row: dict[str, object]) -> tuple[int | None, int | None, str | None]:
        action_time = parse_datetime(row.get("动账时间"))
        if action_time is None:
            return None, None, "动账时间无法解析"
        return action_time.year, action_time.month, None

    @staticmethod
    def _period_failed_result(*, evaluation: TransactionEvaluationResult, message: str | None) -> TransactionEvaluationResult:
        if evaluation.status == "unmatched":
            return evaluation
        return TransactionEvaluationResult(
            row_number=evaluation.row_number,
            status="failed",
            rule_id=evaluation.rule_id,
            subject_id=evaluation.subject_id,
            category_id=evaluation.category_id,
            amount_field=evaluation.amount_field,
            original_amount=Decimal("0"),
            calculated_amount=Decimal("0"),
            error_message=message or "动账时间无法解析",
        )

    @staticmethod
    def _build_transaction_row_error_message(
        *,
        row: dict[str, object],
        evaluation: TransactionEvaluationResult,
        direction_field: str,
        scene_field: str,
        remark_field: str,
        period_error: str | None,
    ) -> str:
        reason = safe_str(evaluation.error_message)
        if evaluation.status == "unmatched" and reason == "未匹配到动账核算规则":
            reason = "未匹配分类（未命中启用规则）"
        if period_error and period_error not in reason:
            reason = f"{period_error}；{reason}" if reason else period_error

        parts = [f"第 {evaluation.row_number} 行：{reason or '处理失败'}"]
        direction = safe_str(row.get(direction_field))
        scene = safe_str(row.get(scene_field))
        remark = safe_str(row.get(remark_field))
        if direction:
            parts.append(f"{direction_field}={direction}")
        if scene:
            parts.append(f"{scene_field}={scene}")
        if remark:
            parts.append(f"{remark_field}={remark}")
        if evaluation.amount_field:
            raw_amount = safe_str(row.get(evaluation.amount_field))
            if raw_amount:
                parts.append(f"{evaluation.amount_field}={raw_amount}")
            else:
                parts.append(f"取数字段={evaluation.amount_field}")
        return "；".join(parts)

    @staticmethod
    def _short_log_text(value: str, *, limit: int = 120) -> str:
        text = safe_str(value)
        if len(text) <= limit:
            return text
        return f"{text[:limit]}..."

    @staticmethod
    def _log_rule_result_summaries(
        *,
        task: TransactionTask,
        upload_file: TransactionUploadFile,
        rule_map: dict[int, TransactionRuleCandidate],
        rule_accumulator: dict[tuple[int, int | None, int | None], dict[str, object]],
    ) -> None:
        for (rule_id, accounting_year, accounting_month), bucket in sorted(rule_accumulator.items()):
            rule = rule_map.get(rule_id)
            if rule is None:
                continue
            logger.info(
                "transaction_accounting.rule_result task_id=%s file_id=%s platform=%s shop=%s rule_id=%s subject_id=%s category_id=%s direction=%s scene=%s match_type=%s remark=%s amount_field=%s result_direction=%s accounting_period=%s row_count=%s total_amount=%s first_row=%s",
                task.id,
                upload_file.id,
                upload_file.platform_code,
                upload_file.shop_name,
                rule.id,
                rule.subject_id,
                rule.category_id,
                rule.transaction_direction,
                safe_str(rule.transaction_scene),
                rule.match_type,
                TransactionAccountingService._short_log_text(rule.remark_pattern),
                rule.amount_field,
                rule.result_direction,
                TransactionAccountingService._format_month(accounting_year, accounting_month),
                int(bucket["count"]),
                bucket["amount"],
                int(bucket["first_row"]),
            )

    @staticmethod
    def _json_safe_raw_row(row: dict[str, object]) -> dict[str, object]:
        return {str(key): TransactionAccountingService._json_safe_value(value) for key, value in row.items()}

    @staticmethod
    def _json_safe_value(value: object) -> object:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, Decimal):
            return str(value)
        if isinstance(value, dict):
            return {str(key): TransactionAccountingService._json_safe_value(item) for key, item in value.items()}
        if isinstance(value, (list, tuple)):
            return [TransactionAccountingService._json_safe_value(item) for item in value]
        return value

    @staticmethod
    def _split_filter_values(value: str | None) -> list[str]:
        if not value:
            return []
        return [item.strip() for item in value.split(",") if item.strip()]

    @staticmethod
    def _split_int_filter_values(value: int | str | None) -> list[int]:
        if value is None or value == "":
            return []
        if isinstance(value, int):
            return [value]
        values: list[int] = []
        for item in str(value).split(","):
            item = item.strip()
            if not item:
                continue
            values.append(int(item))
        return values

    @staticmethod
    async def _get_latest_task_for_upload_file(db: AsyncSession, *, file_id: int) -> TransactionTask | None:
        stmt = (
            select(TransactionTask)
            .where(
                TransactionTask.file_id == file_id,
                TransactionTask.is_deleted.is_(False),
            )
            .order_by(TransactionTask.id.desc())
        )
        return (await db.execute(stmt)).scalars().first()

    @staticmethod
    def _latest_result_task_ids_select():
        return (
            select(func.max(TransactionTask.id).label("task_id"))
            .select_from(TransactionTask)
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionTask.file_id)
            .where(
                TransactionTask.is_deleted.is_(False),
                TransactionUploadFile.is_deleted.is_(False),
                TransactionTask.status.in_(TRANSACTION_RESULT_TASK_STATUSES),
            )
            .group_by(
                TransactionUploadFile.org_id,
                TransactionUploadFile.platform_code,
                TransactionUploadFile.shop_id,
                TransactionUploadFile.accounting_year,
                TransactionUploadFile.accounting_month,
            )
        )

    @staticmethod
    def _cash_flow_group_subquery():
        parent = aliased(CashFlowItem)
        return (
            select(
                CashFlowItem.name.label("subject_name"),
                func.min(parent.name).label("cash_flow_group_name"),
            )
            .join(parent, parent.id == CashFlowItem.parent_id)
            .where(
                CashFlowItem.is_deleted.is_(False),
                parent.is_deleted.is_(False),
                CashFlowItem.level == 2,
                parent.level == 1,
            )
            .group_by(CashFlowItem.name)
            .subquery()
        )

    @staticmethod
    async def _load_cash_flow_items(db: AsyncSession) -> list[CashFlowItem]:
        result = await db.execute(select(CashFlowItem).where(CashFlowItem.is_deleted.is_(False), CashFlowItem.status == 1).order_by(CashFlowItem.sort_order, CashFlowItem.id))
        return list(result.scalars().all())

    @staticmethod
    async def _load_annual_summary_amounts(
        db: AsyncSession,
        *,
        user: User,
        year: int,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        major_category_id: int | str | None = None,
        subject_id: int | str | None = None,
        category_id: int | str | None = None,
        transaction_direction: str | None = None,
        upload_accounting_year: int | None = None,
        upload_accounting_month: int | None = None,
        upload_accounting_start_year: int | None = None,
        upload_accounting_start_month: int | None = None,
        upload_accounting_end_year: int | None = None,
        upload_accounting_end_month: int | None = None,
        keyword: str | None = None,
    ) -> list[AnnualSummaryAmount]:
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        filters = [
            TransactionSummaryRow.is_deleted.is_(False),
            TransactionUploadFile.is_deleted.is_(False),
            TransactionSummaryRow.accounting_year == year,
            TransactionSummaryRow.accounting_month.is_not(None),
        ]
        if org_ids is not None:
            filters.append(TransactionSummaryRow.org_id.in_(org_ids))
        if task_id is not None:
            filters.append(TransactionSummaryRow.task_id == task_id)
        else:
            filters.append(TransactionSummaryRow.task_id.in_(TransactionAccountingService._latest_result_task_ids_select()))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(TransactionSummaryRow.platform_code.in_(platform_codes))
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(TransactionSummaryRow.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(TransactionSummaryRow.shop_id.in_(shop_id_list))
        major_category_ids = TransactionAccountingService._split_int_filter_values(major_category_id)
        if major_category_ids:
            filters.append(TransactionSummaryRow.major_category_id.in_(major_category_ids))
        subject_ids = TransactionAccountingService._split_int_filter_values(subject_id)
        if subject_ids:
            filters.append(TransactionSummaryRow.subject_id.in_(subject_ids))
        category_ids = TransactionAccountingService._split_int_filter_values(category_id)
        if category_ids:
            filters.append(TransactionSummaryRow.category_id.in_(category_ids))
        if transaction_direction:
            filters.append(TransactionSummaryRow.transaction_direction == transaction_direction)
        if upload_accounting_year is not None:
            filters.append(TransactionUploadFile.accounting_year == upload_accounting_year)
        if upload_accounting_month is not None:
            filters.append(TransactionUploadFile.accounting_month == upload_accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                TransactionUploadFile.accounting_year * 100 + TransactionUploadFile.accounting_month,
                start_year=upload_accounting_start_year,
                start_month=upload_accounting_start_month,
                end_year=upload_accounting_end_year,
                end_month=upload_accounting_end_month,
            )
        )
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    TransactionSummaryRow.major_category_name.ilike(like_pattern),
                    TransactionSummaryRow.subject_name.ilike(like_pattern),
                    TransactionSummaryRow.category_name.ilike(like_pattern),
                    TransactionSummaryRow.shop_name.ilike(like_pattern),
                    TransactionSummaryRow.platform_code.ilike(like_pattern),
                )
            )

        stmt = (
            select(
                TransactionSummaryRow.cash_flow_item_id,
                TransactionSummaryRow.accounting_month,
                func.sum(TransactionSummaryRow.total_amount).label("total_amount"),
            )
            .join(TransactionUploadFile, TransactionUploadFile.id == TransactionSummaryRow.file_id)
            .where(*filters, TransactionSummaryRow.cash_flow_item_id.is_not(None))
            .group_by(TransactionSummaryRow.cash_flow_item_id, TransactionSummaryRow.accounting_month)
        )
        result = await db.execute(stmt)
        return [
            AnnualSummaryAmount(
                cash_flow_item_id=int(row.cash_flow_item_id),
                accounting_month=int(row.accounting_month),
                total_amount=row.total_amount or Decimal("0"),
            )
            for row in result.all()
            if (
                row.cash_flow_item_id is not None
                and row.accounting_month is not None
                and 1 <= int(row.accounting_month) <= 12
            )
        ]

    @staticmethod
    def _build_annual_report_rows(
        *,
        year: int,
        cash_items: list[CashFlowItem],
        amounts: list[AnnualSummaryAmount],
    ) -> list[TransactionAnnualSummaryReportRow]:
        month_keys = TransactionAccountingService._annual_month_keys(year)
        amounts_by_cash_item_id: dict[int, dict[str, Decimal]] = defaultdict(
            lambda: {month: Decimal("0") for month in month_keys}
        )
        for amount in amounts:
            month_key = f"{year}{int(amount.accounting_month):02d}"
            if month_key in amounts_by_cash_item_id[amount.cash_flow_item_id]:
                amounts_by_cash_item_id[amount.cash_flow_item_id][month_key] += amount.total_amount or Decimal("0")

        active_cash_items = [item for item in cash_items if not getattr(item, "is_deleted", False)]
        children_by_parent: dict[int, list[CashFlowItem]] = defaultdict(list)
        for item in active_cash_items:
            if item.parent_id is not None:
                children_by_parent[int(item.parent_id)].append(item)

        rows_by_id: dict[int, TransactionAnnualSummaryReportRow] = {}
        ordered_items = sorted(active_cash_items, key=lambda item: (item.sort_order, item.id))
        for item in ordered_items:
            months = {month: Decimal("0") for month in month_keys}
            if item.level == 2:
                months = dict(amounts_by_cash_item_id.get(int(item.id), months))
            row = TransactionAnnualSummaryReportRow(
                code=item.code,
                name=item.name,
                parent_code=TransactionAccountingService._cash_parent_code(item, active_cash_items),
                level=item.level,
                item_type=item.item_type,
                flow_section=item.flow_section,
                flow_direction=item.flow_direction,
                summary_method=item.summary_method,
                months=months,
                total_amount=sum(months.values(), Decimal("0")),
            )
            rows_by_id[item.id] = row

        for item in sorted(ordered_items, key=lambda cash_item: (cash_item.level, cash_item.sort_order), reverse=True):
            row = rows_by_id[item.id]
            if item.summary_method != "sum_children":
                row.total_amount = sum(row.months.values(), Decimal("0"))
                continue
            children = children_by_parent.get(item.id, [])
            child_rows = [rows_by_id[child.id] for child in children if child.id in rows_by_id]
            if child_rows:
                for month in month_keys:
                    row.months[month] = sum((child.months.get(month, Decimal("0")) for child in child_rows), Decimal("0"))
            row.total_amount = sum(row.months.values(), Decimal("0"))

        TransactionAccountingService._apply_annual_formula_rows(
            month_keys=month_keys,
            ordered_items=ordered_items,
            rows_by_id=rows_by_id,
            children_by_parent=children_by_parent,
        )
        return [rows_by_id[item.id] for item in ordered_items]

    @staticmethod
    def _apply_annual_formula_rows(
        *,
        month_keys: list[str],
        ordered_items: list[CashFlowItem],
        rows_by_id: dict[int, TransactionAnnualSummaryReportRow],
        children_by_parent: dict[int, list[CashFlowItem]],
    ) -> None:
        for net_item in ordered_items:
            if net_item.summary_method != "formula" or net_item.flow_direction != "net" or net_item.parent_id is not None:
                continue
            related_groups = [
                item
                for item in ordered_items
                if item.parent_id is None
                and item.flow_section == net_item.flow_section
                and item.flow_direction in {"inflow", "outflow"}
                and item.id in rows_by_id
                and item.sort_order < net_item.sort_order
            ]
            if not related_groups:
                continue
            net_months: dict[str, Decimal] = {}
            for month in month_keys:
                inflow = sum(
                    (rows_by_id[item.id].months.get(month, Decimal("0")) for item in related_groups if item.flow_direction == "inflow"),
                    Decimal("0"),
                )
                outflow = sum(
                    (rows_by_id[item.id].months.get(month, Decimal("0")) for item in related_groups if item.flow_direction == "outflow"),
                    Decimal("0"),
                )
                net_months[month] = inflow - outflow

            net_row = rows_by_id[net_item.id]
            net_row.months = net_months
            net_row.total_amount = sum(net_months.values(), Decimal("0"))

            formula_children = [
                child for child in children_by_parent.get(net_item.id, []) if child.summary_method == "formula" and child.flow_direction == "net" and child.id in rows_by_id
            ]
            if formula_children:
                target_child = formula_children[-1]
                child_row = rows_by_id[target_child.id]
                child_row.months = dict(net_months)
                child_row.total_amount = net_row.total_amount

    @staticmethod
    def _cash_parent_code(item: CashFlowItem, cash_items: list[CashFlowItem]) -> str | None:
        if item.parent_id is None:
            return None
        parent_by_id = {cash_item.id: cash_item for cash_item in cash_items}
        parent = parent_by_id.get(item.parent_id)
        return parent.code if parent is not None else None

    @staticmethod
    def _annual_month_keys(year: int) -> list[str]:
        return [f"{year}{month:02d}" for month in range(1, 13)]

    @staticmethod
    def _period_filters(
        period_expr,
        *,
        start_year: int | None = None,
        start_month: int | None = None,
        end_year: int | None = None,
        end_month: int | None = None,
    ) -> list:
        filters = []
        if start_year is not None and start_month is not None:
            filters.append(period_expr >= int(start_year) * 100 + int(start_month))
        if end_year is not None and end_month is not None:
            filters.append(period_expr <= int(end_year) * 100 + int(end_month))
        return filters

    @staticmethod
    def _format_month(year: int | None, month: int | None) -> str:
        if not year or not month:
            return ""
        return f"{int(year)}-{int(month):02d}"

    @staticmethod
    def _format_datetime(value: datetime | None) -> str:
        if value is None:
            return ""
        if value.tzinfo:
            value = value.astimezone()
        return value.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _format_platform(platform_code: str | None) -> str:
        if not platform_code:
            return ""
        key = platform_code.strip()
        return TRANSACTION_PLATFORM_LABELS.get(key) or TRANSACTION_PLATFORM_LABELS.get(key.lower(), key)

    @staticmethod
    def _write_only_header_row(sheet, headers: list[str]) -> list[WriteOnlyCell]:
        cells: list[WriteOnlyCell] = []
        for label in headers:
            cell = WriteOnlyCell(sheet, value=label)
            cell.font = Font(bold=True)
            cells.append(cell)
        return cells

    @staticmethod
    def _build_detail_workbook(rows: list[TransactionDetail]) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="动账汇总明细")
        worksheet.append(TransactionAccountingService._write_only_header_row(worksheet, DETAIL_EXPORT_HEADERS))
        for index, detail in enumerate(rows, start=1):
            worksheet.append(
                [
                    index,
                    TransactionAccountingService._format_month(detail.accounting_year, detail.accounting_month),
                    TransactionAccountingService._format_platform(detail.platform_code),
                    detail.shop_name or "",
                    getattr(detail, "cash_flow_group_name", None) or "",
                    getattr(detail, "subject_name", None) or "",
                    getattr(detail, "category_name", None) or "",
                    float(detail.calculated_amount or 0),
                ]
            )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_summary_workbook(rows: list[TransactionSummaryRow]) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="动账汇总报表")
        worksheet.append(TransactionAccountingService._write_only_header_row(worksheet, SUMMARY_EXPORT_HEADERS))
        for index, row in enumerate(rows, start=1):
            worksheet.append(
                [
                    index,
                    TransactionAccountingService._format_month(
                        getattr(row, "upload_accounting_year", None),
                        getattr(row, "upload_accounting_month", None),
                    ),
                    TransactionAccountingService._format_month(row.accounting_year, row.accounting_month),
                    TransactionAccountingService._format_platform(row.platform_code),
                    row.shop_name or "",
                    getattr(row, "cash_flow_group_name", None) or "",
                    row.subject_name,
                    float(row.total_amount or 0),
                ]
            )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_annual_summary_workbook(*, year: int, rows: list[TransactionAnnualSummaryReportRow]) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title=f"{year}资金报表")
        month_keys = TransactionAccountingService._annual_month_keys(year)
        worksheet.append(TransactionAccountingService._write_only_header_row(worksheet, ["序号", f"{year}年", *month_keys, "合计"]))
        for row in rows:
            worksheet.append(
                [
                    row.code,
                    row.name,
                    *[float(row.months.get(month, Decimal("0")) or Decimal("0")) for month in month_keys],
                    float(row.total_amount or Decimal("0")),
                ]
            )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def _create_aggregated_result_rows(
        db: AsyncSession,
        *,
        task: TransactionTask,
        upload_file: TransactionUploadFile,
        summary_accumulator: dict[tuple[int, int, int | None, int | None], dict[str, object]],
    ) -> None:
        if not summary_accumulator:
            return
        subject_ids = {subject_id for subject_id, _, _, _ in summary_accumulator}
        category_ids = {category_id for _, category_id, _, _ in summary_accumulator}
        subject_result = await db.execute(select(TransactionSubject).where(TransactionSubject.id.in_(subject_ids)))
        category_result = await db.execute(select(TransactionCategory).where(TransactionCategory.id.in_(category_ids)))
        subjects = {item.id: item for item in subject_result.scalars().all()}
        categories = {item.id: item for item in category_result.scalars().all()}
        major_category_ids = {
            int(item.major_category_id)
            for item in subjects.values()
            if item.major_category_id is not None
        }
        major_category_names: dict[int, str] = {}
        if major_category_ids:
            major_category_result = await db.execute(
                select(TransactionMajorCategory.id, TransactionMajorCategory.name).where(
                    TransactionMajorCategory.id.in_(major_category_ids)
                )
            )
            major_category_names = {
                int(item_id): item_name
                for item_id, item_name in major_category_result.all()
            }
        detail_rows = []
        summary_rows = []
        for (subject_id, category_id, accounting_year, accounting_month), bucket in summary_accumulator.items():
            subject = subjects[subject_id]
            category = categories[category_id]
            major_category_name = (
                major_category_names.get(int(subject.major_category_id))
                if subject.major_category_id is not None
                else None
            )
            amount = bucket["amount"]
            row_count = int(bucket["count"])
            first_row = int(bucket["first_row"])
            logger.info(
                "transaction_accounting.category_result task_id=%s file_id=%s platform=%s shop=%s subject_id=%s subject=%s category_id=%s category=%s accounting_period=%s row_count=%s total_amount=%s first_row=%s",
                task.id,
                upload_file.id,
                upload_file.platform_code,
                upload_file.shop_name,
                subject_id,
                subject.name,
                category_id,
                category.name,
                TransactionAccountingService._format_month(accounting_year, accounting_month),
                row_count,
                amount,
                first_row,
            )
            detail_rows.append(
                TransactionDetail(
                    task_id=task.id,
                    file_id=upload_file.id,
                    org_id=task.org_id,
                    major_category_id=subject.major_category_id,
                    major_category_name=major_category_name,
                    subject_id=subject_id,
                    category_id=category_id,
                    rule_id=None,
                    row_number=first_row,
                    shop_id=upload_file.shop_id,
                    platform_code=upload_file.platform_code,
                    shop_name=upload_file.shop_name,
                    accounting_year=accounting_year,
                    accounting_month=accounting_month,
                    transaction_direction=None,
                    remark="聚合明细",
                    amount_field="汇总金额",
                    original_amount=amount,
                    calculated_amount=amount,
                    status="matched",
                    error_message=None,
                    raw_row={
                        "明细类型": "聚合明细",
                        "原始匹配明细数": row_count,
                        "核算年月": TransactionAccountingService._format_month(upload_file.accounting_year, upload_file.accounting_month),
                        "业务年月": TransactionAccountingService._format_month(accounting_year, accounting_month),
                    },
                )
            )
            summary_rows.append(
                TransactionSummaryRow(
                    task_id=task.id,
                    file_id=upload_file.id,
                    org_id=task.org_id,
                    major_category_id=subject.major_category_id,
                    major_category_name=major_category_name,
                    subject_id=subject_id,
                    category_id=category_id,
                    cash_flow_item_id=subject.cash_flow_item_id,
                    subject_name=subject.name,
                    category_name=category.name,
                    transaction_direction=None,
                    shop_id=upload_file.shop_id,
                    platform_code=upload_file.platform_code,
                    shop_name=upload_file.shop_name,
                    accounting_year=accounting_year,
                    accounting_month=accounting_month,
                    row_count=row_count,
                    total_amount=amount,
                )
            )
        db.add_all(detail_rows + summary_rows)
