"""Service layer for Douyin merchant reconciliation."""

from __future__ import annotations

import io
import hashlib
import re
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import exists, func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.douyin_dongzhang_detail import DouyinDongzhangDetail
from app.models.merchant_reconciliation import (
    MerchantBankFlowFile,
    MerchantBankFlowRow,
    MerchantOpeningBalance,
    MerchantRedSheet,
    MerchantRedSheetPayment,
    MerchantRedSheetPurchase,
)
from app.models.organization import Organization
from app.models.shop import Shop
from app.models.summary import FinancialSummary
from app.models.user import User
from app.schemas.merchant_reconciliation import (
    MerchantBankFlowImportResult,
    MerchantOpeningBalanceBatchResult,
    MerchantOpeningBalanceBatchUpsert,
    MerchantRedSheetImportResult,
    MerchantReconciliationStatsOut,
)
from app.services.audit_service import AuditService
from app.services.shop_visibility import active_shop_filter
from app.services.shop_service import ShopService
from app.tasks.processors.base import canonical_header, open_tabular_rows, parse_datetime, safe_str
from app.utils.live_code import extract_live_code
from app.utils.money import ZERO_MONEY, safe_decimal
from app.utils.product_code import extract_product_code
from app.utils.query_filters import resolve_org_ids, split_int_filter_values

PURCHASE_HEADERS = [
    "直播间",
    "商家",
    "直播日期",
    "借/退货单号",
    "借/退货日期",
    "直播编号",
    "匹配",
    "备注",
    "店铺",
    "主体",
    "摘要",
    "货品名称",
    "件/元",
    "克/元",
    "卖价",
    "借货数量",
    "借货重量g",
    "借货金额",
    "退货数量",
    "退货重量g",
    "退货金额",
    "预计结款日期",
    "工费/克",
    "工费/件",
]

PAYMENT_HEADERS = [
    "序号",
    "直播间",
    "直播日期",
    "商家",
    "借货总金额",
    "退货总金额",
    "冲减业务费用",
    "冲减金额",
    "应付货款金额",
    "退货率",
    "结算主体",
    "收款主体",
    "成本主体",
    "应付款金额",
    "主体回款金额",
    "收款商家",
    "回款商家",
    "是否已结款",
    "是否已回款",
    "结算状态",
    "付款截图",
    "结算日期",
    "回款日期",
    "冲减备注",
    "待解决事项",
    "是否修改收款商家",
    "是否修改收款金额",
    "修改月份",
    "申请日期",
    "已付",
    "借-退",
]

RECONCILIATION_EXPORT_COLUMNS: tuple[tuple[str, str, bool], ...] = (
    ("platform_label", "平台", False),
    ("shop_name", "店铺", False),
    ("accounting_date", "业务年月", False),
    ("transaction_time", "动账时间", False),
    ("transaction_flow_no", "动账流水号", False),
    ("transaction_direction", "动账方向", False),
    ("transaction_amount", "动账金额", True),
    ("transaction_scene", "动账场景", False),
    ("sub_order_no", "子订单号", False),
    ("order_no", "订单号", False),
    ("order_time", "下单时间", False),
    ("product_id", "商品ID", False),
    ("product_code", "商品编码", False),
    ("product_name", "商品名称", False),
    ("author_name", "达人名称", False),
    ("gmv", "实收GMV", True),
    ("allocated_bic", "分摊BIC", True),
    ("allocated_insurance_fee", "分摊运费险", True),
    ("live_amount", "直播款", True),
    ("major_merchant_name", "大商家名称", False),
    ("merchant_receipt_subject", "商家收款主体", False),
    ("receipt_merchant", "收款商家", False),
    ("live_room", "直播间", False),
    ("live_date_text", "直播日期", False),
    ("match_status", "匹配状态", False),
    ("match_error", "匹配失败原因", False),
)

SUMMARY_EXPORT_COLUMNS: tuple[tuple[str, str, bool], ...] = (
    ("org_name", "组织", False),
    ("accounting_date", "业务年月", False),
    ("our_subject", "我方主体", False),
    ("merchant_receipt_subject", "商家收款主体", False),
    ("gmv", "实收GMV", True),
    ("merchant_payable_net_amount", "应付商家净额", True),
    ("opening_balance", "期初余额", True),
    ("business_fee_deduction", "冲减业务费用", True),
    ("other_deduction_amount", "其他冲减金额", True),
    ("payable_goods_balance", "应付货款余额", True),
    ("paid_flow_amount", "已付流水", True),
    ("unpaid_flow_amount", "未付流水", True),
    ("bank_flow_amount", "银行流水汇总", True),
    ("bank_payment_diff", "银行付款差", True),
    ("bank_status", "银行流水状态", False),
)

MONEY_QUANT = Decimal("0.01")
PAYMENT_SETTLEMENT_STATUSES = ("已核算", "已结算", "未结算")
PAYMENT_SETTLEMENT_STATUS_SET = set(PAYMENT_SETTLEMENT_STATUSES)
PAYMENT_SUMMARY_SETTLED_STATUS = "已结算"
MERCHANT_BANK_FLOW_TYPE = "银行流水"
MERCHANT_BANK_FLOW_FILENAME_PATTERN = re.compile(
    r"(?P<year>\d{4})(?:\s*年\s*|[-/.])?(?P<month>\d{1,2})(?:\s*月)?[_\-\s]+银行流水",
    re.IGNORECASE,
)
COMPACT_BANK_DATETIME_PATTERN = re.compile(r"^(?P<date>\d{8})\s+(?P<time>\d{1,2}:\d{2}(?::\d{2})?)$")
LIVE_DATE_YMD_RANGE_PATTERN = re.compile(
    r"(?P<year>\d{4})[-/.年](?P<month>\d{1,2})[-/.月](?P<start>\d{1,2})\s*(?:日)?"
    r"(?:\s*[-~至到—－]\s*(?:(?P<end_year>\d{4})[-/.年])?(?:(?P<end_month>\d{1,2})[-/.月])?(?P<end>\d{1,2})\s*(?:日)?)?"
)
LIVE_DATE_RANGE_PATTERN = re.compile(
    r"(?P<month>\d{1,2})\s*月\s*(?P<start>\d{1,2})\s*(?:日)?"
    r"\s*(?:[-~至到—－]\s*(?:(?P<end_month>\d{1,2})\s*月\s*)?(?P<end>\d{1,2})\s*日?)?"
)
LIVE_DATE_ISO_PATTERN = re.compile(
    r"(?P<year>\d{4})[-/.年](?P<month>\d{1,2})[-/.月](?P<day>\d{1,2})"
)
SQL_PRODUCT_CODE_EXTRACT_PATTERN = (
    r"(CAN[0-9]+(?:\+CAN[0-9]+)*|T[0-9]{4,}|V[0-9]+(?:-[A-Z0-9]+)?|"
    r"RE[0-9]+(?:-[A-Z0-9]+)?|[A-Z]{1,2}[0-9]{4,}(?:-[A-Z0-9]+)?)"
)


@dataclass(frozen=True)
class _RedSheetContext:
    purchases_by_code: dict[str, MerchantRedSheetPurchase]
    payments_by_key: dict[tuple[str, str, str], MerchantRedSheetPayment]


@dataclass(frozen=True)
class _ReconciliationLoadContext:
    org_id: int
    total_gmv: Decimal
    total_bic: Decimal
    total_insurance: Decimal
    red_sheet_context: _RedSheetContext


@dataclass(frozen=True)
class MerchantDerivedFields:
    product_code: str
    major_merchant_name: str
    our_subject: str
    merchant_receipt_subject: str
    receipt_merchant: str
    live_room: str
    live_date: str
    live_date_text: str
    red_sheet_payment_id: int | None
    allocated_bic: Decimal
    allocated_insurance_fee: Decimal
    live_amount: Decimal
    match_status: str
    match_error: str


def _period(year: int, month: int) -> int:
    return int(year) * 100 + int(month)


def _normalize_text(value: object) -> str:
    return safe_str(value).strip()


def _match_text(value: object) -> str:
    return _normalize_text(value).upper()


def _entity_match_text(value: object) -> str:
    return (
        _normalize_text(value)
        .replace("（", "(")
        .replace("）", ")")
        .replace(" ", "")
        .replace("　", "")
        .upper()
    )


def _date_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalize_text(value)


def _date_key(value: object) -> str:
    return MerchantReconciliationService.normalize_live_date_text(value) or _date_text(value)


def _parse_date(value: object) -> date | None:
    parsed = parse_datetime(value)
    if parsed is not None:
        return parsed.date()
    return None


def _money(value: object) -> Decimal:
    return safe_decimal(value).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _allocation_amount(row_gmv: Decimal, total_gmv: Decimal, total_amount: Decimal) -> Decimal:
    if total_gmv == ZERO_MONEY or total_amount == ZERO_MONEY:
        return ZERO_MONEY
    return (row_gmv * total_amount / total_gmv).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _write_only_header_row(sheet, headers: list[str]) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for label in headers:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = Font(bold=True)
        cells.append(cell)
    return cells


class MerchantReconciliationService:
    @staticmethod
    def build_red_sheet_template(*, accounting_year: int, accounting_month: int) -> io.BytesIO:
        workbook = Workbook()
        payment_sheet = workbook.active
        payment_sheet.title = f"{accounting_year}{accounting_month:02d}货款"
        purchase_sheet = workbook.create_sheet(f"{accounting_year}{accounting_month:02d}采购")
        MerchantReconciliationService._style_template_sheet(payment_sheet, PAYMENT_HEADERS)
        MerchantReconciliationService._style_template_sheet(purchase_sheet, PURCHASE_HEADERS)

        note_sheet = workbook.create_sheet("导入说明")
        note_rows = [
            ("月份", f"当前模板月份为 {accounting_year}-{accounting_month:02d}，上传时也会按该月份校验 sheet 名。"),
            ("Sheet", "必须包含 YYYYMM货款、YYYYMM采购 两个 sheet。"),
            ("匹配", "商品编码从抖音动账商品名称提取，再匹配采购 sheet 的直播编号。"),
            ("收款", "商家收款主体按商家 + 直播日期 + 直播间匹配货款 sheet。"),
        ]
        for row_index, (title, text_value) in enumerate(note_rows, start=1):
            note_sheet.cell(row=row_index, column=1, value=title).font = Font(bold=True)
            note_sheet.cell(row=row_index, column=2, value=text_value)
        note_sheet.column_dimensions["A"].width = 14
        note_sheet.column_dimensions["B"].width = 90

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _style_template_sheet(worksheet, headers: list[str]) -> None:
        required_fill = PatternFill("solid", fgColor="FCE7F3")
        header_font = Font(bold=True, color="111827")
        for index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=index, value=header)
            cell.font = header_font
            cell.fill = required_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.column_dimensions[cell.column_letter].width = max(12, min(24, len(header) + 6))
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(headers)).coordinate}"

    @staticmethod
    async def import_red_sheet(
        db: AsyncSession,
        *,
        content: bytes,
        filename: str,
        accounting_year: int,
        accounting_month: int,
        operator: User,
        org_id: str | int | None = None,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> MerchantRedSheetImportResult:
        MerchantReconciliationService._validate_month(accounting_year, accounting_month)
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("请上传 .xlsx 或 .xlsm 格式的红单文件")
        if not content:
            raise ValueError("红单文件不能为空")
        import_org_id = MerchantReconciliationService._resolve_import_org_id(operator=operator, requested_org_id=org_id)

        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("红单文件无法读取，请下载模板重新填写") from exc

        period = _period(accounting_year, accounting_month)
        payment_sheet_name = f"{period}货款"
        purchase_sheet_name = f"{period}采购"
        if payment_sheet_name not in workbook.sheetnames:
            raise ValueError(f"红单缺少 sheet：{payment_sheet_name}")
        if purchase_sheet_name not in workbook.sheetnames:
            raise ValueError(f"红单缺少 sheet：{purchase_sheet_name}")

        purchases = MerchantReconciliationService._parse_purchase_sheet(workbook[purchase_sheet_name])
        payments = MerchantReconciliationService._parse_payment_sheet(workbook[payment_sheet_name])
        shop_mapping = await MerchantReconciliationService._resolve_import_shops(
            db,
            org_id=import_org_id,
            operator=operator,
            purchases=purchases,
        )
        purchase_shop_lookup = MerchantReconciliationService._build_purchase_shop_lookup(purchases)
        payment_shop_assignments: dict[int, str] = {}
        warnings: list[str] = []
        for payment_row in payments:
            # 货款 sheet 没有独立店铺列，这里回用采购 sheet 的匹配键来推断所属店铺。
            payment_shop_name = MerchantReconciliationService._match_payment_shop_name(payment_row, purchase_shop_lookup)
            if not payment_shop_name:
                warnings.append(MerchantReconciliationService._payment_shop_warning(payment_row))
                continue
            payment_shop_assignments[id(payment_row)] = payment_shop_name

        file_hash = hashlib.sha256(content).hexdigest()
        now = datetime.now(timezone.utc)

        shop_ids = [shop.id for shop in shop_mapping.values()]
        overlapping_red_sheet = or_(
            MerchantRedSheet.file_hash == file_hash,
            MerchantRedSheet.shop_id.in_(shop_ids),
            exists(
                select(1).where(
                    MerchantRedSheetPurchase.red_sheet_id == MerchantRedSheet.id,
                    MerchantRedSheetPurchase.shop_id.in_(shop_ids),
                    MerchantRedSheetPurchase.is_deleted.is_(False),
                )
            ),
            exists(
                select(1).where(
                    MerchantRedSheetPayment.red_sheet_id == MerchantRedSheet.id,
                    MerchantRedSheetPayment.shop_id.in_(shop_ids),
                    MerchantRedSheetPayment.is_deleted.is_(False),
                )
            ),
        )
        existing_rows = (
            await db.execute(
                select(MerchantRedSheet).where(
                    MerchantRedSheet.org_id == import_org_id,
                    MerchantRedSheet.platform_code == "douyin",
                    MerchantRedSheet.accounting_period == period,
                    MerchantRedSheet.is_deleted.is_(False),
                    overlapping_red_sheet,
                )
            )
        ).scalars().all()
        existing_ids: list[int] = []
        for existing in existing_rows:
            # 同期间重新导入时，旧批次与其明细整体软删除，只保留最新红单结果。
            existing_ids.append(existing.id)
            existing.is_deleted = True
            existing.deleted_at = now
        if existing_ids:
            await db.execute(
                update(MerchantRedSheetPurchase)
                .where(
                    MerchantRedSheetPurchase.red_sheet_id.in_(existing_ids),
                    MerchantRedSheetPurchase.is_deleted.is_(False),
                )
                .values(is_deleted=True, deleted_at=now)
            )
            await db.execute(
                update(MerchantRedSheetPayment)
                .where(
                    MerchantRedSheetPayment.red_sheet_id.in_(existing_ids),
                    MerchantRedSheetPayment.is_deleted.is_(False),
                )
                .values(is_deleted=True, deleted_at=now)
            )

        red_sheet = MerchantRedSheet(
            org_id=import_org_id,
            user_id=operator.id,
            shop_id=None,
            platform_code="douyin",
            shop_name="多店铺" if len(shop_mapping) > 1 else next(iter(shop_mapping.values())).shop_name,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            accounting_period=period,
            original_name=filename,
            file_size=len(content),
            file_hash=file_hash,
            status="success",
            purchase_rows=len(purchases),
            payment_rows=len(payments),
            result_summary={
                "采购行数": len(purchases),
                "货款行数": len(payments),
                "店铺数": len(shop_mapping),
                "店铺": [shop.shop_name for shop in shop_mapping.values()],
                "warnings": warnings[:50],
            },
        )
        db.add(red_sheet)
        await db.flush()

        db.add_all(
            [
                MerchantReconciliationService._purchase_model(
                    row,
                    red_sheet_id=red_sheet.id,
                    org_id=shop_mapping[_normalize_text(row.get("source_shop_name"))].org_id,
                    shop_id=shop_mapping[_normalize_text(row.get("source_shop_name"))].id,
                    shop_name=shop_mapping[_normalize_text(row.get("source_shop_name"))].shop_name,
                    accounting_period=period,
                )
                for row in purchases
            ]
        )
        payment_models: list[MerchantRedSheetPayment] = []
        for row in payments:
            payment_shop_name = payment_shop_assignments.get(id(row))
            if payment_shop_name:
                payment_shop = shop_mapping[payment_shop_name]
                payment_models.append(
                    MerchantReconciliationService._payment_model(
                        row,
                        red_sheet_id=red_sheet.id,
                        org_id=payment_shop.org_id,
                        shop_id=payment_shop.id,
                        shop_name=payment_shop.shop_name,
                        accounting_period=period,
                    )
                )
                continue
            payment_models.append(
                MerchantReconciliationService._payment_model(
                    row,
                    red_sheet_id=red_sheet.id,
                    org_id=import_org_id,
                    shop_id=None,
                    shop_name="",
                    accounting_period=period,
                )
            )
        db.add_all(payment_models)

        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=import_org_id,
            module="merchant_reconciliation",
            action="import",
            description=f"导入商家对账红单 [{filename}]，采购 {len(purchases)} 行，货款 {len(payments)} 行",
            target_type="merchant_red_sheet",
            target_id=red_sheet.id,
            target_name=filename,
            ip=ip,
            user_agent=user_agent,
            new_value={
                "shop_ids": shop_ids,
                "shop_names": [shop.shop_name for shop in shop_mapping.values()],
                "accounting_year": accounting_year,
                "accounting_month": accounting_month,
                "purchase_rows": len(purchases),
                "payment_rows": len(payments),
                "warnings": warnings[:50],
            },
        )
        await db.flush()
        return MerchantRedSheetImportResult(
            red_sheet_id=red_sheet.id,
            purchase_rows=len(purchases),
            payment_rows=len(payments),
            errors=[],
            warnings=warnings,
        )

    @staticmethod
    async def import_bank_flow(
        db: AsyncSession,
        *,
        file_path: str | Path,
        filename: str,
        operator: User,
        org_id: str | int | None = None,
        file_size: int | None = None,
        file_hash: str | None = None,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> MerchantBankFlowImportResult:
        parsed_month = MerchantReconciliationService.parse_bank_flow_filename(filename)
        if parsed_month is None:
            raise ValueError("银行流水文件名必须符合：YYYYMM_银行流水_xxxx")
        accounting_year, accounting_month = parsed_month
        MerchantReconciliationService._validate_month(accounting_year, accounting_month)
        import_org_id = MerchantReconciliationService._resolve_import_org_id(operator=operator, requested_org_id=org_id)
        rows, bank_name, account_name = MerchantReconciliationService._parse_bank_flow_file(
            file_path=file_path,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        if not rows:
            raise ValueError("银行流水文件没有可导入明细")
        if not account_name:
            raise ValueError("银行流水未识别到账户名称")

        period = _period(accounting_year, accounting_month)
        now = datetime.now(timezone.utc)
        existing_rows = (
            await db.execute(
                select(MerchantBankFlowFile).where(
                    MerchantBankFlowFile.org_id == import_org_id,
                    MerchantBankFlowFile.accounting_period == period,
                    MerchantBankFlowFile.account_name == account_name,
                    MerchantBankFlowFile.is_deleted.is_(False),
                )
            )
        ).scalars().all()
        existing_ids: list[int] = []
        for existing in existing_rows:
            existing_ids.append(existing.id)
            existing.is_deleted = True
            existing.deleted_at = now
        if existing_ids:
            await db.execute(
                update(MerchantBankFlowRow)
                .where(
                    MerchantBankFlowRow.bank_flow_file_id.in_(existing_ids),
                    MerchantBankFlowRow.is_deleted.is_(False),
                )
                .values(is_deleted=True, deleted_at=now)
            )

        matched_row_count = sum(1 for row in rows if _normalize_text(row.get("live_date")))
        bank_flow_file = MerchantBankFlowFile(
            org_id=import_org_id,
            user_id=operator.id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            accounting_period=period,
            original_name=filename,
            file_size=int(file_size or Path(file_path).stat().st_size),
            file_hash=file_hash,
            bank_name=bank_name,
            account_name=account_name,
            status="success",
            row_count=len(rows),
            matched_row_count=matched_row_count,
            result_summary={
                "银行": bank_name,
                "账户名称": account_name,
                "流水行数": len(rows),
                "已解析直播日期": matched_row_count,
            },
        )
        db.add(bank_flow_file)
        await db.flush()
        db.add_all(
            [
                MerchantReconciliationService._bank_flow_row_model(
                    row,
                    bank_flow_file_id=bank_flow_file.id,
                    org_id=import_org_id,
                    accounting_period=period,
                )
                for row in rows
            ]
        )
        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=import_org_id,
            module="merchant_reconciliation",
            action="import_bank_flow",
            description=f"导入商家对账银行流水 [{filename}]，共 {len(rows)} 行",
            target_type="merchant_bank_flow",
            target_id=bank_flow_file.id,
            target_name=filename,
            ip=ip,
            user_agent=user_agent,
            new_value={
                "accounting_year": accounting_year,
                "accounting_month": accounting_month,
                "bank_name": bank_name,
                "account_name": account_name,
                "row_count": len(rows),
                "matched_row_count": matched_row_count,
            },
        )
        await db.flush()
        return MerchantBankFlowImportResult(
            bank_flow_file_id=bank_flow_file.id,
            row_count=len(rows),
            matched_row_count=matched_row_count,
            errors=[],
            warnings=[],
        )

    @staticmethod
    def parse_bank_flow_filename(filename: str) -> tuple[int, int] | None:
        match = MERCHANT_BANK_FLOW_FILENAME_PATTERN.search(Path(filename).stem)
        if not match:
            return None
        year = int(match.group("year"))
        month = int(match.group("month"))
        if month < 1 or month > 12:
            return None
        return year, month

    @staticmethod
    def _resolve_import_org_id(*, operator: User, requested_org_id: str | int | None) -> int:
        if operator.role == "superadmin":
            resolved = requested_org_id if requested_org_id is not None else operator.org_id
            if resolved is None:
                raise ValueError("请选择组织后导入红单")
            return int(resolved)
        if operator.org_id is None:
            raise ValueError("当前账号没有所属组织，无法导入")
        if requested_org_id is not None and int(requested_org_id) != int(operator.org_id):
            raise ValueError("无权导入其他组织的红单")
        return int(operator.org_id)

    @staticmethod
    async def _resolve_import_shops(
        db: AsyncSession,
        *,
        org_id: int,
        operator: User,
        purchases: list[dict[str, object]],
    ) -> dict[str, Shop]:
        missing_rows = [
            int(row.get("source_row_number") or 0)
            for row in purchases
            if not _normalize_text(row.get("source_shop_name"))
        ]
        if missing_rows:
            raise ValueError(f"采购 sheet 第 {missing_rows[0]} 行缺少店铺")
        shop_names = {
            _normalize_text(row.get("source_shop_name"))
            for row in purchases
        }
        if not shop_names:
            raise ValueError("采购 sheet 缺少店铺信息，无法识别店铺")
        mappings: dict[str, Shop] = {}
        for source_shop_name in shop_names:
            shop = await MerchantReconciliationService._resolve_single_import_shop(
                db,
                org_id=org_id,
                operator=operator,
                source_shop_name=source_shop_name,
            )
            mappings[source_shop_name] = shop
        return mappings

    @staticmethod
    async def _resolve_single_import_shop(
        db: AsyncSession,
        *,
        org_id: int,
        operator: User,
        source_shop_name: str,
    ) -> Shop:
        _ = operator
        stmt = select(Shop).where(
            Shop.is_deleted.is_(False),
            Shop.org_id == org_id,
            func.lower(Shop.shop_name) == source_shop_name.lower(),
            Shop.platform_name.in_(["抖音", "douyin", "抖店"]),
        )
        candidates = list((await db.execute(stmt)).scalars().all())
        if not candidates:
            return await ShopService.get_or_create_shop(
                db,
                org_id=org_id,
                platform_name="douyin",
                shop_name=source_shop_name,
            )
        if len(candidates) > 1:
            raise ValueError(f"店铺【{source_shop_name}】匹配到多个店铺，请先清理店铺管理数据")
        return candidates[0]

    @staticmethod
    def _build_purchase_shop_lookup(
        purchases: list[dict[str, object]],
    ) -> dict[tuple[str, str, str], str]:
        lookup: dict[tuple[str, str, str], str] = {}
        for row in purchases:
            shop_name = _normalize_text(row.get("source_shop_name"))
            if not shop_name:
                continue
            # 采购 sheet 的店铺信息最稳定，后续货款匹配也统一基于这组键值。
            key = (
                _match_text(row.get("merchant")),
                _date_key(row.get("live_date")),
                _match_text(row.get("live_room")),
            )
            if key != ("", "", ""):
                lookup.setdefault(key, shop_name)
        return lookup

    @staticmethod
    def _match_payment_shop_name(
        payment_row: dict[str, object],
        purchase_shop_lookup: dict[tuple[str, str, str], str],
    ) -> str:
        key = (
            _match_text(payment_row.get("merchant")),
            _date_key(payment_row.get("live_date")),
            _match_text(payment_row.get("live_room")),
        )
        return purchase_shop_lookup.get(key, "")

    @staticmethod
    def _payment_shop_warning(payment_row: dict[str, object]) -> str:
        row_number = int(payment_row.get("source_row_number") or 0)
        merchant = _normalize_text(payment_row.get("merchant")) or "-"
        live_room = _normalize_text(payment_row.get("live_room")) or "-"
        return (
            f"货款 sheet 第 {row_number} 行未识别到店铺："
            f"商家 {merchant}，直播间 {live_room}"
        )

    @staticmethod
    async def list_red_sheets(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        shop_ids: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[tuple[MerchantRedSheet, str | None, str | None]], int]:
        filters = [MerchantRedSheet.is_deleted.is_(False), active_shop_filter(MerchantRedSheet.shop_id)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(MerchantRedSheet.org_id.in_(org_ids))
        shop_id_values = split_int_filter_values(shop_ids)
        if shop_id_values:
            filters.append(
                or_(
                    MerchantRedSheet.shop_id.in_(shop_id_values),
                    exists(
                        select(1).where(
                            MerchantRedSheetPurchase.red_sheet_id == MerchantRedSheet.id,
                            MerchantRedSheetPurchase.shop_id.in_(shop_id_values),
                            MerchantRedSheetPurchase.is_deleted.is_(False),
                        )
                    ),
                    exists(
                        select(1).where(
                            MerchantRedSheetPayment.red_sheet_id == MerchantRedSheet.id,
                            MerchantRedSheetPayment.shop_id.in_(shop_id_values),
                            MerchantRedSheetPayment.is_deleted.is_(False),
                        )
                    ),
                )
            )
        if accounting_year is not None:
            filters.append(MerchantRedSheet.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(MerchantRedSheet.accounting_month == accounting_month)
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(or_(MerchantRedSheet.original_name.ilike(like), MerchantRedSheet.shop_name.ilike(like)))

        total = (await db.execute(select(func.count()).select_from(MerchantRedSheet).where(*filters))).scalar() or 0
        rows = (
            await db.execute(
                select(MerchantRedSheet, Organization.name.label("org_name"), Shop.shop_color.label("shop_color"))
                .outerjoin(Organization, Organization.id == MerchantRedSheet.org_id)
                .outerjoin(Shop, Shop.id == MerchantRedSheet.shop_id)
                .where(*filters)
                .order_by(MerchantRedSheet.created_at.desc(), MerchantRedSheet.id.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [(row[0], row[1], row[2]) for row in rows], int(total)

    @staticmethod
    async def list_payment_details(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        shop_ids: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict[str, object]], int]:
        filters = MerchantReconciliationService._red_sheet_detail_filters(
            user=user,
            model=MerchantRedSheetPayment,
            org_id=org_id,
            shop_ids=shop_ids,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    MerchantRedSheetPayment.live_room.ilike(like),
                    MerchantRedSheetPayment.merchant.ilike(like),
                    MerchantRedSheetPayment.receipt_subject.ilike(like),
                    MerchantRedSheetPayment.receipt_merchant.ilike(like),
                    MerchantRedSheetPayment.collection_merchant.ilike(like),
                    MerchantRedSheetPayment.remark.ilike(like),
                )
            )
        total = (await db.execute(select(func.count()).select_from(MerchantRedSheetPayment).where(*filters))).scalar() or 0
        rows = (
            await db.execute(
                select(
                    MerchantRedSheetPayment,
                    Organization.name.label("org_name"),
                    Shop.shop_color.label("shop_color"),
                )
                .outerjoin(Organization, Organization.id == MerchantRedSheetPayment.org_id)
                .outerjoin(Shop, Shop.id == MerchantRedSheetPayment.shop_id)
                .where(*filters)
                .order_by(
                    MerchantRedSheetPayment.accounting_period.desc(),
                    MerchantRedSheetPayment.source_row_number.asc(),
                    MerchantRedSheetPayment.id.asc(),
                )
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [
            MerchantReconciliationService._red_sheet_detail_payload(
                item,
                org_name=org_name,
                shop_color=shop_color,
            )
            for item, org_name, shop_color in rows
        ], int(total)

    @staticmethod
    async def list_purchase_details(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        shop_ids: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict[str, object]], int]:
        filters = MerchantReconciliationService._red_sheet_detail_filters(
            user=user,
            model=MerchantRedSheetPurchase,
            org_id=org_id,
            shop_ids=shop_ids,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    MerchantRedSheetPurchase.live_room.ilike(like),
                    MerchantRedSheetPurchase.merchant.ilike(like),
                    MerchantRedSheetPurchase.live_code.ilike(like),
                    MerchantRedSheetPurchase.loan_return_order_no.ilike(like),
                    MerchantRedSheetPurchase.product_name.ilike(like),
                    MerchantRedSheetPurchase.remark.ilike(like),
                )
            )
        total = (await db.execute(select(func.count()).select_from(MerchantRedSheetPurchase).where(*filters))).scalar() or 0
        rows = (
            await db.execute(
                select(
                    MerchantRedSheetPurchase,
                    Organization.name.label("org_name"),
                    Shop.shop_color.label("shop_color"),
                )
                .outerjoin(Organization, Organization.id == MerchantRedSheetPurchase.org_id)
                .outerjoin(Shop, Shop.id == MerchantRedSheetPurchase.shop_id)
                .where(*filters)
                .order_by(
                    MerchantRedSheetPurchase.accounting_period.desc(),
                    MerchantRedSheetPurchase.source_row_number.asc(),
                    MerchantRedSheetPurchase.id.asc(),
                )
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [
            MerchantReconciliationService._red_sheet_detail_payload(
                item,
                org_name=org_name,
                shop_color=shop_color,
            )
            for item, org_name, shop_color in rows
        ], int(total)

    @staticmethod
    async def list_bank_flow_files(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[tuple[MerchantBankFlowFile, str | None]], int]:
        filters = [MerchantBankFlowFile.is_deleted.is_(False)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(MerchantBankFlowFile.org_id.in_(org_ids))
        if accounting_year is not None:
            filters.append(MerchantBankFlowFile.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(MerchantBankFlowFile.accounting_month == accounting_month)
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    MerchantBankFlowFile.original_name.ilike(like),
                    MerchantBankFlowFile.account_name.ilike(like),
                    MerchantBankFlowFile.bank_name.ilike(like),
                )
            )

        total = (await db.execute(select(func.count()).select_from(MerchantBankFlowFile).where(*filters))).scalar() or 0
        rows = (
            await db.execute(
                select(MerchantBankFlowFile, Organization.name.label("org_name"))
                .outerjoin(Organization, Organization.id == MerchantBankFlowFile.org_id)
                .where(*filters)
                .order_by(MerchantBankFlowFile.created_at.desc(), MerchantBankFlowFile.id.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [(row[0], row[1]) for row in rows], int(total)

    @staticmethod
    async def list_bank_flow_rows(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict[str, object]], int]:
        filters = [MerchantBankFlowRow.is_deleted.is_(False)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(MerchantBankFlowRow.org_id.in_(org_ids))
        if accounting_year is not None and accounting_month is not None:
            filters.append(MerchantBankFlowRow.accounting_period == _period(accounting_year, accounting_month))
        elif accounting_year is not None:
            filters.append(MerchantBankFlowRow.accounting_period.between(_period(accounting_year, 1), _period(accounting_year, 12)))
        elif accounting_month is not None:
            filters.append((MerchantBankFlowRow.accounting_period % 100) == accounting_month)
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    MerchantBankFlowRow.account_name.ilike(like),
                    MerchantBankFlowRow.counterparty_name.ilike(like),
                    MerchantBankFlowRow.purpose.ilike(like),
                    MerchantBankFlowRow.remark.ilike(like),
                    MerchantBankFlowRow.summary.ilike(like),
                    MerchantBankFlowRow.transaction_flow_no.ilike(like),
                )
            )

        total = (await db.execute(select(func.count()).select_from(MerchantBankFlowRow).where(*filters))).scalar() or 0
        rows = (
            await db.execute(
                select(MerchantBankFlowRow, Organization.name.label("org_name"))
                .outerjoin(Organization, Organization.id == MerchantBankFlowRow.org_id)
                .where(*filters)
                .order_by(
                    MerchantBankFlowRow.accounting_period.desc(),
                    MerchantBankFlowRow.transaction_time.desc().nullslast(),
                    MerchantBankFlowRow.source_row_number.asc(),
                    MerchantBankFlowRow.id.asc(),
                )
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [
            MerchantReconciliationService._bank_flow_row_payload(row, org_name=org_name)
            for row, org_name in rows
        ], int(total)

    @staticmethod
    async def list_summary(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None = None,
        org_id: str | int | None = None,
        keyword: str | None = None,
        bank_status: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict[str, object]], int]:
        summary_rows = await MerchantReconciliationService._load_summary_rows(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
            keyword=keyword,
            bank_status=bank_status,
            include_opening_balances=True,
        )
        total = len(summary_rows)
        paged_rows = summary_rows[(page - 1) * page_size : page * page_size]
        return paged_rows, total

    @staticmethod
    async def _load_summary_rows(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None = None,
        org_id: str | int | None = None,
        keyword: str | None = None,
        bank_status: str | None = None,
        include_opening_balances: bool = True,
        include_bank_flows: bool = True,
    ) -> list[dict[str, object]]:
        grouped_rows = await MerchantReconciliationService._load_matched_summary_groups(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
        )
        detail_totals, payment_group_weights = MerchantReconciliationService._collect_summary_detail_totals(grouped_rows)
        payment_adjustments = await MerchantReconciliationService._load_payment_adjustments(
            db,
            user=user,
            org_id=org_id,
            shop_id=shop_id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            payment_group_weights=payment_group_weights,
            payment_ids=payment_group_weights.keys(),
        )
        summary_rows = MerchantReconciliationService._build_summary_rows_from_aggregates(
            detail_totals=detail_totals,
            payment_adjustments=payment_adjustments,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        if include_opening_balances:
            opening_balance_rows = await MerchantReconciliationService._load_opening_balance_records(
                db,
                user=user,
                org_id=org_id,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
            )
            MerchantReconciliationService._merge_opening_balance_rows(
                summary_rows,
                opening_balance_rows,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                append_missing=shop_id is None,
            )
        if include_bank_flows:
            bank_flow_totals = await MerchantReconciliationService._load_bank_flow_summary_totals(
                db,
                user=user,
                org_id=org_id,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                payment_group_weights=payment_group_weights,
            )
            MerchantReconciliationService._merge_bank_flow_totals(summary_rows, bank_flow_totals)
        if keyword:
            keyword_text = keyword.strip().lower()
            summary_rows = [
                row
                for row in summary_rows
                if keyword_text in str(row.get("our_subject") or "").lower()
                or keyword_text in str(row.get("merchant_receipt_subject") or "").lower()
            ]
        summary_rows = MerchantReconciliationService._filter_summary_rows_by_bank_status(summary_rows, bank_status=bank_status)
        return MerchantReconciliationService._sort_summary_rows(summary_rows)

    @staticmethod
    async def list_opening_balances_for_summary(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None = None,
        org_id: str | int | None = None,
        keyword: str | None = None,
    ) -> list[dict[str, object]]:
        summary_rows = await MerchantReconciliationService._load_summary_rows(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
            include_opening_balances=False,
            include_bank_flows=False,
        )
        opening_balance_rows = await MerchantReconciliationService._load_opening_balance_records(
            db,
            user=user,
            org_id=org_id,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        opening_by_key = {
            MerchantReconciliationService._opening_balance_key(
                org_id=balance.org_id,
                our_subject=balance.our_subject,
                receipt_merchant=balance.receipt_merchant,
            ): (balance, org_name)
            for balance, org_name in opening_balance_rows
        }
        accounting_period = _period(accounting_year, accounting_month)
        rows: list[dict[str, object]] = []
        used_keys: set[tuple[int, str, str]] = set()
        for summary_row in summary_rows:
            summary_org_id = MerchantReconciliationService._optional_int(summary_row.get("org_id"))
            if summary_org_id is None:
                continue
            our_subject = _normalize_text(summary_row.get("our_subject"))
            receipt_merchant = _normalize_text(summary_row.get("merchant_receipt_subject") or summary_row.get("receipt_merchant"))
            if not our_subject or not receipt_merchant:
                continue
            key = MerchantReconciliationService._opening_balance_key(
                org_id=summary_org_id,
                our_subject=our_subject,
                receipt_merchant=receipt_merchant,
            )
            balance_pair = opening_by_key.get(key)
            balance = balance_pair[0] if balance_pair else None
            rows.append(
                {
                    "id": balance.id if balance is not None else None,
                    "org_id": summary_org_id,
                    "org_name": summary_row.get("org_name"),
                    "platform_code": "douyin",
                    "accounting_year": accounting_year,
                    "accounting_month": accounting_month,
                    "accounting_period": accounting_period,
                    "our_subject": our_subject,
                    "receipt_merchant": receipt_merchant,
                    "opening_balance": safe_decimal(balance.opening_balance if balance is not None else ZERO_MONEY),
                    "remark": safe_str(balance.remark if balance is not None else ""),
                    "updated_at": balance.updated_at if balance is not None else None,
                }
            )
            used_keys.add(key)

        if shop_id is None:
            for balance, org_name in opening_balance_rows:
                key = MerchantReconciliationService._opening_balance_key(
                    org_id=balance.org_id,
                    our_subject=balance.our_subject,
                    receipt_merchant=balance.receipt_merchant,
                )
                if key in used_keys:
                    continue
                rows.append(MerchantReconciliationService._opening_balance_payload(balance, org_name=org_name))

        if keyword:
            keyword_text = keyword.strip().lower()
            rows = [
                row
                for row in rows
                if keyword_text in str(row.get("our_subject") or "").lower()
                or keyword_text in str(row.get("receipt_merchant") or "").lower()
            ]
        return sorted(
            rows,
            key=lambda row: (
                str(row.get("org_name") or ""),
                str(row.get("our_subject") or ""),
                str(row.get("receipt_merchant") or ""),
            ),
        )

    @staticmethod
    async def upsert_opening_balances(
        db: AsyncSession,
        *,
        data: MerchantOpeningBalanceBatchUpsert,
        operator: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> MerchantOpeningBalanceBatchResult:
        MerchantReconciliationService._validate_month(data.accounting_year, data.accounting_month)
        accounting_period = _period(data.accounting_year, data.accounting_month)
        platform_code = _normalize_text(data.platform_code) or "douyin"
        allowed_org_ids = resolve_org_ids(user_role=operator.role, user_org_id=operator.org_id)

        deduped_items: dict[tuple[int, str, str], object] = {}
        for item in data.items:
            if allowed_org_ids is not None and item.org_id not in allowed_org_ids:
                raise ValueError("不能维护其他组织的期初金额")
            key = (int(item.org_id), _normalize_text(item.our_subject), _normalize_text(item.receipt_merchant))
            deduped_items[key] = item
        if not deduped_items:
            return MerchantOpeningBalanceBatchResult(created_count=0, updated_count=0, total_count=0)

        org_ids = sorted({key[0] for key in deduped_items})
        existing_rows = (
            await db.execute(
                select(MerchantOpeningBalance).where(
                    MerchantOpeningBalance.org_id.in_(org_ids),
                    MerchantOpeningBalance.platform_code == platform_code,
                    MerchantOpeningBalance.accounting_period == accounting_period,
                    MerchantOpeningBalance.is_deleted.is_(False),
                )
            )
        ).scalars().all()
        existing_by_key = {
            (int(row.org_id), _normalize_text(row.our_subject), _normalize_text(row.receipt_merchant)): row
            for row in existing_rows
        }

        created_count = 0
        updated_count = 0
        for key, item in deduped_items.items():
            org_id_value, our_subject, receipt_merchant = key
            opening_balance = _money(item.opening_balance)
            remark = safe_str(item.remark)
            existing = existing_by_key.get(key)
            if existing is not None:
                existing.opening_balance = opening_balance
                existing.remark = remark
                existing.updated_by = operator.id
                updated_count += 1
                continue
            if opening_balance == ZERO_MONEY and not remark:
                continue
            db.add(
                MerchantOpeningBalance(
                    org_id=org_id_value,
                    platform_code=platform_code,
                    accounting_year=data.accounting_year,
                    accounting_month=data.accounting_month,
                    accounting_period=accounting_period,
                    our_subject=our_subject,
                    receipt_merchant=receipt_merchant,
                    opening_balance=opening_balance,
                    remark=remark,
                    created_by=operator.id,
                    updated_by=operator.id,
                )
            )
            created_count += 1

        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=operator.org_id,
            module="merchant_reconciliation",
            action="upsert_opening_balance",
            description=f"维护商家对账期初金额 {data.accounting_year}-{data.accounting_month:02d}，新增 {created_count} 条，更新 {updated_count} 条",
            target_type="merchant_opening_balance",
            target_name=f"{data.accounting_year}{data.accounting_month:02d}",
            ip=ip,
            user_agent=user_agent,
            new_value={
                "platform_code": platform_code,
                "accounting_year": data.accounting_year,
                "accounting_month": data.accounting_month,
                "created_count": created_count,
                "updated_count": updated_count,
                "total_count": created_count + updated_count,
            },
        )
        await db.flush()
        return MerchantOpeningBalanceBatchResult(
            created_count=created_count,
            updated_count=updated_count,
            total_count=created_count + updated_count,
        )

    @staticmethod
    async def export_summary(
        db: AsyncSession,
        *,
        user: User,
        output_path: Path | None = None,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None = None,
        org_id: str | int | None = None,
        keyword: str | None = None,
        bank_status: str | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ) -> io.BytesIO | int:
        rows, _total = await MerchantReconciliationService.list_summary(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
            keyword=keyword,
            bank_status=bank_status,
            page=page or 1,
            page_size=page_size or 1000,
        )
        if output_path is not None:
            return MerchantReconciliationService._save_summary_workbook(rows, output_path=output_path)
        return MerchantReconciliationService._build_summary_workbook(rows)

    @staticmethod
    async def export_summary_to_file(db: AsyncSession, *, user: User, output_path: Path, **kwargs) -> int:
        result = await MerchantReconciliationService.export_summary(db, user=user, output_path=output_path, **kwargs)
        return int(result)

    @staticmethod
    async def list_details(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int,
        org_id: str | int | None = None,
        keyword: str | None = None,
        match_status: str | None = None,
        ids: list[int] | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[dict[str, object]], int, MerchantReconciliationStatsOut]:
        rows, total, stats = await MerchantReconciliationService._load_reconciliation_rows(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
            keyword=keyword,
            match_status=match_status,
            ids=ids,
            page=page,
            page_size=page_size,
            include_all=False,
        )
        return rows, total, stats

    @staticmethod
    def _red_sheet_detail_filters(
        *,
        user: User,
        model: type[MerchantRedSheetPayment] | type[MerchantRedSheetPurchase],
        org_id: str | int | None,
        shop_ids: str | int | None,
        accounting_year: int | None,
        accounting_month: int | None,
    ) -> list[Any]:
        filters: list[Any] = [model.is_deleted.is_(False), active_shop_filter(model.shop_id)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(model.org_id.in_(org_ids))
        shop_id_values = split_int_filter_values(shop_ids)
        if shop_id_values:
            filters.append(model.shop_id.in_(shop_id_values))
        if accounting_year is not None and accounting_month is not None:
            filters.append(model.accounting_period == _period(accounting_year, accounting_month))
        elif accounting_year is not None:
            filters.append(model.accounting_period.between(_period(accounting_year, 1), _period(accounting_year, 12)))
        elif accounting_month is not None:
            filters.append((model.accounting_period % 100) == accounting_month)
        return filters

    @staticmethod
    def _red_sheet_detail_payload(
        item: MerchantRedSheetPayment | MerchantRedSheetPurchase,
        *,
        org_name: str | None,
        shop_color: str | None,
    ) -> dict[str, object]:
        payload = dict(item.__dict__)
        payload.pop("_sa_instance_state", None)
        payload["org_name"] = org_name
        payload["shop_name"] = safe_str(payload.get("shop_name"))
        payload["shop_color"] = shop_color
        if isinstance(item, MerchantRedSheetPayment):
            payload["settlement_status"] = safe_str(payload.get("remark"))
        return payload

    @staticmethod
    def _empty_summary_row(
        *,
        org_name: str | None,
        accounting_year: int,
        accounting_month: int,
        our_subject: str,
        receipt_subject: str,
        org_id: int | None = None,
        adjustment: dict[str, Decimal] | None = None,
    ) -> dict[str, object]:
        adjustment = adjustment or {}
        paid_flow_amount = ZERO_MONEY
        bank_flow_amount = ZERO_MONEY
        paid_flow_amount = safe_decimal(adjustment.get("paid_flow_amount"))
        row = {
            "key": f"{org_id or ''}|{our_subject}|{receipt_subject}",
            "org_id": org_id,
            "org_name": org_name,
            "accounting_year": accounting_year,
            "accounting_month": accounting_month,
            "accounting_date": f"{int(accounting_year):04d}-{int(accounting_month):02d}",
            "our_subject": our_subject,
            "merchant_receipt_subject": receipt_subject,
            "receipt_merchant": receipt_subject,
            "gmv": ZERO_MONEY,
            "merchant_payable_net_amount": ZERO_MONEY,
            "opening_balance": ZERO_MONEY,
            "business_fee_deduction": safe_decimal(adjustment.get("business_fee_deduction")),
            "other_deduction_amount": safe_decimal(adjustment.get("other_deduction_amount")),
            "payable_goods_balance": ZERO_MONEY,
            "paid_flow_amount": paid_flow_amount,
            "unpaid_flow_amount": ZERO_MONEY,
            "bank_flow_amount": bank_flow_amount,
            "bank_payment_diff": ZERO_MONEY,
            "row_count": 0,
            "bank_status": "pending",
        }
        MerchantReconciliationService._refresh_summary_flow_amounts(row)
        return row

    @staticmethod
    def _summary_group_key_from_detail_row(row: dict[str, object]) -> tuple[str, str, str]:
        org_id = MerchantReconciliationService._optional_int(row.get("org_id"))
        return (
            str(org_id) if org_id is not None else _normalize_text(row.get("org_name")),
            _normalize_text(row.get("merchant_name") or row.get("our_subject") or row.get("shop_name")),
            _normalize_text(row.get("receipt_merchant") or row.get("merchant_receipt_subject")),
        )

    @staticmethod
    async def _load_matched_summary_groups(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None,
        org_id: str | int | None,
    ) -> list[dict[str, object]]:
        MerchantReconciliationService._validate_month(accounting_year, accounting_month)
        accounting_period = _period(accounting_year, accounting_month)
        filters = [
            DouyinDongzhangDetail.is_deleted.is_(False),
            DouyinDongzhangDetail.source_platform_code == "douyin",
            DouyinDongzhangDetail.summary_year == accounting_year,
            DouyinDongzhangDetail.summary_month == accounting_month,
            active_shop_filter(DouyinDongzhangDetail.shop_id),
        ]
        if shop_id is not None:
            filters.append(DouyinDongzhangDetail.shop_id == shop_id)
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            if not org_ids:
                return []
            filters.append(DouyinDongzhangDetail.org_id.in_(org_ids))

        red_sheet_filters = [
            MerchantRedSheet.is_deleted.is_(False),
            MerchantRedSheet.platform_code == "douyin",
            MerchantRedSheet.accounting_period == accounting_period,
            active_shop_filter(MerchantRedSheet.shop_id),
        ]
        if shop_id is not None:
            red_sheet_filters.append(
                or_(
                    MerchantRedSheet.shop_id == shop_id,
                    exists(
                        select(1).where(
                            MerchantRedSheetPurchase.red_sheet_id == MerchantRedSheet.id,
                            MerchantRedSheetPurchase.shop_id == shop_id,
                            MerchantRedSheetPurchase.is_deleted.is_(False),
                        )
                    ),
                    exists(
                        select(1).where(
                            MerchantRedSheetPayment.red_sheet_id == MerchantRedSheet.id,
                            MerchantRedSheetPayment.shop_id == shop_id,
                            MerchantRedSheetPayment.is_deleted.is_(False),
                        )
                    ),
                )
            )
        if org_ids is not None:
            red_sheet_filters.append(MerchantRedSheet.org_id.in_(org_ids))
        has_red_sheet = (
            await db.execute(
                select(exists(select(1).select_from(MerchantRedSheet).where(*red_sheet_filters)))
            )
        ).scalar()
        if not has_red_sheet:
            return []

        product_code_expr = func.upper(
            func.coalesce(
                func.nullif(DouyinDongzhangDetail.product_code, ""),
                func.substring(DouyinDongzhangDetail.product_name, SQL_PRODUCT_CODE_EXTRACT_PATTERN),
            )
        )
        filters.append(product_code_expr != "")
        grouped_detail_rows = (
            await db.execute(
                select(
                    Organization.name.label("org_name"),
                    DouyinDongzhangDetail.org_id,
                    DouyinDongzhangDetail.shop_id,
                    DouyinDongzhangDetail.shop_name,
                    product_code_expr.label("product_code"),
                    DouyinDongzhangDetail.merchant_name,
                    func.coalesce(func.sum(DouyinDongzhangDetail.gmv), 0).label("gmv"),
                    func.coalesce(func.sum(func.round(DouyinDongzhangDetail.gmv * Decimal("0.7"), 2)), 0).label("live_amount"),
                    func.count(DouyinDongzhangDetail.id).label("row_count"),
                )
                .outerjoin(Organization, Organization.id == DouyinDongzhangDetail.org_id)
                .where(*filters)
                .group_by(
                    Organization.name,
                    DouyinDongzhangDetail.org_id,
                    DouyinDongzhangDetail.shop_id,
                    DouyinDongzhangDetail.shop_name,
                    product_code_expr,
                    DouyinDongzhangDetail.merchant_name,
                )
                .order_by(
                    Organization.name.asc(),
                    DouyinDongzhangDetail.shop_id.asc(),
                    DouyinDongzhangDetail.merchant_name.asc(),
                    product_code_expr.asc(),
                )
            )
        ).all()
        if not grouped_detail_rows:
            return []

        red_sheet_contexts: dict[tuple[int, int], _RedSheetContext] = {}
        for row in grouped_detail_rows:
            if row.org_id is None or row.shop_id is None:
                continue
            context_key = (int(row.org_id), int(row.shop_id))
            if context_key not in red_sheet_contexts:
                red_sheet_contexts[context_key] = await MerchantReconciliationService._load_red_sheet_context(
                    db,
                    org_id=context_key[0],
                    shop_id=context_key[1],
                    accounting_period=accounting_period,
                    payment_settlement_status=PAYMENT_SUMMARY_SETTLED_STATUS,
                )

        rows: list[dict[str, object]] = []
        for row in grouped_detail_rows:
            if row.org_id is None or row.shop_id is None:
                continue
            red_sheet_context = red_sheet_contexts.get((int(row.org_id), int(row.shop_id)))
            if red_sheet_context is None:
                continue
            purchase = next(
                (
                    red_sheet_context.purchases_by_code.get(code)
                    for code in MerchantReconciliationService._split_product_codes(row.product_code)
                    if red_sheet_context.purchases_by_code.get(code)
                ),
                None,
            )
            if purchase is None:
                continue
            payment = red_sheet_context.payments_by_key.get(
                (_match_text(purchase.merchant), _date_key(purchase.live_date), _match_text(purchase.live_room))
            )
            if payment is None or not _normalize_text(payment.receipt_merchant):
                continue
            rows.append(
                {
                    "org_id": int(row.org_id),
                    "org_name": row.org_name,
                    "shop_name": row.shop_name,
                    "merchant_name": row.merchant_name,
                    "our_subject": payment.settlement_subject,
                    "merchant_receipt_subject": payment.receipt_subject,
                    "receipt_merchant": payment.receipt_merchant,
                    "gmv": safe_decimal(row.gmv),
                    "live_amount": safe_decimal(row.live_amount),
                    "row_count": int(row.row_count or 0),
                    "red_sheet_payment_id": int(payment.id) if payment.id is not None else None,
                }
            )
        return rows

    @staticmethod
    def _payment_group_key_from_red_sheet(payment: MerchantRedSheetPayment, *, org_name: str | None) -> tuple[str, str, str]:
        org_id = MerchantReconciliationService._optional_int(getattr(payment, "org_id", None))
        return (
            str(org_id) if org_id is not None else _normalize_text(org_name),
            _normalize_text(payment.settlement_subject),
            _normalize_text(payment.receipt_merchant or payment.receipt_subject),
        )

    @staticmethod
    def _collect_summary_detail_totals(
        rows: Iterable[dict[str, object]],
    ) -> tuple[dict[tuple[str, str, str], dict[str, object]], dict[int, dict[tuple[str, str, str], Decimal]]]:
        detail_totals: dict[tuple[str, str, str], dict[str, object]] = {}
        payment_group_weights: dict[int, dict[tuple[str, str, str], Decimal]] = {}
        for row in rows:
            key = MerchantReconciliationService._summary_group_key_from_detail_row(row)
            item = detail_totals.setdefault(
                key,
                {
                    "org_id": MerchantReconciliationService._optional_int(row.get("org_id")),
                    "org_name": row.get("org_name"),
                    "gmv": ZERO_MONEY,
                    "merchant_payable_net_amount": ZERO_MONEY,
                    "row_count": 0,
                },
            )
            gmv = safe_decimal(row.get("gmv"))
            live_amount = safe_decimal(row.get("live_amount"))
            item["gmv"] = safe_decimal(item["gmv"]) + gmv
            item["merchant_payable_net_amount"] = safe_decimal(item["merchant_payable_net_amount"]) + live_amount
            item["row_count"] = int(item["row_count"]) + int(row.get("row_count") or 1)
            if not item.get("org_name") and row.get("org_name"):
                item["org_name"] = row.get("org_name")

            payment_id = MerchantReconciliationService._optional_int(row.get("red_sheet_payment_id"))
            if payment_id is None:
                continue
            group_weights = payment_group_weights.setdefault(payment_id, {})
            group_weights[key] = group_weights.get(key, ZERO_MONEY) + live_amount
        return detail_totals, payment_group_weights

    @staticmethod
    def _build_summary_rows_from_aggregates(
        *,
        detail_totals: dict[tuple[str, str, str], dict[str, object]],
        payment_adjustments: dict[tuple[str, str, str], dict[str, Decimal]],
        accounting_year: int,
        accounting_month: int,
    ) -> list[dict[str, object]]:
        summary_by_key: dict[tuple[str, str, str], dict[str, object]] = {}
        for key, adjustment in payment_adjustments.items():
            org_token, our_subject, receipt_merchant = key
            row_org_id = MerchantReconciliationService._optional_int(adjustment.get("org_id")) or MerchantReconciliationService._optional_int(org_token)
            org_name = safe_str(adjustment.get("org_name")) or (None if row_org_id is not None else org_token)
            summary_by_key[key] = MerchantReconciliationService._empty_summary_row(
                org_id=row_org_id,
                org_name=org_name or None,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                our_subject=our_subject,
                receipt_subject=receipt_merchant,
                adjustment=adjustment,
            )
        for key, totals in detail_totals.items():
            org_token, our_subject, receipt_merchant = key
            row_org_id = MerchantReconciliationService._optional_int(totals.get("org_id")) or MerchantReconciliationService._optional_int(org_token)
            org_name = safe_str(totals.get("org_name")) or (None if row_org_id is not None else org_token)
            adjustment = payment_adjustments.get(key, {})
            item = summary_by_key.setdefault(
                key,
                MerchantReconciliationService._empty_summary_row(
                    org_id=row_org_id,
                    org_name=org_name or None,
                    accounting_year=accounting_year,
                    accounting_month=accounting_month,
                    our_subject=our_subject,
                    receipt_subject=receipt_merchant,
                    adjustment=adjustment,
                ),
            )
            merchant_payable = safe_decimal(totals.get("merchant_payable_net_amount"))
            if item.get("org_id") is None and row_org_id is not None:
                item["org_id"] = row_org_id
                item["key"] = f"{row_org_id}|{our_subject}|{receipt_merchant}"
            if not item.get("org_name") and org_name:
                item["org_name"] = org_name
            item["gmv"] = safe_decimal(item["gmv"]) + safe_decimal(totals.get("gmv"))
            item["merchant_payable_net_amount"] = safe_decimal(item["merchant_payable_net_amount"]) + merchant_payable
            item["row_count"] = int(item["row_count"]) + int(totals.get("row_count") or 0)
            MerchantReconciliationService._refresh_summary_flow_amounts(item)

        return MerchantReconciliationService._sort_summary_rows(list(summary_by_key.values()))

    @staticmethod
    def _refresh_summary_flow_amounts(item: dict[str, object]) -> None:
        item["payable_goods_balance"] = (
            safe_decimal(item.get("merchant_payable_net_amount"))
            + safe_decimal(item.get("opening_balance"))
            - safe_decimal(item.get("business_fee_deduction"))
            - safe_decimal(item.get("other_deduction_amount"))
        )
        item["unpaid_flow_amount"] = safe_decimal(item["payable_goods_balance"]) - safe_decimal(item["paid_flow_amount"])
        item["bank_payment_diff"] = safe_decimal(item["unpaid_flow_amount"]) - safe_decimal(item["bank_flow_amount"])
        bank_flow_amount = safe_decimal(item["bank_flow_amount"])
        bank_payment_diff = safe_decimal(item["bank_payment_diff"])
        if bank_flow_amount == ZERO_MONEY:
            item["bank_status"] = "pending"
        elif bank_payment_diff == ZERO_MONEY:
            item["bank_status"] = "matched"
        else:
            item["bank_status"] = "diff"

    @staticmethod
    def _filter_summary_rows_by_bank_status(
        rows: list[dict[str, object]],
        *,
        bank_status: str | None,
    ) -> list[dict[str, object]]:
        if not bank_status:
            return rows
        return [row for row in rows if row.get("bank_status") == bank_status]

    @staticmethod
    def _sort_summary_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
        return sorted(
            rows,
            key=lambda item: (
                str(item.get("org_name") or ""),
                str(item.get("our_subject") or ""),
                str(item.get("merchant_receipt_subject") or item.get("receipt_merchant") or ""),
            ),
        )

    @staticmethod
    async def _load_opening_balance_records(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None,
        accounting_year: int,
        accounting_month: int,
    ) -> list[tuple[MerchantOpeningBalance, str | None]]:
        filters = [
            MerchantOpeningBalance.is_deleted.is_(False),
            MerchantOpeningBalance.platform_code == "douyin",
            MerchantOpeningBalance.accounting_period == _period(accounting_year, accounting_month),
        ]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            if not org_ids:
                return []
            filters.append(MerchantOpeningBalance.org_id.in_(org_ids))
        return list(
            (
                await db.execute(
                    select(MerchantOpeningBalance, Organization.name.label("org_name"))
                    .outerjoin(Organization, Organization.id == MerchantOpeningBalance.org_id)
                    .where(*filters)
                )
            ).all()
        )

    @staticmethod
    def _opening_balance_key(*, org_id: int, our_subject: object, receipt_merchant: object) -> tuple[int, str, str]:
        return (
            int(org_id),
            _entity_match_text(our_subject),
            _entity_match_text(receipt_merchant),
        )

    @staticmethod
    def _opening_balance_payload(balance: MerchantOpeningBalance, *, org_name: str | None) -> dict[str, object]:
        return {
            "id": balance.id,
            "org_id": int(balance.org_id),
            "org_name": org_name,
            "platform_code": balance.platform_code,
            "accounting_year": balance.accounting_year,
            "accounting_month": balance.accounting_month,
            "accounting_period": balance.accounting_period,
            "our_subject": balance.our_subject,
            "receipt_merchant": balance.receipt_merchant,
            "opening_balance": safe_decimal(balance.opening_balance),
            "remark": safe_str(balance.remark),
            "updated_at": balance.updated_at,
        }

    @staticmethod
    def _merge_opening_balance_rows(
        summary_rows: list[dict[str, object]],
        opening_balance_rows: list[tuple[MerchantOpeningBalance, str | None]],
        *,
        accounting_year: int,
        accounting_month: int,
        append_missing: bool,
    ) -> None:
        balance_by_key = {
            MerchantReconciliationService._opening_balance_key(
                org_id=balance.org_id,
                our_subject=balance.our_subject,
                receipt_merchant=balance.receipt_merchant,
            ): (balance, org_name)
            for balance, org_name in opening_balance_rows
        }
        used_keys: set[tuple[int, str, str]] = set()
        for row in summary_rows:
            row_org_id = MerchantReconciliationService._optional_int(row.get("org_id"))
            if row_org_id is None:
                continue
            key = MerchantReconciliationService._opening_balance_key(
                org_id=row_org_id,
                our_subject=row.get("our_subject"),
                receipt_merchant=row.get("merchant_receipt_subject") or row.get("receipt_merchant"),
            )
            balance_pair = balance_by_key.get(key)
            if balance_pair is None:
                continue
            balance, _org_name = balance_pair
            row["opening_balance"] = safe_decimal(balance.opening_balance)
            MerchantReconciliationService._refresh_summary_flow_amounts(row)
            used_keys.add(key)

        if not append_missing:
            return
        for balance, org_name in opening_balance_rows:
            key = MerchantReconciliationService._opening_balance_key(
                org_id=balance.org_id,
                our_subject=balance.our_subject,
                receipt_merchant=balance.receipt_merchant,
            )
            if key in used_keys:
                continue
            row = MerchantReconciliationService._empty_summary_row(
                org_id=int(balance.org_id),
                org_name=org_name,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                our_subject=balance.our_subject,
                receipt_subject=balance.receipt_merchant,
            )
            row["opening_balance"] = safe_decimal(balance.opening_balance)
            MerchantReconciliationService._refresh_summary_flow_amounts(row)
            summary_rows.append(row)

    @staticmethod
    async def _load_bank_flow_summary_totals(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None,
        accounting_year: int,
        accounting_month: int,
        payment_group_weights: dict[int, dict[tuple[str, str, str], Decimal]] | None = None,
    ) -> dict[tuple[str, str, str], Decimal]:
        payment_group_weights = payment_group_weights or {}
        accounting_period = _period(accounting_year, accounting_month)
        filters = [
            MerchantBankFlowRow.is_deleted.is_(False),
            MerchantBankFlowRow.accounting_period == accounting_period,
            MerchantBankFlowRow.live_date != "",
            MerchantBankFlowRow.flow_amount > 0,
        ]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(MerchantBankFlowRow.org_id.in_(org_ids))
        bank_rows = (
            await db.execute(
                select(MerchantBankFlowRow, Organization.name.label("org_name"))
                .outerjoin(Organization, Organization.id == MerchantBankFlowRow.org_id)
                .where(*filters)
            )
        ).all()

        payment_filters = [
            MerchantRedSheetPayment.is_deleted.is_(False),
            MerchantRedSheetPayment.accounting_period == accounting_period,
        ]
        if org_ids is not None:
            payment_filters.append(MerchantRedSheetPayment.org_id.in_(org_ids))
        payment_rows = (
            await db.execute(
                select(MerchantRedSheetPayment).where(*payment_filters).order_by(MerchantRedSheetPayment.id.desc())
            )
        ).scalars().all()
        payments_by_bank_key: dict[tuple[int, str, str, str], MerchantRedSheetPayment] = {}
        for payment in payment_rows:
            key = (
                int(payment.org_id),
                _entity_match_text(payment.settlement_subject),
                _entity_match_text(payment.receipt_merchant),
                _date_key(payment.live_date),
            )
            if key[1] and key[2] and key[3]:
                payments_by_bank_key.setdefault(key, payment)

        totals: dict[tuple[str, str, str], Decimal] = {}
        for bank_row, org_name in bank_rows:
            amount = safe_decimal(bank_row.flow_amount)
            payment = payments_by_bank_key.get(
                (
                    int(bank_row.org_id),
                    _entity_match_text(bank_row.account_name),
                    _entity_match_text(bank_row.counterparty_name),
                    _date_key(bank_row.live_date),
                )
            )
            payment_id = MerchantReconciliationService._optional_int(payment.id if payment is not None else None)
            weighted_keys = sorted((payment_group_weights.get(payment_id or 0) or {}).items(), key=lambda item: item[0])
            if weighted_keys:
                allocated = MerchantReconciliationService._allocate_money_by_weight(amount, weighted_keys)
                for key, value in allocated.items():
                    totals[key] = totals.get(key, ZERO_MONEY) + value
                continue
            key = (
                str(int(bank_row.org_id or 0)),
                _entity_match_text(bank_row.account_name),
                _entity_match_text(bank_row.counterparty_name),
            )
            totals[key] = totals.get(key, ZERO_MONEY) + amount
        return totals

    @staticmethod
    def _merge_bank_flow_totals(
        summary_rows: list[dict[str, object]],
        bank_flow_totals: dict[tuple[str, str, str], Decimal],
    ) -> None:
        for row in summary_rows:
            row_org_id = MerchantReconciliationService._optional_int(row.get("org_id"))
            org_token = str(row_org_id) if row_org_id is not None else _normalize_text(row.get("org_name"))
            key = (
                org_token,
                _entity_match_text(row.get("our_subject")),
                _entity_match_text(row.get("merchant_receipt_subject") or row.get("receipt_merchant")),
            )
            row["bank_flow_amount"] = bank_flow_totals.get(key, ZERO_MONEY)
            MerchantReconciliationService._refresh_summary_flow_amounts(row)

    @staticmethod
    def _build_summary_workbook(rows: Iterable[dict[str, object]]) -> io.BytesIO:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="商家对账汇总")
        ws.append(_write_only_header_row(ws, [label for _field, label, _money_flag in SUMMARY_EXPORT_COLUMNS]))
        for row in rows:
            ws.append([MerchantReconciliationService._format_export_value(row.get(field), money=money_flag) for field, _label, money_flag in SUMMARY_EXPORT_COLUMNS])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _save_summary_workbook(rows: Iterable[dict[str, object]], *, output_path: Path) -> int:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="商家对账汇总")
        ws.append(_write_only_header_row(ws, [label for _field, label, _money_flag in SUMMARY_EXPORT_COLUMNS]))
        row_count = 0
        for row in rows:
            ws.append([MerchantReconciliationService._format_export_value(row.get(field), money=money_flag) for field, _label, money_flag in SUMMARY_EXPORT_COLUMNS])
            row_count += 1
        wb.save(output_path)
        return row_count

    @staticmethod
    async def export_details(
        db: AsyncSession,
        *,
        user: User,
        output_path: Path | None = None,
        accounting_year: int,
        accounting_month: int,
        shop_id: int,
        org_id: str | int | None = None,
        keyword: str | None = None,
        match_status: str | None = None,
        ids: list[int] | None = None,
        page: int | None = None,
        page_size: int | None = None,
    ) -> io.BytesIO | int:
        rows, _total, _stats = await MerchantReconciliationService._load_reconciliation_rows(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
            keyword=keyword,
            match_status=match_status,
            ids=ids,
            page=page or 1,
            page_size=page_size or 100,
            include_all=page is None or page_size is None,
        )
        if output_path is not None:
            return MerchantReconciliationService._save_detail_workbook(rows, output_path=output_path)
        return MerchantReconciliationService._build_detail_workbook(rows)

    @staticmethod
    async def export_details_to_file(db: AsyncSession, *, user: User, output_path: Path, **kwargs) -> int:
        result = await MerchantReconciliationService.export_details(db, user=user, output_path=output_path, **kwargs)
        return int(result)

    @staticmethod
    async def _load_reconciliation_rows(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None,
        org_id: str | int | None,
        keyword: str | None,
        match_status: str | None,
        ids: list[int] | None,
        page: int,
        page_size: int,
        include_all: bool,
    ) -> tuple[list[dict[str, object]], int, MerchantReconciliationStatsOut]:
        MerchantReconciliationService._validate_month(accounting_year, accounting_month)
        base_filters = [
            DouyinDongzhangDetail.is_deleted.is_(False),
            DouyinDongzhangDetail.source_platform_code == "douyin",
            DouyinDongzhangDetail.summary_year == accounting_year,
            DouyinDongzhangDetail.summary_month == accounting_month,
            active_shop_filter(DouyinDongzhangDetail.shop_id),
        ]
        if shop_id is not None:
            base_filters.append(DouyinDongzhangDetail.shop_id == shop_id)
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            base_filters.append(DouyinDongzhangDetail.org_id.in_(org_ids))
        filters = list(base_filters)
        if ids is not None:
            filters.append(DouyinDongzhangDetail.id.in_(ids or [-1]))
        if keyword:
            like = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    DouyinDongzhangDetail.transaction_flow_no.ilike(like),
                    DouyinDongzhangDetail.product_code.ilike(like),
                    DouyinDongzhangDetail.product_name.ilike(like),
                    DouyinDongzhangDetail.order_no.ilike(like),
                    DouyinDongzhangDetail.sub_order_no.ilike(like),
                    DouyinDongzhangDetail.author_name.ilike(like),
                )
            )

        total = (await db.execute(select(func.count()).select_from(DouyinDongzhangDetail).where(*filters))).scalar() or 0
        detail_contexts = await MerchantReconciliationService._build_reconciliation_contexts(
            db,
            user=user,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            shop_id=shop_id,
            org_id=org_id,
        )

        dynamic_filter = match_status in {"matched", "unmatched"}
        stmt = (
            select(DouyinDongzhangDetail, Organization.name.label("org_name"), Shop.shop_color.label("shop_color"))
            .outerjoin(Organization, Organization.id == DouyinDongzhangDetail.org_id)
            .outerjoin(Shop, Shop.id == DouyinDongzhangDetail.shop_id)
            .where(*filters)
            .order_by(DouyinDongzhangDetail.source_row_number.asc(), DouyinDongzhangDetail.id.asc())
        )
        if not include_all and not dynamic_filter:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        details = list((await db.execute(stmt)).all())
        loaded_rows = [
            MerchantReconciliationService._build_detail_payload(
                detail,
                org_name=org_name,
                shop_color=shop_color,
                load_context=detail_contexts.get(int(detail.shop_id)) or MerchantReconciliationService._empty_load_context(detail),
            )
            for detail, org_name, shop_color in details
        ]
        if dynamic_filter:
            loaded_rows = [row for row in loaded_rows if row["match_status"] == match_status]
            total = len(loaded_rows)
            rows = loaded_rows if include_all else loaded_rows[(page - 1) * page_size : page * page_size]
            stats_source = loaded_rows
        else:
            rows = loaded_rows
            stats_source = rows if include_all else await MerchantReconciliationService._load_all_payloads_for_stats(
                db,
                filters=filters,
                detail_contexts=detail_contexts,
            )
        stats = MerchantReconciliationService._build_stats(
            stats_source,
            total_bic=sum((ctx.total_bic for ctx in detail_contexts.values()), ZERO_MONEY),
            total_insurance=sum((ctx.total_insurance for ctx in detail_contexts.values()), ZERO_MONEY),
        )
        return rows, int(total), stats

    @staticmethod
    async def _load_all_payloads_for_stats(
        db: AsyncSession,
        *,
        filters: list[Any],
        detail_contexts: dict[int, _ReconciliationLoadContext],
    ) -> list[dict[str, object]]:
        stmt = select(DouyinDongzhangDetail).where(*filters)
        details = list((await db.execute(stmt)).scalars().all())
        return [
            MerchantReconciliationService._build_detail_payload(
                detail,
                org_name=None,
                shop_color=None,
                load_context=detail_contexts.get(int(detail.shop_id)) or MerchantReconciliationService._empty_load_context(detail),
            )
            for detail in details
        ]

    @staticmethod
    async def _build_reconciliation_contexts(
        db: AsyncSession,
        *,
        user: User,
        accounting_year: int,
        accounting_month: int,
        shop_id: int | None,
        org_id: str | int | None,
    ) -> dict[int, _ReconciliationLoadContext]:
        base_filters = [
            DouyinDongzhangDetail.is_deleted.is_(False),
            DouyinDongzhangDetail.source_platform_code == "douyin",
            DouyinDongzhangDetail.summary_year == accounting_year,
            DouyinDongzhangDetail.summary_month == accounting_month,
            active_shop_filter(DouyinDongzhangDetail.shop_id),
        ]
        if shop_id is not None:
            base_filters.append(DouyinDongzhangDetail.shop_id == shop_id)
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            base_filters.append(DouyinDongzhangDetail.org_id.in_(org_ids))

        summary_rows = (
            await db.execute(
                select(
                    DouyinDongzhangDetail.shop_id,
                    func.coalesce(func.sum(DouyinDongzhangDetail.gmv), 0),
                    func.min(DouyinDongzhangDetail.org_id),
                )
                .where(*base_filters)
                .group_by(DouyinDongzhangDetail.shop_id)
            )
        ).all()

        contexts: dict[int, _ReconciliationLoadContext] = {}
        accounting_period = _period(accounting_year, accounting_month)
        for detail_shop_id, total_gmv, detail_org_id in summary_rows:
            if detail_shop_id is None:
                continue
            resolved_org_id = int(detail_org_id or (user.org_id or 0))
            total_bic, total_insurance = await MerchantReconciliationService._load_summary_totals(
                db,
                org_id=resolved_org_id,
                shop_id=int(detail_shop_id),
                accounting_year=accounting_year,
                accounting_month=accounting_month,
            )
            red_sheet_context = await MerchantReconciliationService._load_red_sheet_context(
                db,
                org_id=resolved_org_id,
                shop_id=int(detail_shop_id),
                accounting_period=accounting_period,
            )
            contexts[int(detail_shop_id)] = _ReconciliationLoadContext(
                org_id=resolved_org_id,
                total_gmv=safe_decimal(total_gmv),
                total_bic=total_bic,
                total_insurance=total_insurance,
                red_sheet_context=red_sheet_context,
            )
        return contexts

    @staticmethod
    def _empty_load_context(detail: DouyinDongzhangDetail) -> _ReconciliationLoadContext:
        return _ReconciliationLoadContext(
            org_id=int(detail.org_id or 0),
            total_gmv=ZERO_MONEY,
            total_bic=ZERO_MONEY,
            total_insurance=ZERO_MONEY,
            red_sheet_context=_RedSheetContext(purchases_by_code={}, payments_by_key={}),
        )

    @staticmethod
    async def _load_summary_totals(
        db: AsyncSession,
        *,
        org_id: int,
        shop_id: int,
        accounting_year: int,
        accounting_month: int,
        source_year: int | None = None,
        source_month: int | None = None,
    ) -> tuple[Decimal, Decimal]:
        filters = [
            FinancialSummary.org_id == org_id,
            FinancialSummary.shop_id == shop_id,
            FinancialSummary.source_platform_code == "douyin",
            FinancialSummary.summary_year == accounting_year,
            FinancialSummary.summary_month == accounting_month,
            FinancialSummary.is_deleted.is_(False),
        ]
        if source_year is not None:
            filters.append(FinancialSummary.source_year == source_year)
        if source_month is not None:
            filters.append(FinancialSummary.source_month == source_month)
        row = (
            await db.execute(
                select(
                    func.coalesce(func.sum(FinancialSummary.bic), 0),
                    func.coalesce(func.sum(FinancialSummary.insurance_fee), 0),
                ).where(*filters)
            )
        ).one()
        return safe_decimal(row[0]), safe_decimal(row[1])

    @staticmethod
    async def build_detail_derived_field_map(
        db: AsyncSession,
        *,
        details: Iterable[DouyinDongzhangDetail],
    ) -> dict[int, MerchantDerivedFields]:
        detail_list = [detail for detail in details if detail.id is not None]
        if not detail_list:
            return {}
        detail_by_shop_period: dict[tuple[int, int, int, int, int, int], list[DouyinDongzhangDetail]] = {}
        for detail in detail_list:
            detail_by_shop_period.setdefault(
                (
                    int(detail.org_id),
                    int(detail.shop_id),
                    int(detail.source_year),
                    int(detail.source_month),
                    int(detail.summary_year),
                    int(detail.summary_month),
                ),
                [],
            ).append(detail)

        context_by_key: dict[tuple[int, int, int, int, int, int], _ReconciliationLoadContext] = {}
        for (org_id, shop_id, source_year, source_month, accounting_year, accounting_month), group in detail_by_shop_period.items():
            _ = group
            total_gmv = await MerchantReconciliationService._load_douyin_detail_total_gmv(
                db,
                org_id=org_id,
                shop_id=shop_id,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                source_year=source_year,
                source_month=source_month,
            )
            total_bic, total_insurance = await MerchantReconciliationService._load_summary_totals(
                db,
                org_id=org_id,
                shop_id=shop_id,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
                source_year=source_year,
                source_month=source_month,
            )
            red_sheet_context = await MerchantReconciliationService._load_red_sheet_context(
                db,
                org_id=org_id,
                shop_id=shop_id,
                accounting_period=_period(source_year, source_month),
            )
            context_by_key[(org_id, shop_id, source_year, source_month, accounting_year, accounting_month)] = _ReconciliationLoadContext(
                org_id=org_id,
                total_gmv=total_gmv,
                total_bic=total_bic,
                total_insurance=total_insurance,
                red_sheet_context=red_sheet_context,
            )

        return {
            int(detail.id): MerchantReconciliationService._build_detail_derived_fields(
                detail,
                load_context=context_by_key[
                    (
                        int(detail.org_id),
                        int(detail.shop_id),
                        int(detail.source_year),
                        int(detail.source_month),
                        int(detail.summary_year),
                        int(detail.summary_month),
                    )
                ],
            )
            for detail in detail_list
        }

    @staticmethod
    async def _load_douyin_detail_total_gmv(
        db: AsyncSession,
        *,
        org_id: int,
        shop_id: int,
        accounting_year: int,
        accounting_month: int,
        source_year: int | None = None,
        source_month: int | None = None,
    ) -> Decimal:
        filters = [
            DouyinDongzhangDetail.org_id == org_id,
            DouyinDongzhangDetail.shop_id == shop_id,
            DouyinDongzhangDetail.source_platform_code == "douyin",
            DouyinDongzhangDetail.summary_year == accounting_year,
            DouyinDongzhangDetail.summary_month == accounting_month,
            DouyinDongzhangDetail.is_deleted.is_(False),
        ]
        if source_year is not None:
            filters.append(DouyinDongzhangDetail.source_year == source_year)
        if source_month is not None:
            filters.append(DouyinDongzhangDetail.source_month == source_month)
        total = (
            await db.execute(
                select(func.coalesce(func.sum(DouyinDongzhangDetail.gmv), 0)).where(*filters)
            )
        ).scalar()
        return safe_decimal(total)

    @staticmethod
    async def _load_red_sheet_context(
        db: AsyncSession,
        *,
        org_id: int,
        shop_id: int,
        accounting_period: int,
        payment_settlement_status: str | None = None,
    ) -> _RedSheetContext:
        # 对账计算始终只读取同组织、同店铺、同期间的最新一版红单。
        red_sheet_has_shop_rows = or_(
            MerchantRedSheet.shop_id == shop_id,
            exists(
                select(1).where(
                    MerchantRedSheetPurchase.red_sheet_id == MerchantRedSheet.id,
                    MerchantRedSheetPurchase.shop_id == shop_id,
                    MerchantRedSheetPurchase.is_deleted.is_(False),
                )
            ),
            exists(
                select(1).where(
                    MerchantRedSheetPayment.red_sheet_id == MerchantRedSheet.id,
                    MerchantRedSheetPayment.shop_id == shop_id,
                    MerchantRedSheetPayment.is_deleted.is_(False),
                )
            ),
        )
        latest = (
            await db.execute(
                select(MerchantRedSheet.id)
                .where(
                    MerchantRedSheet.org_id == org_id,
                    MerchantRedSheet.platform_code == "douyin",
                    MerchantRedSheet.accounting_period == accounting_period,
                    MerchantRedSheet.is_deleted.is_(False),
                    red_sheet_has_shop_rows,
                )
                .order_by(MerchantRedSheet.created_at.desc(), MerchantRedSheet.id.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        if latest is None:
            return _RedSheetContext(purchases_by_code={}, payments_by_key={})

        purchases = (
            await db.execute(
                select(MerchantRedSheetPurchase).where(
                    MerchantRedSheetPurchase.red_sheet_id == latest,
                    MerchantRedSheetPurchase.shop_id == shop_id,
                    MerchantRedSheetPurchase.is_deleted.is_(False),
                )
            )
        ).scalars().all()
        payment_filters = [
            MerchantRedSheetPayment.red_sheet_id == latest,
            MerchantRedSheetPayment.is_deleted.is_(False),
        ]
        if payment_settlement_status is not None:
            payment_filters.append(MerchantRedSheetPayment.remark == payment_settlement_status)
        payments = (
            await db.execute(
                select(MerchantRedSheetPayment).where(*payment_filters)
            )
        ).scalars().all()
        purchases_by_code: dict[str, MerchantRedSheetPurchase] = {}
        for purchase in purchases:
            for code_value in (purchase.normalized_live_code, purchase.live_code):
                for code in MerchantReconciliationService._split_product_codes(code_value):
                    purchases_by_code.setdefault(code, purchase)
        purchase_payment_keys = {
            (_match_text(purchase.merchant), _date_key(purchase.live_date), _match_text(purchase.live_room))
            for purchase in purchases
        }
        payments_by_key: dict[tuple[str, str, str], MerchantRedSheetPayment] = {}
        for payment in payments:
            key = (_match_text(payment.merchant), _date_key(payment.live_date), _match_text(payment.live_room))
            if payment.shop_id == shop_id or (payment.shop_id is None and key in purchase_payment_keys):
                payments_by_key.setdefault(key, payment)
        return _RedSheetContext(purchases_by_code=purchases_by_code, payments_by_key=payments_by_key)

    @staticmethod
    def _build_detail_derived_fields(
        detail: DouyinDongzhangDetail,
        *,
        load_context: _ReconciliationLoadContext,
    ) -> MerchantDerivedFields:
        product_code = detail.product_code or extract_product_code(detail.product_name)
        product_codes = MerchantReconciliationService._split_product_codes(product_code)
        purchase = next(
            (
                load_context.red_sheet_context.purchases_by_code.get(code)
                for code in product_codes
                if load_context.red_sheet_context.purchases_by_code.get(code)
            ),
            None,
        )
        payment = None
        errors: list[str] = []
        if not product_code:
            errors.append("商品名称未提取到商品编码")
        if purchase is None:
            errors.append("未匹配到红单采购")
        else:
            payment = load_context.red_sheet_context.payments_by_key.get(
                (_match_text(purchase.merchant), _date_key(purchase.live_date), _match_text(purchase.live_room))
            )
            if payment is None:
                errors.append("未匹配到红单货款")

        gmv = safe_decimal(detail.gmv)
        live_date = purchase.live_date if purchase is not None else ""
        return MerchantDerivedFields(
            product_code=product_code,
            major_merchant_name=purchase.merchant if purchase is not None else "",
            our_subject=payment.settlement_subject if payment is not None else "",
            merchant_receipt_subject=payment.receipt_subject if payment is not None else "",
            receipt_merchant=payment.receipt_merchant if payment is not None else "",
            live_room=purchase.live_room if purchase is not None else "",
            live_date=live_date,
            live_date_text=_date_key(live_date),
            red_sheet_payment_id=int(payment.id) if payment is not None and payment.id is not None else None,
            allocated_bic=_allocation_amount(gmv, load_context.total_gmv, load_context.total_bic),
            allocated_insurance_fee=_allocation_amount(gmv, load_context.total_gmv, load_context.total_insurance),
            live_amount=(gmv * Decimal("0.7")).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP),
            match_status="matched" if not errors else "unmatched",
            match_error="；".join(errors),
        )

    @staticmethod
    def _build_detail_payload(
        detail: DouyinDongzhangDetail,
        *,
        org_name: str | None,
        shop_color: str | None,
        load_context: _ReconciliationLoadContext,
    ) -> dict[str, object]:
        derived = MerchantReconciliationService._build_detail_derived_fields(detail, load_context=load_context)
        gmv = safe_decimal(detail.gmv)
        payload: dict[str, object] = {
            "id": detail.id,
            "org_id": detail.org_id,
            "org_name": org_name,
            "shop_id": detail.shop_id,
            "shop_name": detail.shop_name,
            "shop_color": shop_color,
            "accounting_year": detail.summary_year,
            "accounting_month": detail.summary_month,
            "accounting_period": _period(detail.summary_year, detail.summary_month),
            "accounting_date": f"{int(detail.summary_year):04d}-{int(detail.summary_month):02d}",
            "platform_code": detail.source_platform_code,
            "platform_label": "抖音",
            "source_row_number": detail.source_row_number,
            "transaction_time": MerchantReconciliationService._format_datetime(detail.transaction_time),
            "transaction_flow_no": detail.transaction_flow_no,
            "transaction_direction": detail.transaction_direction,
            "transaction_amount": safe_decimal(detail.transaction_amount),
            "transaction_scene": detail.transaction_scene,
            "sub_order_no": detail.sub_order_no,
            "order_no": detail.order_no,
            "order_time": MerchantReconciliationService._format_datetime(detail.order_time),
            "product_id": detail.product_id,
            "product_code": derived.product_code,
            "product_name": detail.product_name,
            "author_name": detail.author_name,
            "merchant_name": getattr(detail, "merchant_name", ""),
            "gmv": gmv,
            "allocated_bic": derived.allocated_bic,
            "allocated_insurance_fee": derived.allocated_insurance_fee,
            "live_amount": derived.live_amount,
            "major_merchant_name": derived.major_merchant_name,
            "our_subject": derived.our_subject,
            "merchant_receipt_subject": derived.merchant_receipt_subject,
            "receipt_merchant": derived.receipt_merchant,
            "red_sheet_payment_id": derived.red_sheet_payment_id,
            "live_room": derived.live_room,
            "live_date": derived.live_date,
            "live_date_text": derived.live_date_text,
            "match_status": derived.match_status,
            "match_error": derived.match_error,
        }
        return payload

    @staticmethod
    def _build_stats(rows: list[dict[str, object]], *, total_bic: Decimal, total_insurance: Decimal) -> MerchantReconciliationStatsOut:
        matched_rows = sum(1 for row in rows if row.get("match_status") == "matched")
        total_gmv = sum((safe_decimal(row.get("gmv")) for row in rows), ZERO_MONEY)
        total_allocated_bic = sum((safe_decimal(row.get("allocated_bic")) for row in rows), ZERO_MONEY)
        total_allocated_insurance = sum((safe_decimal(row.get("allocated_insurance_fee")) for row in rows), ZERO_MONEY)
        total_live_amount = sum((safe_decimal(row.get("live_amount")) for row in rows), ZERO_MONEY)
        return MerchantReconciliationStatsOut(
            total_gmv=total_gmv,
            total_bic=total_bic,
            total_allocated_bic=total_allocated_bic,
            total_insurance_fee=total_insurance,
            total_allocated_insurance_fee=total_allocated_insurance,
            total_live_amount=total_live_amount,
            matched_rows=matched_rows,
            unmatched_rows=len(rows) - matched_rows,
        )

    @staticmethod
    async def _load_payment_adjustments(
        db: AsyncSession,
        *,
        user: User,
        org_id: str | int | None,
        shop_id: int | None,
        accounting_year: int,
        accounting_month: int,
        payment_group_weights: dict[int, dict[tuple[str, str, str], Decimal]] | None = None,
        payment_ids: Iterable[int] | None = None,
    ) -> dict[tuple[str, str, str], dict[str, Decimal]]:
        payment_group_weights = payment_group_weights or {}
        payment_id_values: list[int] | None = None
        if payment_ids is not None:
            payment_id_values = [int(payment_id) for payment_id in payment_ids if payment_id is not None]
            if not payment_id_values:
                return {}
        filters = MerchantReconciliationService._red_sheet_detail_filters(
            user=user,
            model=MerchantRedSheetPayment,
            org_id=org_id,
            shop_ids=None if payment_id_values is not None else str(shop_id) if shop_id is not None else None,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
        )
        if payment_id_values is not None:
            filters.append(MerchantRedSheetPayment.id.in_(payment_id_values))
        rows = (
            await db.execute(
                select(MerchantRedSheetPayment, Organization.name.label("org_name"))
                .outerjoin(Organization, Organization.id == MerchantRedSheetPayment.org_id)
                .where(*filters)
            )
        ).all()
        adjustments: dict[tuple[str, str, str], dict[str, Decimal]] = {}
        for payment, org_name in rows:
            payment_id = MerchantReconciliationService._optional_int(payment.id)
            if payment_id is None:
                key = MerchantReconciliationService._payment_group_key_from_red_sheet(payment, org_name=org_name)
                MerchantReconciliationService._add_payment_adjustment(
                    adjustments,
                    key,
                    business_fee_deduction=safe_decimal(payment.business_fee_deduction),
                    other_deduction_amount=safe_decimal(payment.deduction_amount),
                    payable_goods_balance=safe_decimal(payment.payable_goods_amount),
                    paid_flow_amount=safe_decimal(getattr(payment, "paid_amount", ZERO_MONEY)),
                )
                continue
            weighted_keys = sorted((payment_group_weights.get(payment_id) or {}).items(), key=lambda item: item[0])
            if not weighted_keys:
                key = MerchantReconciliationService._payment_group_key_from_red_sheet(payment, org_name=org_name)
                MerchantReconciliationService._add_payment_adjustment(
                    adjustments,
                    key,
                    business_fee_deduction=safe_decimal(payment.business_fee_deduction),
                    other_deduction_amount=safe_decimal(payment.deduction_amount),
                    payable_goods_balance=safe_decimal(payment.payable_goods_amount),
                    paid_flow_amount=safe_decimal(getattr(payment, "paid_amount", ZERO_MONEY)),
                )
                continue

            business_fee_by_key = MerchantReconciliationService._allocate_money_by_weight(
                safe_decimal(payment.business_fee_deduction),
                weighted_keys,
            )
            deduction_by_key = MerchantReconciliationService._allocate_money_by_weight(
                safe_decimal(payment.deduction_amount),
                weighted_keys,
            )
            payable_by_key = MerchantReconciliationService._allocate_money_by_weight(
                safe_decimal(payment.payable_goods_amount),
                weighted_keys,
            )
            paid_by_key = MerchantReconciliationService._allocate_money_by_weight(
                safe_decimal(getattr(payment, "paid_amount", ZERO_MONEY)),
                weighted_keys,
            )
            for key, _weight in weighted_keys:
                MerchantReconciliationService._add_payment_adjustment(
                    adjustments,
                    key,
                    business_fee_deduction=business_fee_by_key.get(key, ZERO_MONEY),
                    other_deduction_amount=deduction_by_key.get(key, ZERO_MONEY),
                    payable_goods_balance=payable_by_key.get(key, ZERO_MONEY),
                    paid_flow_amount=paid_by_key.get(key, ZERO_MONEY),
                )
        return adjustments

    @staticmethod
    def _add_payment_adjustment(
        adjustments: dict[tuple[str, str, str], dict[str, Decimal]],
        key: tuple[str, str, str],
        *,
        business_fee_deduction: Decimal,
        other_deduction_amount: Decimal,
        payable_goods_balance: Decimal,
        paid_flow_amount: Decimal = ZERO_MONEY,
    ) -> None:
        item = adjustments.setdefault(
            key,
            {
                "business_fee_deduction": ZERO_MONEY,
                "other_deduction_amount": ZERO_MONEY,
                "payable_goods_balance": ZERO_MONEY,
                "paid_flow_amount": ZERO_MONEY,
            },
        )
        item["business_fee_deduction"] += business_fee_deduction
        item["other_deduction_amount"] += other_deduction_amount
        item["payable_goods_balance"] += payable_goods_balance
        item["paid_flow_amount"] += paid_flow_amount

    @staticmethod
    def _allocate_money_by_weight(
        amount: Decimal,
        weighted_keys: list[tuple[tuple[str, str, str], Decimal]],
    ) -> dict[tuple[str, str, str], Decimal]:
        if not weighted_keys:
            return {}
        if len(weighted_keys) == 1:
            return {weighted_keys[0][0]: amount}

        weights = [(key, abs(safe_decimal(weight))) for key, weight in weighted_keys]
        total_weight = sum((weight for _key, weight in weights), ZERO_MONEY)
        if total_weight == ZERO_MONEY:
            weights = [(key, Decimal("1")) for key, _weight in weights]
            total_weight = Decimal(len(weights))

        allocations: dict[tuple[str, str, str], Decimal] = {}
        remaining = amount
        for index, (key, weight) in enumerate(weights):
            if index == len(weights) - 1:
                allocations[key] = remaining
                break
            value = (amount * weight / total_weight).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
            allocations[key] = value
            remaining -= value
        return allocations

    @staticmethod
    def _optional_int(value: object) -> int | None:
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _build_detail_workbook(rows: Iterable[dict[str, object]]) -> io.BytesIO:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="商家对账明细")
        ws.append(_write_only_header_row(ws, [label for _field, label, _money_flag in RECONCILIATION_EXPORT_COLUMNS]))
        for row in rows:
            MerchantReconciliationService._append_export_row(ws, row)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _save_detail_workbook(rows: Iterable[dict[str, object]], *, output_path: Path) -> int:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="商家对账明细")
        ws.append(_write_only_header_row(ws, [label for _field, label, _money_flag in RECONCILIATION_EXPORT_COLUMNS]))
        row_count = 0
        for row in rows:
            MerchantReconciliationService._append_export_row(ws, row)
            row_count += 1
        wb.save(output_path)
        return row_count

    @staticmethod
    def _append_export_row(ws, row: dict[str, object]) -> None:
        ws.append([MerchantReconciliationService._format_export_value(row.get(field), money=money_flag) for field, _label, money_flag in RECONCILIATION_EXPORT_COLUMNS])

    @staticmethod
    def _format_export_value(value: object, *, money: bool = False) -> object:
        if money:
            return float(safe_decimal(value))
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return MerchantReconciliationService._format_datetime(value)
        if isinstance(value, date):
            return value.isoformat()
        return value

    @staticmethod
    def _format_datetime(value: object) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value or "")

    @staticmethod
    def _parse_bank_flow_file(
        *,
        file_path: str | Path,
        accounting_year: int,
        accounting_month: int,
    ) -> tuple[list[dict[str, object]], str, str]:
        with open_tabular_rows(str(file_path)) as table_rows:
            if table_rows is None:
                raise ValueError("银行流水文件无法读取")
            rows = [list(row) for row in table_rows]
        if not rows:
            return [], "", ""

        header_index = MerchantReconciliationService._find_bank_flow_header_index(rows)
        if header_index is None:
            raise ValueError("银行流水未识别到表头")
        header_row = rows[header_index]
        header_map = {
            canonical_header(value): index
            for index, value in enumerate(header_row)
            if canonical_header(value)
        }
        bank_name = "中国邮政" if "用途" in header_map and "附言" in header_map else "企业网银"
        account_name = MerchantReconciliationService._extract_bank_flow_account_name(
            rows=rows,
            header_map=header_map,
            header_index=header_index,
        )
        parsed_rows: list[dict[str, object]] = []
        for row_number, row_values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(_normalize_text(value) for value in row_values):
                continue
            row = MerchantReconciliationService._parse_bank_flow_row(
                row_values,
                header_map=header_map,
                source_row_number=row_number,
                bank_name=bank_name,
                fallback_account_name=account_name,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
            )
            if row is not None:
                parsed_rows.append(row)
        return parsed_rows, bank_name, account_name

    @staticmethod
    def _find_bank_flow_header_index(rows: list[list[object]]) -> int | None:
        required_sets = (
            ("账户名称", "交易时间", "对方户名", "备注"),
            ("交易日期", "交易时间", "对方户名", "用途"),
        )
        for index, row in enumerate(rows[:30]):
            normalized = {canonical_header(value) for value in row if canonical_header(value)}
            for required in required_sets:
                if all(canonical_header(header) in normalized for header in required):
                    return index
        return None

    @staticmethod
    def _extract_bank_flow_account_name(
        *,
        rows: list[list[object]],
        header_map: dict[str, int],
        header_index: int,
    ) -> str:
        account_header = canonical_header("账户名称")
        if account_header in header_map:
            for row_values in rows[header_index + 1 :]:
                value = MerchantReconciliationService._bank_cell(row_values, header_map, "账户名称")
                if value:
                    return value

        # 邮政导出里 F2 是“户名”的值，B/F 等列属于合并区域的主单元格。
        for row_values in rows[:header_index]:
            for index, value in enumerate(row_values):
                if canonical_header(value) == canonical_header("户名"):
                    for candidate_index in range(index + 1, min(index + 5, len(row_values))):
                        candidate = MerchantReconciliationService._cell_text(row_values[candidate_index])
                        if candidate:
                            return candidate
        if len(rows) > 1 and len(rows[1]) > 5:
            return MerchantReconciliationService._cell_text(rows[1][5])
        return ""

    @staticmethod
    def _parse_bank_flow_row(
        row_values: list[object],
        *,
        header_map: dict[str, int],
        source_row_number: int,
        bank_name: str,
        fallback_account_name: str,
        accounting_year: int,
        accounting_month: int,
    ) -> dict[str, object] | None:
        account_name = MerchantReconciliationService._bank_cell(row_values, header_map, "账户名称") or fallback_account_name
        counterparty_name = MerchantReconciliationService._bank_cell(row_values, header_map, "对方户名")
        purpose_parts = [
            MerchantReconciliationService._bank_cell(row_values, header_map, "用途"),
            MerchantReconciliationService._bank_cell(row_values, header_map, "备注"),
            MerchantReconciliationService._bank_cell(row_values, header_map, "附言"),
        ]
        purpose = "；".join(part for part in purpose_parts if part)
        summary = MerchantReconciliationService._bank_cell(row_values, header_map, "摘要")
        debit_amount = _money(
            MerchantReconciliationService._bank_cell(row_values, header_map, "借方发生额（支取）")
            or MerchantReconciliationService._bank_cell(row_values, header_map, "支出金额")
        )
        credit_amount = _money(
            MerchantReconciliationService._bank_cell(row_values, header_map, "贷方发生额（收入）")
            or MerchantReconciliationService._bank_cell(row_values, header_map, "收入金额")
        )
        if not counterparty_name and debit_amount == ZERO_MONEY and credit_amount == ZERO_MONEY:
            return None
        if "直播" not in purpose:
            return None
        transaction_time = MerchantReconciliationService._parse_bank_transaction_time(row_values, header_map)
        transaction_date = transaction_time.date() if transaction_time is not None else _parse_date(MerchantReconciliationService._bank_cell(row_values, header_map, "交易日期"))
        raw_row = MerchantReconciliationService._bank_raw_row(row_values, header_map)
        return {
            "source_row_number": source_row_number,
            "bank_name": bank_name,
            "account_no": MerchantReconciliationService._bank_cell(row_values, header_map, "账号"),
            "account_name": account_name,
            "transaction_date": transaction_date,
            "transaction_time": transaction_time,
            "debit_amount": debit_amount,
            "credit_amount": credit_amount,
            "flow_amount": (debit_amount - credit_amount).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP),
            "balance": _money(MerchantReconciliationService._bank_cell(row_values, header_map, "余额")),
            "counterparty_account_no": MerchantReconciliationService._bank_cell(row_values, header_map, "对方账号"),
            "counterparty_name": counterparty_name,
            "counterparty_bank": (
                MerchantReconciliationService._bank_cell(row_values, header_map, "对方开户机构")
                or MerchantReconciliationService._bank_cell(row_values, header_map, "对方行名")
            ),
            "summary": summary,
            "purpose": purpose,
            "remark": MerchantReconciliationService._bank_cell(row_values, header_map, "备注"),
            "live_date": MerchantReconciliationService.parse_live_date_from_bank_text(
                purpose,
                accounting_year=accounting_year,
                accounting_month=accounting_month,
            ),
            "transaction_flow_no": (
                MerchantReconciliationService._bank_cell(row_values, header_map, "账户明细编号-交易流水号")
                or MerchantReconciliationService._bank_cell(row_values, header_map, "交易流水号")
                or MerchantReconciliationService._bank_cell(row_values, header_map, "全局路由号")
            ),
            "raw_row": raw_row,
        }

    @staticmethod
    def parse_live_date_from_bank_text(
        value: object,
        *,
        accounting_year: int,
        accounting_month: int,
    ) -> str:
        text = _normalize_text(value)
        if not text:
            return ""
        range_match = LIVE_DATE_RANGE_PATTERN.search(text)
        if range_match:
            live_month = int(range_match.group("month"))
            start_day = int(range_match.group("start"))
            end_day = int(range_match.group("end") or start_day)
            end_month = int(range_match.group("end_month") or live_month)
            if end_day == start_day:
                return f"{live_month}月{start_day}日"
            if end_month != live_month:
                return f"{live_month}月{start_day}日-{end_month}月{end_day}日"
            return f"{live_month}月{start_day}-{end_day}日"
        return MerchantReconciliationService.normalize_live_date_text(text)

    @staticmethod
    def normalize_live_date_text(value: object) -> str:
        text = _normalize_text(value)
        if not text:
            return ""
        ymd_range_match = LIVE_DATE_YMD_RANGE_PATTERN.search(text)
        if ymd_range_match:
            month = int(ymd_range_match.group("month"))
            start_day = int(ymd_range_match.group("start"))
            end_month = int(ymd_range_match.group("end_month") or month)
            end_day = int(ymd_range_match.group("end") or start_day)
            if start_day == end_day and month == end_month:
                return f"{month}月{start_day}日"
            if month == end_month:
                return f"{month}月{start_day}-{end_day}日"
            return f"{month}月{start_day}日-{end_month}月{end_day}日"
        range_match = LIVE_DATE_RANGE_PATTERN.search(text)
        if range_match:
            month = int(range_match.group("month"))
            start_day = int(range_match.group("start"))
            end_month = int(range_match.group("end_month") or month)
            end_day = int(range_match.group("end") or start_day)
            if start_day == end_day and month == end_month:
                return f"{month}月{start_day}日"
            if month == end_month:
                return f"{month}月{start_day}-{end_day}日"
            return f"{month}月{start_day}日-{end_month}月{end_day}日"
        iso_match = LIVE_DATE_ISO_PATTERN.search(text)
        if iso_match:
            return f"{int(iso_match.group('month'))}月{int(iso_match.group('day'))}日"
        return ""

    @staticmethod
    def _parse_bank_transaction_time(row_values: list[object], header_map: dict[str, int]) -> datetime | None:
        raw_datetime = MerchantReconciliationService._bank_cell(row_values, header_map, "交易时间")
        raw_date = MerchantReconciliationService._bank_cell(row_values, header_map, "交易日期")
        compact_match = COMPACT_BANK_DATETIME_PATTERN.match(raw_datetime)
        if compact_match:
            compact_value = f"{compact_match.group('date')[:4]}-{compact_match.group('date')[4:6]}-{compact_match.group('date')[6:8]} {compact_match.group('time')}"
            compact_parsed = parse_datetime(compact_value)
            if compact_parsed is not None:
                return compact_parsed
        if raw_date and raw_datetime:
            combined = parse_datetime(f"{raw_date} {raw_datetime}")
            if combined is not None:
                return combined
        parsed = parse_datetime(raw_datetime)
        if parsed is not None and raw_date:
            parsed_date = _parse_date(raw_date)
            if parsed_date is not None and parsed.year == 1900:
                return datetime.combine(parsed_date, parsed.time())
            return parsed
        if parsed is not None:
            return parsed
        if raw_date:
            parsed_date = _parse_date(raw_date)
            if parsed_date is not None:
                return datetime.combine(parsed_date, datetime.min.time())
        return None

    @staticmethod
    def _bank_cell(row_values: list[object], header_map: dict[str, int], header: str) -> str:
        index = header_map.get(canonical_header(header))
        if index is None:
            return ""
        return MerchantReconciliationService._cell_text(row_values[index] if index < len(row_values) else None)

    @staticmethod
    def _bank_raw_row(row_values: list[object], header_map: dict[str, int]) -> dict[str, str]:
        raw: dict[str, str] = {}
        for canonical, index in header_map.items():
            if not canonical:
                continue
            raw[canonical] = MerchantReconciliationService._cell_text(row_values[index] if index < len(row_values) else None)
        return raw

    @staticmethod
    def _split_product_codes(value: object) -> list[str]:
        codes: list[str] = []
        normalized = safe_str(value).replace("，", ",").replace("＋", ",").replace("+", ",")
        for part in normalized.split(","):
            part = part.strip().upper()
            if part:
                codes.append(part)
        return codes

    @staticmethod
    def _parse_purchase_sheet(worksheet) -> list[dict[str, object]]:
        header_map = MerchantReconciliationService._parse_headers(worksheet, PURCHASE_HEADERS, "采购")
        rows: list[dict[str, object]] = []
        for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(_normalize_text(value) for value in row_values):
                continue
            rows.append(
                {
                    "source_row_number": row_number,
                    "raw_row": MerchantReconciliationService._raw_row(row_values, PURCHASE_HEADERS, header_map),
                    "live_room": MerchantReconciliationService._cell(row_values, header_map, "直播间"),
                    "merchant": MerchantReconciliationService._cell(row_values, header_map, "商家"),
                    "live_date": MerchantReconciliationService._cell(row_values, header_map, "直播日期"),
                    "loan_return_order_no": MerchantReconciliationService._cell(row_values, header_map, "借/退货单号"),
                    "loan_return_date": _parse_date(MerchantReconciliationService._cell(row_values, header_map, "借/退货日期")),
                    "live_code": MerchantReconciliationService._cell(row_values, header_map, "直播编号"),
                    "normalized_live_code": extract_live_code(MerchantReconciliationService._cell(row_values, header_map, "直播编号")),
                    "match_status": MerchantReconciliationService._cell(row_values, header_map, "匹配"),
                    "remark": MerchantReconciliationService._cell(row_values, header_map, "备注"),
                    "source_shop_name": MerchantReconciliationService._cell(row_values, header_map, "店铺"),
                    "subject": MerchantReconciliationService._cell(row_values, header_map, "主体"),
                    "summary": MerchantReconciliationService._cell(row_values, header_map, "摘要"),
                    "product_name": MerchantReconciliationService._cell(row_values, header_map, "货品名称"),
                    "piece_price": _money(MerchantReconciliationService._cell(row_values, header_map, "件/元")),
                    "gram_price": _money(MerchantReconciliationService._cell(row_values, header_map, "克/元")),
                    "sale_price": _money(MerchantReconciliationService._cell(row_values, header_map, "卖价")),
                    "borrow_quantity": _money(MerchantReconciliationService._cell(row_values, header_map, "借货数量")),
                    "borrow_weight_g": _money(MerchantReconciliationService._cell(row_values, header_map, "借货重量g")),
                    "borrow_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "借货金额")),
                    "return_quantity": _money(MerchantReconciliationService._cell(row_values, header_map, "退货数量")),
                    "return_weight_g": _money(MerchantReconciliationService._cell(row_values, header_map, "退货重量g")),
                    "return_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "退货金额")),
                    "estimated_settlement_date": _parse_date(MerchantReconciliationService._cell(row_values, header_map, "预计结款日期")),
                    "labor_fee_per_gram": _money(MerchantReconciliationService._cell(row_values, header_map, "工费/克")),
                    "labor_fee_per_piece": _money(MerchantReconciliationService._cell(row_values, header_map, "工费/件")),
                }
            )
        return rows

    @staticmethod
    def _parse_payment_sheet(worksheet) -> list[dict[str, object]]:
        header_map = MerchantReconciliationService._parse_headers(worksheet, PAYMENT_HEADERS, "货款")
        rows: list[dict[str, object]] = []
        for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(_normalize_text(value) for value in row_values):
                continue
            settlement_status = MerchantReconciliationService._cell(row_values, header_map, "结算状态")
            MerchantReconciliationService._validate_payment_settlement_status(settlement_status, row_number=row_number)
            rows.append(
                {
                    "source_row_number": row_number,
                    "raw_row": MerchantReconciliationService._raw_row(row_values, PAYMENT_HEADERS, header_map),
                    "sequence_no": MerchantReconciliationService._cell(row_values, header_map, "序号"),
                    "live_room": MerchantReconciliationService._cell(row_values, header_map, "直播间"),
                    "live_date": MerchantReconciliationService._cell(row_values, header_map, "直播日期"),
                    "merchant": MerchantReconciliationService._cell(row_values, header_map, "商家"),
                    "borrow_total_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "借货总金额")),
                    "return_total_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "退货总金额")),
                    "business_fee_deduction": _money(MerchantReconciliationService._cell(row_values, header_map, "冲减业务费用")),
                    "deduction_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "冲减金额")),
                    "payable_goods_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "应付货款金额")),
                    "return_rate": MerchantReconciliationService._cell(row_values, header_map, "退货率"),
                    "settlement_subject": MerchantReconciliationService._cell(row_values, header_map, "结算主体"),
                    "receipt_subject": MerchantReconciliationService._cell(row_values, header_map, "收款主体"),
                    "cost_subject": MerchantReconciliationService._cell(row_values, header_map, "成本主体"),
                    "payable_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "应付款金额")),
                    "subject_collection_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "主体回款金额")),
                    "receipt_merchant": MerchantReconciliationService._cell(row_values, header_map, "收款商家"),
                    "collection_merchant": MerchantReconciliationService._cell(row_values, header_map, "回款商家"),
                    "is_settled": MerchantReconciliationService._cell(row_values, header_map, "是否已结款"),
                    "is_collected": MerchantReconciliationService._cell(row_values, header_map, "是否已回款"),
                    "remark": settlement_status,
                    "payment_screenshot": MerchantReconciliationService._cell(row_values, header_map, "付款截图"),
                    "settlement_date": _parse_date(MerchantReconciliationService._cell(row_values, header_map, "结算日期")),
                    "collection_date": _parse_date(MerchantReconciliationService._cell(row_values, header_map, "回款日期")),
                    "deduction_remark": MerchantReconciliationService._cell(row_values, header_map, "冲减备注"),
                    "pending_issue": MerchantReconciliationService._cell(row_values, header_map, "待解决事项"),
                    "is_receipt_merchant_modified": MerchantReconciliationService._cell(row_values, header_map, "是否修改收款商家"),
                    "is_receipt_amount_modified": MerchantReconciliationService._cell(row_values, header_map, "是否修改收款金额"),
                    "modified_month": MerchantReconciliationService._cell(row_values, header_map, "修改月份"),
                    "application_date": _parse_date(MerchantReconciliationService._cell(row_values, header_map, "申请日期")),
                    "paid_amount": _money(MerchantReconciliationService._cell(row_values, header_map, "已付")),
                    "borrow_minus_return": _money(MerchantReconciliationService._cell(row_values, header_map, "借-退")),
                }
            )
        return rows

    @staticmethod
    def _parse_headers(worksheet, expected_headers: list[str], sheet_label: str) -> dict[str, int]:
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            raise ValueError(f"{sheet_label} sheet 缺少表头")
        normalized_to_index = {canonical_header(value): index for index, value in enumerate(header_cells) if canonical_header(value)}
        header_map: dict[str, int] = {}
        missing: list[str] = []
        for header in expected_headers:
            normalized = canonical_header(header)
            if normalized not in normalized_to_index:
                missing.append(header)
            else:
                header_map[header] = normalized_to_index[normalized]
        if missing:
            raise ValueError(f"{sheet_label} sheet 缺少表头：{'、'.join(missing)}")
        return header_map

    @staticmethod
    def _cell(row_values: tuple[object, ...], header_map: dict[str, int], header: str) -> str:
        index = header_map[header]
        return MerchantReconciliationService._cell_text(row_values[index] if index < len(row_values) else None)

    @staticmethod
    def _validate_payment_settlement_status(value: str, *, row_number: int) -> None:
        if value not in PAYMENT_SETTLEMENT_STATUS_SET:
            allowed = "、".join(PAYMENT_SETTLEMENT_STATUSES)
            raise ValueError(f"货款 sheet 第 {row_number} 行结算状态必须是：{allowed}")

    @staticmethod
    def _raw_row(row_values: tuple[object, ...], headers: list[str], header_map: dict[str, int]) -> dict[str, str]:
        raw: dict[str, str] = {}
        for header in headers:
            index = header_map[header]
            raw[header] = MerchantReconciliationService._cell_text(row_values[index] if index < len(row_values) else None)
        return raw

    @staticmethod
    def _cell_text(value: object) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return _normalize_text(value)

    @staticmethod
    def _purchase_model(
        row: dict[str, object],
        *,
        red_sheet_id: int,
        org_id: int,
        shop_id: int,
        shop_name: str,
        accounting_period: int,
    ) -> MerchantRedSheetPurchase:
        return MerchantRedSheetPurchase(
            red_sheet_id=red_sheet_id,
            org_id=org_id,
            shop_id=shop_id,
            shop_name=shop_name,
            accounting_period=accounting_period,
            **row,
        )

    @staticmethod
    def _payment_model(
        row: dict[str, object],
        *,
        red_sheet_id: int,
        org_id: int,
        shop_id: int,
        shop_name: str,
        accounting_period: int,
    ) -> MerchantRedSheetPayment:
        return MerchantRedSheetPayment(
            red_sheet_id=red_sheet_id,
            org_id=org_id,
            shop_id=shop_id,
            shop_name=shop_name,
            accounting_period=accounting_period,
            **row,
        )

    @staticmethod
    def _bank_flow_row_model(
        row: dict[str, object],
        *,
        bank_flow_file_id: int,
        org_id: int,
        accounting_period: int,
    ) -> MerchantBankFlowRow:
        return MerchantBankFlowRow(
            bank_flow_file_id=bank_flow_file_id,
            org_id=org_id,
            accounting_period=accounting_period,
            **row,
        )

    @staticmethod
    def _bank_flow_row_payload(item: MerchantBankFlowRow, *, org_name: str | None) -> dict[str, object]:
        payload = dict(item.__dict__)
        payload.pop("_sa_instance_state", None)
        payload["org_name"] = org_name
        return payload

    @staticmethod
    def _validate_month(year: int, month: int) -> None:
        if year < 2000 or year > 2100 or month < 1 or month > 12:
            raise ValueError("月份不正确")
