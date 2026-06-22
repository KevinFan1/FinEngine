from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql.elements import ColumnElement

from app.models.organization import Organization
from app.models.platform import Platform
from app.models.shop import Shop
from app.models.user import User
from app.schemas.shop import ShopCreate, ShopImportError, ShopImportResult, ShopUpdate
from app.services.audit_service import AuditService
from app.services.shop_color_service import assign_default_shop_color, normalize_shop_color
from app.services.platform_profile_service import resolve_platform_profile
from app.utils.query_filters import split_int_filter_values


SHOP_DUPLICATE_MESSAGE = "该平台下已存在同名店铺，请勿重复创建"
SHOP_IMPORT_UPDATE_FIELDS = (
    "tax_no",
    "merchant",
    "registered_address",
    "legal_person",
    "previous_name",
    "store_long_id",
    "store_short_id",
    "settlement_period",
    "primary_account",
    "anchor",
    "shop_type",
    "purpose",
    "former_name",
)
SHOP_IMPORT_HEADERS = (
    "平台",
    "店铺名",
    "税号",
    "商户",
    "注册地址",
    "法人",
    "曾用名",
    "store_long_id",
    "store_short_id",
    "settlement_period",
    "primary_account",
    "主播",
    "类型",
    "purpose",
    "former_name",
)
SHOP_IMPORT_HEADER_ALIASES = {
    "平台": "platform_name",
    "平台名称": "platform_name",
    "platform": "platform_name",
    "platform_name": "platform_name",
    "店铺": "shop_name",
    "店铺名": "shop_name",
    "店铺名称": "shop_name",
    "shop": "shop_name",
    "shop_name": "shop_name",
    "税号": "tax_no",
    "tax_no": "tax_no",
    "商户": "merchant",
    "merchant": "merchant",
    "注册地址": "registered_address",
    "registered_address": "registered_address",
    "法人": "legal_person",
    "legal_person": "legal_person",
    "曾用名": "previous_name",
    "previous_name": "previous_name",
    "store_long_id": "store_long_id",
    "store_short_id": "store_short_id",
    "settlement_period": "settlement_period",
    "primary_account": "primary_account",
    "主播": "anchor",
    "anchor": "anchor",
    "类型": "shop_type",
    "type": "shop_type",
    "shop_type": "shop_type",
    "purpose": "purpose",
    "former_name": "former_name",
}
SHOP_IMPORT_FIELD_LIMITS = {
    "platform_name": 50,
    "shop_name": 200,
    "tax_no": 100,
    "merchant": 200,
    "registered_address": 500,
    "legal_person": 100,
    "previous_name": 200,
    "store_long_id": 100,
    "store_short_id": 100,
    "settlement_period": 100,
    "primary_account": 200,
    "anchor": 100,
    "shop_type": 100,
    "purpose": 200,
    "former_name": 200,
}


class ShopService:
    @staticmethod
    def active_filter() -> ColumnElement[bool]:
        return Shop.is_deleted.is_(False)

    @staticmethod
    def validate_shop_scope(*, shop: Shop, operator: User) -> None:
        if operator.role == "superadmin":
            return
        if shop.org_id != operator.org_id:
            raise ValueError("店铺不存在")

    @staticmethod
    async def list_shops(
        db: AsyncSession,
        *,
        org_id: str | int | None = None,
        ids: str | int | None = None,
        keyword: str | None = None,
        platform_name: str | None = None,
        page: int | None = 1,
        page_size: int | None = 20,
    ) -> tuple[list[tuple[Shop, str | None]], int]:
        stmt = (
            select(Shop, Organization.name.label("org_name"))
            .outerjoin(
                Organization,
                and_(Organization.id == Shop.org_id, Organization.is_deleted.is_(False)),
            )
            .where(ShopService.active_filter())
            .order_by(Shop.updated_at.desc(), Shop.id.desc())
        )
        count_stmt = select(func.count()).select_from(Shop).where(ShopService.active_filter())
        org_ids = split_int_filter_values(org_id)
        if org_ids:
            stmt = stmt.where(Shop.org_id.in_(org_ids))
            count_stmt = count_stmt.where(Shop.org_id.in_(org_ids))
        shop_ids = split_int_filter_values(ids)
        if shop_ids:
            stmt = stmt.where(Shop.id.in_(shop_ids))
            count_stmt = count_stmt.where(Shop.id.in_(shop_ids))
        if platform_name:
            platform_values = [item.strip() for item in platform_name.split(",") if item.strip()]
            report_platform_codes = set()
            for platform_value in platform_values:
                profile = await resolve_platform_profile(db, platform_value)
                report_platform_codes.add(profile.report_platform_code)
            if report_platform_codes:
                stmt = stmt.where(Shop.platform_name.in_(report_platform_codes))
                count_stmt = count_stmt.where(Shop.platform_name.in_(report_platform_codes))
        if keyword:
            like = f"%{keyword}%"
            cond = or_(
                Shop.shop_name.ilike(like),
                Shop.tax_no.ilike(like),
                Shop.merchant.ilike(like),
                Shop.registered_address.ilike(like),
                Shop.legal_person.ilike(like),
                Shop.previous_name.ilike(like),
                Shop.store_long_id.ilike(like),
                Shop.store_short_id.ilike(like),
                Shop.primary_account.ilike(like),
                Shop.anchor.ilike(like),
                Shop.shop_type.ilike(like),
                Shop.purpose.ilike(like),
                Shop.former_name.ilike(like),
            )
            stmt = stmt.where(cond)
            count_stmt = count_stmt.where(cond)
        total = (await db.execute(count_stmt)).scalar() or 0
        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        items = list((await db.execute(stmt)).all())
        return items, total

    @staticmethod
    async def get_shop(db: AsyncSession, shop_id: int) -> Shop | None:
        return (await db.execute(select(Shop).where(Shop.id == shop_id, ShopService.active_filter()))).scalar_one_or_none()

    @staticmethod
    async def create_shop(db: AsyncSession, *, data: ShopCreate, org_id: int, operator: User, ip: str | None = None, user_agent: str | None = None) -> Shop:
        shop_color = normalize_shop_color(data.shop_color) or assign_default_shop_color(
            org_id=org_id,
            platform_name=data.platform_name,
            shop_name=data.shop_name,
        )
        stmt = (
            insert(Shop)
            .values(
                org_id=org_id,
                platform_name=data.platform_name,
                shop_name=data.shop_name,
                shop_color=shop_color,
                tax_no=data.tax_no,
                merchant=data.merchant,
                registered_address=data.registered_address,
                legal_person=data.legal_person,
                previous_name=data.previous_name,
                store_long_id=data.store_long_id,
                store_short_id=data.store_short_id,
                settlement_period=data.settlement_period,
                primary_account=data.primary_account,
                anchor=data.anchor,
                shop_type=data.shop_type,
                purpose=data.purpose,
                former_name=data.former_name,
                remark=data.remark,
                status=1,
                is_deleted=False,
                deleted_at=None,
            )
            .on_conflict_do_nothing(
                index_elements=[Shop.org_id, Shop.platform_name, Shop.shop_name],
                index_where=text("is_deleted = false"),
            )
            .returning(Shop)
        )
        result = await db.execute(stmt)
        shop = result.scalar_one_or_none()
        if shop is None:
            raise ValueError(SHOP_DUPLICATE_MESSAGE)
        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=org_id,
            module="shop",
            action="create",
            description=f"创建店铺 [{data.platform_name}] [{data.shop_name}]",
            target_type="shop",
            target_id=shop.id,
            target_name=data.shop_name,
            ip=ip,
            user_agent=user_agent,
        )
        return shop

    @staticmethod
    async def update_shop(db: AsyncSession, *, shop_id: int, data: ShopUpdate, operator: User, ip: str | None = None, user_agent: str | None = None) -> Shop | None:
        shop = await ShopService.get_shop(db, shop_id)
        if not shop:
            return None
        ShopService.validate_shop_scope(shop=shop, operator=operator)
        old = {
            "platform_name": shop.platform_name,
            "shop_name": shop.shop_name,
            "shop_color": shop.shop_color,
            "tax_no": shop.tax_no,
            "merchant": shop.merchant,
            "registered_address": shop.registered_address,
            "legal_person": shop.legal_person,
            "previous_name": shop.previous_name,
            "store_long_id": shop.store_long_id,
            "store_short_id": shop.store_short_id,
            "settlement_period": shop.settlement_period,
            "primary_account": shop.primary_account,
            "anchor": shop.anchor,
            "shop_type": shop.shop_type,
            "purpose": shop.purpose,
            "former_name": shop.former_name,
            "remark": shop.remark,
        }
        update_data = data.model_dump(exclude_unset=True)
        if "shop_color" in update_data:
            update_data["shop_color"] = normalize_shop_color(update_data["shop_color"])
        if "platform_name" in update_data or "shop_name" in update_data:
            next_platform_name = update_data.get("platform_name") or shop.platform_name
            next_shop_name = update_data.get("shop_name") or shop.shop_name
            existing = await db.execute(
                select(Shop).where(
                    Shop.org_id == shop.org_id,
                    Shop.platform_name == next_platform_name,
                    Shop.shop_name == next_shop_name,
                    Shop.id != shop.id,
                    ShopService.active_filter(),
                )
            )
            if existing.scalar_one_or_none():
                raise ValueError(SHOP_DUPLICATE_MESSAGE)
        if not update_data.get("shop_color") and (
            "platform_name" in update_data or "shop_name" in update_data
        ):
            next_platform_name = update_data.get("platform_name") or shop.platform_name
            next_shop_name = update_data.get("shop_name") or shop.shop_name
            update_data["shop_color"] = assign_default_shop_color(
                org_id=shop.org_id,
                platform_name=next_platform_name,
                shop_name=next_shop_name,
            )
        for k, v in update_data.items():
            setattr(shop, k, v)
        await db.flush()
        await db.refresh(shop)
        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=shop.org_id,
            module="shop",
            action="update",
            description=f"修改店铺 [{shop.shop_name}]",
            target_type="shop",
            target_id=shop.id,
            target_name=shop.shop_name,
            ip=ip,
            user_agent=user_agent,
            old_value=old,
            new_value=update_data,
        )
        return shop

    @staticmethod
    async def delete_shop(db: AsyncSession, *, shop_id: int, operator: User, ip: str | None = None, user_agent: str | None = None) -> bool:
        shop = await ShopService.get_shop(db, shop_id)
        if not shop:
            return False
        ShopService.validate_shop_scope(shop=shop, operator=operator)
        from datetime import datetime, timezone

        shop.is_deleted = True
        shop.deleted_at = datetime.now(timezone.utc)
        shop.status = 0
        await db.flush()
        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=shop.org_id,
            module="shop",
            action="delete",
            description=f"删除店铺 [{shop.platform_name}] [{shop.shop_name}]",
            target_type="shop",
            target_id=shop_id,
            target_name=shop.shop_name,
            ip=ip,
            user_agent=user_agent,
        )
        return True

    @staticmethod
    async def get_or_create_shop(db: AsyncSession, *, org_id: int, platform_name: str, shop_name: str) -> Shop:
        """Get existing shop or create new one. Used during upload callback."""
        shop_color = assign_default_shop_color(
            org_id=org_id,
            platform_name=platform_name,
            shop_name=shop_name,
        )
        stmt = (
            insert(Shop)
            .values(
                org_id=org_id,
                platform_name=platform_name,
                shop_name=shop_name,
                shop_color=shop_color,
                is_deleted=False,
                deleted_at=None,
            )
            .on_conflict_do_nothing(
                index_elements=[Shop.org_id, Shop.platform_name, Shop.shop_name],
                index_where=text("is_deleted = false"),
            )
            .returning(Shop)
        )
        result = await db.execute(stmt)
        shop = result.scalar_one_or_none()
        if shop is not None:
            return shop

        result = await db.execute(
            select(Shop).where(
                Shop.org_id == org_id,
                Shop.platform_name == platform_name,
                Shop.shop_name == shop_name,
                ShopService.active_filter(),
            )
        )
        shop = result.scalar_one_or_none()
        if shop is None:
            raise ValueError("店铺创建失败，请稍后重试")
        if not shop.shop_color:
            shop.shop_color = shop_color
            await db.flush()
        return shop

    @staticmethod
    def build_import_template() -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "店铺导入模板"

        required_fill = PatternFill("solid", fgColor="FCE7F3")
        optional_fill = PatternFill("solid", fgColor="E0F2FE")
        header_font = Font(bold=True, color="111827")
        for index, header in enumerate(SHOP_IMPORT_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=index, value=header)
            cell.font = header_font
            cell.fill = required_fill if header in {"平台", "店铺名"} else optional_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if header in {"平台", "店铺名"}:
                cell.comment = Comment(
                    "当前组织内，平台 + 店铺名 是唯一定位键。导入时两者一致会更新资料字段，两者不同会新增店铺。",
                    "FinEngine",
                )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(SHOP_IMPORT_HEADERS)).coordinate}"
        widths = {
            "平台": 14,
            "店铺名": 24,
            "注册地址": 32,
            "primary_account": 22,
            "settlement_period": 20,
            "store_long_id": 20,
            "store_short_id": 20,
        }
        for index, header in enumerate(SHOP_IMPORT_HEADERS, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = widths.get(header, 16)

        note_sheet = workbook.create_sheet("导入说明")
        note_rows = [
            ("重要", "当前组织内，平台 + 店铺名 是唯一键，名称请与系统内店铺名称保持完全一致。"),
            ("更新规则", "导入时如果唯一键已存在，只更新模板中的资料字段；不存在则新增店铺。"),
            ("上传新增", "上传文件自动新增店铺的逻辑不变，不会写入或覆盖这些资料字段。"),
            ("必填列", "平台、店铺名。系统会自动使用当前登录用户所属组织。"),
            ("可选列", "税号、商户、注册地址、法人、曾用名、store_long_id、store_short_id、settlement_period、primary_account、主播、类型、purpose、former_name。"),
        ]
        for row_index, (title, text_value) in enumerate(note_rows, start=1):
            note_sheet.cell(row=row_index, column=1, value=title).font = Font(bold=True)
            note_sheet.cell(row=row_index, column=2, value=text_value)
        note_sheet.column_dimensions["A"].width = 16
        note_sheet.column_dimensions["B"].width = 92

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def build_export_workbook(rows: list[tuple[Shop, str | None]]) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "店铺信息"

        header_font = Font(bold=True, color="111827")
        header_fill = PatternFill("solid", fgColor="E0F2FE")
        for index, header in enumerate(SHOP_IMPORT_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        export_fields = (
            "platform_name",
            "shop_name",
            "tax_no",
            "merchant",
            "registered_address",
            "legal_person",
            "previous_name",
            "store_long_id",
            "store_short_id",
            "settlement_period",
            "primary_account",
            "anchor",
            "shop_type",
            "purpose",
            "former_name",
        )
        for row_index, (shop, _org_name) in enumerate(rows, start=2):
            for column_index, field in enumerate(export_fields, start=1):
                worksheet.cell(row=row_index, column=column_index, value=getattr(shop, field, None) or "")

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(SHOP_IMPORT_HEADERS)).coordinate}"
        widths = {
            "平台": 14,
            "店铺名": 24,
            "注册地址": 32,
            "primary_account": 22,
            "settlement_period": 20,
            "store_long_id": 20,
            "store_short_id": 20,
        }
        for index, header in enumerate(SHOP_IMPORT_HEADERS, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = widths.get(header, 16)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    async def import_shops(
        db: AsyncSession,
        *,
        content: bytes,
        filename: str,
        operator: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> ShopImportResult:
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("请上传 .xlsx 或 .xlsm 格式的店铺导入模板")

        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("文件无法读取，请使用店铺导入模板重新填写") from exc

        worksheet = workbook[workbook.sheetnames[0]]
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map = ShopService._parse_import_headers(header_cells)
        result = ShopImportResult(total=0, created=0, updated=0, skipped=0, errors=[])

        for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(ShopService._normalize_cell(value) for value in row_values):
                continue
            result.total += 1
            row_data = ShopService._read_import_row(row_values, header_map)
            row_error = await ShopService._import_shop_row(db, row_data=row_data, operator=operator)
            if row_error:
                result.skipped += 1
                result.errors.append(ShopImportError(row=row_number, message=row_error))
                continue
            if row_data["_created"]:
                result.created += 1
            else:
                result.updated += 1

        await AuditService.log(
            db,
            user_id=operator.id,
            username=operator.username,
            display_name=operator.display_name,
            org_id=operator.org_id,
            module="shop",
            action="import",
            description=f"导入店铺资料：新增 {result.created}，更新 {result.updated}，跳过 {result.skipped}",
            target_type="shop",
            ip=ip,
            user_agent=user_agent,
            new_value=result.model_dump(),
        )
        return result

    @staticmethod
    def _parse_import_headers(header_cells: tuple[object, ...] | None) -> dict[str, int]:
        if not header_cells:
            raise ValueError("模板缺少表头，请下载最新店铺导入模板")
        header_map: dict[str, int] = {}
        for index, header in enumerate(header_cells):
            normalized = ShopService._normalize_header(header)
            field = SHOP_IMPORT_HEADER_ALIASES.get(normalized)
            if field and field not in header_map:
                header_map[field] = index

        missing = [label for label, field in (("平台", "platform_name"), ("店铺名", "shop_name")) if field not in header_map]
        if missing:
            raise ValueError(f"模板缺少必填列：{'、'.join(missing)}")
        return header_map

    @staticmethod
    def _read_import_row(row_values: tuple[object, ...], header_map: dict[str, int]) -> dict[str, str | None | bool]:
        row_data: dict[str, str | None | bool] = {}
        for field, index in header_map.items():
            row_data[field] = ShopService._normalize_cell(row_values[index] if index < len(row_values) else None)
        row_data["_created"] = False
        return row_data

    @staticmethod
    async def _import_shop_row(db: AsyncSession, *, row_data: dict[str, str | None | bool], operator: User) -> str | None:
        org_id = operator.org_id
        if not org_id:
            return "当前账号没有所属组织，无法导入店铺"

        platform_name = str(row_data.get("platform_name") or "").strip()
        shop_name = str(row_data.get("shop_name") or "").strip()
        if not platform_name:
            return "平台不能为空"
        if not shop_name:
            return "店铺名不能为空"

        platform_name = await ShopService._resolve_import_platform_name(db, platform_name)
        row_data["platform_name"] = platform_name
        row_data["shop_name"] = shop_name
        field_error = ShopService._validate_import_field_lengths(row_data)
        if field_error:
            return field_error

        existing = await db.execute(
            select(Shop).where(
                Shop.org_id == org_id,
                Shop.platform_name == platform_name,
                Shop.shop_name == shop_name,
                ShopService.active_filter(),
            )
        )
        shop = existing.scalar_one_or_none()
        field_values = {field: row_data.get(field) for field in SHOP_IMPORT_UPDATE_FIELDS}
        if shop is None:
            shop = Shop(
                org_id=org_id,
                platform_name=platform_name,
                shop_name=shop_name,
                shop_color=assign_default_shop_color(
                    org_id=org_id,
                    platform_name=platform_name,
                    shop_name=shop_name,
                ),
                status=1,
                is_deleted=False,
                deleted_at=None,
                **field_values,
            )
            db.add(shop)
            row_data["_created"] = True
        else:
            for field, value in field_values.items():
                setattr(shop, field, value)
            row_data["_created"] = False
        await db.flush()
        return None

    @staticmethod
    async def _resolve_import_platform_name(db: AsyncSession, value: str) -> str:
        result = await db.execute(
            select(Platform).where(
                or_(Platform.code == value, Platform.name == value),
                Platform.is_deleted.is_(False),
            )
        )
        platform = result.scalar_one_or_none()
        profile = await resolve_platform_profile(db, platform.code if platform else value)
        return profile.report_platform_code

    @staticmethod
    def _validate_import_field_lengths(row_data: dict[str, str | None | bool]) -> str | None:
        for field, max_length in SHOP_IMPORT_FIELD_LIMITS.items():
            value = row_data.get(field)
            if isinstance(value, str) and len(value) > max_length:
                return f"{field} 超过 {max_length} 个字符"
        return None

    @staticmethod
    def _normalize_cell(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            text_value = value.strip()
        elif isinstance(value, float) and value.is_integer():
            text_value = str(int(value))
        else:
            text_value = str(value).strip()
        return text_value or None

    @staticmethod
    def _normalize_header(value: object) -> str:
        return str(value or "").strip()
