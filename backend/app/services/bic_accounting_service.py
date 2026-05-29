"""Service layer for independent BIC accounting."""
import io
import tempfile
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl.cell import WriteOnlyCell
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import delete, func, literal, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import register_after_commit
from app.models.bic_accounting import BicDetail, BicSourceRow, BicTask, BicUploadFile
from app.models.organization import Organization
from app.models.shop import Shop
from app.models.upload import UploadFile
from app.models.user import User
from app.services.audit_service import AuditService
from app.services.oss_service import SOURCE_FILE_UNAVAILABLE_MESSAGE, is_oss_object_unavailable_error, oss_service
from app.services.partition_service import BIC_SOURCE_PARTITION, ensure_month_partition
from app.services.shop_service import ShopService
from app.services.shop_visibility import active_shop_filter
from app.services.transaction_accounting_service import TransactionAccountingService
from app.services.upload_period_service import extract_upload_period, resolve_upload_period_header
from app.tasks.processors.base import FinancialSummaryExcelProcessorMixin, open_tabular_rows, parse_datetime, safe_str
from app.tasks.processors.douyin import DOUYIN_BIC_HEADERS
from app.utils.query_filters import datetime_range_filters, resolve_org_ids
from app.utils.money import safe_decimal

BIC_TARGET_FEE_ITEM = "质检费(通过)"
BIC_DETAIL_EXPORT_HEADERS = ["序号", "平台", "服务商", "店铺id", "店铺名称", "QIC仓", "公司名称", "税号", "抬头类型", "注册地址", "结算金额"]
BIC_SOURCE_EXPORT_HEADERS = DOUYIN_BIC_HEADERS
BIC_RECONCILIATION_SOURCE_EXPORT_HEADERS = ["店铺", *DOUYIN_BIC_HEADERS]
BIC_RESULT_TASK_STATUSES = ("success", "partial_success", "failed", "expired")
BIC_ERROR_SAMPLE_LIMIT = 10
BIC_EXPORT_BATCH_SIZE = 5000


def _capture_bic_result_state(task: BicTask) -> dict[str, object]:
    return {
        "processed_rows": task.processed_rows,
        "success_rows": task.success_rows,
        "failed_rows": task.failed_rows,
        "result_summary": task.result_summary,
    }


def _restore_bic_result_state(task: BicTask, state: dict[str, object]) -> None:
    task.processed_rows = int(state.get("processed_rows") or 0)
    task.success_rows = int(state.get("success_rows") or 0)
    task.failed_rows = int(state.get("failed_rows") or 0)
    task.result_summary = state.get("result_summary")  # type: ignore[assignment]


def _service_provider_filter(column, service_provider: str | None):
    normalized = service_provider.replace("，", ",") if service_provider else service_provider
    values = TransactionAccountingService._split_filter_values(normalized)
    if not values:
        return None
    return or_(*(column.ilike(f"%{value}%") for value in values))


def _normalize_bic_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return safe_str(value)


def _format_source_value(value: object) -> str:
    return _normalize_bic_text(value)


def _format_source_datetime(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return _format_source_value(value)


def _parse_source_datetime(value: object) -> datetime | None:
    return parse_datetime(value)


def _bic_accounting_period(year: int | None, month: int | None) -> int:
    return int(year or 0) * 100 + int(month or 0)


def _source_row_payload(
    *,
    task: BicTask,
    upload_file: BicUploadFile,
    detail_id: int,
    shop_id: int | None,
    shop_name: str,
    platform_code: str,
    source_row: dict[str, object],
    source_row_number: int,
) -> BicSourceRow:
    return BicSourceRow(
        task_id=task.id,
        file_id=upload_file.id,
        detail_id=detail_id,
        org_id=task.org_id,
        shop_id=shop_id,
        platform_code=platform_code,
        shop_name=shop_name,
        accounting_year=int(upload_file.accounting_year),
        accounting_month=int(upload_file.accounting_month),
        accounting_period=_bic_accounting_period(upload_file.accounting_year, upload_file.accounting_month),
        service_provider=_format_source_value(source_row.get("service_provider")) or "-",
        qic_warehouse=_format_source_value(source_row.get("qic_warehouse")) or "-",
        source_row_number=source_row_number,
        settlement_no=_format_source_value(source_row.get("settlement_no")),
        order_code=_format_source_value(source_row.get("order_code")),
        related_order_no=_format_source_value(source_row.get("related_order_no")),
        related_waybill_no=_format_source_value(source_row.get("related_waybill_no")),
        fee_item=_format_source_value(source_row.get("fee_item")),
        settlement_amount=safe_decimal(source_row.get("amount")),
        billing_params=_format_source_value(source_row.get("billing_params")),
        billing_completed_time=_parse_source_datetime(source_row.get("billing_completed_time")),
        business_node=_format_source_value(source_row.get("business_node")),
        business_occurred_time=_parse_source_datetime(source_row.get("business_occurred_time")),
        settled_at=_parse_source_datetime(source_row.get("settled_at")),
        status=_format_source_value(source_row.get("status")),
        transaction_account=_format_source_value(source_row.get("transaction_account")),
        transaction_flow_no=_format_source_value(source_row.get("transaction_flow_no")),
        remark=_format_source_value(source_row.get("remark")),
        is_mudaibao=_format_source_value(source_row.get("is_mudaibao")),
        is_child_order=_format_source_value(source_row.get("is_child_order")),
    )


def _dedupe_bic_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    deduped: list[dict[str, object]] = []
    positions: dict[str, int] = {}
    for row in rows:
        flow_no = safe_str(row.get("transaction_flow_no"))
        if not flow_no:
            deduped.append(row)
            continue
        if flow_no in positions:
            deduped[positions[flow_no]] = row
        else:
            positions[flow_no] = len(deduped)
            deduped.append(row)
    return deduped


_SOURCE_ROW_OVERWRITE_FIELDS = (
    "task_id",
    "file_id",
    "detail_id",
    "org_id",
    "shop_id",
    "platform_code",
    "shop_name",
    "accounting_year",
    "accounting_month",
    "service_provider",
    "qic_warehouse",
    "source_row_number",
    "settlement_no",
    "order_code",
    "related_order_no",
    "related_waybill_no",
    "fee_item",
    "settlement_amount",
    "billing_params",
    "billing_completed_time",
    "business_node",
    "business_occurred_time",
    "settled_at",
    "status",
    "transaction_account",
    "transaction_flow_no",
    "remark",
    "is_mudaibao",
    "is_child_order",
)


def _overwrite_source_row(target: BicSourceRow, source: BicSourceRow) -> BicSourceRow:
    for field in _SOURCE_ROW_OVERWRITE_FIELDS:
        setattr(target, field, getattr(source, field))
    target.is_deleted = False
    target.deleted_at = None
    return target


async def _upsert_source_rows_by_flow(
    db: AsyncSession,
    *,
    platform_code: str,
    accounting_period: int,
    source_rows: Iterable[BicSourceRow],
) -> list[BicSourceRow]:
    rows = list(source_rows)
    flow_nos = sorted({safe_str(row.transaction_flow_no) for row in rows if safe_str(row.transaction_flow_no)})
    existing_by_flow: dict[str, BicSourceRow] = {}

    chunk_size = 500
    for index in range(0, len(flow_nos), chunk_size):
        chunk = flow_nos[index:index + chunk_size]
        result = await db.execute(
            select(BicSourceRow).where(
                BicSourceRow.is_deleted.is_(False),
                BicSourceRow.accounting_period == accounting_period,
                BicSourceRow.platform_code == platform_code,
                BicSourceRow.transaction_flow_no.in_(chunk),
            )
        )
        for existing in result.scalars().all():
            existing_by_flow[safe_str(existing.transaction_flow_no)] = existing

    synced_rows: list[BicSourceRow] = []
    new_rows: list[BicSourceRow] = []
    for row in rows:
        flow_no = safe_str(row.transaction_flow_no)
        existing = existing_by_flow.get(flow_no) if flow_no else None
        if existing is None:
            new_rows.append(row)
            synced_rows.append(row)
        else:
            synced_rows.append(_overwrite_source_row(existing, row))

    if new_rows:
        db.add_all(new_rows)
    return synced_rows


def _apply_sheet_style(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    max_columns = sheet.max_column
    for column_index in range(1, max_columns + 1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        max_length = len(str(sheet.cell(row=1, column=column_index).value or ""))
        for row_index in range(2, min(sheet.max_row, 200) + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)
    sheet.freeze_panes = "A2"


def _write_only_header_row(sheet, headers: list[str]) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for label in headers:
        cell = WriteOnlyCell(sheet, value=label)
        cell.font = Font(bold=True)
        cells.append(cell)
    return cells


def _ensure_export_row_limit(label: str, row_count: int) -> None:
    return None


class BicAccountingService:
    @staticmethod
    def resolve_org_id(*, user: User, requested_org_id: int | None = None) -> int | None:
        if user.role == "superadmin":
            return requested_org_id
        return user.org_id

    @staticmethod
    def validate_org_scope(*, org_id: int, user: User) -> None:
        if user.role != "superadmin" and org_id != user.org_id:
            raise ValueError("数据不存在或无权访问")

    @staticmethod
    async def create_from_shared_upload(
        db: AsyncSession,
        *,
        upload_file: UploadFile,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> BicTask:
        if (upload_file.parsed_type or "").lower() != "bic":
            raise ValueError("仅 BIC 文件可创建 BIC 任务")
        if (upload_file.source_platform_code or upload_file.detected_platform or "").lower() != "douyin":
            raise ValueError("BIC 核算当前仅支持抖音文件")

        source_platform = (upload_file.source_platform_code or upload_file.detected_platform or "").lower()
        shop_id = upload_file.shop_id
        shop_name = upload_file.parsed_shop
        if shop_id is None and shop_name:
            shop = await ShopService.get_or_create_shop(
                db,
                org_id=upload_file.org_id,
                platform_name=source_platform,
                shop_name=shop_name,
            )
            shop_id = shop.id
            shop_name = shop.shop_name

        existing_stmt = select(BicUploadFile).where(
            BicUploadFile.source_upload_file_id == upload_file.id,
            BicUploadFile.is_deleted.is_(False),
        )
        existing = (await db.execute(existing_stmt)).scalar_one_or_none()
        bic_upload_file = existing
        if bic_upload_file is not None and shop_id is not None and bic_upload_file.shop_id is None:
            bic_upload_file.shop_id = shop_id
        if existing is not None:
            task_stmt = (
                select(BicTask)
                .where(
                    BicTask.file_id == existing.id,
                    BicTask.is_deleted.is_(False),
                )
                .order_by(BicTask.id.desc())
            )
            current_task = (await db.execute(task_stmt)).scalars().first()
            if current_task is not None:
                return current_task

        if bic_upload_file is None:
            bic_upload_file = BicUploadFile(
                org_id=upload_file.org_id,
                user_id=user.id,
                shop_id=shop_id,
                source_upload_file_id=upload_file.id,
                original_name=upload_file.original_name,
                oss_key=upload_file.oss_key,
                file_size=upload_file.file_size,
                file_hash=upload_file.file_hash,
                platform_code=source_platform,
                shop_name=shop_name,
                accounting_year=upload_file.parsed_year,
                accounting_month=upload_file.parsed_month,
                status="uploaded",
            )
            db.add(bic_upload_file)
            await db.flush()

        task = BicTask(
            file_id=bic_upload_file.id,
            org_id=bic_upload_file.org_id,
            user_id=user.id,
            status="queued",
            progress=0,
        )
        db.add(task)
        await db.flush()

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=bic_upload_file.org_id,
            module="bic_accounting",
            action="upload_file",
            description=f"上传 BIC 文件 [{bic_upload_file.original_name}]",
            target_type="bic_upload_file",
            target_id=bic_upload_file.id,
            target_name=bic_upload_file.original_name,
            ip=ip,
            user_agent=user_agent,
            extra_data={
                "task_id": task.id,
                "file_size": bic_upload_file.file_size,
                "source_upload_file_id": upload_file.id,
            },
        )

        BicAccountingService.dispatch_task_after_commit(
            db,
            task=task,
            upload_file=bic_upload_file,
        )
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def rerun_task(
        db: AsyncSession,
        *,
        task_id: int,
        user: User,
        ip: str | None = None,
        user_agent: str | None = None,
    ) -> BicTask | None:
        task = await db.get(BicTask, task_id)
        if task is None or task.is_deleted:
            return None
        BicAccountingService.validate_org_scope(org_id=task.org_id, user=user)
        if task.status == "expired":
            raise ValueError("源文件已过期或不存在，不能重新统计，请重新上传文件")
        upload_file = await db.get(BicUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("BIC上传文件不存在")

        task.status = "queued"
        task.progress = 0
        task.celery_task_id = None
        task.error_message = None
        task.started_at = None
        task.finished_at = None
        upload_file.status = "uploaded"
        upload_file.error_message = None

        await AuditService.log(
            db,
            user_id=user.id,
            username=user.username,
            display_name=user.display_name,
            org_id=task.org_id,
            module="bic_accounting",
            action="task_rerun",
            description=f"重跑 BIC 任务 [{task.id}]",
            target_type="bic_task",
            target_id=task.id,
            ip=ip,
            user_agent=user_agent,
        )

        BicAccountingService.dispatch_task_after_commit(
            db,
            task=task,
            upload_file=upload_file,
        )
        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    def dispatch_task_after_commit(
        db: AsyncSession,
        *,
        task: BicTask,
        upload_file: BicUploadFile | None,
    ) -> None:
        async def dispatch() -> None:
            try:
                from app.tasks.bic_accounting import run_bic_accounting_task

                async_result = run_bic_accounting_task.delay(task.id)
                task.celery_task_id = async_result.id
            except Exception as exc:
                task.status = "failed"
                task.progress = 100
                task.error_message = f"BIC任务投递失败: {exc}"
                task.finished_at = datetime.now(timezone.utc)
                if upload_file is not None:
                    upload_file.status = "failed"
                    upload_file.error_message = task.error_message
            await db.flush()

        register_after_commit(db, dispatch)

    @staticmethod
    async def _get_latest_task_for_upload_file(db: AsyncSession, *, file_id: int) -> BicTask | None:
        stmt = (
            select(BicTask)
            .where(
                BicTask.file_id == file_id,
                BicTask.is_deleted.is_(False),
            )
            .order_by(BicTask.id.desc())
        )
        return (await db.execute(stmt)).scalars().first()

    @staticmethod
    def _latest_result_task_ids_select():
        return (
            select(func.max(BicTask.id).label("task_id"))
            .select_from(BicTask)
            .join(BicUploadFile, BicUploadFile.id == BicTask.file_id)
            .where(
                BicTask.is_deleted.is_(False),
                BicUploadFile.is_deleted.is_(False),
                active_shop_filter(BicUploadFile.shop_id),
                BicTask.status.in_(BIC_RESULT_TASK_STATUSES),
            )
            .group_by(
                BicUploadFile.org_id,
                BicUploadFile.platform_code,
                BicUploadFile.shop_id,
                BicUploadFile.accounting_year,
                BicUploadFile.accounting_month,
            )
        )

    @staticmethod
    async def list_tasks(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        status: str | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        keyword: str | None = None,
        created_start_time: datetime | None = None,
        created_end_time: datetime | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[tuple[BicTask, BicUploadFile, str | None, str | None]], int]:
        filters = [BicTask.is_deleted.is_(False), BicUploadFile.is_deleted.is_(False), active_shop_filter(BicUploadFile.shop_id)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(BicTask.org_id.in_(org_ids))
        if status:
            filters.append(BicTask.status.in_(TransactionAccountingService._split_filter_values(status)))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(BicUploadFile.platform_code.in_(platform_codes))
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(BicUploadFile.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(BicUploadFile.shop_id.in_(shop_id_list))
        filters.extend(
            TransactionAccountingService._period_filters(
                BicUploadFile.accounting_year * 100 + BicUploadFile.accounting_month,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )
        if keyword:
            like_pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    BicUploadFile.original_name.ilike(like_pattern),
                    BicUploadFile.shop_name.ilike(like_pattern),
                    BicTask.error_message.ilike(like_pattern),
                )
            )
        filters.extend(
            datetime_range_filters(
                BicTask.created_at,
                start_time=created_start_time,
                end_time=created_end_time,
            )
        )

        stmt = (
            select(BicTask, BicUploadFile, Shop.shop_color, Organization.name.label("org_name"))
            .join(BicUploadFile, BicUploadFile.id == BicTask.file_id)
            .outerjoin(Shop, Shop.id == BicUploadFile.shop_id)
            .outerjoin(Organization, Organization.id == BicTask.org_id)
            .where(*filters)
            .order_by(BicTask.id.desc())
        )
        count_stmt = select(func.count()).select_from(BicTask).join(BicUploadFile, BicUploadFile.id == BicTask.file_id).where(*filters)
        total = (await db.execute(count_stmt)).scalar() or 0
        result = await db.execute(stmt.offset((page - 1) * page_size).limit(page_size))
        return list(result.all()), total

    @staticmethod
    async def list_details(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        service_provider: str | None = None,
        qic_warehouse: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        ids: list[int] | None = None,
        page: int | None = 1,
        page_size: int | None = 50,
    ) -> tuple[list[dict[str, object]], int]:
        filters = [BicDetail.is_deleted.is_(False), active_shop_filter(BicDetail.shop_id)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(BicDetail.org_id.in_(org_ids))
        if ids is not None:
            if not ids:
                return [], 0
            filters.append(BicDetail.id.in_(ids))
        if task_id is not None:
            filters.append(BicDetail.task_id == task_id)
        else:
            filters.append(BicDetail.task_id.in_(BicAccountingService._latest_result_task_ids_select()))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(BicDetail.platform_code.in_(platform_codes))
        service_provider_filter = _service_provider_filter(BicDetail.service_provider, service_provider)
        if service_provider_filter is not None:
            filters.append(service_provider_filter)
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(BicDetail.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(BicDetail.shop_id.in_(shop_id_list))
        qic_values = TransactionAccountingService._split_filter_values(qic_warehouse)
        if qic_values:
            filters.append(BicDetail.qic_warehouse.in_(qic_values))
        if accounting_year is not None:
            filters.append(BicDetail.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(BicDetail.accounting_month == accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                BicDetail.accounting_year * 100 + BicDetail.accounting_month,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )

        stmt = (
            select(
                BicDetail.id.label("id"),
                BicDetail.task_id.label("task_id"),
                BicDetail.file_id.label("file_id"),
                BicDetail.org_id.label("org_id"),
                Organization.name.label("org_name"),
                BicDetail.shop_id.label("shop_id"),
                BicDetail.platform_code.label("platform_code"),
                BicDetail.service_provider.label("service_provider"),
                BicDetail.shop_name.label("shop_name"),
                BicDetail.accounting_year.label("accounting_year"),
                BicDetail.accounting_month.label("accounting_month"),
                BicDetail.qic_warehouse.label("qic_warehouse"),
                BicDetail.row_count.label("row_count"),
                BicDetail.total_amount.label("total_amount"),
                BicDetail.created_at.label("created_at"),
                Shop.store_short_id.label("store_short_id"),
                Shop.merchant.label("merchant"),
                Shop.tax_no.label("tax_no"),
                Shop.shop_type.label("shop_type"),
                Shop.registered_address.label("registered_address"),
            )
            .outerjoin(Shop, Shop.id == BicDetail.shop_id)
            .outerjoin(Organization, Organization.id == BicDetail.org_id)
            .where(*filters)
            .order_by(
                BicDetail.accounting_year.desc(),
                BicDetail.accounting_month.desc(),
                BicDetail.platform_code,
                BicDetail.service_provider,
                BicDetail.shop_id,
                BicDetail.shop_name,
                BicDetail.qic_warehouse,
            )
        )
        count_stmt = select(func.count()).select_from(BicDetail).where(*filters)
        total = (await db.execute(count_stmt)).scalar() or 0
        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(stmt)
        return [dict(row) for row in result.mappings().all()], total

    @staticmethod
    async def execute_task(db: AsyncSession, *, task_id: int) -> BicTask:
        task = await db.get(BicTask, task_id)
        if task is None or task.is_deleted:
            raise ValueError("BIC任务不存在")
        upload_file = await db.get(BicUploadFile, task.file_id)
        if upload_file is None or upload_file.is_deleted:
            raise ValueError("BIC上传文件不存在")

        previous_result_state = _capture_bic_result_state(task)
        task.status = "processing"
        task.progress = 5
        task.started_at = datetime.now(timezone.utc)
        task.finished_at = None
        task.error_message = None
        await db.flush()

        try:
            suffix = Path(upload_file.original_name).suffix or ".xlsx"
            with tempfile.NamedTemporaryFile(suffix=suffix) as tmp:
                oss_service.download_to_temp(upload_file.oss_key, tmp.name)
                period_header = await resolve_upload_period_header(db, upload_file.platform_code, "bic")
                upload_period = extract_upload_period(
                    tmp.name,
                    platform_code=upload_file.platform_code,
                    type_code="bic",
                    header_name=period_header,
                )
                upload_file.accounting_year = upload_period.year
                upload_file.accounting_month = upload_period.month
                summary = await BicAccountingService.persist_task_result(
                    db,
                    task=task,
                    upload_file=upload_file,
                    file_path=tmp.name,
                    platform_code=upload_file.platform_code or "douyin",
                    shop_id=upload_file.shop_id,
                    shop_name=upload_file.shop_name or "",
                )

            task.processed_rows = int(summary.get("总行数", 0))
            task.success_rows = int(summary.get("符合条件行数", 0))
            task.failed_rows = int(summary.get("失败行数", 0))
            task.result_summary = summary
            task.progress = 100
            if summary.get("文件解析失败"):
                task.status = "failed"
            else:
                task.status = "success" if task.failed_rows == 0 else "partial_success"
            error_messages = summary.get("错误明细")
            task.error_message = "\n".join(map(str, error_messages)) if isinstance(error_messages, list) and error_messages else None
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "failed" if task.status == "failed" else "processed"
            upload_file.error_message = task.error_message
        except Exception as exc:
            await db.rollback()
            task = await db.get(BicTask, task_id)
            if task is None or task.is_deleted:
                raise ValueError("BIC任务不存在") from exc
            upload_file = await db.get(BicUploadFile, task.file_id)
            if upload_file is None or upload_file.is_deleted:
                raise ValueError("BIC上传文件不存在") from exc
            is_expired = is_oss_object_unavailable_error(exc)
            _restore_bic_result_state(task, previous_result_state)
            task.status = "expired" if is_expired else "failed"
            task.progress = 100
            task.error_message = SOURCE_FILE_UNAVAILABLE_MESSAGE if is_expired else str(exc)
            task.finished_at = datetime.now(timezone.utc)
            upload_file.status = "expired" if is_expired else "failed"
            upload_file.error_message = task.error_message

        await db.flush()
        await db.refresh(task)
        return task

    @staticmethod
    async def persist_task_result(
        db: AsyncSession,
        *,
        task: BicTask,
        upload_file: BicUploadFile,
        file_path: str,
        platform_code: str,
        shop_id: int | None,
        shop_name: str,
    ) -> dict:
        if upload_file.accounting_year is None or upload_file.accounting_month is None:
            raise ValueError("BIC文件缺少所属年月")
        shop_name = shop_name or upload_file.shop_name or ""
        if not shop_name:
            raise ValueError("BIC文件缺少店铺名称")
        if shop_id is None:
            shop = await ShopService.get_or_create_shop(
                db,
                org_id=task.org_id,
                platform_name=platform_code,
                shop_name=shop_name,
            )
            shop_id = shop.id
            upload_file.shop_id = shop_id
            upload_file.shop_name = shop.shop_name

        old_detail_result = await db.execute(select(BicDetail.id).where(BicDetail.task_id == task.id))
        old_detail_ids = list(old_detail_result.scalars().all())

        parse_result = BicAccountingService.parse_file(file_path)
        detail_rows: list[BicDetail] = []
        source_rows: list[BicSourceRow] = []
        if parse_result.get("fatal_error"):
            await db.execute(delete(BicSourceRow).where(BicSourceRow.task_id == task.id))
            if old_detail_ids:
                await db.execute(delete(BicDetail).where(BicDetail.id.in_(old_detail_ids)))
        else:
            await db.execute(delete(BicSourceRow).where(BicSourceRow.task_id == task.id))
            if old_detail_ids:
                await db.execute(delete(BicDetail).where(BicDetail.id.in_(old_detail_ids)))

            parsed_rows = _dedupe_bic_rows(parse_result.get("bic_rows", []))
            detail_rows = BicAccountingService._build_detail_rows(
                task=task,
                upload_file=upload_file,
                rows=parsed_rows,
                platform_code=platform_code,
                shop_id=shop_id,
                shop_name=shop_name,
            )
            db.add_all(detail_rows)
            await db.flush()

            source_rows = BicAccountingService._build_source_rows(
                task=task,
                upload_file=upload_file,
                detail_rows=detail_rows,
                rows=parsed_rows,
                platform_code=platform_code,
                shop_id=shop_id,
                shop_name=shop_name,
            )
            await ensure_month_partition(
                db,
                spec=BIC_SOURCE_PARTITION,
                period=_bic_accounting_period(upload_file.accounting_year, upload_file.accounting_month),
            )
            source_rows = await _upsert_source_rows_by_flow(
                db,
                platform_code=platform_code,
                accounting_period=_bic_accounting_period(upload_file.accounting_year, upload_file.accounting_month),
                source_rows=source_rows,
            )
            await db.flush()
        await db.flush()

        summary = {
            "文件类型": "BIC",
            "总行数": parse_result.get("total_rows", 0),
            "符合条件行数": parse_result.get("success_rows", 0),
            "失败行数": parse_result.get("failed_rows", 0),
            "明细分组数": len(detail_rows),
            "源数据行数": len(source_rows),
        }
        errors = parse_result.get("errors", [])[:BIC_ERROR_SAMPLE_LIMIT]
        if errors:
            summary["错误明细"] = errors
        if parse_result.get("fatal_error"):
            summary["文件解析失败"] = "是"
        warnings = parse_result.get("warnings", [])
        if warnings:
            summary["处理提示"] = warnings
        return summary

    @staticmethod
    def parse_file(file_path: str) -> dict:
        result = {
            "total_rows": 0,
            "success_rows": 0,
            "failed_rows": 0,
            "errors": [],
            "warnings": [],
            "fatal_error": False,
            "bic_rows": [],
        }

        def add_file_error(message: str) -> dict:
            result["errors"].append(message)
            result["failed_rows"] = max(int(result["failed_rows"]), 1)
            result["fatal_error"] = True
            return result

        required_headers = tuple(DOUYIN_BIC_HEADERS)
        with open_tabular_rows(file_path) as rows:
            if rows is None:
                return add_file_error("无法打开表格文件")

            row_iter = iter(rows)
            header_result = FinancialSummaryExcelProcessorMixin._find_header_row(row_iter, required_headers)
            if header_result is None:
                return add_file_error("无法读取BIC表头")

            headers, header_row_number = header_result
            col_idx = FinancialSummaryExcelProcessorMixin._build_col_idx(headers, required_headers)
            missing = FinancialSummaryExcelProcessorMixin._missing_headers(col_idx, required_headers)
            if missing:
                return add_file_error(f"缺少BIC必要表头: {', '.join(missing)}")

            for row in row_iter:
                result["total_rows"] += 1
                try:
                    vals = FinancialSummaryExcelProcessorMixin._row_to_values(row, col_idx)
                    if safe_str(vals.get("费用项")) != BIC_TARGET_FEE_ITEM:
                        continue
                    result["bic_rows"].append(
                        {
                            "source_row_number": result["total_rows"] + header_row_number,
                            "settlement_no": safe_str(vals.get("结算单号")),
                            "order_code": safe_str(vals.get("订单码")),
                            "related_order_no": safe_str(vals.get("关联订单号")),
                            "related_waybill_no": safe_str(vals.get("关联运单号")),
                            "fee_item": safe_str(vals.get("费用项")),
                            "service_provider": safe_str(vals.get("服务商")) or "-",
                            "qic_warehouse": safe_str(vals.get("QIC仓")) or "-",
                            "amount": safe_decimal(vals.get("结算金额")),
                            "billing_params": safe_str(vals.get("计费参数")),
                            "billing_completed_time": parse_datetime(vals.get("计费完成时间")),
                            "business_node": safe_str(vals.get("业务节点")),
                            "business_occurred_time": parse_datetime(vals.get("业务发生时间")),
                            "settled_at": parse_datetime(vals.get("结算时间")),
                            "status": safe_str(vals.get("状态")),
                            "transaction_account": safe_str(vals.get("动账账户")),
                            "transaction_flow_no": safe_str(vals.get("动账流水号")),
                            "remark": safe_str(vals.get("备注")),
                            "is_mudaibao": safe_str(vals.get("是否木带宝")),
                            "is_child_order": safe_str(vals.get("是否子单")),
                        }
                    )
                    result["success_rows"] += 1
                except Exception as exc:
                    result["failed_rows"] += 1
                    result["errors"].append(f"Row {result['total_rows'] + header_row_number}: {exc}")
        if result["total_rows"] > 0 and result["success_rows"] == 0 and result["failed_rows"] == 0:
            result["warnings"].append(f"未找到费用项为“{BIC_TARGET_FEE_ITEM}”的记录")
        return result

    @staticmethod
    def _build_detail_rows(
        *,
        task: BicTask,
        upload_file: BicUploadFile,
        rows: Iterable[dict],
        platform_code: str,
        shop_id: int | None,
        shop_name: str,
    ) -> list[BicDetail]:
        groups: dict[tuple[str, str], dict[str, object]] = {}
        for row in rows:
            service_provider = safe_str(row.get("service_provider")) or "-"
            qic_warehouse = safe_str(row.get("qic_warehouse")) or "-"
            group_key = (service_provider, qic_warehouse)
            group = groups.setdefault(group_key, {"amount": Decimal("0"), "row_count": 0})
            group["amount"] = safe_decimal(group["amount"]) + safe_decimal(row.get("amount"))
            group["row_count"] = int(group["row_count"]) + 1

        return [
            BicDetail(
                task_id=task.id,
                file_id=upload_file.id,
                org_id=task.org_id,
                shop_id=shop_id,
                platform_code=platform_code,
                service_provider=service_provider,
                shop_name=shop_name,
                accounting_year=int(upload_file.accounting_year),
                accounting_month=int(upload_file.accounting_month),
                qic_warehouse=qic_warehouse,
                row_count=int(group["row_count"]),
                total_amount=safe_decimal(group["amount"]),
            )
            for (service_provider, qic_warehouse), group in groups.items()
        ]

    @staticmethod
    def _build_source_rows(
        *,
        task: BicTask,
        upload_file: BicUploadFile,
        detail_rows: list[BicDetail],
        rows: Iterable[dict],
        platform_code: str,
        shop_id: int | None,
        shop_name: str,
    ) -> list[BicSourceRow]:
        detail_map = {(row.service_provider, row.qic_warehouse): row.id for row in detail_rows}
        source_rows: list[BicSourceRow] = []
        for row in rows:
            service_provider = safe_str(row.get("service_provider")) or "-"
            qic_warehouse = safe_str(row.get("qic_warehouse")) or "-"
            detail_id = detail_map.get((service_provider, qic_warehouse))
            if detail_id is None:
                continue
            source_rows.append(
                _source_row_payload(
                    task=task,
                    upload_file=upload_file,
                    detail_id=detail_id,
                    shop_id=shop_id,
                    shop_name=shop_name,
                    platform_code=platform_code,
                    source_row=row,
                    source_row_number=int(row.get("source_row_number") or 0),
                )
            )
        return source_rows

    @staticmethod
    async def export_details(
        db: AsyncSession,
        *,
        scope: str = "all",
        page: int | None = None,
        page_size: int | None = None,
        **kwargs,
    ) -> io.BytesIO:
        if scope == "current_page":
            rows, _ = await BicAccountingService.list_details(db, page=page or 1, page_size=page_size or 50, **kwargs)
            _ensure_export_row_limit("BIC汇总", len(rows))
        else:
            _, total = await BicAccountingService.list_details(db, page=1, page_size=1, **kwargs)
            _ensure_export_row_limit("BIC汇总", total)
            rows, _ = await BicAccountingService.list_details(db, page=None, page_size=None, **kwargs)
            _ensure_export_row_limit("BIC汇总", len(rows))
        return BicAccountingService._build_detail_workbook(rows)

    @staticmethod
    async def export_source_rows(
        db: AsyncSession,
        *,
        scope: str = "current_page",
        page: int | None = None,
        page_size: int | None = None,
        **kwargs,
    ) -> io.BytesIO:
        if scope == "current_page":
            rows, _ = await BicAccountingService.list_source_rows(db, page=page or 1, page_size=page_size or 50, **kwargs)
            _ensure_export_row_limit("BIC源明细", len(rows))
        else:
            _, total = await BicAccountingService.list_source_rows(db, page=1, page_size=1, **kwargs)
            _ensure_export_row_limit("BIC源明细", total)
            rows, _ = await BicAccountingService.list_source_rows(db, page=None, page_size=None, **kwargs)
            _ensure_export_row_limit("BIC源明细", len(rows))
        return BicAccountingService._build_source_workbook(rows, title="BIC源明细")

    @staticmethod
    async def export_source_rows_to_file(
        db: AsyncSession,
        *,
        output_path: Path,
        scope: str = "current_page",
        page: int | None = None,
        page_size: int | None = None,
        **kwargs,
    ) -> int:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="BIC源明细")
        worksheet.append(_write_only_header_row(worksheet, BIC_SOURCE_EXPORT_HEADERS))
        row_count = 0

        if scope == "current_page":
            rows, _ = await BicAccountingService.list_source_rows(db, page=page or 1, page_size=page_size or 50, **kwargs)
            _ensure_export_row_limit("BIC源明细", len(rows))
            for row in rows:
                BicAccountingService._append_source_row(worksheet, row)
                row_count += 1
            workbook.save(output_path)
            return row_count

        _, total = await BicAccountingService.list_source_rows(db, page=1, page_size=1, **kwargs)
        _ensure_export_row_limit("BIC源明细", total)
        current_page = 1
        while True:
            rows, _ = await BicAccountingService.list_source_rows(
                db,
                page=current_page,
                page_size=BIC_EXPORT_BATCH_SIZE,
                **kwargs,
            )
            if not rows:
                break
            for row in rows:
                BicAccountingService._append_source_row(worksheet, row)
                row_count += 1
            current_page += 1
        _ensure_export_row_limit("BIC源明细", row_count)
        workbook.save(output_path)
        return row_count

    @staticmethod
    async def list_source_rows(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        detail_id: int | None = None,
        task_id: int | None = None,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        service_provider: str | None = None,
        exact_service_provider: bool = False,
        qic_warehouse: str | None = None,
        accounting_year: int | None = None,
        accounting_month: int | None = None,
        accounting_start_year: int | None = None,
        accounting_start_month: int | None = None,
        accounting_end_year: int | None = None,
        accounting_end_month: int | None = None,
        ids: list[int] | None = None,
        page: int | None = 1,
        page_size: int | None = 50,
    ) -> tuple[list[dict[str, object]], int]:
        filters = [BicSourceRow.is_deleted.is_(False), active_shop_filter(BicSourceRow.shop_id)]
        org_ids = resolve_org_ids(user_role=user.role, user_org_id=user.org_id, requested_org_id=org_id)
        if org_ids is not None:
            filters.append(BicSourceRow.org_id.in_(org_ids))
        if ids is not None:
            if not ids:
                return [], 0
            filters.append(BicSourceRow.id.in_(ids))
        if detail_id is not None:
            filters.append(BicSourceRow.detail_id == detail_id)
        elif task_id is not None:
            filters.append(BicSourceRow.task_id == task_id)
        else:
            filters.append(BicSourceRow.task_id.in_(BicAccountingService._latest_result_task_ids_select()))
        platform_codes = TransactionAccountingService._split_filter_values(platform_code)
        if platform_codes:
            filters.append(BicSourceRow.platform_code.in_(platform_codes))
        service_provider_values = TransactionAccountingService._split_filter_values(
            service_provider.replace("，", ",") if service_provider else service_provider
        )
        if exact_service_provider and service_provider_values:
            filters.append(BicSourceRow.service_provider == service_provider_values[0])
        else:
            service_provider_filter = _service_provider_filter(BicSourceRow.service_provider, service_provider)
            if service_provider_filter is not None:
                filters.append(service_provider_filter)
        shop_names = TransactionAccountingService._split_filter_values(shop_name)
        if shop_names:
            filters.append(BicSourceRow.shop_name.in_(shop_names))
        shop_id_list = TransactionAccountingService._split_int_filter_values(shop_ids)
        if shop_id_list:
            filters.append(BicSourceRow.shop_id.in_(shop_id_list))
        qic_values = TransactionAccountingService._split_filter_values(qic_warehouse)
        if qic_values:
            filters.append(BicSourceRow.qic_warehouse.in_(qic_values))
        if accounting_year is not None:
            filters.append(BicSourceRow.accounting_year == accounting_year)
        if accounting_month is not None:
            filters.append(BicSourceRow.accounting_month == accounting_month)
        filters.extend(
            TransactionAccountingService._period_filters(
                BicSourceRow.accounting_period,
                start_year=accounting_start_year,
                start_month=accounting_start_month,
                end_year=accounting_end_year,
                end_month=accounting_end_month,
            )
        )

        stmt = (
            select(
                BicSourceRow.id.label("id"),
                BicSourceRow.task_id.label("task_id"),
                BicSourceRow.file_id.label("file_id"),
                BicSourceRow.detail_id.label("detail_id"),
                BicSourceRow.org_id.label("org_id"),
                Organization.name.label("org_name"),
                BicSourceRow.shop_id.label("shop_id"),
                BicSourceRow.platform_code.label("platform_code"),
                BicSourceRow.service_provider.label("service_provider"),
                BicSourceRow.shop_name.label("shop_name"),
                BicSourceRow.accounting_year.label("accounting_year"),
                BicSourceRow.accounting_month.label("accounting_month"),
                BicSourceRow.accounting_period.label("accounting_period"),
                BicSourceRow.qic_warehouse.label("qic_warehouse"),
                BicSourceRow.source_row_number.label("source_row_number"),
                BicSourceRow.settlement_no.label("settlement_no"),
                BicSourceRow.order_code.label("order_code"),
                BicSourceRow.related_order_no.label("related_order_no"),
                BicSourceRow.related_waybill_no.label("related_waybill_no"),
                BicSourceRow.fee_item.label("fee_item"),
                BicSourceRow.settlement_amount.label("settlement_amount"),
                BicSourceRow.billing_params.label("billing_params"),
                func.coalesce(
                    func.to_char(BicSourceRow.billing_completed_time, literal("YYYY-MM-DD HH24:MI:SS")),
                    literal(""),
                ).label("billing_completed_time"),
                BicSourceRow.business_node.label("business_node"),
                func.coalesce(
                    func.to_char(BicSourceRow.business_occurred_time, literal("YYYY-MM-DD HH24:MI:SS")),
                    literal(""),
                ).label("business_occurred_time"),
                func.coalesce(
                    func.to_char(BicSourceRow.settled_at, literal("YYYY-MM-DD HH24:MI:SS")),
                    literal(""),
                ).label("settled_at"),
                BicSourceRow.status.label("status"),
                BicSourceRow.transaction_account.label("transaction_account"),
                BicSourceRow.transaction_flow_no.label("transaction_flow_no"),
                BicSourceRow.remark.label("remark"),
                BicSourceRow.is_mudaibao.label("is_mudaibao"),
                BicSourceRow.is_child_order.label("is_child_order"),
                BicSourceRow.created_at.label("created_at"),
                Shop.store_short_id.label("store_short_id"),
                Shop.merchant.label("merchant"),
                Shop.tax_no.label("tax_no"),
                Shop.shop_type.label("shop_type"),
                Shop.registered_address.label("registered_address"),
            )
            .outerjoin(Shop, Shop.id == BicSourceRow.shop_id)
            .outerjoin(Organization, Organization.id == BicSourceRow.org_id)
            .where(*filters)
            .order_by(
                BicSourceRow.accounting_year.desc(),
                BicSourceRow.accounting_month.desc(),
                BicSourceRow.platform_code,
                BicSourceRow.service_provider,
                BicSourceRow.shop_id,
                BicSourceRow.shop_name,
                BicSourceRow.qic_warehouse,
                BicSourceRow.source_row_number,
            )
        )
        count_stmt = select(func.count()).select_from(BicSourceRow).where(*filters)
        total = (await db.execute(count_stmt)).scalar() or 0
        if page is not None and page_size is not None:
            stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(stmt)
        return [dict(row) for row in result.mappings().all()], total

    @staticmethod
    async def export_reconciliation(
        db: AsyncSession,
        *,
        user: User,
        org_id: int | None = None,
        accounting_year: int,
        accounting_month: int,
        service_provider: str,
        platform_code: str | None = None,
        shop_name: str | None = None,
        shop_ids: str | None = None,
        qic_warehouse: str | None = None,
    ) -> io.BytesIO:
        service_provider_values = TransactionAccountingService._split_filter_values(service_provider.replace("，", ","))
        if len(service_provider_values) != 1:
            raise ValueError("请选择服务商")
        normalized_provider = service_provider_values[0]

        _, total = await BicAccountingService.list_source_rows(
            db,
            user=user,
            org_id=org_id,
            platform_code=platform_code,
            shop_name=shop_name,
            shop_ids=shop_ids,
            service_provider=normalized_provider,
            qic_warehouse=qic_warehouse,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            page=1,
            page_size=1,
        )
        if total == 0:
            raise ValueError("未找到符合条件的BIC对账数据，请检查月份、服务商和筛选条件")
        _ensure_export_row_limit("BIC对账表", total)

        rows, _ = await BicAccountingService.list_source_rows(
            db,
            user=user,
            org_id=org_id,
            platform_code=platform_code,
            shop_name=shop_name,
            shop_ids=shop_ids,
            service_provider=normalized_provider,
            qic_warehouse=qic_warehouse,
            accounting_year=accounting_year,
            accounting_month=accounting_month,
            page=None,
            page_size=None,
        )
        _ensure_export_row_limit("BIC对账表", len(rows))
        return BicAccountingService._build_reconciliation_workbook(rows, accounting_year, accounting_month, normalized_provider)

    @staticmethod
    def _source_export_values(row: dict[str, object], *, include_shop_name: bool = False) -> list[object]:
        values: list[object] = []
        if include_shop_name:
            values.append(row.get("shop_name") or "")
        values.extend(
            [
                row.get("settlement_no") or "",
                row.get("order_code") or "",
                row.get("related_order_no") or "",
                row.get("related_waybill_no") or "",
                row.get("fee_item") or "",
                row.get("service_provider") or "",
                row.get("qic_warehouse") or "",
                float(row.get("settlement_amount") or 0),
                row.get("billing_params") or "",
                _format_source_datetime(row.get("billing_completed_time")),
                row.get("business_node") or "",
                _format_source_datetime(row.get("business_occurred_time")),
                _format_source_datetime(row.get("settled_at")),
                row.get("status") or "",
                row.get("transaction_account") or "",
                row.get("transaction_flow_no") or "",
                row.get("remark") or "",
                row.get("is_mudaibao") or "",
                row.get("is_child_order") or "",
            ]
        )
        return values

    @staticmethod
    def _append_source_row(worksheet, row: dict[str, object], *, include_shop_name: bool = False) -> None:
        worksheet.append(BicAccountingService._source_export_values(row, include_shop_name=include_shop_name))

    @staticmethod
    def _build_source_workbook(rows: list[dict[str, object]], *, title: str) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title=title[:31] or "Sheet")
        worksheet.append(_write_only_header_row(worksheet, BIC_SOURCE_EXPORT_HEADERS))
        for row in rows:
            BicAccountingService._append_source_row(worksheet, row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_detail_workbook(rows: list[dict[str, object]]) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="BIC汇总")
        worksheet.append(_write_only_header_row(worksheet, BIC_DETAIL_EXPORT_HEADERS))
        for index, row in enumerate(rows, start=1):
            worksheet.append(
                [
                    index,
                    TransactionAccountingService._format_platform(str(row.get("platform_code") or "")),
                    row.get("service_provider") or "",
                    row.get("store_short_id") or "",
                    row.get("shop_name") or "",
                    row.get("qic_warehouse") or "",
                    row.get("merchant") or "",
                    row.get("tax_no") or "",
                    row.get("shop_type") or "",
                    row.get("registered_address") or "",
                    float(row.get("total_amount") or 0),
                ]
            )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_reconciliation_workbook(
        rows: list[dict[str, object]],
        accounting_year: int,
        accounting_month: int,
        service_provider: str,
    ) -> io.BytesIO:
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet(title="汇总")
        worksheet.append(_write_only_header_row(worksheet, ["店铺id", "店铺名称", "QIC仓", "公司名称", "税号", "抬头类型", "注册地址", "结算金额"]))
        source_worksheet = workbook.create_sheet(title="明细")
        source_worksheet.append(_write_only_header_row(source_worksheet, BIC_RECONCILIATION_SOURCE_EXPORT_HEADERS))

        summary_map: dict[tuple[str, str], dict[str, object]] = {}
        for row in rows:
            shop_key = safe_str(row.get("shop_name")) or "-"
            qic_key = safe_str(row.get("qic_warehouse")) or "-"
            shop_sheet_key = safe_str(row.get("shop_id")) or shop_key
            group_key = (shop_sheet_key, qic_key)
            group = summary_map.setdefault(
                group_key,
                {
                    "shop_id": row.get("shop_id"),
                    "store_short_id": row.get("store_short_id"),
                    "shop_name": shop_key,
                    "qic_warehouse": qic_key,
                    "merchant": row.get("merchant") or "",
                    "tax_no": row.get("tax_no") or "",
                    "shop_type": row.get("shop_type") or "",
                    "registered_address": row.get("registered_address") or "",
                    "amount": Decimal("0"),
                },
            )
            group["amount"] = safe_decimal(group["amount"]) + safe_decimal(row.get("settlement_amount"))

        for group_key in sorted(summary_map, key=lambda key: (safe_str(summary_map[key]["shop_name"]), safe_str(summary_map[key]["qic_warehouse"]))):
            group = summary_map[group_key]
            worksheet.append(
                [
                    group["store_short_id"] or group["shop_id"] or "",
                    group["shop_name"] or "",
                    group["qic_warehouse"] or "",
                    group["merchant"] or "",
                    group["tax_no"] or "",
                    group["shop_type"] or "",
                    group["registered_address"] or "",
                    float(group["amount"] or 0),
                ]
            )

        for row in rows:
            BicAccountingService._append_source_row(
                source_worksheet,
                {**row, "service_provider": row.get("service_provider") or service_provider},
                include_shop_name=True,
            )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
