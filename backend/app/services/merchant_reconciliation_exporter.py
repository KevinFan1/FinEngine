"""商家对账导出工具。"""

from __future__ import annotations

import io
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from app.utils.money import safe_decimal


class MerchantReconciliationExporter:
    """统一处理商家对账明细和汇总的工作簿输出。"""

    @staticmethod
    def build_workbook(
        rows: Iterable[dict[str, object]],
        *,
        title: str,
        columns: tuple[tuple[str, str, bool], ...],
    ) -> io.BytesIO:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=title)
        ws.append(MerchantReconciliationExporter.header_row(ws, [label for _field, label, _money_flag in columns]))
        for row in rows:
            MerchantReconciliationExporter.append_row(ws, row, columns=columns)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def save_workbook(
        rows: Iterable[dict[str, object]],
        *,
        output_path: Path,
        title: str,
        columns: tuple[tuple[str, str, bool], ...],
    ) -> int:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=title)
        ws.append(MerchantReconciliationExporter.header_row(ws, [label for _field, label, _money_flag in columns]))
        row_count = 0
        for row in rows:
            MerchantReconciliationExporter.append_row(ws, row, columns=columns)
            row_count += 1
        wb.save(output_path)
        return row_count

    @staticmethod
    def append_sheet(
        wb: Workbook,
        rows: Iterable[dict[str, object]],
        *,
        title: str,
        columns: tuple[tuple[str, str, bool], ...],
    ) -> int:
        ws = wb.create_sheet(title=title)
        ws.append(MerchantReconciliationExporter.header_row(ws, [label for _field, label, _money_flag in columns]))
        row_count = 0
        for row in rows:
            MerchantReconciliationExporter.append_row(ws, row, columns=columns)
            row_count += 1
        return row_count

    @staticmethod
    def header_row(sheet, headers: list[str]) -> list[WriteOnlyCell]:
        cells: list[WriteOnlyCell] = []
        for label in headers:
            cell = WriteOnlyCell(sheet, value=label)
            cell.font = Font(bold=True)
            cells.append(cell)
        return cells

    @staticmethod
    def append_row(
        ws,
        row: dict[str, object],
        *,
        columns: tuple[tuple[str, str, bool], ...],
    ) -> None:
        ws.append(
            [
                MerchantReconciliationExporter.format_export_value(row.get(field), money=money_flag)
                for field, _label, money_flag in columns
            ]
        )

    @staticmethod
    def format_export_value(value: object, *, money: bool = False) -> object:
        if money:
            return float(safe_decimal(value))
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.isoformat()
        return value
