from __future__ import annotations

import importlib.util
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path("/Users/kevinfan/Documents/对账单需求/新逻辑.xlsx")
CFG = ROOT / "backend/app/config/transaction_accounting_rules.py"


def load_config():
    spec = importlib.util.spec_from_file_location("transaction_accounting_rules_for_compare", CFG)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load config: {CFG}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def split_patterns(pattern_text: str) -> tuple[str, ...]:
    text = str(pattern_text or "")
    if not text:
        return ()
    for separator in ("\r\n", "\n", "\r", ",", "，", "、", ";", "；"):
        text = text.replace(separator, "\n")
    return tuple(part.strip() for part in text.split("\n") if part.strip())


def parse_extra(extra: object) -> tuple[str, str]:
    text = str(extra or "").strip()
    if not text:
        return "none", ""
    if "找不到例子" in text:
        return "ignored", text
    if text.startswith("备注 ="):
        return "exact", text.split("=", 1)[1].strip()
    if text.startswith("备注 - 包含 -"):
        return "contains", text.split("备注 - 包含 -", 1)[1].strip()
    if text.startswith("备注 - 不包含 -"):
        return "not_contains", text.split("备注 - 不包含 -", 1)[1].strip()
    return "unknown", text


def normalize_scene(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "(空白动账场景)":
        return ""
    if text.startswith("动账场景="):
        return text.split("=", 1)[1].strip()
    return text


def normalize_sign(value: object) -> str:
    text = str(value or "").strip()
    return {"正": "positive", "负": "negative"}.get(text, text)


def load_new_rules() -> list[dict[str, object]]:
    workbook = load_workbook(XLSX, data_only=True)
    sheet = workbook.active
    subject = None
    category = None
    rows: list[dict[str, object]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        if values[0]:
            subject = str(values[0]).strip()
        if values[1]:
            category = str(values[1]).strip()
        match_type, remark_pattern = parse_extra(values[4])
        rows.append(
            {
                "row": row_number,
                "subject": subject,
                "category": category,
                "direction": str(values[2] or "").strip(),
                "scene": normalize_scene(values[3]),
                "match_type": match_type,
                "remark_pattern": remark_pattern,
                "amount_field": str(values[5] or "").strip(),
                "result_direction": normalize_sign(values[6]),
                "old_remark": str(values[7] or "").strip(),
                "note": str(values[8] or "").strip(),
                "extra": str(values[4] or "").strip(),
            }
        )
    return rows


def load_config_rules() -> list[dict[str, object]]:
    module = load_config()
    return [
        {
            "subject": rule.subject,
            "category": rule.category,
            "direction": rule.transaction_direction,
            "scene": rule.transaction_scene,
            "match_type": rule.match_type,
            "remark_pattern": rule.remark_pattern,
            "remark_exclude_pattern": rule.remark_exclude_pattern,
            "amount_field": rule.amount_field,
            "result_direction": rule.result_direction,
            "source_rows": rule.source_rows,
        }
        for rule in module.TRANSACTION_ACCOUNTING_RULE_SPECS
    ]


def key(rule: dict[str, object]) -> tuple[object, ...]:
    return (
        rule["subject"],
        rule["category"],
        rule["direction"],
        rule["scene"],
        rule["match_type"],
        split_patterns(str(rule["remark_pattern"])),
        rule["amount_field"],
        rule["result_direction"],
    )


def base_key(rule: dict[str, object]) -> tuple[object, ...]:
    return (
        rule["subject"],
        rule["category"],
        rule["direction"],
        rule["scene"],
        rule["amount_field"],
    )


def describe(rule: dict[str, object]) -> str:
    scene = rule["scene"]
    if scene is None:
        scene_label = "不限"
    elif scene == "":
        scene_label = "空场景"
    else:
        scene_label = str(scene)
    return (
        f"{rule['subject']} / {rule['category']} / {rule['direction']} / {scene_label} / "
        f"{rule['match_type']}:{rule['remark_pattern']} / {rule['amount_field']} / {rule['result_direction']}"
    )


def main() -> None:
    new_rules = load_new_rules()
    config_rules = load_config_rules()
    new_active = [rule for rule in new_rules if rule["match_type"] != "ignored"]
    new_complete = [
        rule
        for rule in new_active
        if rule["direction"]
        and rule["amount_field"]
        and rule["result_direction"] in {"positive", "negative"}
        and rule["match_type"] != "unknown"
    ]
    print("counts", {"xlsx_rows": len(new_rules), "active_rows": len(new_active), "complete_rows": len(new_complete), "config_rules": len(config_rules)})

    duplicates: defaultdict[tuple[object, ...], list[int]] = defaultdict(list)
    for rule in new_complete:
        duplicates[key(rule)].append(int(rule["row"]))
    print("\nduplicate exact keys in xlsx:")
    for signature, rows in duplicates.items():
        if len(rows) > 1:
            print(rows, signature)

    config_keys = {key(rule): rule for rule in config_rules}
    new_keys = {key(rule): rule for rule in new_complete}

    print("\nnew complete rows not exact in config:")
    for rule in new_complete:
        if key(rule) in config_keys:
            continue
        candidates = [candidate for candidate in config_rules if base_key(candidate) == base_key(rule)]
        print(
            rule["row"],
            describe(rule),
            "candidates=",
            [(candidate["match_type"], candidate["remark_pattern"], candidate["result_direction"], candidate["source_rows"]) for candidate in candidates],
        )

    print("\nconfig rows not exact in new complete:")
    for rule in config_rules:
        if key(rule) in new_keys:
            continue
        candidates = [candidate for candidate in new_complete if base_key(candidate) == base_key(rule)]
        print(
            describe(rule),
            "source_rows=",
            rule["source_rows"],
            "new_candidates=",
            [(candidate["row"], candidate["match_type"], candidate["remark_pattern"], candidate["result_direction"]) for candidate in candidates],
        )

    print("\nincomplete/ignored/unknown new rows:")
    for rule in new_rules:
        if rule not in new_complete:
            print(rule["row"], describe(rule), "extra=", rule["extra"], "old=", rule["old_remark"], "note=", rule["note"])

    print("\nnotes:")
    for rule in new_rules:
        if rule["note"]:
            print(rule["row"], describe(rule), "note=", rule["note"])


if __name__ == "__main__":
    main()
