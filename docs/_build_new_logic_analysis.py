from __future__ import annotations

import importlib.util
import sys
from collections import defaultdict, OrderedDict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
XLSX = Path("/Users/kevinfan/Documents/对账单需求/新逻辑.xlsx")
CFG = ROOT / "backend/app/config/transaction_accounting_rules.py"
OUT_MD = ROOT / "docs/transaction_accounting_new_logic_analysis.md"
OUT_OVERVIEW = ROOT / "docs/assets/transaction_accounting_new_logic_overview.png"
OUT_DETAIL = ROOT / "docs/assets/transaction_accounting_new_logic_payment_other_detail.png"


def load_config():
    spec = importlib.util.spec_from_file_location("transaction_accounting_rules_for_analysis", CFG)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load config: {CFG}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def split_patterns(pattern_text: str) -> tuple[str, ...]:
    text = str(pattern_text or "")
    if not text:
        return ()
    for separator in ("\r\n", "\n", "\r", ",", "，", "、", ";", "；"):
        text = text.replace(separator, "\n")
    return tuple(part.strip() for part in text.split("\n") if part.strip())


def parse_extra(extra: object) -> tuple[str, str]:
    text = str(extra or "").strip()
    if not text:
        return "none", ""
    if "找不到例子" in text:
        return "ignored", text
    if text.startswith("备注 ="):
        return "exact", text.split("=", 1)[1].strip()
    if text.startswith("备注 - 包含 -"):
        return "contains", text.split("备注 - 包含 -", 1)[1].strip()
    if text.startswith("备注 - 不包含 -"):
        return "not_contains", text.split("备注 - 不包含 -", 1)[1].strip()
    return "unknown", text


def normalize_scene(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "(空白动账场景)":
        return ""
    if text.startswith("动账场景="):
        return text.split("=", 1)[1].strip()
    return text


def normalize_sign(value: object) -> str:
    text = str(value or "").strip()
    return {"正": "positive", "负": "negative"}.get(text, text)


def scene_label(scene: str | None) -> str:
    if scene is None:
        return "不限"
    if scene == "":
        return "空场景"
    return scene


def status_for_category(name: str) -> str:
    if name in {"评价有礼"}:
        return "red"
    if name in {"月付贴息费用"}:
        return "red"
    if name in {"赔付", "小额打款", "捐赠", "可提现充值千川", "BIC"}:
        return "amber"
    return "green"


def status_for_subject(name: str) -> str:
    if name in {"支付其他与经营相关的支出", "支付抖音平台服务费", "收到抖音分账款", "支付线上BIC费用"}:
        return "amber"
    return "green"


def status_tag(status: str) -> str:
    return {"green": "=", "amber": "Δ", "red": "?"}.get(status, "?")


def status_colors(status: str) -> tuple[str, str, str]:
    if status == "green":
        return "#d1fae5", "#065f46", "#10b981"
    if status == "red":
        return "#fee2e2", "#991b1b", "#ef4444"
    return "#fef3c7", "#92400e", "#f59e0b"


def font(size: int, bold: bool = False):
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/Supplemental/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            try:
                return ImageFont.truetype(path, size=size, index=1 if bold else 0)
            except Exception:
                continue
    return ImageFont.load_default()


def draw_round(draw: ImageDraw.ImageDraw, box, fill, outline=None, width=1, radius=18):
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def draw_text(draw: ImageDraw.ImageDraw, pos, text, fnt, fill, max_width=None, line_gap=6):
    x, y = pos
    if max_width is None:
        draw.text((x, y), text, font=fnt, fill=fill)
        bbox = draw.textbbox((x, y), text, font=fnt)
        return bbox[3]
    lines = []
    current = ""
    for char in text:
        if char == "\n":
            if current:
                lines.append(current)
            current = ""
            continue
        candidate = current + char
        if not current or draw.textlength(candidate, font=fnt) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = char
    if current:
        lines.append(current)
    for line in lines:
        draw.text((x, y), line, font=fnt, fill=fill)
        bbox = draw.textbbox((x, y), line, font=fnt)
        y = bbox[3] + line_gap
    return y


def draw_chip(draw: ImageDraw.ImageDraw, x: int, y: int, text: str, status: str, font_size: int = 18):
    bg, fg, border = status_colors(status)
    fnt = font(font_size, True)
    bbox = draw.textbbox((0, 0), text, font=fnt)
    w = bbox[2] - bbox[0] + 26
    h = bbox[3] - bbox[1] + 14
    draw.rounded_rectangle((x, y, x + w, y + h), radius=h // 2, fill=bg, outline=border, width=2)
    draw.text((x + 13, y + 6), text, font=fnt, fill=fg)
    return w, h


def load_new_rules():
    workbook = load_workbook(XLSX, data_only=True)
    sheet = workbook.active
    subject = None
    category = None
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        if values[0]:
            subject = str(values[0]).strip()
        if values[1]:
            category = str(values[1]).strip()
        match_type, remark_pattern = parse_extra(values[4])
        rows.append(
            {
                "row": row_number,
                "subject": subject,
                "category": category,
                "direction": str(values[2] or "").strip(),
                "scene": normalize_scene(values[3]),
                "match_type": match_type,
                "remark_pattern": remark_pattern,
                "amount_field": str(values[5] or "").strip(),
                "result_direction": normalize_sign(values[6]),
                "old_remark": str(values[7] or "").strip(),
                "note": str(values[8] or "").strip(),
                "extra": str(values[4] or "").strip(),
            }
        )
    return rows


def load_config_rules():
    module = load_config()
    rules = []
    for rule in module.TRANSACTION_ACCOUNTING_RULE_SPECS:
        rules.append(
            {
                "subject": rule.subject,
                "category": rule.category,
                "direction": rule.transaction_direction,
                "scene": rule.transaction_scene,
                "match_type": rule.match_type,
                "remark_pattern": rule.remark_pattern,
                "amount_field": rule.amount_field,
                "result_direction": rule.result_direction,
            }
        )
    return rules


def key(rule: dict[str, object]) -> tuple[object, ...]:
    return (
        rule["subject"],
        rule["category"],
        rule["direction"],
        rule["scene"],
        rule["match_type"],
        split_patterns(str(rule["remark_pattern"])),
        rule["amount_field"],
        rule["result_direction"],
    )


def base_key(rule: dict[str, object]) -> tuple[object, ...]:
    return (rule["subject"], rule["category"], rule["direction"], rule["scene"], rule["amount_field"])


def compute_analysis(new_rules, config_rules):
    new_active = [rule for rule in new_rules if rule["match_type"] != "ignored"]
    new_complete = [
        rule
        for rule in new_active
        if rule["direction"] and rule["amount_field"] and rule["result_direction"] in {"positive", "negative"}
    ]
    config_keys = {key(rule) for rule in config_rules}
    new_keys = {key(rule) for rule in new_complete}

    exact_matches = sum(1 for rule in new_complete if key(rule) in config_keys)
    duplicates = defaultdict(list)
    for rule in new_complete:
        duplicates[key(rule)].append(rule["row"])

    note_rows = [rule for rule in new_rules if rule["note"]]
    difference_rows = [rule for rule in new_complete if key(rule) not in config_keys]
    ignored_rows = [rule for rule in new_rules if rule["match_type"] == "ignored"]
    incomplete_rows = [rule for rule in new_rules if rule not in new_complete and rule["match_type"] != "ignored"]
    return {
        "new_active": new_active,
        "new_complete": new_complete,
        "exact_matches": exact_matches,
        "duplicates": {sig: rows for sig, rows in duplicates.items() if len(rows) > 1},
        "difference_rows": difference_rows,
        "ignored_rows": ignored_rows,
        "incomplete_rows": incomplete_rows,
        "note_rows": note_rows,
        "config_keys": config_keys,
        "new_keys": new_keys,
    }


def category_status(new_rules, config_rules, subject, category):
    rules = [r for r in new_rules if r["subject"] == subject and r["category"] == category]
    cfg = [r for r in config_rules if r["subject"] == subject and r["category"] == category]
    if any(r["match_type"] == "ignored" for r in rules):
        return "amber"
    if any(r["note"] for r in rules):
        return "red"
    new_keys = {key(r) for r in rules if r["result_direction"] in {"positive", "negative"} and r["amount_field"]}
    cfg_keys = {key(r) for r in cfg}
    if new_keys == cfg_keys:
        return "green"
    if cfg_keys and any(base_key(r) == base_key(cfg_rule) for r in rules for cfg_rule in cfg):
        return "amber"
    return "amber"


def subject_status(new_rules, config_rules, subject):
    categories = sorted({r["category"] for r in new_rules if r["subject"] == subject})
    if any(category_status(new_rules, config_rules, subject, c) == "red" for c in categories):
        return "red"
    if any(category_status(new_rules, config_rules, subject, c) == "amber" for c in categories):
        return "amber"
    return "green"


def create_overview_image(new_rules, config_rules, analysis):
    width, height = 2200, 1440
    img = Image.new("RGB", (width, height), "#f7f9fc")
    draw = ImageDraw.Draw(img)
    title_f = font(44, True)
    subtitle_f = font(20)
    subject_f = font(24, True)
    meta_f = font(18)
    small_f = font(16)

    draw_round(draw, (40, 30, width - 40, height - 40), "#ffffff", "#d6e0ef", 3, 28)
    draw.text((70, 64), "动账重分类新逻辑 - 总览脑图", fill="#0f172a", font=title_f)
    draw.text(
        (70, 130),
        f"新逻辑表 58 行 / 56 条完整规则；config rules 53 条；重复 3 行、待确认 4 行、忽略 1 行。",
        fill="#475569",
        font=subtitle_f,
    )

    root_x = width // 2 - 230
    draw_round(draw, (root_x, 195, root_x + 460, 285), "#dbeafe", "#2563eb", 3, 28)
    draw.text((root_x + 120, 224), "动账重分类", fill="#1d4ed8", font=subject_f)

    subjects = [
        "收到抖音分账款",
        "收到其他与经营相关的收入",
        "支付达人分账佣金",
        "支付抖音平台服务费",
        "支付抖音平台运费险",
        "支付抖音提现",
        "支付其他与经营相关的支出",
        "支付线上BIC费用",
    ]
    positions = {
        0: (70, 360),
        1: (610, 360),
        2: (1150, 360),
        3: (1690, 360),
        4: (70, 610),
        5: (610, 610),
        6: (1150, 610),
        7: (1690, 610),
    }
    colors = {
        "green": ("#dcfce7", "#166534", "#22c55e"),
        "amber": ("#fef3c7", "#92400e", "#f59e0b"),
        "red": ("#fee2e2", "#991b1b", "#ef4444"),
    }
    for idx, subject in enumerate(subjects):
        x, y = positions[idx]
        status = subject_status(new_rules, config_rules, subject)
        bg, fg, border = colors[status]
        draw_round(draw, (x, y, x + 470, y + 180), bg, border, 3, 24)
        draw.text((x + 22, y + 18), subject, fill="#111827", font=subject_f)
        rules = [r for r in new_rules if r["subject"] == subject and r["match_type"] != "ignored"]
        categories = sorted({r["category"] for r in rules})
        draw.text((x + 22, y + 62), f"{len(rules)} 条规则 / {len(categories)} 个小分类", fill="#334155", font=meta_f)
        draw_chip(draw, x + 22, y + 102, status_tag(status), status)
        cat_sample = "、".join(categories[:3])
        if len(categories) > 3:
            cat_sample += " ..."
        draw.text((x + 70, y + 108), cat_sample, fill="#475569", font=small_f)
        draw.line((root_x + 230, 285, x + 235, y), fill="#94a3b8", width=3)

    draw_round(draw, (70, 840, width - 70, 1360), "#eef6ff", "#9ec7ff", 2, 24)
    draw.text((100, 868), "图例", fill="#1d4ed8", font=subject_f)
    legend_items = [
        ("=", "与 config 基本一致", "green"),
        ("Δ", "有结构变化 / 新增分支 / 可能多重命中", "amber"),
        ("?", "待业务确认 / 缺字段 / 明确写了错误逻辑", "red"),
    ]
    lx = 100
    for symbol, label, status in legend_items:
        draw_chip(draw, lx, 920, symbol, status)
        draw.text((lx + 60, 924), label, fill="#0f172a", font=subtitle_f)
        lx += 620
    draw.text(
        (100, 1000),
        "重点看三块：收到抖音分账款的方向变化，支付其他与经营相关的支出里的新增/分歧分支，"
        "以及支付线上BIC费用里 broad + specific 同时存在的重复匹配风险。",
        fill="#334155",
        font=subtitle_f,
    )
    draw.text(
        (100, 1050),
        "如果要跟业务讲清楚，建议先看这张总览，再看下面的重点分支图。",
        fill="#475569",
        font=subtitle_f,
    )

    img.save(OUT_OVERVIEW)


def category_summary(new_rules, category):
    rules = [r for r in new_rules if r["category"] == category]
    lines = []
    for rule in rules:
        if rule["match_type"] == "ignored":
            continue
        if rule["match_type"] == "none":
            remark = "不限备注"
        elif rule["match_type"] == "exact":
            remark = f"备注={rule['remark_pattern']}"
        elif rule["match_type"] == "contains":
            remark = f"备注包含 {rule['remark_pattern']}"
        elif rule["match_type"] == "not_contains":
            remark = f"备注不含 {rule['remark_pattern']}"
        else:
            remark = rule["remark_pattern"]
        scene = scene_label(rule["scene"])
        lines.append(f"{rule['direction']} / {scene} / {remark} / {rule['amount_field']} / {rule['result_direction']}")
    return lines


def create_detail_image(new_rules, config_rules):
    subject = "支付其他与经营相关的支出"
    categories = [
        "拦截费",
        "赔付",
        "退款转赔付",
        "月付贴息费用",
        "小额打款",
        "捐赠",
        "可提现充值千川",
        "返现",
        "评价有礼",
        "可提现充值保证金",
    ]
    width, height = 2200, 2280
    img = Image.new("RGB", (width, height), "#f7f9fc")
    draw = ImageDraw.Draw(img)
    title_f = font(42, True)
    subject_f = font(26, True)
    body_f = font(18)
    small_f = font(16)
    chip_f = font(18, True)

    draw_round(draw, (40, 30, width - 40, height - 40), "#ffffff", "#d6e0ef", 3, 28)
    draw.text((70, 64), "重点分支脑图：支付其他与经营相关的支出", fill="#0f172a", font=title_f)
    draw.text((70, 130), "这个分支最复杂，建议用“分类 -> 条件 -> 金额方向”的方式和业务拆开讲。", fill="#475569", font=body_f)

    draw_round(draw, (70, 185, width - 70, 260), "#e0f2fe", "#0284c7", 3, 20)
    draw.text((100, 210), subject, fill="#0369a1", font=subject_f)
    draw.text((420, 212), "10 个小分类 / 25 条规则", fill="#075985", font=body_f)

    col_w = 1000
    card_w = 960
    card_h = 170
    start_y = 310
    row_gap = 184
    col_xs = [100, 1140]
    category_status_map = {
        "拦截费": "green",
        "赔付": "amber",
        "退款转赔付": "green",
        "月付贴息费用": "red",
        "小额打款": "amber",
        "捐赠": "amber",
        "可提现充值千川": "amber",
        "返现": "green",
        "评价有礼": "red",
        "可提现充值保证金": "green",
    }
    summaries = {
        "拦截费": "空场景 + 备注包含“拦截费”",
        "赔付": "消费者赔付有正/反向两支；补了“打款,订单号”反向分支",
        "退款转赔付": "退转付、极速退二次售后、小额打款三类",
        "月付贴息费用": "两条规则旁边都标了“错误逻辑”",
        "小额打款": "新增“平台赔付 / 仲裁申诉通过打款”；原表里 3 行完全重复",
        "捐赠": "空场景 + 备注包含“公益商家佣金捐赠”",
        "可提现充值千川": "空场景 + 备注精确匹配",
        "返现": "按三种活动分别做入/出账反向",
        "评价有礼": "入账方向和原逻辑相反，需重点确认",
        "可提现充值保证金": "单条规则，出账进入保证金",
    }

    for idx, category in enumerate(categories):
        row = idx // 2
        col = idx % 2
        x = col_xs[col]
        y = start_y + row * row_gap
        status = category_status_map[category]
        bg, fg, border = status_colors(status)
        draw_round(draw, (x, y, x + card_w, y + card_h), bg, border, 3, 22)
        draw.text((x + 22, y + 18), category, fill="#111827", font=subject_f)
        rules = [r for r in new_rules if r["subject"] == subject and r["category"] == category and r["match_type"] != "ignored"]
        cfg = [r for r in config_rules if r["subject"] == subject and r["category"] == category]
        draw_chip(draw, x + 20, y + 62, status_tag(status), status, font_size=17)
        draw.text((x + 70, y + 67), f"{len(rules)} 条规则 / config {len(cfg)} 条", fill="#334155", font=body_f)
        draw.text((x + 22, y + 102), summaries[category], fill="#475569", font=small_f)
        sample_lines = category_summary(new_rules, category)[:2]
        yy = y + 126
        for line in sample_lines:
            yy = draw_text(draw, (x + 22, yy), f"• {line}", small_f, "#0f172a", max_width=card_w - 44, line_gap=2)
            if yy > y + card_h - 18:
                break

    draw_round(draw, (70, 2040, width - 70, 2210), "#fff7ed", "#fb923c", 2, 20)
    draw.text((100, 2070), "本分支最需要业务确认的点：月付贴息费用、评价有礼、以及小额打款 / BIC 是否允许 broad + specific 同时命中。", fill="#9a3412", font=body_f)
    draw.text((100, 2110), "如果这些规则要直接进引擎，建议先把重复行、缺 sign、以及明确写成“错误逻辑”的行收口后再上线。", fill="#9a3412", font=body_f)

    img.save(OUT_DETAIL)


def build_markdown(new_rules, config_rules, analysis):
    subjects = OrderedDict()
    for rule in new_rules:
        if rule["match_type"] == "ignored":
            continue
        subjects.setdefault(rule["subject"], set()).add(rule["category"])
    subject_lines = []
    for subject, cats in subjects.items():
        subject_lines.append(f"- {subject}: {len(cats)} 个小分类，{len([r for r in new_rules if r['subject'] == subject and r['match_type'] != 'ignored'])} 条规则")

    diff_rows = [
        "## 主要差异",
        "",
        f"- 新逻辑表共有 {len(new_rules)} 行，去掉 1 行“找不到例子”后剩 {len(analysis['new_active'])} 行可用；其中 {len(analysis['new_complete'])} 行是完整规则。",
        f"- config rules 只有 {len(config_rules)} 条；新逻辑里有 3 行完全重复（40-42），如果直接导入会造成同一条记录被重复命中。",
        "- `收到抖音分账款 / 订单结算`：新逻辑把“订单退款”改成了 `出账`，而 config 里还是 `入账`，这是最明显的方向变化。",
        "- `支付抖音平台服务费`：new logic 把 row21 的 `服务费返还` 备注条件拿掉了；row22 备注条件还在，但 `取值`/方向字段缺失，需要补齐。",
        "- `支付其他与经营相关的支出`：新增了 `平台赔付 / 仲裁申诉通过打款`、`消费者赔付 / 打款,订单号` 反向分支，以及 `评价有礼` 方向反转。",
        "- `捐赠`、`可提现充值千川`：new logic 显式要求空场景，config 里是 `不限场景`，新逻辑更严格。",
        "- `支付线上BIC费用`：new logic 里既保留了 broad 的 `供应链QIC费用`，又加了 contains 分支，存在多重命中风险。",
        "",
        "## 需要业务确认",
        "",
        "- row 22：`支付抖音平台服务费 / 退款-订单退款触发-分账 / 备注=服务费返还`，`取值` 为空。",
        "- row 36-37：`月付贴息费用` 被标了“错误逻辑”，需要确认是否保留。",
        "- row 52：`评价有礼` 当前是入账负数，和原逻辑相反。",
        "- row 57：`商家集运物流费：找不到例子` 建议按你前面的要求直接忽略。",
        "",
        "## 分支概览",
        "",
        *subject_lines,
        "",
        "## 图片",
        "",
        f"![总览脑图]({OUT_OVERVIEW})",
        "",
        f"![重点分支]({OUT_DETAIL})",
        "",
    ]
    OUT_MD.write_text("\n".join(diff_rows), encoding="utf-8")


def main():
    new_rules = load_new_rules()
    config_rules = load_config_rules()
    analysis = compute_analysis(new_rules, config_rules)
    create_overview_image(new_rules, config_rules, analysis)
    create_detail_image(new_rules, config_rules)
    build_markdown(new_rules, config_rules, analysis)
    print(OUT_MD)
    print(OUT_OVERVIEW)
    print(OUT_DETAIL)


if __name__ == "__main__":
    main()
